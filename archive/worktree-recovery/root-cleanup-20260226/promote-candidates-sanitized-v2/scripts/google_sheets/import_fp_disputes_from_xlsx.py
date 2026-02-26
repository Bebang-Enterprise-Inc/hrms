#!/usr/bin/env python3
"""Import FoodPanda disputes from workbook xlsx into FP_DISPUTES tab.

Expected input format:
- Workbook with sheet name: DISPUTE
- Header columns (A:H): [blank], DISPUTE, STORE NAME, ORDER, CANCELLED AT, REASON, STATUS, Subtotal

Output target:
- Google Sheet FP_DISPUTES columns A:K
  logged_at, order_id, store_code, original_business_date, detected_date, delta_gross,
  reason, evidence_link, status, approved_by, notes

Behavior:
- Accepts raw files as-is (no manual reformatting needed).
- Imports only valid disputed rows (cancelled + mapped store + valid order date + positive subtotal).
- Converts subtotal into negative adjustment (deduction).
- Deduplicates by order_id against existing FP_DISPUTES rows.
- Writes status as Approved so FP_DISPUTE_DAILY and SALES_FACT_DAILY_FINAL auto-deduct.
"""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from google.oauth2 import service_account
from googleapiclient.discovery import build
from openpyxl import load_workbook


DEFAULT_SPREADSHEET_ID = "1F9Zqn_5r42iLSWkHZqGaFr-a6-zXj5eOg52DJ3Oac78"
SERVICE_ACCOUNT_FILE = "credentials/task-manager-service.json"
IMPERSONATE_USER = "sam@bebang.ph"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

EXPECTED_HEADER = ["", "DISPUTE", "STORE NAME", "ORDER", "CANCELLED AT", "REASON", "STATUS", "Subtotal"]
VALID_SOURCE_STATUSES = {"cancelled", "canceled"}
EXCEL_EPOCH = datetime(1899, 12, 30)

# Known label variants between source file and SUMMARY mapping.
STORE_ALIASES = {
    "Bebang Halo-Halo - CTTM Square Tomas Morato": "Bebang Halo-halo - CTTM Square Tomas Morato",
    "Bebang Halo-Halo - Lucky Chinatown": "Bebang Halo-Halo - Lucky Chinatown Mall",
    "Bebang Halo-Halo - Gateway Mall": "Bebang Halo-Halo - Gateway Mall 1",
    "Bebang Halo-Halo - SM Sta. Rosa": "Bebang Halo-Halo - SM City Santa Rosa",
    "Bebang Halo-Halo - D'Verde Calamba": "Bebang Halo-Halo - D'Verde Extension",
}


@dataclass
class SourceRow:
    row_num: int
    dispute_id: str
    store_name: str
    order_at_raw: Any
    cancelled_at_raw: Any
    reason: str
    status: str
    subtotal_raw: Any


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_blank_or_na(value: Any) -> bool:
    text = normalize_text(value)
    return text == "" or text.upper() in {"#N/A", "N/A", "NA"}


def parse_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, (int, float)):
        try:
            return EXCEL_EPOCH + timedelta(days=float(value))
        except (TypeError, ValueError, OverflowError):
            return None

    text = normalize_text(value)
    if is_blank_or_na(text):
        return None

    for fmt in (
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def parse_positive_amount(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if float(value) > 0 else None

    text = normalize_text(value)
    if is_blank_or_na(text):
        return None

    cleaned = re.sub(r"[^\d.\-]", "", text)
    if cleaned in {"", "-", ".", "-."}:
        return None
    try:
        amount = float(cleaned)
    except ValueError:
        return None
    return amount if amount > 0 else None


def get_service():
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES
    ).with_subject(IMPERSONATE_USER)
    return build("sheets", "v4", credentials=creds, cache_discovery=False)


def load_summary_store_map(service, spreadsheet_id: str) -> dict[str, str]:
    values = (
        service.spreadsheets()
        .values()
        .get(
            spreadsheetId=spreadsheet_id,
            range="SUMMARY!A2:D200",
            valueRenderOption="FORMATTED_VALUE",
        )
        .execute()
        .get("values", [])
    )

    mapping: dict[str, str] = {}
    for row in values:
        if len(row) < 3:
            continue
        fp_label = normalize_text(row[0])
        store_code = normalize_text(row[2])
        if fp_label and store_code:
            mapping[fp_label] = store_code

    # Add alias keys if canonical targets exist.
    for alias, canonical in STORE_ALIASES.items():
        if canonical in mapping and alias not in mapping:
            mapping[alias] = mapping[canonical]

    return mapping


def load_existing_dispute_order_ids(service, spreadsheet_id: str) -> set[str]:
    values = (
        service.spreadsheets()
        .values()
        .get(
            spreadsheetId=spreadsheet_id,
            range="FP_DISPUTES!B2:B",
            valueRenderOption="FORMATTED_VALUE",
        )
        .execute()
        .get("values", [])
    )
    ids: set[str] = set()
    for row in values:
        if not row:
            continue
        order_id = normalize_text(row[0])
        if order_id:
            ids.add(order_id)
    return ids


def detect_header_row(sheet) -> int:
    for r in range(1, 25):
        values = [normalize_text(sheet.cell(r, c).value) for c in range(1, 12)]
        if "DISPUTE" in values and "STORE NAME" in values and "STATUS" in values:
            return r
    return 1


def extract_source_rows(xlsx_path: Path) -> tuple[list[SourceRow], dict[str, Any]]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if "DISPUTE" not in wb.sheetnames:
        raise RuntimeError("Sheet 'DISPUTE' not found in workbook")

    ws = wb["DISPUTE"]
    header_row = detect_header_row(ws)
    header = [normalize_text(ws.cell(header_row, c).value) for c in range(1, 9)]

    rows: list[SourceRow] = []
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, min_col=1, max_col=8, values_only=True),
        start=header_row + 1,
    ):
        if not any(not is_blank_or_na(v) for v in row):
            continue
        rows.append(
            SourceRow(
                row_num=row_idx,
                dispute_id=normalize_text(row[1]),
                store_name=normalize_text(row[2]),
                order_at_raw=row[3],
                cancelled_at_raw=row[4],
                reason=normalize_text(row[5]),
                status=normalize_text(row[6]),
                subtotal_raw=row[7],
            )
        )

    metadata = {
        "header_row": header_row,
        "header": header,
        "header_matches_expected": header == EXPECTED_HEADER,
        "source_rows_non_empty": len(rows),
    }
    return rows, metadata


def transform_rows(
    source_rows: list[SourceRow],
    store_map: dict[str, str],
    existing_order_ids: set[str],
    source_file_name: str,
    logged_at_iso: str,
) -> tuple[list[list[Any]], dict[str, Any]]:
    accepted: list[list[Any]] = []
    rejected: list[dict[str, Any]] = []
    accepted_ids_in_batch: set[str] = set()
    total_deduction = 0.0
    fallback_order_date_from_cancelled = 0

    reject_counters: dict[str, int] = {}

    def reject(row: SourceRow, reason: str):
        reject_counters[reason] = reject_counters.get(reason, 0) + 1
        rejected.append(
            {
                "row_num": row.row_num,
                "dispute_id": row.dispute_id,
                "store_name": row.store_name,
                "status": row.status,
                "subtotal": row.subtotal_raw,
                "reject_reason": reason,
            }
        )

    for row in source_rows:
        if is_blank_or_na(row.dispute_id):
            reject(row, "MISSING_DISPUTE_ID")
            continue
        if row.dispute_id in existing_order_ids:
            reject(row, "DUPLICATE_EXISTING_ORDER_ID")
            continue
        if row.dispute_id in accepted_ids_in_batch:
            reject(row, "DUPLICATE_WITHIN_FILE")
            continue
        if is_blank_or_na(row.store_name):
            reject(row, "MISSING_STORE_NAME")
            continue
        if row.store_name not in store_map:
            reject(row, "UNMAPPED_STORE_NAME")
            continue

        status_norm = row.status.lower().strip()
        if status_norm not in VALID_SOURCE_STATUSES:
            reject(row, "SOURCE_STATUS_NOT_CANCELLED")
            continue

        cancelled_at = parse_datetime(row.cancelled_at_raw)
        order_at = parse_datetime(row.order_at_raw)
        if order_at is None:
            if cancelled_at is None:
                reject(row, "INVALID_ORDER_AND_CANCELLED_DATETIME")
                continue
            order_at = cancelled_at
            fallback_order_date_from_cancelled += 1

        detected_at = cancelled_at or order_at

        subtotal = parse_positive_amount(row.subtotal_raw)
        if subtotal is None:
            reject(row, "INVALID_SUBTOTAL")
            continue

        store_code = store_map[row.store_name]
        delta_gross = -round(abs(subtotal), 2)
        total_deduction += delta_gross

        evidence_link = f"{source_file_name}|DISPUTE!R{row.row_num}"
        notes = (
            f"AUTO_IMPORT from {source_file_name} row {row.row_num}; "
            f"source_status={row.status or 'N/A'}"
        )
        if parse_datetime(row.order_at_raw) is None and cancelled_at is not None:
            notes += "; order_date_fallback=cancelled_at"

        accepted.append(
            [
                logged_at_iso,
                row.dispute_id,
                store_code,
                order_at.date().isoformat(),
                detected_at.date().isoformat(),
                delta_gross,
                row.reason,
                evidence_link,
                "Approved",
                "AUTO_IMPORT",
                notes,
            ]
        )
        accepted_ids_in_batch.add(row.dispute_id)

    report = {
        "accepted_rows": len(accepted),
        "rejected_rows": len(rejected),
        "total_deduction_php": round(total_deduction, 2),
        "fallback_order_date_from_cancelled": fallback_order_date_from_cancelled,
        "reject_counts": reject_counters,
        "rejected_sample": rejected[:50],
    }
    return accepted, report


def append_rows(service, spreadsheet_id: str, rows: list[list[Any]]) -> int:
    if not rows:
        return 0
    (
        service.spreadsheets()
        .values()
        .append(
            spreadsheetId=spreadsheet_id,
            range="FP_DISPUTES!A2:K2",
            valueInputOption="USER_ENTERED",
            insertDataOption="INSERT_ROWS",
            body={"values": rows},
        )
        .execute()
    )
    return len(rows)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Import FoodPanda dispute adjustments from Delivery Sales Summary xlsx to FP_DISPUTES"
    )
    parser.add_argument(
        "--xlsx",
        required=True,
        help="Path to source workbook (expects DISPUTE sheet).",
    )
    parser.add_argument(
        "--spreadsheet-id",
        default=DEFAULT_SPREADSHEET_ID,
        help=f"Google Sheets id (default: {DEFAULT_SPREADSHEET_ID})",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Write accepted rows into FP_DISPUTES. Default is dry-run.",
    )
    parser.add_argument(
        "--report-json",
        help="Optional path to save import report JSON.",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX file not found: {xlsx_path}")

    service = get_service()
    store_map = load_summary_store_map(service, args.spreadsheet_id)
    existing_ids = load_existing_dispute_order_ids(service, args.spreadsheet_id)
    source_rows, source_meta = extract_source_rows(xlsx_path)

    now = datetime.now().replace(microsecond=0).isoformat(sep=" ")
    accepted_rows, transform_report = transform_rows(
        source_rows=source_rows,
        store_map=store_map,
        existing_order_ids=existing_ids,
        source_file_name=xlsx_path.name,
        logged_at_iso=now,
    )

    written = 0
    if args.apply:
        written = append_rows(service, args.spreadsheet_id, accepted_rows)

    report = {
        "mode": "APPLY" if args.apply else "DRY_RUN",
        "spreadsheet_id": args.spreadsheet_id,
        "source_file": str(xlsx_path),
        "source_meta": source_meta,
        "existing_order_ids": len(existing_ids),
        "store_map_size": len(store_map),
        "accepted_rows_prewrite": len(accepted_rows),
        "written_rows": written,
        "transform_report": transform_report,
    }

    output = json.dumps(report, ensure_ascii=True, indent=2)
    print(output)

    if args.report_json:
        report_path = Path(args.report_json)
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(output + "\n", encoding="utf-8")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
