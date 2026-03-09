"""Discount identity monitoring APIs for Accounting and Audit."""

from __future__ import annotations

import io
import json
import os
from collections import Counter
from datetime import date, datetime, timedelta, timezone
from itertools import pairwise
from typing import Any
from zoneinfo import ZoneInfo

import requests

import frappe

MANILA_TZ = ZoneInfo("Asia/Manila")
DEFAULT_NOTIFICATION_GO_LIVE_DATE = date(2026, 3, 10)
REPORT_FILENAME_PREFIX = "Discount_Identity_Audit_Report_"
PORTAL_QUEUE_URL = "https://my.bebang.ph/dashboard/accounting/discount-abuse"
ALLOWED_ROLES = {
	"HQ User",
	"HQ Finance",
	"Accounts User",
	"Accounts Manager",
	"System Manager",
	"Administrator",
}
SAME_DAY_QUEUE_ORDER = (
	"severity_rank.asc,rapid_rank.asc,min_gap_minutes.asc.nullslast,"
	"discount_amount_total.desc,order_count.desc,store_name.asc"
)
ROLLING_QUEUE_ORDER = (
	"severity.asc,order_count.desc,active_day_count.desc,distinct_counterparty_count.desc,store_name.asc"
)
SEVERITY_RANK = {"critical": 1, "high": 2, "medium": 3, "review": 4}
QUEUE_BUCKET_RANK = {"same_day": 1, "rolling_30d": 2}
STATUTORY_CATEGORIES = ("SC", "PWD")
BENCHMARK_DENOMINATOR_SCOPES = {"pos_original_gross", "pos_post_discount_gross", "all_channel_gross"}
EXECUTIVE_CATEGORY_SCOPES = {"SC", "PWD", "BOTH"}
STORE_SCOPE = "store"
CHAIN_SCOPE = "chain"
CHAIN_LOCATION_ID = 0
CHAIN_STORE_NAME = "Chainwide"
CHAIN_PLACEHOLDER = "CHAINWIDE"
MIN_PEER_GROUP_SIZE = 5
HIGH_CONFIDENCE_SEVERITIES = {"critical", "high"}
WEIGHTED_SCORE_WEIGHTS = {
	"repeat_name": 2.0,
	"repeat_reference": 2.0,
	"same_reference_different_name": 7.0,
	"rapid_repeat_name": 3.0,
	"cross_store_overlap_critical": 1.5,
}
WEIGHTED_RATE_WEIGHTS = {
	"repeat_name": 2.0,
	"repeat_reference": 2.0,
	"same_reference_different_name": 5.0,
	"rapid_repeat_name": 3.0,
}
EXECUTIVE_KPI_CARD_ORDER = (
	"recorded_sc_pct_of_sales",
	"recorded_pwd_pct_of_sales",
	"effective_statutory_benefit_pct",
	"same_day_high_confidence_incidents",
	"flagged_discount_amount",
	"top_outlier_store",
)


def _conf_get(key: str, default: Any = None) -> Any:
	conf = getattr(frappe, "conf", {}) or {}
	if hasattr(conf, "get"):
		return conf.get(key, default)
	return default


def _permission_error_class():
	return getattr(frappe, "PermissionError", Exception)


def _is_guest() -> bool:
	return getattr(getattr(frappe, "session", None), "user", "Guest") == "Guest"


def _check_discount_audit_role() -> None:
	roles = set(frappe.get_roles() or [])
	if not roles.intersection(ALLOWED_ROLES):
		frappe.throw(
			frappe._("Only Accounting / Audit users can access discount monitoring."),
			_permission_error_class(),
		)


def _manila_now() -> datetime:
	return datetime.now(timezone.utc).astimezone(MANILA_TZ)


def _default_business_date() -> date:
	return _manila_now().date() - timedelta(days=1)


def _coerce_date(value: Any, default: date | None = None) -> date:
	if isinstance(value, date):
		return value
	if isinstance(value, datetime):
		return value.date()
	if isinstance(value, str) and value.strip():
		return date.fromisoformat(value.strip()[:10])
	if default is not None:
		return default
	raise ValueError("Business date is required")


def _format_display_date(value: date) -> str:
	return value.strftime("%B %d, %Y")


def _get_supabase_url() -> str:
	return (
		os.environ.get("SUPABASE_URL")
		or _conf_get("supabase_url")
		or "https://csnniykjrychgajfrgua.supabase.co"
	).rstrip("/")


def _get_supabase_service_key() -> str:
	return (
		os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
		or _conf_get("supabase_service_role_key")
		or _conf_get("supabase_service_role_key".upper())
		or ""
	)


def _supabase_headers() -> dict[str, str]:
	key = _get_supabase_service_key()
	if not key:
		raise RuntimeError("SUPABASE_SERVICE_ROLE_KEY is not configured")
	return {
		"apikey": key,
		"Authorization": f"Bearer {key}",
		"Content-Type": "application/json",
	}


def _supabase_get(
	resource: str,
	params: dict[str, Any] | list[tuple[str, Any]] | None = None,
) -> list[dict[str, Any]]:
	response = requests.get(
		f"{_get_supabase_url()}/rest/v1/{resource}",
		headers=_supabase_headers(),
		params=params or {},
		timeout=30,
	)
	if not response.ok:
		raise RuntimeError(
			f"Supabase GET failed for {resource}: {response.status_code} {response.text[:300]}"
		)
	payload = response.json()
	if isinstance(payload, list):
		return payload
	raise RuntimeError(f"Unexpected Supabase payload for {resource}")


def _supabase_merge_params(
	params: dict[str, Any] | list[tuple[str, Any]] | None,
	extra: list[tuple[str, Any]],
) -> dict[str, Any] | list[tuple[str, Any]]:
	if params is None:
		return list(extra)
	if isinstance(params, dict):
		merged = dict(params)
		for key, value in extra:
			merged[key] = value
		return merged
	merged = [(key, value) for key, value in params if key not in {"limit", "offset"}]
	merged.extend(extra)
	return merged


def _supabase_get_all(
	resource: str,
	params: dict[str, Any] | list[tuple[str, Any]] | None = None,
	page_size: int = 1000,
) -> list[dict[str, Any]]:
	all_rows: list[dict[str, Any]] = []
	offset = 0
	while True:
		page_params = _supabase_merge_params(
			params,
			[("limit", str(page_size)), ("offset", str(offset))],
		)
		rows = _supabase_get(resource, page_params)
		all_rows.extend(rows)
		if len(rows) < page_size:
			break
		offset += page_size
	return all_rows


def _supabase_patch(
	resource: str,
	filters: dict[str, str],
	payload: dict[str, Any],
) -> list[dict[str, Any]]:
	headers = _supabase_headers()
	headers["Prefer"] = "return=representation"
	response = requests.patch(
		f"{_get_supabase_url()}/rest/v1/{resource}",
		headers=headers,
		params=filters,
		json=payload,
		timeout=30,
	)
	if not response.ok:
		raise RuntimeError(
			f"Supabase PATCH failed for {resource}: {response.status_code} {response.text[:300]}"
		)
	result = response.json()
	return result if isinstance(result, list) else []


def _supabase_post(
	resource: str,
	payload: list[dict[str, Any]] | dict[str, Any],
	*,
	on_conflict: str | None = None,
	return_rows: bool = False,
) -> list[dict[str, Any]]:
	headers = _supabase_headers()
	headers["Prefer"] = (
		"resolution=merge-duplicates,return=representation"
		if return_rows
		else "resolution=merge-duplicates,return=minimal"
	)
	url = f"{_get_supabase_url()}/rest/v1/{resource}"
	params: dict[str, str] = {}
	if on_conflict:
		params["on_conflict"] = on_conflict
	response = requests.post(
		url,
		headers=headers,
		params=params,
		json=payload,
		timeout=60,
	)
	if not response.ok:
		raise RuntimeError(
			f"Supabase POST failed for {resource}: {response.status_code} {response.text[:300]}"
		)
	if not return_rows:
		return []
	result = response.json()
	return result if isinstance(result, list) else []


def _supabase_delete(resource: str, filters: dict[str, str]) -> None:
	headers = _supabase_headers()
	headers["Prefer"] = "return=minimal"
	response = requests.delete(
		f"{_get_supabase_url()}/rest/v1/{resource}",
		headers=headers,
		params=filters,
		timeout=60,
	)
	if not response.ok:
		raise RuntimeError(
			f"Supabase DELETE failed for {resource}: {response.status_code} {response.text[:300]}"
		)


def _parse_email_recipients() -> list[str]:
	raw = os.environ.get("DISCOUNT_ALERT_EMAIL_RECIPIENTS") or _conf_get(
		"discount_alert_email_recipients", ""
	)
	if not raw:
		return []
	if isinstance(raw, list | tuple):
		return [str(item).strip() for item in raw if str(item).strip()]
	return [item.strip() for item in str(raw).split(",") if item.strip()]


def _get_report_owner(default_user: str | None = None) -> str:
	explicit = os.environ.get("DISCOUNT_ALERT_REPORT_OWNER") or _conf_get("discount_alert_report_owner")
	if explicit:
		return explicit
	if default_user and default_user != "Guest":
		return default_user
	return "Administrator"


def _get_notification_go_live_date() -> date:
	raw = os.environ.get("DISCOUNT_ALERT_NOTIFY_ENABLE_FROM") or _conf_get(
		"discount_alert_notify_enable_from"
	)
	if not raw:
		return DEFAULT_NOTIFICATION_GO_LIVE_DATE
	try:
		return _coerce_date(raw)
	except Exception:
		return DEFAULT_NOTIFICATION_GO_LIVE_DATE


def _notifications_enabled_for_day(target_day: date) -> bool:
	return target_day >= _get_notification_go_live_date()


def _to_number(value: Any) -> float:
	if value in (None, "", False):
		return 0.0
	try:
		return float(value)
	except (TypeError, ValueError):
		return 0.0


def _normalize_text(value: Any) -> str:
	if value is None:
		return ""
	return " ".join(str(value).strip().upper().split())


def _split_pipe_values(value: Any) -> list[str]:
	if value is None:
		return []
	if isinstance(value, list | tuple | set):
		values: list[str] = []
		for entry in value:
			values.extend(_split_pipe_values(entry))
		return values
	text = str(value).strip()
	if not text:
		return []
	return [part.strip() for part in text.split("|") if part.strip()]


def _dedupe_preserve_order(values: list[str]) -> list[str]:
	seen: set[str] = set()
	result: list[str] = []
	for value in values:
		key = _normalize_text(value)
		if not key or key in seen:
			continue
		seen.add(key)
		result.append(key)
	return result


def _dedupe_preserve_display(values: list[str]) -> list[str]:
	seen: set[str] = set()
	result: list[str] = []
	for value in values:
		text = str(value).strip()
		key = _normalize_text(text)
		if not key or key in seen:
			continue
		seen.add(key)
		result.append(text)
	return result


def _canonical_discount_category(
	category: Any,
	discount_name_normalized: Any = None,
	discount_name: Any = None,
) -> str:
	category_text = _normalize_text(category)
	if category_text in {"SC", "PWD"}:
		return category_text
	name_text = _normalize_text(discount_name_normalized or discount_name)
	if "PWD" in name_text:
		return "PWD"
	if "SENIOR" in name_text or "SC" == name_text:
		return "SC"
	return ""


def _parse_iso_datetime(value: Any) -> datetime | None:
	if not value:
		return None
	text = str(value).strip()
	if not text:
		return None
	try:
		return datetime.fromisoformat(text.replace("Z", "+00:00"))
	except ValueError:
		return None


def _safe_json(value: Any) -> dict[str, Any]:
	if isinstance(value, dict):
		return value
	if isinstance(value, str) and value.strip():
		try:
			parsed = json.loads(value)
			return parsed if isinstance(parsed, dict) else {}
		except json.JSONDecodeError:
			return {}
	return {}


def _normalize_queue_row(row: dict[str, Any]) -> dict[str, Any]:
	normalized = dict(row)
	normalized["discount_amount_total"] = _to_number(row.get("discount_amount_total"))
	normalized["min_gap_minutes"] = (
		None if row.get("min_gap_minutes") in (None, "") else _to_number(row.get("min_gap_minutes"))
	)
	normalized["order_count"] = int(row.get("order_count") or 0)
	normalized["store_count"] = int(row.get("store_count") or 0)
	normalized["active_day_count"] = int(row.get("active_day_count") or 0)
	normalized["distinct_counterparty_count"] = int(row.get("distinct_counterparty_count") or 0)
	normalized["rapid_within_4h"] = bool(row.get("rapid_within_4h"))
	normalized["details"] = _safe_json(row.get("details"))
	normalized["event_date"] = str(row.get("event_date") or "")
	normalized["window_start"] = str(row.get("window_start") or "")
	normalized["window_end"] = str(row.get("window_end") or "")
	return normalized


def _sort_same_day_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
	return sorted(
		rows,
		key=lambda row: (
			SEVERITY_RANK.get(str(row.get("severity") or ""), 9),
			0 if row.get("rapid_within_4h") else 1,
			row.get("min_gap_minutes") if row.get("min_gap_minutes") is not None else 999999,
			-_to_number(row.get("discount_amount_total")),
			-(int(row.get("order_count") or 0)),
			str(row.get("store_name") or ""),
			str(row.get("identity_key") or ""),
		),
	)


def _sort_rolling_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
	return sorted(
		rows,
		key=lambda row: (
			SEVERITY_RANK.get(str(row.get("severity") or ""), 9),
			-(int(row.get("order_count") or 0)),
			-(int(row.get("active_day_count") or 0)),
			-(int(row.get("distinct_counterparty_count") or 0)),
			str(row.get("store_name") or ""),
			str(row.get("identity_key") or ""),
		),
	)


def _query_same_day_rows(business_date: date, severity: str | None = None) -> list[dict[str, Any]]:
	params: dict[str, Any] = {
		"select": "*",
		"queue_bucket": "eq.same_day",
		"event_date": f"eq.{business_date.isoformat()}",
		"order": SAME_DAY_QUEUE_ORDER,
		"limit": "1000",
	}
	if severity and severity != "all":
		params["severity"] = f"eq.{severity}"
	return [_normalize_queue_row(row) for row in _supabase_get("v_discount_identity_audit_queue", params)]


def _query_rolling_rows(business_date: date, severity: str | None = None) -> list[dict[str, Any]]:
	queue_params: dict[str, Any] = {
		"select": "*",
		"queue_bucket": "eq.rolling_30d",
		"event_date": f"eq.{business_date.isoformat()}",
		"order": "severity_rank.asc,discount_amount_total.desc,order_count.desc,store_name.asc",
		"limit": "1000",
	}
	if severity and severity != "all":
		queue_params["severity"] = f"eq.{severity}"
	rows = _supabase_get("v_discount_identity_audit_queue", queue_params)
	if rows:
		return [_normalize_queue_row(row) for row in rows]

	snapshot_params: dict[str, Any] = {
		"select": (
			"id,as_of_date,window_start,window_end,scope,scope_key,location_id,store_name,"
			"discount_name,discount_bir_category,identity_type,identity_key,customer_name,"
			"reference_number,order_count,active_day_count,distinct_counterparty_count,"
			"detection_type,severity,details,created_at,updated_at"
		),
		"as_of_date": f"eq.{business_date.isoformat()}",
		"order": ROLLING_QUEUE_ORDER,
		"limit": "1000",
	}
	if severity and severity != "all":
		snapshot_params["severity"] = f"eq.{severity}"
	snapshot_rows = _supabase_get("discount_identity_30d_snapshots", snapshot_params)
	normalized_rows: list[dict[str, Any]] = []
	for row in snapshot_rows:
		details = _safe_json(row.get("details"))
		normalized_rows.append(
			_normalize_queue_row(
				{
					"queue_bucket": "rolling_30d",
					"queue_bucket_rank": 2,
					"severity_rank": 4 if row.get("severity") == "review" else 2,
					"rapid_rank": 1,
					"event_date": row.get("as_of_date"),
					"window_start": row.get("window_start"),
					"window_end": row.get("window_end"),
					"scope": row.get("scope"),
					"scope_key": row.get("scope_key"),
					"location_id": row.get("location_id"),
					"store_name": row.get("store_name"),
					"discount_name": row.get("discount_name"),
					"discount_bir_category": row.get("discount_bir_category"),
					"identity_type": row.get("identity_type"),
					"identity_key": row.get("identity_key"),
					"customer_name": row.get("customer_name"),
					"reference_number": row.get("reference_number"),
					"order_count": row.get("order_count"),
					"store_count": 1,
					"active_day_count": row.get("active_day_count"),
					"distinct_counterparty_count": row.get("distinct_counterparty_count"),
					"detection_type": row.get("detection_type"),
					"severity": row.get("severity"),
					"rapid_within_4h": False,
					"min_gap_minutes": None,
					"discount_amount_total": details.get("discount_amount_total"),
					"resolved": False,
					"notified_at": None,
					"created_at": row.get("created_at"),
					"updated_at": row.get("updated_at"),
					"details": details,
				}
			)
		)
	return normalized_rows


def _filter_rows(
	rows: list[dict[str, Any]],
	search: str | None = None,
	severity: str | None = None,
	store_name: str | None = None,
	category: str | None = None,
) -> list[dict[str, Any]]:
	filtered = list(rows)
	if severity and severity != "all":
		filtered = [row for row in filtered if str(row.get("severity") or "") == severity]
	if store_name and store_name != "all":
		filtered = [row for row in filtered if str(row.get("store_name") or "") == store_name]
	if category and category != "all":
		filtered = [
			row for row in filtered if str(row.get("discount_bir_category") or "").upper() == category.upper()
		]
	if search:
		needle = search.strip().lower()
		filtered = [
			row
			for row in filtered
			if needle
			in " | ".join(
				[
					str(row.get("store_name") or ""),
					str(row.get("discount_name") or ""),
					str(row.get("identity_key") or ""),
					str(row.get("customer_name") or ""),
					str(row.get("reference_number") or ""),
					str(row.get("detection_type") or ""),
				]
			).lower()
		]
	return filtered


def _make_summary(same_day_rows: list[dict[str, Any]], rolling_rows: list[dict[str, Any]]) -> dict[str, int]:
	same_day_counts = Counter(row.get("severity") for row in same_day_rows)
	rolling_counts = Counter(row.get("severity") for row in rolling_rows)
	return {
		"same_day_critical": int(same_day_counts.get("critical", 0)),
		"same_day_high": int(same_day_counts.get("high", 0)),
		"same_day_medium": int(same_day_counts.get("medium", 0)),
		"rolling_high": int(rolling_counts.get("high", 0)),
		"rolling_review": int(rolling_counts.get("review", 0)),
	}


def _canonical_bundle(values: list[str]) -> str:
	return " | ".join(sorted({_normalize_text(value) for value in values if _normalize_text(value)}))


def _extract_alert_store_names(row: dict[str, Any]) -> list[str]:
	details = _safe_json(row.get("details"))
	return _dedupe_preserve_display(
		_split_pipe_values(row.get("store_name")) + _split_pipe_values(details.get("store_names"))
	)


def _extract_alert_order_ids(row: dict[str, Any]) -> list[str]:
	details = _safe_json(row.get("details"))
	return _dedupe_preserve_display(_split_pipe_values(details.get("order_ids")))


def _resolve_identity_fallback(
	row: dict[str, Any], names: list[str], references: list[str]
) -> tuple[list[str], list[str]]:
	identity_type = _normalize_text(row.get("identity_type"))
	identity_key = _normalize_text(row.get("identity_key"))
	resolved_names = list(names)
	resolved_references = list(references)
	if identity_key and identity_type == "CUSTOMER_NAME" and not resolved_names:
		resolved_names = [identity_key]
	if identity_key and identity_type == "REFERENCE_NUMBER" and not resolved_references:
		resolved_references = [identity_key]
	return resolved_names, resolved_references


def _build_resolution_target(row: dict[str, Any]) -> dict[str, Any]:
	return {
		"event_date": str(row.get("event_date") or ""),
		"scope": str(row.get("scope") or ""),
		"scope_key": int(row.get("scope_key") or 0),
		"discount_bir_category": str(row.get("discount_bir_category") or ""),
		"identity_type": str(row.get("identity_type") or ""),
		"identity_key": str(row.get("identity_key") or ""),
		"detection_type": str(row.get("detection_type") or ""),
	}


def _resolution_target_key(target: dict[str, Any]) -> str:
	return "::".join(
		[
			str(target.get("event_date") or ""),
			str(target.get("scope") or ""),
			str(target.get("scope_key") or ""),
			str(target.get("discount_bir_category") or ""),
			str(target.get("identity_type") or ""),
			str(target.get("identity_key") or ""),
			str(target.get("detection_type") or ""),
		]
	)


def _build_incident_cluster_key(row: dict[str, Any]) -> str:
	queue_bucket = str(row.get("queue_bucket") or "same_day")
	associations = _extract_alert_associations(row)
	store_names = _extract_alert_store_names(row)
	names, references = _resolve_identity_fallback(row, associations["names"], associations["references"])
	bills = associations["bill_numbers"]
	if queue_bucket == "rolling_30d":
		return "::".join(
			[
				queue_bucket,
				str(row.get("window_start") or ""),
				str(row.get("window_end") or ""),
				str(row.get("scope") or ""),
				str(int(row.get("scope_key") or 0)),
				_canonical_bundle(store_names),
				str(row.get("discount_bir_category") or ""),
				str(row.get("identity_type") or ""),
				_canonical_bundle(names),
				_canonical_bundle(references),
			]
		)
	return "::".join(
		[
			queue_bucket,
			str(row.get("event_date") or ""),
			str(row.get("scope") or ""),
			str(int(row.get("scope_key") or 0)),
			_canonical_bundle(store_names),
			str(row.get("discount_bir_category") or ""),
			_canonical_bundle(bills),
			_canonical_bundle(names),
			_canonical_bundle(references),
		]
	)


def _pick_cluster_identity_key(names: list[str], references: list[str], fallback: str) -> str:
	if references:
		return " | ".join(references)
	if names:
		return " | ".join(names)
	return fallback


def _cluster_queue_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
	clusters: dict[str, dict[str, Any]] = {}
	for row in rows:
		cluster_key = _build_incident_cluster_key(row)
		associations = _extract_alert_associations(row)
		names, references = _resolve_identity_fallback(row, associations["names"], associations["references"])
		store_names = _extract_alert_store_names(row)
		bill_numbers = associations["bill_numbers"]
		business_dates = associations["business_dates"] or [str(row.get("event_date") or "")]
		order_ids = _extract_alert_order_ids(row)
		cluster = clusters.setdefault(
			cluster_key,
			{
				"cluster_id": cluster_key,
				"cluster_key": cluster_key,
				"queue_bucket": str(row.get("queue_bucket") or "same_day"),
				"queue_bucket_rank": QUEUE_BUCKET_RANK.get(str(row.get("queue_bucket") or "same_day"), 9),
				"event_date": str(row.get("event_date") or ""),
				"window_start": str(row.get("window_start") or ""),
				"window_end": str(row.get("window_end") or ""),
				"scope": str(row.get("scope") or ""),
				"scope_key": int(row.get("scope_key") or 0),
				"location_id": int(row.get("location_id") or 0)
				if row.get("location_id") not in (None, "")
				else None,
				"discount_name": row.get("discount_name"),
				"discount_bir_category": str(row.get("discount_bir_category") or ""),
				"severity": str(row.get("severity") or "review"),
				"rapid_within_4h": False,
				"min_gap_minutes": None,
				"discount_amount_total": 0.0,
				"order_count": 0,
				"store_count": 0,
				"active_day_count": 0,
				"distinct_counterparty_count": 0,
				"detection_types": [],
				"identity_types": [],
				"store_names": [],
				"customer_names": [],
				"reference_numbers": [],
				"bill_numbers": [],
				"business_dates": [],
				"order_ids": [],
				"resolution_targets": [],
				"resolve_scope_policy": "all_underlying_rows",
				"raw_row_count": 0,
				"resolved": False,
				"notified_at": None,
				"created_at": row.get("created_at"),
				"updated_at": row.get("updated_at"),
			},
		)
		cluster["raw_row_count"] += 1
		cluster["rapid_within_4h"] = bool(cluster["rapid_within_4h"] or row.get("rapid_within_4h"))
		row_min_gap = row.get("min_gap_minutes")
		if row_min_gap is not None:
			cluster["min_gap_minutes"] = (
				row_min_gap
				if cluster["min_gap_minutes"] is None
				else min(float(cluster["min_gap_minutes"]), float(row_min_gap))
			)
		cluster["discount_amount_total"] = max(
			float(cluster["discount_amount_total"]),
			_to_number(row.get("discount_amount_total")),
		)
		cluster["order_count"] = max(int(cluster["order_count"]), int(row.get("order_count") or 0))
		cluster["store_count"] = max(int(cluster["store_count"]), int(row.get("store_count") or 0))
		cluster["active_day_count"] = max(
			int(cluster["active_day_count"]), int(row.get("active_day_count") or 0)
		)
		cluster["distinct_counterparty_count"] = max(
			int(cluster["distinct_counterparty_count"]),
			int(row.get("distinct_counterparty_count") or 0),
		)
		if SEVERITY_RANK.get(str(row.get("severity") or ""), 9) < SEVERITY_RANK.get(
			str(cluster["severity"] or ""), 9
		):
			cluster["severity"] = str(row.get("severity") or "review")
		cluster["detection_types"] = _dedupe_preserve_display(
			cluster["detection_types"] + _split_pipe_values(row.get("detection_type"))
		)
		cluster["identity_types"] = _dedupe_preserve_display(
			cluster["identity_types"] + _split_pipe_values(row.get("identity_type"))
		)
		cluster["store_names"] = _dedupe_preserve_display(cluster["store_names"] + store_names)
		cluster["customer_names"] = _dedupe_preserve_order(cluster["customer_names"] + names)
		cluster["reference_numbers"] = _dedupe_preserve_order(cluster["reference_numbers"] + references)
		cluster["bill_numbers"] = _dedupe_preserve_display(cluster["bill_numbers"] + bill_numbers)
		cluster["business_dates"] = _dedupe_preserve_display(cluster["business_dates"] + business_dates)
		cluster["order_ids"] = _dedupe_preserve_display(cluster["order_ids"] + order_ids)
		target = _build_resolution_target(row)
		target_keys = {_resolution_target_key(entry) for entry in cluster["resolution_targets"]}
		target_key = _resolution_target_key(target)
		if target_key not in target_keys:
			cluster["resolution_targets"].append(target)

	final_rows: list[dict[str, Any]] = []
	for cluster in clusters.values():
		identity_key = _pick_cluster_identity_key(
			cluster["customer_names"],
			cluster["reference_numbers"],
			str(cluster.get("identity_key") or ""),
		)
		cluster["store_name"] = " | ".join(cluster["store_names"])
		cluster["customer_name"] = cluster["customer_names"][0] if cluster["customer_names"] else None
		cluster["reference_number"] = (
			cluster["reference_numbers"][0] if cluster["reference_numbers"] else None
		)
		cluster["customer_names_display"] = " | ".join(cluster["customer_names"])
		cluster["reference_numbers_display"] = " | ".join(cluster["reference_numbers"])
		cluster["bill_numbers_display"] = " | ".join(cluster["bill_numbers"])
		cluster["business_dates_display"] = " | ".join(cluster["business_dates"])
		cluster["identity_key"] = identity_key
		cluster["identity_type"] = (
			cluster["identity_types"][0] if len(cluster["identity_types"]) == 1 else "mixed"
		)
		cluster["detection_type"] = (
			cluster["detection_types"][0] if cluster["detection_types"] else "incident_cluster"
		)
		cluster["details"] = {
			"order_ids": cluster["order_ids"],
			"store_names": cluster["store_names"],
			"customer_names": cluster["customer_names"],
			"reference_numbers": cluster["reference_numbers"],
			"bill_numbers": cluster["bill_numbers"],
			"business_dates": cluster["business_dates"],
			"raw_detection_types": cluster["detection_types"],
			"raw_identity_types": cluster["identity_types"],
			"raw_row_count": cluster["raw_row_count"],
			"distinct_name_count": len(cluster["customer_names"]),
			"distinct_reference_count": len(cluster["reference_numbers"]),
		}
		final_rows.append(cluster)
	if any(row.get("queue_bucket") == "rolling_30d" for row in final_rows):
		same_day_rows = [row for row in final_rows if row.get("queue_bucket") == "same_day"]
		rolling_rows = [row for row in final_rows if row.get("queue_bucket") == "rolling_30d"]
		return _sort_same_day_rows(same_day_rows) + _sort_rolling_rows(rolling_rows)
	return _sort_same_day_rows(final_rows)


def _filter_incident_clusters(
	rows: list[dict[str, Any]],
	search: str | None = None,
	severity: str | None = None,
	store_name: str | None = None,
	category: str | None = None,
) -> list[dict[str, Any]]:
	filtered = list(rows)
	if severity and severity != "all":
		filtered = [row for row in filtered if str(row.get("severity") or "") == severity]
	if category and category != "all":
		needle = _normalize_text(category)
		filtered = [row for row in filtered if _normalize_text(row.get("discount_bir_category")) == needle]
	if store_name:
		filtered = [row for row in filtered if _row_mentions_store(row, store_name)]
	if search:
		needle = search.strip().lower()
		filtered = [
			row
			for row in filtered
			if needle
			in " | ".join(
				[
					str(row.get("store_name") or ""),
					str(row.get("identity_key") or ""),
					str(row.get("customer_names_display") or ""),
					str(row.get("reference_numbers_display") or ""),
					str(row.get("bill_numbers_display") or ""),
					" | ".join(_split_pipe_values(row.get("detection_type")))
					if row.get("detection_type")
					else "",
					" | ".join(row.get("detection_types") or []),
				]
			).lower()
		]
	return filtered


def _format_display_window(start_day: date, end_day: date) -> str:
	if start_day == end_day:
		return _format_display_date(start_day)
	return f"{_format_display_date(start_day)} to {_format_display_date(end_day)}"


def _parse_store_names(value: str | None) -> list[str]:
	if not value:
		return []
	return [part.strip() for part in str(value).split(",") if part.strip()]


def _parse_category_filter(value: str | None) -> set[str] | None:
	text = _normalize_text(value)
	if not text or text == "ALL":
		return None
	return {part.strip() for part in text.split(",") if part.strip() in {"SC", "PWD"}}


def _get_store_directory() -> dict[int, str]:
	return {
		location_id: str(row.get("store_name") or location_id)
		for location_id, row in _get_store_catalog().items()
	}


def _get_store_catalog() -> dict[int, dict[str, Any]]:
	rows = _supabase_get_all(
		"stores",
		{
			"select": "location_id,store_name,legal_entity,store_type",
			"order": "store_name.asc",
		},
		page_size=200,
	)
	catalog: dict[int, dict[str, Any]] = {}
	for row in rows:
		location_id = int(row.get("location_id") or 0)
		if not location_id:
			continue
		catalog[location_id] = {
			"location_id": location_id,
			"store_name": str(row.get("store_name") or location_id),
			"legal_entity": str(row.get("legal_entity") or ""),
			"store_type": str(row.get("store_type") or ""),
		}
	return catalog


def _month_start(value: date) -> date:
	return value.replace(day=1)


def _month_end(value: date) -> date:
	if value.month == 12:
		return date(value.year, 12, 31)
	return date(value.year, value.month + 1, 1) - timedelta(days=1)


def _iter_days(start_day: date, end_day: date) -> list[date]:
	return [start_day + timedelta(days=offset) for offset in range((end_day - start_day).days + 1)]


def _normalize_category_scope(value: Any) -> str:
	text = _normalize_text(value)
	if text not in EXECUTIVE_CATEGORY_SCOPES:
		return "BOTH"
	return text


def _normalize_denominator_scope(value: Any) -> str:
	text = str(value or "pos_original_gross").strip().lower()
	if text not in BENCHMARK_DENOMINATOR_SCOPES:
		return "pos_original_gross"
	return text


def _normalize_peer_mode(value: Any) -> str:
	text = str(value or "auto").strip().lower()
	if text not in {"auto", "store_type_legal_entity", "store_type", "chainwide"}:
		return "auto"
	return text


def _round_metric(value: float, digits: int = 2) -> float:
	return round(float(value or 0), digits)


def _safe_divide(numerator: float, denominator: float) -> float:
	if not denominator:
		return 0.0
	return float(numerator) / float(denominator)


def _pct(numerator: float, denominator: float) -> float:
	return _round_metric(_safe_divide(numerator, denominator) * 100, 4)


def _query_store_daily_closing_rows(start_day: date, end_day: date) -> list[dict[str, Any]]:
	params: list[tuple[str, Any]] = [
		(
			"select",
			",".join(
				[
					"location_id",
					"store_name",
					"legal_entity",
					"store_type",
					"business_date",
					"pos_orders",
					"pos_original_gross",
					"pos_after_discount",
					"pos_net_of_vat",
					"pos_discounts",
					"pos_vat",
					"pos_vat_exempt",
				]
			),
		),
		("business_date", f"gte.{start_day.isoformat()}"),
		("business_date", f"lte.{end_day.isoformat()}"),
		("order", "business_date.asc,store_name.asc"),
	]
	return _supabase_get_all("store_daily_closing", params)


def _query_all_channel_daily_rows(start_day: date, end_day: date) -> list[dict[str, Any]]:
	system_params: list[tuple[str, Any]] = [
		(
			"select",
			",".join(
				[
					"business_date",
					"pos_order_count",
					"pos_gross_sales",
					"total_orders",
					"total_gross_sales",
					"web_order_count",
					"web_gross_sales",
				]
			),
		),
		("business_date", f"gte.{start_day.isoformat()}"),
		("business_date", f"lte.{end_day.isoformat()}"),
		("order", "business_date.asc"),
	]
	system_rows = _supabase_get_all("v_system_daily_totals", system_params)
	foodpanda_params: list[tuple[str, Any]] = [
		("select", "business_date,subtotal,order_status"),
		("business_date", f"gte.{start_day.isoformat()}"),
		("business_date", f"lte.{end_day.isoformat()}"),
		("order_status", "ilike.delivered"),
		("order", "business_date.asc"),
	]
	foodpanda_rows = _supabase_get_all("foodpanda_orders", foodpanda_params)
	foodpanda_by_day: dict[str, dict[str, float | int]] = {}
	for row in foodpanda_rows:
		day_key = str(row.get("business_date") or "")
		if not day_key:
			continue
		day_bucket = foodpanda_by_day.setdefault(day_key, {"fp_orders": 0, "fp_gross_sales": 0.0})
		day_bucket["fp_orders"] = int(day_bucket["fp_orders"]) + 1
		day_bucket["fp_gross_sales"] = _to_number(day_bucket["fp_gross_sales"]) + _to_number(
			row.get("subtotal")
		)

	results: list[dict[str, Any]] = []
	for row in system_rows:
		day_key = str(row.get("business_date") or "")
		foodpanda = foodpanda_by_day.get(day_key, {"fp_orders": 0, "fp_gross_sales": 0.0})
		pos_orders = int(row.get("pos_order_count") or 0)
		web_orders = int(row.get("web_order_count") or 0)
		fp_orders = int(foodpanda["fp_orders"])
		pos_gross_sales = _to_number(row.get("pos_gross_sales"))
		web_gross_sales = _to_number(row.get("web_gross_sales"))
		fp_gross_sales = _to_number(foodpanda["fp_gross_sales"])
		results.append(
			{
				"business_date": day_key,
				"pos_orders": pos_orders,
				"pos_gross_sales": pos_gross_sales,
				"web_orders": web_orders,
				"web_gross_sales": web_gross_sales,
				"fp_orders": fp_orders,
				"fp_gross_sales": fp_gross_sales,
				"total_orders": pos_orders + web_orders + fp_orders,
				"total_gross_sales": _round_metric(pos_gross_sales + web_gross_sales + fp_gross_sales, 6),
				"channel_count": sum(1 for count in (pos_orders, web_orders, fp_orders) if count > 0),
				"data_sources": " | ".join(
					source
					for source, count in (
						("POS", pos_orders),
						("Web", web_orders),
						("FoodPanda", fp_orders),
					)
					if count > 0
				),
			}
		)
	return results


def _query_all_channel_daily_rows_legacy(start_day: date, end_day: date) -> list[dict[str, Any]]:
	params: list[tuple[str, Any]] = [
		(
			"select",
			",".join(
				[
					"business_date",
					"pos_orders",
					"pos_gross_sales",
					"pos_net_sales",
					"web_orders",
					"web_gross_sales",
					"web_net_sales",
					"fp_orders",
					"fp_gross_sales",
					"total_orders",
					"total_gross_sales",
					"channel_count",
					"data_sources",
				]
			),
		),
		("business_date", f"gte.{start_day.isoformat()}"),
		("business_date", f"lte.{end_day.isoformat()}"),
		("order", "business_date.asc"),
	]
	return _supabase_get_all("v_all_channel_daily", params)


def _query_store_day_snapshots(start_day: date, end_day: date) -> list[dict[str, Any]]:
	params: list[tuple[str, Any]] = [
		("select", "*"),
		("business_date", f"gte.{start_day.isoformat()}"),
		("business_date", f"lte.{end_day.isoformat()}"),
		("order", "business_date.asc,scope.asc,store_name.asc"),
	]
	return _supabase_get_all("discount_investigation_store_day", params)


def _query_store_month_snapshots(month_start: date) -> list[dict[str, Any]]:
	params = {
		"select": "*",
		"month_start": f"eq.{month_start.isoformat()}",
		"order": "scope.asc,store_name.asc",
	}
	return _supabase_get_all("discount_investigation_store_month", params)


def _resolve_selected_stores(store_names: list[str]) -> tuple[list[int], dict[int, str]]:
	store_directory = _get_store_directory()
	if not store_names:
		return [], {}
	normalized_to_id = {
		_normalize_text(store_name): location_id for location_id, store_name in store_directory.items()
	}
	selected_location_ids: list[int] = []
	selected_store_map: dict[int, str] = {}
	for store_name in store_names:
		location_id = normalized_to_id.get(_normalize_text(store_name))
		if location_id is None:
			continue
		if location_id in selected_store_map:
			continue
		selected_location_ids.append(location_id)
		selected_store_map[location_id] = store_directory[location_id]
	return selected_location_ids, selected_store_map


def _row_mentions_store(row: dict[str, Any], store_name: str) -> bool:
	names = _split_pipe_values(row.get("store_name"))
	if not names and row.get("store_name"):
		names = [str(row["store_name"])]
	needle = _normalize_text(store_name)
	return any(_normalize_text(name) == needle for name in names)


def _query_same_day_rows_range(
	start_day: date,
	end_day: date,
	categories: set[str] | None = None,
) -> list[dict[str, Any]]:
	params: list[tuple[str, Any]] = [
		("select", "*"),
		("queue_bucket", "eq.same_day"),
		("event_date", f"gte.{start_day.isoformat()}"),
		("event_date", f"lte.{end_day.isoformat()}"),
		("order", SAME_DAY_QUEUE_ORDER),
	]
	rows = [_normalize_queue_row(row) for row in _supabase_get_all("v_discount_identity_audit_queue", params)]
	if categories:
		rows = [row for row in rows if _normalize_text(row.get("discount_bir_category")) in categories]
	return rows


def _query_paid_orders_for_range(
	start_day: date, end_day: date, location_ids: list[int]
) -> list[dict[str, Any]]:
	if not location_ids:
		return []
	location_list = ",".join(str(location_id) for location_id in location_ids)
	params: list[tuple[str, Any]] = [
		(
			"select",
			"id,location_id,business_date,bill_number,receipt_number,billed_at,paid_at,payment_status,original_gross_sales,gross_sales,net_sales,total_discounts",
		),
		("business_date", f"gte.{start_day.isoformat()}"),
		("business_date", f"lte.{end_day.isoformat()}"),
		("location_id", f"in.({location_list})"),
		("payment_status", "eq.PAID"),
		("order", "business_date.asc,id.asc"),
	]
	return _supabase_get_all("pos_orders", params)


def _chunk_values(values: list[int], size: int = 200) -> list[list[int]]:
	if not values:
		return []
	return [values[index : index + size] for index in range(0, len(values), size)]


def _query_discount_item_rows_for_orders(
	paid_order_rows: list[dict[str, Any]],
	categories: set[str] | None = None,
) -> list[dict[str, Any]]:
	if not paid_order_rows:
		return []
	order_lookup: dict[int, dict[str, Any]] = {}
	for row in paid_order_rows:
		order_id = int(row.get("id") or 0)
		if order_id:
			order_lookup[order_id] = row
	order_ids = sorted(order_lookup)
	if not order_ids:
		return []
	select_fields = ",".join(
		[
			"order_id",
			"line_number",
			"product_name",
			"quantity",
			"discount_amount",
			"discount_name",
			"discount_name_normalized",
			"discount_bir_category",
			"discount_customer_full_name",
			"discount_customer_full_name_normalized",
			"discount_reference_number",
			"discount_reference_number_normalized",
		]
	)
	result: list[dict[str, Any]] = []
	for chunk in _chunk_values(order_ids):
		params: list[tuple[str, Any]] = [
			("select", select_fields),
			("discount_amount", "gt.0"),
			("order_id", f"in.({','.join(str(order_id) for order_id in chunk)})"),
			("order", "order_id.asc,line_number.asc"),
		]
		if categories:
			params.append(("discount_bir_category", f"in.({','.join(sorted(categories))})"))
		rows = _supabase_get_all("pos_order_items", params)
		for row in rows:
			order_id = int(row.get("order_id") or 0)
			order = order_lookup.get(order_id)
			if not order:
				continue
			location_id = int(order.get("location_id") or 0)
			category = _canonical_discount_category(
				row.get("discount_bir_category"),
				row.get("discount_name_normalized"),
				row.get("discount_name"),
			)
			if not category:
				continue
			if categories and category not in categories:
				continue
			result.append(
				{
					"location_id": location_id,
					"store_name": str(order.get("store_name") or location_id),
					"business_date": str(order.get("business_date") or ""),
					"order_id": order_id,
					"bill_number": str(order.get("bill_number") or ""),
					"receipt_number": str(order.get("receipt_number") or ""),
					"billed_at": order.get("billed_at"),
					"paid_at": order.get("paid_at"),
					"order_gross_sales": _to_number(order.get("original_gross_sales")),
					"order_net_sales_w_vat": _to_number(order.get("gross_sales")),
					"order_net_sales_wo_vat": _to_number(order.get("net_sales")),
					"order_total_discounts": _to_number(order.get("total_discounts")),
					"line_number": int(row.get("line_number") or 0),
					"product_name": str(row.get("product_name") or ""),
					"quantity": int(row.get("quantity") or 0),
					"discount_amount": _to_number(row.get("discount_amount")),
					"discount_name": str(row.get("discount_name") or ""),
					"discount_name_normalized": str(row.get("discount_name_normalized") or ""),
					"discount_bir_category": category,
					"discount_customer_full_name": str(row.get("discount_customer_full_name") or ""),
					"discount_customer_full_name_normalized": _normalize_text(
						row.get("discount_customer_full_name_normalized")
						or row.get("discount_customer_full_name")
					),
					"discount_reference_number": str(row.get("discount_reference_number") or ""),
					"discount_reference_number_normalized": _normalize_text(
						row.get("discount_reference_number_normalized")
						or row.get("discount_reference_number")
					),
				}
			)
	return result


def _min_gap_minutes(rows: list[dict[str, Any]]) -> int | None:
	timestamps = [
		_parse_iso_datetime(row.get("billed_at")) or _parse_iso_datetime(row.get("paid_at")) for row in rows
	]
	parsed = sorted(ts for ts in timestamps if ts is not None)
	if len(parsed) < 2:
		return None
	return min(int((right - left).total_seconds() // 60) for left, right in pairwise(parsed))


def _extract_alert_associations(row: dict[str, Any]) -> dict[str, list[str]]:
	details = _safe_json(row.get("details"))
	names = _dedupe_preserve_order(
		_split_pipe_values(row.get("customer_name"))
		+ _split_pipe_values(row.get("customer_names"))
		+ _split_pipe_values(details.get("customer_name"))
		+ _split_pipe_values(details.get("customer_names"))
	)
	references = _dedupe_preserve_order(
		_split_pipe_values(row.get("reference_number"))
		+ _split_pipe_values(row.get("reference_numbers"))
		+ _split_pipe_values(details.get("reference_number"))
		+ _split_pipe_values(details.get("reference_numbers"))
	)
	bills = _dedupe_preserve_order(
		_split_pipe_values(row.get("bill_numbers")) + _split_pipe_values(details.get("bill_numbers"))
	)
	business_dates = _dedupe_preserve_order(
		_split_pipe_values(row.get("business_dates")) + _split_pipe_values(details.get("business_dates"))
	)
	return {
		"names": names,
		"references": references,
		"bill_numbers": bills,
		"business_dates": business_dates,
	}


def _build_store_sales_summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
	active_days = {str(row.get("business_date") or "") for row in rows if row.get("business_date")}
	order_count = len(rows)
	gross_sales = sum(_to_number(row.get("original_gross_sales")) for row in rows)
	net_sales_w_vat = sum(_to_number(row.get("gross_sales")) for row in rows)
	net_sales_wo_vat = sum(_to_number(row.get("net_sales")) for row in rows)
	total_discounts = sum(_to_number(row.get("total_discounts")) for row in rows)
	return {
		"paid_orders": order_count,
		"active_days": len(active_days),
		"gross_sales": round(gross_sales, 2),
		"net_sales_w_vat": round(net_sales_w_vat, 2),
		"net_sales_wo_vat": round(net_sales_wo_vat, 2),
		"total_discounts": round(total_discounts, 2),
		"avg_order_gross": round(gross_sales / order_count, 2) if order_count else 0.0,
	}


def _rate_per_1000(count: int | float, denominator: int | float) -> float:
	if not denominator:
		return 0.0
	return round((float(count) / float(denominator)) * 1000, 2)


def _build_store_item_summary(
	rows: list[dict[str, Any]],
	selected_categories: set[str] | None,
) -> dict[str, Any]:
	order_ids = {int(row["order_id"]) for row in rows if row.get("order_id")}
	rows_with_name = [
		row for row in rows if _normalize_text(row.get("discount_customer_full_name_normalized"))
	]
	rows_with_reference = [
		row for row in rows if _normalize_text(row.get("discount_reference_number_normalized"))
	]
	breakdown_counter: Counter[str] = Counter(
		_normalize_text(row.get("discount_bir_category")) for row in rows if row.get("discount_bir_category")
	)
	label = "ALL" if selected_categories is None else ",".join(sorted(selected_categories))
	return {
		"category_label": label,
		"item_rows": len(rows),
		"orders": len(order_ids),
		"discount_total": round(sum(_to_number(row.get("discount_amount")) for row in rows), 2),
		"rows_with_name": len(rows_with_name),
		"rows_with_reference": len(rows_with_reference),
		"category_breakdown": dict(sorted(breakdown_counter.items())),
	}


def _build_store_contextual_metrics(
	rows: list[dict[str, Any]],
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
	orders: dict[tuple[str, int], dict[str, Any]] = {}
	for row in rows:
		key = (str(row.get("business_date") or ""), int(row.get("order_id") or 0))
		order_bucket = orders.setdefault(
			key,
			{
				"business_date": key[0],
				"order_id": key[1],
				"store_name": row.get("store_name"),
				"bill_numbers": set(),
				"names": set(),
				"references": set(),
				"discount_amount_total": 0.0,
			},
		)
		if row.get("bill_number"):
			order_bucket["bill_numbers"].add(str(row["bill_number"]))
		name = _normalize_text(row.get("discount_customer_full_name_normalized"))
		ref = _normalize_text(row.get("discount_reference_number_normalized"))
		if name:
			order_bucket["names"].add(name)
		if ref:
			order_bucket["references"].add(ref)
		order_bucket["discount_amount_total"] += _to_number(row.get("discount_amount"))

	name_groups: dict[str, dict[str, set[Any]]] = {}
	ref_groups: dict[str, dict[str, set[Any]]] = {}
	non_four_digit_reference_rows = 0
	for row in rows:
		name = _normalize_text(row.get("discount_customer_full_name_normalized"))
		ref = _normalize_text(row.get("discount_reference_number_normalized"))
		order_id = int(row.get("order_id") or 0)
		if ref and len(ref) != 4:
			non_four_digit_reference_rows += 1
		if name:
			name_bucket = name_groups.setdefault(name, {"references": set(), "orders": set()})
			if ref:
				name_bucket["references"].add(ref)
			name_bucket["orders"].add(order_id)
		if ref:
			ref_bucket = ref_groups.setdefault(ref, {"names": set(), "orders": set()})
			if name:
				ref_bucket["names"].add(name)
			ref_bucket["orders"].add(order_id)

	multi_name_receipts: list[dict[str, Any]] = []
	for order_bucket in orders.values():
		distinct_names = sorted(order_bucket["names"])
		distinct_references = sorted(order_bucket["references"])
		if len(distinct_names) >= 2:
			multi_name_receipts.append(
				{
					"business_date": order_bucket["business_date"],
					"order_id": order_bucket["order_id"],
					"store_name": order_bucket["store_name"],
					"bill_numbers": sorted(order_bucket["bill_numbers"]),
					"names": distinct_names,
					"references": distinct_references,
					"distinct_name_count": len(distinct_names),
					"discount_amount_total": round(order_bucket["discount_amount_total"], 2),
				}
			)

	order_count = len(orders)
	contextual_metrics = {
		"multi_name_receipts": len(multi_name_receipts),
		"multi_name_receipts_3plus": sum(1 for row in multi_name_receipts if row["distinct_name_count"] >= 3),
		"multi_name_receipts_5plus": sum(1 for row in multi_name_receipts if row["distinct_name_count"] >= 5),
		"max_names_on_single_receipt": max(
			(row["distinct_name_count"] for row in multi_name_receipts), default=0
		),
		"multi_name_receipts_pct_of_orders": round((len(multi_name_receipts) / order_count) * 100, 2)
		if order_count
		else 0.0,
		"month_name_multiple_reference_findings": sum(
			1 for bucket in name_groups.values() if len(bucket["references"]) >= 2
		),
		"month_reference_multi_name_findings": sum(
			1 for bucket in ref_groups.values() if len(bucket["names"]) >= 2
		),
		"non_four_digit_reference_rows": non_four_digit_reference_rows,
		"non_four_digit_reference_rate_pct": round((non_four_digit_reference_rows / len(rows)) * 100, 2)
		if rows
		else 0.0,
	}
	return contextual_metrics, multi_name_receipts


def _build_same_day_metrics(
	store_name: str, rows: list[dict[str, Any]]
) -> tuple[dict[str, Any], dict[str, Any]]:
	store_rows = [
		row
		for row in rows
		if _normalize_text(row.get("scope")) == "STORE"
		and _normalize_text(row.get("store_name")) == _normalize_text(store_name)
	]
	chain_rows = [
		row
		for row in rows
		if _normalize_text(row.get("scope")) != "STORE" and _row_mentions_store(row, store_name)
	]
	store_severity = Counter(_normalize_text(row.get("severity")) for row in store_rows)
	chain_severity = Counter(_normalize_text(row.get("severity")) for row in chain_rows)
	store_metrics = {
		"repeat_name_findings": sum(
			1 for row in store_rows if row.get("detection_type") == "same_name_same_day_same_store"
		),
		"repeat_reference_findings": sum(
			1 for row in store_rows if row.get("detection_type") == "same_reference_same_day_same_store"
		),
		"same_reference_different_name_findings": sum(
			1
			for row in store_rows
			if row.get("detection_type") == "same_reference_diff_name_same_day_same_store"
		),
		"same_name_multiple_reference_findings": sum(
			1
			for row in store_rows
			if row.get("detection_type") == "same_name_diff_reference_same_day_same_store"
		),
		"rapid_repeat_name_findings_4h": sum(
			1
			for row in store_rows
			if row.get("detection_type") == "same_name_same_day_same_store" and row.get("rapid_within_4h")
		),
		"rapid_repeat_reference_findings_4h": sum(
			1
			for row in store_rows
			if row.get("detection_type") == "same_reference_same_day_same_store"
			and row.get("rapid_within_4h")
		),
		"same_day_rows_by_severity": {
			"critical": int(store_severity.get("CRITICAL", 0)),
			"high": int(store_severity.get("HIGH", 0)),
			"medium": int(store_severity.get("MEDIUM", 0)),
			"review": int(store_severity.get("REVIEW", 0)),
		},
	}
	chain_metrics = {
		"same_day_chain_rows": len(chain_rows),
		"chain_rows_by_severity": {
			"critical": int(chain_severity.get("CRITICAL", 0)),
			"high": int(chain_severity.get("HIGH", 0)),
			"medium": int(chain_severity.get("MEDIUM", 0)),
			"review": int(chain_severity.get("REVIEW", 0)),
		},
	}
	return store_metrics, chain_metrics


def _build_investigation_summary_payload(
	start_day: date,
	end_day: date,
	selected_store_names: list[str],
	selected_categories: set[str] | None,
	same_day_rows: list[dict[str, Any]],
	item_rows: list[dict[str, Any]],
	paid_order_rows: list[dict[str, Any]],
) -> dict[str, Any]:
	store_summaries: list[dict[str, Any]] = []
	total_repeat_name = 0
	total_same_ref_diff_name = 0
	total_rapid_repeat_name = 0
	total_multi_name_receipts = 0
	total_paid_orders = 0
	for store_name in selected_store_names:
		store_order_rows = [
			row
			for row in paid_order_rows
			if _normalize_text(row.get("store_name")) == _normalize_text(store_name)
		]
		store_item_rows = [
			row for row in item_rows if _normalize_text(row.get("store_name")) == _normalize_text(store_name)
		]
		location_id = int(store_item_rows[0]["location_id"]) if store_item_rows else 0
		sales_summary = _build_store_sales_summary(store_order_rows)
		category_summary = _build_store_item_summary(store_item_rows, selected_categories)
		same_day_metrics, chain_metrics = _build_same_day_metrics(store_name, same_day_rows)
		contextual_metrics, _ = _build_store_contextual_metrics(store_item_rows)
		total_repeat_name += same_day_metrics["repeat_name_findings"]
		total_same_ref_diff_name += same_day_metrics["same_reference_different_name_findings"]
		total_rapid_repeat_name += same_day_metrics["rapid_repeat_name_findings_4h"]
		total_multi_name_receipts += contextual_metrics["multi_name_receipts"]
		total_paid_orders += int(sales_summary["paid_orders"])
		rates_per_1000 = {
			"repeat_name_findings": _rate_per_1000(
				same_day_metrics["repeat_name_findings"], sales_summary["paid_orders"]
			),
			"same_reference_different_name_findings": _rate_per_1000(
				same_day_metrics["same_reference_different_name_findings"],
				sales_summary["paid_orders"],
			),
			"rapid_repeat_name_findings_4h": _rate_per_1000(
				same_day_metrics["rapid_repeat_name_findings_4h"], sales_summary["paid_orders"]
			),
			"contextual_multi_name_receipts": _rate_per_1000(
				contextual_metrics["multi_name_receipts"], sales_summary["paid_orders"]
			),
		}
		store_summaries.append(
			{
				"location_id": location_id,
				"store_name": store_name,
				"sales": sales_summary,
				"category_summary": category_summary,
				"same_day_metrics": same_day_metrics,
				"chain_metrics": chain_metrics,
				"contextual_metrics": contextual_metrics,
				"rates_per_1000_paid_orders": rates_per_1000,
			}
		)

	return {
		"start_date": start_day.isoformat(),
		"end_date": end_day.isoformat(),
		"display_window": _format_display_window(start_day, end_day),
		"selected_store_names": selected_store_names,
		"discount_bir_category": ",".join(sorted(selected_categories)) if selected_categories else "ALL",
		"requires_store_selection": False,
		"methodology_notes": [
			"Same-day metrics are alert-derived from validated discount identity alerts.",
			"Multi-name receipt metrics are contextual receipt-level review metrics, not incident counts.",
			"Rates are normalized per 1,000 paid orders so stores with different sales volume remain comparable.",
		],
		"totals": {
			"same_day_repeat_name_findings": total_repeat_name,
			"same_day_same_reference_different_name_findings": total_same_ref_diff_name,
			"same_day_rapid_repeat_name_findings_4h": total_rapid_repeat_name,
			"contextual_multi_name_receipts": total_multi_name_receipts,
			"paid_orders": total_paid_orders,
			"rates_per_1000_paid_orders": {
				"same_day_repeat_name_findings": _rate_per_1000(total_repeat_name, total_paid_orders),
				"same_day_same_reference_different_name_findings": _rate_per_1000(
					total_same_ref_diff_name, total_paid_orders
				),
				"same_day_rapid_repeat_name_findings_4h": _rate_per_1000(
					total_rapid_repeat_name, total_paid_orders
				),
				"contextual_multi_name_receipts": _rate_per_1000(
					total_multi_name_receipts, total_paid_orders
				),
			},
		},
		"stores": store_summaries,
	}


def _build_investigation_case_rows(
	selected_store_names: list[str],
	same_day_rows: list[dict[str, Any]],
	item_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
	case_rows: list[dict[str, Any]] = []
	for row in _cluster_queue_rows(same_day_rows):
		if not any(_row_mentions_store(row, store_name) for store_name in selected_store_names):
			continue
		associations = {
			"names": list(row.get("customer_names") or []),
			"references": list(row.get("reference_numbers") or []),
			"bill_numbers": list(row.get("bill_numbers") or []),
			"business_dates": list(row.get("business_dates") or []),
		}
		case_rows.append(
			{
				"case_id": f"incident:{row.get('cluster_id')}",
				"case_bucket": "same_day_alert",
				"detection_type": row.get("detection_type"),
				"detection_types": row.get("detection_types") or [row.get("detection_type")],
				"scope": row.get("scope"),
				"store_name": row.get("store_name"),
				"business_date": row.get("event_date"),
				"severity": row.get("severity"),
				"identity_key": row.get("identity_key"),
				"names": associations["names"],
				"references": associations["references"],
				"bill_numbers": associations["bill_numbers"],
				"business_dates": associations["business_dates"] or [str(row.get("event_date") or "")],
				"order_count": int(row.get("order_count") or 0),
				"store_count": int(row.get("store_count") or 0),
				"discount_amount_total": round(_to_number(row.get("discount_amount_total")), 2),
				"rapid_within_4h": bool(row.get("rapid_within_4h")),
				"min_gap_minutes": row.get("min_gap_minutes"),
				"raw_row_count": int(row.get("raw_row_count") or 0),
				"context_note": "",
			}
		)

	contextual_metrics_by_store: dict[str, list[dict[str, Any]]] = {}
	for store_name in selected_store_names:
		store_item_rows = [
			row for row in item_rows if _normalize_text(row.get("store_name")) == _normalize_text(store_name)
		]
		_, multi_name_receipts = _build_store_contextual_metrics(store_item_rows)
		contextual_metrics_by_store[store_name] = multi_name_receipts

	for store_name, receipt_rows in contextual_metrics_by_store.items():
		for receipt in receipt_rows:
			if receipt["distinct_name_count"] < 3:
				continue
			case_rows.append(
				{
					"case_id": f"context:{receipt['business_date']}:{receipt['order_id']}",
					"case_bucket": "contextual_receipt",
					"detection_type": "context_multi_name_receipt",
					"detection_types": ["context_multi_name_receipt"],
					"scope": "store",
					"store_name": store_name,
					"business_date": receipt["business_date"],
					"severity": "high" if receipt["distinct_name_count"] >= 5 else "review",
					"identity_key": receipt["bill_numbers"][0]
					if receipt["bill_numbers"]
					else str(receipt["order_id"]),
					"names": receipt["names"],
					"references": receipt["references"],
					"bill_numbers": receipt["bill_numbers"],
					"business_dates": [receipt["business_date"]],
					"order_count": 1,
					"store_count": 1,
					"discount_amount_total": receipt["discount_amount_total"],
					"rapid_within_4h": False,
					"min_gap_minutes": None,
					"raw_row_count": 1,
					"context_note": "Contextual receipt metric - multiple distinct names on one discounted receipt.",
				}
			)

	case_rows.sort(
		key=lambda row: (
			SEVERITY_RANK.get(str(row.get("severity") or ""), 9),
			0 if row.get("case_bucket") == "same_day_alert" else 1,
			str(row.get("business_date") or ""),
			str(row.get("store_name") or ""),
			str("|".join(row.get("detection_types") or [str(row.get("detection_type") or "")])),
			str(row.get("identity_key") or ""),
		)
	)
	return case_rows


def _build_critical_notification_message(
	business_date: date,
	rows: list[dict[str, Any]],
) -> str:
	line_count = len(rows)
	store_names = sorted(
		{str(row.get("store_name") or "Unknown Store") for row in rows if row.get("store_name")}
	)
	header = [
		"Senior/PWD discount alert digest",
		f"Business date: {_format_display_date(business_date)}",
		f"{line_count} critical clusters require audit review",
	]
	if store_names:
		header.append("Stores: " + ", ".join(store_names))

	body = []
	for row in rows[:12]:
		amount = _to_number(row.get("discount_amount_total"))
		body.append(
			"- {store} | {detection} | {identity} | orders {orders} | amount PHP {amount:,.2f}".format(
				store=row.get("store_name") or "Unknown Store",
				detection=row.get("detection_type") or "unknown_detection",
				identity=row.get("identity_key") or "-",
				orders=int(row.get("order_count") or 0),
				amount=amount,
			)
		)

	footer = [
		"",
		f"Review queue: {PORTAL_QUEUE_URL}",
	]
	return "\n".join(header + body + footer)


def _set_header_style(worksheet, row_number: int = 1) -> None:
	from openpyxl.styles import Alignment, Font, PatternFill

	fill = PatternFill("solid", fgColor="1F4E78")
	font = Font(color="FFFFFF", bold=True)
	alignment = Alignment(vertical="center")
	for cell in worksheet[row_number]:
		cell.fill = fill
		cell.font = font
		cell.alignment = alignment


def _auto_size_columns(worksheet) -> None:
	for column_cells in worksheet.columns:
		max_length = 0
		column_letter = column_cells[0].column_letter
		for cell in column_cells:
			value = "" if cell.value is None else str(cell.value)
			max_length = max(max_length, len(value))
		worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)


def _write_tab(worksheet, rows: list[dict[str, Any]], columns: list[tuple[str, str]]) -> None:
	headers = [label for label, _ in columns]
	worksheet.append(headers)
	_set_header_style(worksheet)
	for row in rows:
		worksheet.append([row.get(key) for _, key in columns])
	worksheet.freeze_panes = "A2"
	_auto_size_columns(worksheet)


def _build_audit_workbook_bytes(
	business_date: date,
	same_day_rows: list[dict[str, Any]],
	rolling_rows: list[dict[str, Any]],
	summary: dict[str, int],
) -> bytes:
	import openpyxl

	workbook = openpyxl.Workbook()
	overview = workbook.active
	overview.title = "Overview"
	overview.append(["Metric", "Value"])
	_set_header_style(overview)
	overview_rows = [
		("Business Date", business_date.isoformat()),
		("Generated At PHT", _manila_now().strftime("%Y-%m-%d %H:%M:%S")),
		("Same-Day Critical", summary.get("same_day_critical", 0)),
		("Same-Day High", summary.get("same_day_high", 0)),
		("Same-Day Medium", summary.get("same_day_medium", 0)),
		("Rolling 30D High", summary.get("rolling_high", 0)),
		("Rolling 30D Review", summary.get("rolling_review", 0)),
		("Same-Day Rows", len(same_day_rows)),
		("Rolling Rows", len(rolling_rows)),
	]
	for metric, value in overview_rows:
		overview.append([metric, value])
	overview.freeze_panes = "A2"
	_auto_size_columns(overview)

	columns = [
		("Bucket", "queue_bucket"),
		("Event Date", "event_date"),
		("Store", "store_name"),
		("Severity", "severity"),
		("Category", "discount_bir_category"),
		("Detection", "detection_type"),
		("Identity Type", "identity_type"),
		("Identity Key", "identity_key"),
		("Customer Name", "customer_name"),
		("Reference No.", "reference_number"),
		("Order Count", "order_count"),
		("Store Count", "store_count"),
		("Rapid <4h", "rapid_within_4h"),
		("Min Gap Minutes", "min_gap_minutes"),
		("Discount Amount", "discount_amount_total"),
		("Window Start", "window_start"),
		("Window End", "window_end"),
	]
	rolling_columns = [
		("Bucket", "queue_bucket"),
		("Event Date", "event_date"),
		("Store", "store_name"),
		("Severity", "severity"),
		("Category", "discount_bir_category"),
		("Detection", "detection_type"),
		("Identity Type", "identity_type"),
		("Identity Key", "identity_key"),
		("Customer Name", "customer_name"),
		("Reference No.", "reference_number"),
		("Order Count", "order_count"),
		("Active Days", "active_day_count"),
		("Distinct Counterparties", "distinct_counterparty_count"),
		("Discount Amount", "discount_amount_total"),
		("Window Start", "window_start"),
		("Window End", "window_end"),
	]

	same_day_sheet = workbook.create_sheet("SameDay_All")
	_write_tab(same_day_sheet, same_day_rows, columns)

	same_day_critical = workbook.create_sheet("SameDay_Critical")
	_write_tab(
		same_day_critical,
		[row for row in same_day_rows if str(row.get("severity") or "") == "critical"],
		columns,
	)

	rolling_all = workbook.create_sheet("Rolling30D_All")
	_write_tab(rolling_all, rolling_rows, rolling_columns)

	rolling_high = workbook.create_sheet("Rolling30D_High")
	_write_tab(
		rolling_high,
		[row for row in rolling_rows if str(row.get("severity") or "") == "high"],
		rolling_columns,
	)

	buffer = io.BytesIO()
	workbook.save(buffer)
	return buffer.getvalue()


def _alert_filters(row: dict[str, Any]) -> dict[str, str]:
	return {
		"business_date": f"eq.{row['event_date']}",
		"scope": f"eq.{row['scope']}",
		"scope_key": f"eq.{int(row['scope_key'])}",
		"discount_bir_category": f"eq.{row['discount_bir_category']}",
		"identity_type": f"eq.{row['identity_type']}",
		"identity_key": f"eq.{row['identity_key']}",
		"detection_type": f"eq.{row['detection_type']}",
	}


def _mark_alerts_notified(rows: list[dict[str, Any]]) -> int:
	notified_at = _manila_now().isoformat()
	updated = 0
	for row in rows:
		result = _supabase_patch(
			"discount_abuse_alerts",
			_alert_filters(row),
			{"notified_at": notified_at, "updated_at": notified_at},
		)
		if result:
			updated += 1
	return updated


@frappe.whitelist()
def get_discount_audit_dashboard(business_date: str | None = None) -> dict[str, Any]:
	_check_discount_audit_role()
	target_day = _coerce_date(business_date, default=_default_business_date())
	same_day_rows = _query_same_day_rows(target_day)
	rolling_rows = _query_rolling_rows(target_day)
	same_day_clusters = _cluster_queue_rows(same_day_rows)
	rolling_clusters = _cluster_queue_rows(rolling_rows)
	summary = _make_summary(same_day_clusters, rolling_clusters)
	raw_summary = _make_summary(same_day_rows, rolling_rows)
	return {
		"success": True,
		"data": {
			"business_date": target_day.isoformat(),
			"display_business_date": _format_display_date(target_day),
			"summary": summary,
			"raw_summary": raw_summary,
			"same_day_total": len(same_day_clusters),
			"rolling_total": len(rolling_clusters),
			"same_day_raw_total": len(same_day_rows),
			"rolling_raw_total": len(rolling_rows),
			"critical_same_day_rows": same_day_rows[:10],
			"top_rolling_rows": rolling_rows[:10],
			"critical_same_day_clusters": [
				row for row in same_day_clusters if str(row.get("severity") or "") == "critical"
			][:10],
			"top_rolling_clusters": rolling_clusters[:10],
			"notifications_enabled": _notifications_enabled_for_day(target_day),
			"notification_go_live_date": _get_notification_go_live_date().isoformat(),
			"latest_reports": list_discount_audit_reports(limit=5).get("data", []),
		},
	}


@frappe.whitelist()
def get_discount_audit_queue(
	business_date: str | None = None,
	queue_bucket: str | None = None,
	severity: str | None = None,
	search: str | None = None,
	store_name: str | None = None,
	discount_bir_category: str | None = None,
) -> dict[str, Any]:
	_check_discount_audit_role()
	target_day = _coerce_date(business_date, default=_default_business_date())
	bucket = (queue_bucket or "all").strip()

	same_day_rows = _query_same_day_rows(target_day) if bucket in ("all", "same_day") else []
	rolling_rows = _query_rolling_rows(target_day) if bucket in ("all", "rolling_30d") else []

	same_day_rows = _filter_rows(
		_sort_same_day_rows(same_day_rows),
		search=search,
		severity=severity if bucket != "rolling_30d" else None,
		store_name=store_name,
		category=discount_bir_category,
	)
	rolling_rows = _filter_rows(
		_sort_rolling_rows(rolling_rows),
		search=search,
		severity=severity if bucket != "same_day" else None,
		store_name=store_name,
		category=discount_bir_category,
	)

	data = same_day_rows + rolling_rows
	stores = sorted(_get_store_directory().values())
	return {
		"success": True,
		"data": {
			"business_date": target_day.isoformat(),
			"rows": data,
			"summary": _make_summary(same_day_rows, rolling_rows),
			"stores": stores,
		},
	}


@frappe.whitelist()
def get_discount_audit_incident_queue(
	business_date: str | None = None,
	queue_bucket: str | None = None,
	severity: str | None = None,
	search: str | None = None,
	store_name: str | None = None,
	discount_bir_category: str | None = None,
) -> dict[str, Any]:
	_check_discount_audit_role()
	target_day = _coerce_date(business_date, default=_default_business_date())
	bucket = (queue_bucket or "all").strip()

	same_day_rows = _query_same_day_rows(target_day) if bucket in ("all", "same_day") else []
	rolling_rows = _query_rolling_rows(target_day) if bucket in ("all", "rolling_30d") else []
	same_day_clusters = _cluster_queue_rows(same_day_rows)
	rolling_clusters = _cluster_queue_rows(rolling_rows)
	same_day_clusters = _filter_incident_clusters(
		_sort_same_day_rows(same_day_clusters),
		search=search,
		severity=severity if bucket != "rolling_30d" else None,
		store_name=store_name,
		category=discount_bir_category,
	)
	rolling_clusters = _filter_incident_clusters(
		_sort_rolling_rows(rolling_clusters),
		search=search,
		severity=severity if bucket != "same_day" else None,
		store_name=store_name,
		category=discount_bir_category,
	)
	cluster_rows = same_day_clusters + rolling_clusters
	stores = sorted(_get_store_directory().values())
	return {
		"success": True,
		"data": {
			"business_date": target_day.isoformat(),
			"rows": cluster_rows,
			"summary": _make_summary(same_day_clusters, rolling_clusters),
			"raw_summary": _make_summary(same_day_rows, rolling_rows),
			"parity": {
				"same_day_raw_rows": len(same_day_rows),
				"same_day_clusters": len(_cluster_queue_rows(same_day_rows)),
				"rolling_raw_rows": len(rolling_rows),
				"rolling_clusters": len(_cluster_queue_rows(rolling_rows)),
			},
			"stores": stores,
		},
	}


def _get_investigation_payload(
	start_date: str | None,
	end_date: str | None,
	store_names: str | None,
	discount_bir_category: str | None,
) -> tuple[
	date, date, list[str], set[str] | None, list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]
]:
	start_day = _coerce_date(start_date)
	end_day = _coerce_date(end_date)
	if end_day < start_day:
		frappe.throw(frappe._("End date cannot be earlier than start date."))
	selected_store_names = _parse_store_names(store_names)
	selected_categories = _parse_category_filter(discount_bir_category)
	location_ids, selected_store_map = _resolve_selected_stores(selected_store_names)
	resolved_store_names = [selected_store_map[location_id] for location_id in location_ids]
	if not resolved_store_names:
		return start_day, end_day, [], selected_categories, [], [], []
	same_day_rows = _query_same_day_rows_range(start_day, end_day, selected_categories)
	paid_order_rows = _query_paid_orders_for_range(start_day, end_day, location_ids)
	store_map = {location_id: store_name for location_id, store_name in selected_store_map.items()}
	for row in paid_order_rows:
		location_id = int(row.get("location_id") or 0)
		row["store_name"] = store_map.get(location_id, str(location_id))
	item_rows = _query_discount_item_rows_for_orders(paid_order_rows, selected_categories)
	return (
		start_day,
		end_day,
		resolved_store_names,
		selected_categories,
		same_day_rows,
		item_rows,
		paid_order_rows,
	)


@frappe.whitelist()
def get_discount_investigation_summary(
	start_date: str,
	end_date: str,
	store_names: str | None = None,
	discount_bir_category: str | None = "SC",
) -> dict[str, Any]:
	_check_discount_audit_role()
	(
		start_day,
		end_day,
		resolved_store_names,
		selected_categories,
		same_day_rows,
		item_rows,
		paid_order_rows,
	) = _get_investigation_payload(start_date, end_date, store_names, discount_bir_category)
	if not resolved_store_names:
		return {
			"success": True,
			"data": {
				"start_date": start_day.isoformat(),
				"end_date": end_day.isoformat(),
				"display_window": _format_display_window(start_day, end_day),
				"discount_bir_category": ",".join(sorted(selected_categories))
				if selected_categories
				else "ALL",
				"selected_store_names": [],
				"requires_store_selection": True,
				"totals": {},
				"stores": [],
				"methodology_notes": [
					"Select at least one store to load investigation analytics.",
				],
			},
		}
	return {
		"success": True,
		"data": _build_investigation_summary_payload(
			start_day,
			end_day,
			resolved_store_names,
			selected_categories,
			same_day_rows,
			item_rows,
			paid_order_rows,
		),
	}


@frappe.whitelist()
def get_discount_investigation_cases(
	start_date: str,
	end_date: str,
	store_names: str | None = None,
	discount_bir_category: str | None = "SC",
) -> dict[str, Any]:
	_check_discount_audit_role()
	(
		start_day,
		end_day,
		resolved_store_names,
		selected_categories,
		same_day_rows,
		item_rows,
		_paid_order_rows,
	) = _get_investigation_payload(start_date, end_date, store_names, discount_bir_category)
	if not resolved_store_names:
		return {
			"success": True,
			"data": {
				"start_date": start_day.isoformat(),
				"end_date": end_day.isoformat(),
				"display_window": _format_display_window(start_day, end_day),
				"discount_bir_category": ",".join(sorted(selected_categories))
				if selected_categories
				else "ALL",
				"selected_store_names": [],
				"requires_store_selection": True,
				"summary": {"same_day_alert_cases": 0, "contextual_cases": 0},
				"rows": [],
			},
		}
	case_rows = _build_investigation_case_rows(resolved_store_names, same_day_rows, item_rows)
	return {
		"success": True,
		"data": {
			"start_date": start_day.isoformat(),
			"end_date": end_day.isoformat(),
			"display_window": _format_display_window(start_day, end_day),
			"discount_bir_category": ",".join(sorted(selected_categories)) if selected_categories else "ALL",
			"selected_store_names": resolved_store_names,
			"requires_store_selection": False,
			"summary": {
				"same_day_alert_cases": sum(1 for row in case_rows if row["case_bucket"] == "same_day_alert"),
				"contextual_cases": sum(1 for row in case_rows if row["case_bucket"] == "contextual_receipt"),
			},
			"sections": {
				"selected_store_incidents": [
					row
					for row in case_rows
					if row["case_bucket"] == "same_day_alert" and _normalize_text(row.get("scope")) == "STORE"
				],
				"cross_store_overlaps": [
					row
					for row in case_rows
					if row["case_bucket"] == "same_day_alert" and _normalize_text(row.get("scope")) != "STORE"
				],
				"contextual_receipt_review": [
					row for row in case_rows if row["case_bucket"] == "contextual_receipt"
				],
			},
			"rows": case_rows,
		},
	}


def _category_prefix(category: str) -> str:
	return category.lower()


def _category_field(category: str, suffix: str) -> str:
	return f"{_category_prefix(category)}_{suffix}"


def _base_snapshot_row(
	*,
	scope: str,
	location_id: int,
	store_name: str,
	business_date: date | None = None,
	month_start: date | None = None,
	legal_entity: str = "",
	store_type: str = "",
	peer_group_key: str = "",
) -> dict[str, Any]:
	row = {
		"scope": scope,
		"location_id": int(location_id),
		"store_name": store_name,
		"legal_entity": legal_entity,
		"store_type": store_type,
		"peer_group_key": peer_group_key,
		"business_date": business_date.isoformat() if business_date else None,
		"month_start": month_start.isoformat() if month_start else None,
		"active_days": 0,
		"paid_orders": 0,
		"pos_original_gross_sales": 0.0,
		"pos_post_discount_gross_sales": 0.0,
		"pos_net_sales_without_vat": 0.0,
		"pos_total_discounts": 0.0,
		"all_channel_total_orders": 0,
		"all_channel_gross_sales": 0.0,
		"statutory_recorded_discount_amount": 0.0,
		"statutory_vat_relief_total": 0.0,
		"effective_statutory_benefit_total": 0.0,
	}
	for category in STATUTORY_CATEGORIES:
		row[_category_field(category, "recorded_discount_amount")] = 0.0
		row[_category_field(category, "statutory_vat_relief")] = 0.0
		row[_category_field(category, "effective_statutory_benefit")] = 0.0
		row[_category_field(category, "repeat_name_findings")] = 0
		row[_category_field(category, "repeat_reference_findings")] = 0
		row[_category_field(category, "same_reference_different_name_findings")] = 0
		row[_category_field(category, "same_name_multiple_reference_findings")] = 0
		row[_category_field(category, "rapid_repeat_name_findings_4h")] = 0
		row[_category_field(category, "rapid_repeat_reference_findings_4h")] = 0
		row[_category_field(category, "cross_store_overlap_count")] = 0
		row[_category_field(category, "cross_store_overlap_critical")] = 0
		row[_category_field(category, "cross_store_overlap_high")] = 0
		row[_category_field(category, "same_day_alert_cases")] = 0
		row[_category_field(category, "high_confidence_same_day_incidents")] = 0
		row[_category_field(category, "same_day_flagged_discount_total")] = 0.0
		row[_category_field(category, "weighted_risk_score")] = 0.0
		row[_category_field(category, "weighted_risk_rate")] = 0.0
	return row


def _peer_group_key_for(store_type: str, legal_entity: str) -> str:
	normalized_store_type = _normalize_text(store_type)
	normalized_legal_entity = _normalize_text(legal_entity)
	if normalized_store_type and normalized_legal_entity:
		return f"{normalized_store_type}::{normalized_legal_entity}"
	if normalized_store_type:
		return normalized_store_type
	return CHAIN_PLACEHOLDER


def _sum_money_fields(target: dict[str, Any], source: dict[str, Any], fields: list[str]) -> None:
	for field in fields:
		target[field] = _to_number(target.get(field)) + _to_number(source.get(field))


def _sum_count_fields(target: dict[str, Any], source: dict[str, Any], fields: list[str]) -> None:
	for field in fields:
		target[field] = int(target.get(field) or 0) + int(source.get(field) or 0)


def _build_order_finance_buckets(
	paid_order_rows: list[dict[str, Any]],
	item_rows: list[dict[str, Any]],
) -> dict[int, dict[str, Any]]:
	buckets: dict[int, dict[str, Any]] = {}
	for order in paid_order_rows:
		order_id = int(order.get("id") or 0)
		if not order_id:
			continue
		buckets[order_id] = {
			"order_id": order_id,
			"location_id": int(order.get("location_id") or 0),
			"store_name": str(order.get("store_name") or order.get("location_id") or ""),
			"business_date": str(order.get("business_date") or ""),
			"gross_gap_before_vat": max(
				0.0,
				_to_number(order.get("original_gross_sales")) - _to_number(order.get("gross_sales")),
			),
			"order_total_discounts": _to_number(order.get("total_discounts")),
			"discounts": {category: 0.0 for category in STATUTORY_CATEGORIES},
			"vat_relief": {category: 0.0 for category in STATUTORY_CATEGORIES},
			"effective_benefit": {category: 0.0 for category in STATUTORY_CATEGORIES},
		}
	for row in item_rows:
		order_id = int(row.get("order_id") or 0)
		category = _canonical_discount_category(
			row.get("discount_bir_category"),
			row.get("discount_name_normalized"),
			row.get("discount_name"),
		)
		if not order_id or category not in STATUTORY_CATEGORIES:
			continue
		bucket = buckets.get(order_id)
		if not bucket:
			continue
		bucket["discounts"][category] += _to_number(row.get("discount_amount"))
	for bucket in buckets.values():
		statutory_discount_total = sum(bucket["discounts"].values())
		order_total_discounts = max(0.0, _to_number(bucket.get("order_total_discounts")))
		gross_gap = max(0.0, _to_number(bucket.get("gross_gap_before_vat")))
		total_vat_relief = max(0.0, gross_gap - order_total_discounts)
		allocatable_vat_relief = 0.0
		if statutory_discount_total > 0 and order_total_discounts > 0:
			allocatable_vat_relief = total_vat_relief * min(
				1.0,
				statutory_discount_total / order_total_discounts,
			)
		for category in STATUTORY_CATEGORIES:
			discount_amount = bucket["discounts"][category]
			share = (discount_amount / statutory_discount_total) if statutory_discount_total else 0.0
			bucket["vat_relief"][category] = allocatable_vat_relief * share
			bucket["effective_benefit"][category] = (
				bucket["discounts"][category] + bucket["vat_relief"][category]
			)
	return buckets


def _build_sales_snapshot_from_orders(order_rows: list[dict[str, Any]]) -> dict[str, Any]:
	sales = _build_store_sales_summary(order_rows)
	return {
		"paid_orders": int(sales["paid_orders"]),
		"pos_original_gross_sales": _to_number(sales["gross_sales"]),
		"pos_post_discount_gross_sales": _to_number(sales["net_sales_w_vat"]),
		"pos_net_sales_without_vat": _to_number(sales["net_sales_wo_vat"]),
		"pos_total_discounts": _to_number(sales["total_discounts"]),
	}


def _category_metrics_for_store(
	*,
	category: str,
	store_name: str,
	paid_order_count: int,
	store_order_ids: set[int],
	same_day_rows: list[dict[str, Any]],
	same_day_clusters: list[dict[str, Any]],
	order_buckets: dict[int, dict[str, Any]],
) -> dict[str, Any]:
	scoped_same_day_rows = [
		row for row in same_day_rows if _normalize_text(row.get("discount_bir_category")) == category
	]
	same_day_metrics, chain_metrics = _build_same_day_metrics(store_name, scoped_same_day_rows)
	store_clusters = [
		row
		for row in same_day_clusters
		if _normalize_text(row.get("discount_bir_category")) == category
		and _normalize_text(row.get("scope")) == "STORE"
		and _row_mentions_store(row, store_name)
	]
	recorded_discount_amount = 0.0
	statutory_vat_relief = 0.0
	effective_statutory_benefit = 0.0
	for order_id in store_order_ids:
		bucket = order_buckets.get(order_id)
		if not bucket:
			continue
		recorded_discount_amount += bucket["discounts"][category]
		statutory_vat_relief += bucket["vat_relief"][category]
		effective_statutory_benefit += bucket["effective_benefit"][category]
	repeat_name_rate = _rate_per_1000(same_day_metrics["repeat_name_findings"], paid_order_count)
	repeat_reference_rate = _rate_per_1000(same_day_metrics["repeat_reference_findings"], paid_order_count)
	same_reference_different_name_rate = _rate_per_1000(
		same_day_metrics["same_reference_different_name_findings"], paid_order_count
	)
	rapid_repeat_name_rate = _rate_per_1000(
		same_day_metrics["rapid_repeat_name_findings_4h"], paid_order_count
	)
	weighted_risk_score = (
		(same_day_metrics["repeat_name_findings"] * WEIGHTED_SCORE_WEIGHTS["repeat_name"])
		+ (same_day_metrics["repeat_reference_findings"] * WEIGHTED_SCORE_WEIGHTS["repeat_reference"])
		+ (
			same_day_metrics["same_reference_different_name_findings"]
			* WEIGHTED_SCORE_WEIGHTS["same_reference_different_name"]
		)
		+ (same_day_metrics["rapid_repeat_name_findings_4h"] * WEIGHTED_SCORE_WEIGHTS["rapid_repeat_name"])
		+ (
			int(chain_metrics["chain_rows_by_severity"].get("critical", 0))
			* WEIGHTED_SCORE_WEIGHTS["cross_store_overlap_critical"]
		)
	)
	return {
		"recorded_discount_amount": _round_metric(recorded_discount_amount, 6),
		"statutory_vat_relief": _round_metric(statutory_vat_relief, 6),
		"effective_statutory_benefit": _round_metric(effective_statutory_benefit, 6),
		"repeat_name_findings": int(same_day_metrics["repeat_name_findings"]),
		"repeat_reference_findings": int(same_day_metrics["repeat_reference_findings"]),
		"same_reference_different_name_findings": int(
			same_day_metrics["same_reference_different_name_findings"]
		),
		"same_name_multiple_reference_findings": int(
			same_day_metrics["same_name_multiple_reference_findings"]
		),
		"rapid_repeat_name_findings_4h": int(same_day_metrics["rapid_repeat_name_findings_4h"]),
		"rapid_repeat_reference_findings_4h": int(same_day_metrics["rapid_repeat_reference_findings_4h"]),
		"cross_store_overlap_count": int(chain_metrics["same_day_chain_rows"]),
		"cross_store_overlap_critical": int(chain_metrics["chain_rows_by_severity"].get("critical", 0)),
		"cross_store_overlap_high": int(chain_metrics["chain_rows_by_severity"].get("high", 0)),
		"same_day_alert_cases": len(store_clusters),
		"high_confidence_same_day_incidents": sum(
			1 for row in store_clusters if str(row.get("severity") or "") in HIGH_CONFIDENCE_SEVERITIES
		),
		"same_day_flagged_discount_total": _round_metric(
			sum(_to_number(row.get("discount_amount_total")) for row in store_clusters), 6
		),
		"weighted_risk_score": _round_metric(weighted_risk_score, 6),
		"weighted_risk_rate": _round_metric(
			(repeat_name_rate * WEIGHTED_RATE_WEIGHTS["repeat_name"])
			+ (repeat_reference_rate * WEIGHTED_RATE_WEIGHTS["repeat_reference"])
			+ (same_reference_different_name_rate * WEIGHTED_RATE_WEIGHTS["same_reference_different_name"])
			+ (rapid_repeat_name_rate * WEIGHTED_RATE_WEIGHTS["rapid_repeat_name"]),
			6,
		),
	}


def _set_category_snapshot_metrics(target: dict[str, Any], category: str, metrics: dict[str, Any]) -> None:
	for suffix, value in metrics.items():
		target[_category_field(category, suffix)] = value


def _build_store_day_snapshot_rows(target_day: date) -> list[dict[str, Any]]:
	catalog = _get_store_catalog()
	location_ids = sorted(catalog)
	store_daily_rows = _query_store_daily_closing_rows(target_day, target_day)
	store_daily_map = {
		int(row.get("location_id") or 0): row for row in store_daily_rows if row.get("location_id")
	}
	all_channel_map = {
		str(row.get("business_date") or ""): row
		for row in _query_all_channel_daily_rows(target_day, target_day)
	}
	paid_order_rows = _query_paid_orders_for_range(target_day, target_day, location_ids)
	for row in paid_order_rows:
		location_id = int(row.get("location_id") or 0)
		row["store_name"] = catalog.get(location_id, {}).get("store_name", str(location_id))
	item_rows = _query_discount_item_rows_for_orders(paid_order_rows, set(STATUTORY_CATEGORIES))
	same_day_rows = _query_same_day_rows_range(target_day, target_day, set(STATUTORY_CATEGORIES))
	same_day_clusters = _cluster_queue_rows(same_day_rows)
	order_buckets = _build_order_finance_buckets(paid_order_rows, item_rows)
	order_rows_by_location: dict[int, list[dict[str, Any]]] = {}
	order_ids_by_location: dict[int, set[int]] = {}
	for row in paid_order_rows:
		location_id = int(row.get("location_id") or 0)
		order_rows_by_location.setdefault(location_id, []).append(row)
		order_id = int(row.get("id") or 0)
		if order_id:
			order_ids_by_location.setdefault(location_id, set()).add(order_id)

	candidate_location_ids = sorted(set(store_daily_map) | set(order_rows_by_location))
	results: list[dict[str, Any]] = []
	for location_id in candidate_location_ids:
		meta = catalog.get(location_id, {})
		store_name = str(meta.get("store_name") or location_id)
		daily_row = store_daily_map.get(location_id)
		order_rows = order_rows_by_location.get(location_id, [])
		sales_snapshot = (
			{
				"paid_orders": int(daily_row.get("pos_orders") or 0),
				"pos_original_gross_sales": _to_number(daily_row.get("pos_original_gross")),
				"pos_post_discount_gross_sales": _to_number(daily_row.get("pos_after_discount")),
				"pos_net_sales_without_vat": _to_number(daily_row.get("pos_net_of_vat")),
				"pos_total_discounts": _to_number(daily_row.get("pos_discounts")),
			}
			if daily_row
			else _build_sales_snapshot_from_orders(order_rows)
		)
		row = _base_snapshot_row(
			scope=STORE_SCOPE,
			location_id=location_id,
			store_name=store_name,
			business_date=target_day,
			month_start=_month_start(target_day),
			legal_entity=str(meta.get("legal_entity") or daily_row.get("legal_entity") or ""),
			store_type=str(meta.get("store_type") or daily_row.get("store_type") or ""),
			peer_group_key=_peer_group_key_for(
				str(meta.get("store_type") or daily_row.get("store_type") or ""),
				str(meta.get("legal_entity") or daily_row.get("legal_entity") or ""),
			),
		)
		row["active_days"] = (
			1 if (sales_snapshot["paid_orders"] or order_ids_by_location.get(location_id)) else 0
		)
		row.update(sales_snapshot)
		store_order_ids = order_ids_by_location.get(location_id, set())
		for category in STATUTORY_CATEGORIES:
			metrics = _category_metrics_for_store(
				category=category,
				store_name=store_name,
				paid_order_count=int(row["paid_orders"]),
				store_order_ids=store_order_ids,
				same_day_rows=same_day_rows,
				same_day_clusters=same_day_clusters,
				order_buckets=order_buckets,
			)
			_set_category_snapshot_metrics(row, category, metrics)
		row["statutory_recorded_discount_amount"] = _round_metric(
			_to_number(row["sc_recorded_discount_amount"]) + _to_number(row["pwd_recorded_discount_amount"]),
			6,
		)
		row["statutory_vat_relief_total"] = _round_metric(
			_to_number(row["sc_statutory_vat_relief"]) + _to_number(row["pwd_statutory_vat_relief"]),
			6,
		)
		row["effective_statutory_benefit_total"] = _round_metric(
			_to_number(row["sc_effective_statutory_benefit"])
			+ _to_number(row["pwd_effective_statutory_benefit"]),
			6,
		)
		results.append(row)

	chain_row = _base_snapshot_row(
		scope=CHAIN_SCOPE,
		location_id=CHAIN_LOCATION_ID,
		store_name=CHAIN_STORE_NAME,
		business_date=target_day,
		month_start=_month_start(target_day),
		legal_entity=CHAIN_PLACEHOLDER,
		store_type=CHAIN_PLACEHOLDER,
		peer_group_key=CHAIN_PLACEHOLDER,
	)
	chain_row["active_days"] = 1 if results else 0
	for store_row in results:
		_sum_count_fields(chain_row, store_row, ["paid_orders"])
		_sum_money_fields(
			chain_row,
			store_row,
			[
				"pos_original_gross_sales",
				"pos_post_discount_gross_sales",
				"pos_net_sales_without_vat",
				"pos_total_discounts",
				"statutory_recorded_discount_amount",
				"statutory_vat_relief_total",
				"effective_statutory_benefit_total",
			],
		)
		for category in STATUTORY_CATEGORIES:
			_sum_count_fields(
				chain_row,
				store_row,
				[
					_category_field(category, "repeat_name_findings"),
					_category_field(category, "repeat_reference_findings"),
					_category_field(category, "same_reference_different_name_findings"),
					_category_field(category, "same_name_multiple_reference_findings"),
					_category_field(category, "rapid_repeat_name_findings_4h"),
					_category_field(category, "rapid_repeat_reference_findings_4h"),
					_category_field(category, "cross_store_overlap_count"),
					_category_field(category, "cross_store_overlap_critical"),
					_category_field(category, "cross_store_overlap_high"),
					_category_field(category, "same_day_alert_cases"),
					_category_field(category, "high_confidence_same_day_incidents"),
				],
			)
			_sum_money_fields(
				chain_row,
				store_row,
				[
					_category_field(category, "recorded_discount_amount"),
					_category_field(category, "statutory_vat_relief"),
					_category_field(category, "effective_statutory_benefit"),
					_category_field(category, "same_day_flagged_discount_total"),
					_category_field(category, "weighted_risk_score"),
				],
			)
	daily_all_channel = all_channel_map.get(target_day.isoformat(), {})
	chain_row["all_channel_total_orders"] = int(daily_all_channel.get("total_orders") or 0)
	chain_row["all_channel_gross_sales"] = _to_number(daily_all_channel.get("total_gross_sales"))
	results.append(chain_row)
	return results


def _aggregate_snapshot_rows_for_period(
	rows: list[dict[str, Any]],
	*,
	start_day: date,
	end_day: date,
) -> list[dict[str, Any]]:
	grouped: dict[tuple[str, int], dict[str, Any]] = {}
	active_days: dict[tuple[str, int], set[str]] = {}
	for source in rows:
		scope = str(source.get("scope") or STORE_SCOPE).lower()
		location_id = int(source.get("location_id") or 0)
		key = (scope, location_id)
		target = grouped.setdefault(
			key,
			_base_snapshot_row(
				scope=scope,
				location_id=location_id,
				store_name=str(source.get("store_name") or location_id),
				month_start=_month_start(start_day),
				legal_entity=str(source.get("legal_entity") or ""),
				store_type=str(source.get("store_type") or ""),
				peer_group_key=str(source.get("peer_group_key") or ""),
			),
		)
		active_days.setdefault(key, set()).add(str(source.get("business_date") or ""))
		_sum_count_fields(target, source, ["paid_orders", "all_channel_total_orders"])
		_sum_money_fields(
			target,
			source,
			[
				"pos_original_gross_sales",
				"pos_post_discount_gross_sales",
				"pos_net_sales_without_vat",
				"pos_total_discounts",
				"all_channel_gross_sales",
				"statutory_recorded_discount_amount",
				"statutory_vat_relief_total",
				"effective_statutory_benefit_total",
			],
		)
		for category in STATUTORY_CATEGORIES:
			_sum_count_fields(
				target,
				source,
				[
					_category_field(category, "repeat_name_findings"),
					_category_field(category, "repeat_reference_findings"),
					_category_field(category, "same_reference_different_name_findings"),
					_category_field(category, "same_name_multiple_reference_findings"),
					_category_field(category, "rapid_repeat_name_findings_4h"),
					_category_field(category, "rapid_repeat_reference_findings_4h"),
					_category_field(category, "cross_store_overlap_count"),
					_category_field(category, "cross_store_overlap_critical"),
					_category_field(category, "cross_store_overlap_high"),
					_category_field(category, "same_day_alert_cases"),
					_category_field(category, "high_confidence_same_day_incidents"),
				],
			)
			_sum_money_fields(
				target,
				source,
				[
					_category_field(category, "recorded_discount_amount"),
					_category_field(category, "statutory_vat_relief"),
					_category_field(category, "effective_statutory_benefit"),
					_category_field(category, "same_day_flagged_discount_total"),
					_category_field(category, "weighted_risk_score"),
				],
			)
		target["business_date"] = end_day.isoformat()
	for row in grouped.values():
		key = (str(row.get("scope") or STORE_SCOPE), int(row.get("location_id") or 0))
		row["active_days"] = len({value for value in active_days.get(key, set()) if value})
	return list(grouped.values())


def _is_single_full_month_window(start_day: date, end_day: date) -> bool:
	return (
		start_day.year == end_day.year
		and start_day.month == end_day.month
		and start_day == _month_start(start_day)
		and end_day == _month_end(start_day)
	)


def _select_snapshot_rows_for_window(
	start_day: date,
	end_day: date,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
	if _is_single_full_month_window(start_day, end_day):
		month_rows = _query_store_month_snapshots(_month_start(start_day))
		if month_rows:
			return month_rows, {
				"snapshot_source": "month",
				"period_granularity": "month",
			}

	day_rows = _query_store_day_snapshots(start_day, end_day)
	if not day_rows:
		return [], {
			"snapshot_source": "day",
			"period_granularity": "day",
		}
	if start_day == end_day:
		return day_rows, {
			"snapshot_source": "day",
			"period_granularity": "day",
		}
	return (
		_aggregate_snapshot_rows_for_period(day_rows, start_day=start_day, end_day=end_day),
		{
			"snapshot_source": "day",
			"period_granularity": "range",
		},
	)


def _category_totals_from_row(row: dict[str, Any], category_scope: str) -> dict[str, float | int]:
	categories = list(STATUTORY_CATEGORIES) if category_scope == "BOTH" else [category_scope]
	totals: dict[str, float | int] = {
		"recorded_discount_amount": 0.0,
		"statutory_vat_relief": 0.0,
		"effective_statutory_benefit": 0.0,
		"repeat_name_findings": 0,
		"repeat_reference_findings": 0,
		"same_reference_different_name_findings": 0,
		"same_name_multiple_reference_findings": 0,
		"rapid_repeat_name_findings_4h": 0,
		"rapid_repeat_reference_findings_4h": 0,
		"cross_store_overlap_count": 0,
		"cross_store_overlap_critical": 0,
		"cross_store_overlap_high": 0,
		"same_day_alert_cases": 0,
		"high_confidence_same_day_incidents": 0,
		"same_day_flagged_discount_total": 0.0,
		"weighted_risk_score": 0.0,
		"weighted_risk_rate": 0.0,
	}
	for category in categories:
		totals["recorded_discount_amount"] = _to_number(totals["recorded_discount_amount"]) + _to_number(
			row.get(_category_field(category, "recorded_discount_amount"))
		)
		totals["statutory_vat_relief"] = _to_number(totals["statutory_vat_relief"]) + _to_number(
			row.get(_category_field(category, "statutory_vat_relief"))
		)
		totals["effective_statutory_benefit"] = _to_number(
			totals["effective_statutory_benefit"]
		) + _to_number(row.get(_category_field(category, "effective_statutory_benefit")))
		for count_suffix in [
			"repeat_name_findings",
			"repeat_reference_findings",
			"same_reference_different_name_findings",
			"same_name_multiple_reference_findings",
			"rapid_repeat_name_findings_4h",
			"rapid_repeat_reference_findings_4h",
			"cross_store_overlap_count",
			"cross_store_overlap_critical",
			"cross_store_overlap_high",
			"same_day_alert_cases",
			"high_confidence_same_day_incidents",
		]:
			totals[count_suffix] = int(totals[count_suffix] or 0) + int(
				row.get(_category_field(category, count_suffix)) or 0
			)
		totals["same_day_flagged_discount_total"] = _to_number(
			totals["same_day_flagged_discount_total"]
		) + _to_number(row.get(_category_field(category, "same_day_flagged_discount_total")))
		totals["weighted_risk_score"] = _to_number(totals["weighted_risk_score"]) + _to_number(
			row.get(_category_field(category, "weighted_risk_score"))
		)
		totals["weighted_risk_rate"] = _to_number(totals["weighted_risk_rate"]) + _to_number(
			row.get(_category_field(category, "weighted_risk_rate"))
		)
	return totals


def _effective_denominator_scope(row: dict[str, Any], denominator_scope: str) -> str:
	requested = _normalize_denominator_scope(denominator_scope)
	row_scope = _normalize_text(row.get("scope"))
	if requested != "all_channel_gross":
		return requested
	if row_scope != "CHAIN" or _to_number(row.get("all_channel_gross_sales")) <= 0:
		return "pos_original_gross"
	return "all_channel_gross"


def _row_denominator_value(row: dict[str, Any], denominator_scope: str) -> float:
	applied_scope = _effective_denominator_scope(row, denominator_scope)
	if applied_scope == "pos_post_discount_gross":
		return _to_number(row.get("pos_post_discount_gross_sales"))
	if applied_scope == "all_channel_gross":
		return _to_number(row.get("all_channel_gross_sales"))
	return _to_number(row.get("pos_original_gross_sales"))


def _build_metric_row(
	row: dict[str, Any],
	*,
	denominator_scope: str,
	category_scope: str,
) -> dict[str, Any]:
	selected = _category_totals_from_row(row, category_scope)
	applied_denominator_scope = _effective_denominator_scope(row, denominator_scope)
	denominator_value = _row_denominator_value(row, denominator_scope)
	paid_orders = int(row.get("paid_orders") or 0)
	return {
		"scope": str(row.get("scope") or STORE_SCOPE),
		"location_id": int(row.get("location_id") or 0),
		"store_name": str(row.get("store_name") or row.get("location_id") or ""),
		"legal_entity": str(row.get("legal_entity") or ""),
		"store_type": str(row.get("store_type") or ""),
		"peer_group_key": str(row.get("peer_group_key") or ""),
		"business_date": str(row.get("business_date") or ""),
		"month_start": str(row.get("month_start") or ""),
		"active_days": int(row.get("active_days") or 0),
		"paid_orders": paid_orders,
		"pos_original_gross_sales": _to_number(row.get("pos_original_gross_sales")),
		"pos_post_discount_gross_sales": _to_number(row.get("pos_post_discount_gross_sales")),
		"pos_net_sales_without_vat": _to_number(row.get("pos_net_sales_without_vat")),
		"pos_total_discounts": _to_number(row.get("pos_total_discounts")),
		"all_channel_total_orders": int(row.get("all_channel_total_orders") or 0),
		"all_channel_gross_sales": _to_number(row.get("all_channel_gross_sales")),
		"sc_recorded_discount_amount": _to_number(row.get("sc_recorded_discount_amount")),
		"pwd_recorded_discount_amount": _to_number(row.get("pwd_recorded_discount_amount")),
		"sc_statutory_vat_relief": _to_number(row.get("sc_statutory_vat_relief")),
		"pwd_statutory_vat_relief": _to_number(row.get("pwd_statutory_vat_relief")),
		"effective_denominator_scope": applied_denominator_scope,
		"denominator_value": denominator_value,
		"recorded_sc_pct_of_sales": _pct(
			_to_number(row.get("sc_recorded_discount_amount")), denominator_value
		),
		"recorded_pwd_pct_of_sales": _pct(
			_to_number(row.get("pwd_recorded_discount_amount")), denominator_value
		),
		"selected_recorded_discount_amount": _to_number(selected["recorded_discount_amount"]),
		"selected_statutory_vat_relief": _to_number(selected["statutory_vat_relief"]),
		"selected_effective_statutory_benefit": _to_number(selected["effective_statutory_benefit"]),
		"effective_statutory_benefit_pct": _pct(
			_to_number(selected["effective_statutory_benefit"]), denominator_value
		),
		"repeat_name_findings": int(selected["repeat_name_findings"] or 0),
		"repeat_reference_findings": int(selected["repeat_reference_findings"] or 0),
		"same_reference_different_name_findings": int(
			selected["same_reference_different_name_findings"] or 0
		),
		"same_name_multiple_reference_findings": int(selected["same_name_multiple_reference_findings"] or 0),
		"rapid_repeat_name_findings_4h": int(selected["rapid_repeat_name_findings_4h"] or 0),
		"rapid_repeat_reference_findings_4h": int(selected["rapid_repeat_reference_findings_4h"] or 0),
		"cross_store_overlap_count": int(selected["cross_store_overlap_count"] or 0),
		"cross_store_overlap_critical": int(selected["cross_store_overlap_critical"] or 0),
		"cross_store_overlap_high": int(selected["cross_store_overlap_high"] or 0),
		"same_day_alert_cases": int(selected["same_day_alert_cases"] or 0),
		"high_confidence_same_day_incidents": int(selected["high_confidence_same_day_incidents"] or 0),
		"same_day_flagged_discount_total": _to_number(selected["same_day_flagged_discount_total"]),
		"flagged_discount_pct_of_discount_total": _pct(
			_to_number(selected["same_day_flagged_discount_total"]),
			_to_number(selected["recorded_discount_amount"]),
		),
		"weighted_risk_score": _to_number(selected["weighted_risk_score"]),
		"weighted_risk_rate": _to_number(selected["weighted_risk_rate"]),
		"repeat_name_per_1000": _rate_per_1000(int(selected["repeat_name_findings"] or 0), paid_orders),
		"repeat_reference_per_1000": _rate_per_1000(
			int(selected["repeat_reference_findings"] or 0), paid_orders
		),
		"same_reference_different_name_per_1000": _rate_per_1000(
			int(selected["same_reference_different_name_findings"] or 0),
			paid_orders,
		),
		"rapid_repeat_name_per_1000": _rate_per_1000(
			int(selected["rapid_repeat_name_findings_4h"] or 0), paid_orders
		),
	}


def _resolve_peer_rows(
	row: dict[str, Any],
	store_rows: list[dict[str, Any]],
	peer_mode: str,
) -> tuple[list[dict[str, Any]], str]:
	normalized_mode = _normalize_peer_mode(peer_mode)
	if normalized_mode == "chainwide":
		return store_rows, "CHAINWIDE"

	if normalized_mode == "store_type_legal_entity":
		target_key = _normalize_text(row.get("peer_group_key"))
		peers = [
			candidate
			for candidate in store_rows
			if _normalize_text(candidate.get("peer_group_key")) == target_key
		]
		return peers, target_key or "CHAINWIDE"

	if normalized_mode == "store_type":
		target_key = _normalize_text(row.get("store_type"))
		peers = [
			candidate
			for candidate in store_rows
			if _normalize_text(candidate.get("store_type")) == target_key
		]
		return peers, target_key or "CHAINWIDE"

	primary_key = _normalize_text(row.get("peer_group_key"))
	primary_peers = [
		candidate
		for candidate in store_rows
		if _normalize_text(candidate.get("peer_group_key")) == primary_key
	]
	if primary_key and len(primary_peers) >= MIN_PEER_GROUP_SIZE:
		return primary_peers, primary_key

	store_type_key = _normalize_text(row.get("store_type"))
	store_type_peers = [
		candidate
		for candidate in store_rows
		if _normalize_text(candidate.get("store_type")) == store_type_key
	]
	if store_type_key and len(store_type_peers) >= MIN_PEER_GROUP_SIZE:
		return store_type_peers, store_type_key

	return store_rows, "CHAINWIDE"


def _derive_range_metrics(
	rows: list[dict[str, Any]],
	*,
	denominator_scope: str,
	category_scope: str,
	peer_mode: str,
) -> dict[str, Any]:
	metric_rows = [
		_build_metric_row(
			row,
			denominator_scope=denominator_scope,
			category_scope=category_scope,
		)
		for row in rows
	]
	store_rows = [row for row in metric_rows if _normalize_text(row.get("scope")) == "STORE"]
	chain_row = next((row for row in metric_rows if _normalize_text(row.get("scope")) == "CHAIN"), None)

	if chain_row is None:
		aggregate_seed = _base_snapshot_row(
			scope=CHAIN_SCOPE,
			location_id=CHAIN_LOCATION_ID,
			store_name=CHAIN_STORE_NAME,
		)
		for source in rows:
			if _normalize_text(source.get("scope")) != "STORE":
				continue
			_sum_count_fields(
				aggregate_seed, source, ["active_days", "paid_orders", "all_channel_total_orders"]
			)
			_sum_money_fields(
				aggregate_seed,
				source,
				[
					"pos_original_gross_sales",
					"pos_post_discount_gross_sales",
					"pos_net_sales_without_vat",
					"pos_total_discounts",
					"all_channel_gross_sales",
					"statutory_recorded_discount_amount",
					"statutory_vat_relief_total",
					"effective_statutory_benefit_total",
				],
			)
			for category in STATUTORY_CATEGORIES:
				_sum_count_fields(
					aggregate_seed,
					source,
					[
						_category_field(category, "repeat_name_findings"),
						_category_field(category, "repeat_reference_findings"),
						_category_field(category, "same_reference_different_name_findings"),
						_category_field(category, "same_name_multiple_reference_findings"),
						_category_field(category, "rapid_repeat_name_findings_4h"),
						_category_field(category, "rapid_repeat_reference_findings_4h"),
						_category_field(category, "cross_store_overlap_count"),
						_category_field(category, "cross_store_overlap_critical"),
						_category_field(category, "cross_store_overlap_high"),
						_category_field(category, "same_day_alert_cases"),
						_category_field(category, "high_confidence_same_day_incidents"),
					],
				)
				_sum_money_fields(
					aggregate_seed,
					source,
					[
						_category_field(category, "recorded_discount_amount"),
						_category_field(category, "statutory_vat_relief"),
						_category_field(category, "effective_statutory_benefit"),
						_category_field(category, "same_day_flagged_discount_total"),
						_category_field(category, "weighted_risk_score"),
						_category_field(category, "weighted_risk_rate"),
					],
				)
		chain_row = _build_metric_row(
			aggregate_seed,
			denominator_scope=denominator_scope,
			category_scope=category_scope,
		)

	chain_denominator_value = _to_number(chain_row.get("denominator_value")) if chain_row else 0.0
	total_incidents = sum(int(row.get("high_confidence_same_day_incidents") or 0) for row in store_rows)

	sorted_store_rows = sorted(
		store_rows,
		key=lambda row: (
			-_to_number(row.get("weighted_risk_rate")),
			-_to_number(row.get("weighted_risk_score")),
			-_to_number(row.get("same_reference_different_name_findings")),
			str(row.get("store_name") or ""),
		),
	)
	for index, row in enumerate(sorted_store_rows, start=1):
		peer_rows, effective_peer_group = _resolve_peer_rows(row, store_rows, peer_mode)
		peer_sorted = sorted(
			peer_rows,
			key=lambda entry: (
				_to_number(entry.get("weighted_risk_rate")),
				_to_number(entry.get("weighted_risk_score")),
				str(entry.get("store_name") or ""),
			),
		)
		peer_rank_desc = sorted(
			peer_rows,
			key=lambda entry: (
				-_to_number(entry.get("weighted_risk_rate")),
				-_to_number(entry.get("weighted_risk_score")),
				str(entry.get("store_name") or ""),
			),
		)
		row["weighted_risk_rank"] = index
		row["peer_group_effective"] = effective_peer_group
		row["peer_group_size"] = len(peer_rows)
		row["peer_percentile"] = (
			_round_metric(
				(
					sum(
						1
						for candidate in peer_sorted
						if _to_number(candidate.get("weighted_risk_rate"))
						<= _to_number(row.get("weighted_risk_rate"))
					)
					/ len(peer_sorted)
				)
				* 100,
				2,
			)
			if peer_sorted
			else 0.0
		)
		row["peer_rank"] = next(
			(
				position
				for position, candidate in enumerate(peer_rank_desc, start=1)
				if int(candidate.get("location_id") or 0) == int(row.get("location_id") or 0)
			),
			1,
		)
		row["sales_share_pct"] = _pct(_to_number(row.get("denominator_value")), chain_denominator_value)
		row["incident_share_pct"] = _pct(
			int(row.get("high_confidence_same_day_incidents") or 0),
			total_incidents,
		)

	top_outlier = sorted_store_rows[0] if sorted_store_rows else None
	return {
		"chain_row": chain_row,
		"store_rows": sorted_store_rows,
		"top_outlier": top_outlier,
		"chain_denominator_value": chain_denominator_value,
		"total_high_confidence_incidents": total_incidents,
	}


def _build_control_trend(day_rows: list[dict[str, Any]], denominator_scope: str) -> list[dict[str, Any]]:
	if not day_rows:
		return []
	grouped: dict[str, list[dict[str, Any]]] = {}
	for row in day_rows:
		day_key = str(row.get("business_date") or "")
		if not day_key:
			continue
		grouped.setdefault(day_key, []).append(row)

	trend: list[dict[str, Any]] = []
	for day_key in sorted(grouped):
		day_group = grouped[day_key]
		day_chain = next((row for row in day_group if _normalize_text(row.get("scope")) == "CHAIN"), None)
		source_row = (
			day_chain
			or _aggregate_snapshot_rows_for_period(
				day_group,
				start_day=_coerce_date(day_key),
				end_day=_coerce_date(day_key),
			)[0]
		)
		metric_row = _build_metric_row(
			source_row,
			denominator_scope=denominator_scope,
			category_scope="BOTH",
		)
		trend.append(
			{
				"business_date": day_key,
				"label": _format_display_date(_coerce_date(day_key)),
				"recorded_sc_pct_of_sales": metric_row["recorded_sc_pct_of_sales"],
				"recorded_pwd_pct_of_sales": metric_row["recorded_pwd_pct_of_sales"],
				"effective_statutory_benefit_pct": metric_row["effective_statutory_benefit_pct"],
				"same_day_high_confidence_incidents": metric_row["high_confidence_same_day_incidents"],
				"flagged_discount_amount": metric_row["same_day_flagged_discount_total"],
			}
		)
	return trend


def _build_executive_summary_payload(
	start_day: date,
	end_day: date,
	*,
	denominator_scope: str,
	category_scope: str,
	peer_mode: str,
) -> dict[str, Any]:
	selected_rows, snapshot_meta = _select_snapshot_rows_for_window(start_day, end_day)
	day_rows = _query_store_day_snapshots(start_day, end_day)
	derived = _derive_range_metrics(
		selected_rows,
		denominator_scope=denominator_scope,
		category_scope=category_scope,
		peer_mode=peer_mode,
	)
	chain_row = derived["chain_row"] or {}
	top_outlier = derived["top_outlier"] or {}
	cards = {
		"recorded_sc_pct_of_sales": _round_metric(
			_to_number(chain_row.get("recorded_sc_pct_of_sales")),
			2,
		),
		"recorded_pwd_pct_of_sales": _round_metric(
			_to_number(chain_row.get("recorded_pwd_pct_of_sales")),
			2,
		),
		"effective_statutory_benefit_pct": _round_metric(
			_to_number(chain_row.get("effective_statutory_benefit_pct")),
			2,
		),
		"same_day_high_confidence_incidents": int(chain_row.get("high_confidence_same_day_incidents") or 0),
		"flagged_discount_amount": _round_metric(
			_to_number(chain_row.get("same_day_flagged_discount_total")),
			2,
		),
		"top_outlier_store": {
			"store_name": str(top_outlier.get("store_name") or "-"),
			"weighted_risk_rate": _round_metric(_to_number(top_outlier.get("weighted_risk_rate")), 2),
			"weighted_risk_rank": int(top_outlier.get("weighted_risk_rank") or 0),
		},
	}
	return {
		"start_date": start_day.isoformat(),
		"end_date": end_day.isoformat(),
		"display_window": _format_display_window(start_day, end_day),
		"category_scope": category_scope,
		"denominator_scope": denominator_scope,
		"effective_chain_denominator_scope": str(
			chain_row.get("effective_denominator_scope") or denominator_scope
		),
		"peer_mode": _normalize_peer_mode(peer_mode),
		"snapshot_source": snapshot_meta["snapshot_source"],
		"period_granularity": snapshot_meta["period_granularity"],
		"cards": cards,
		"control_trend": _build_control_trend(day_rows, denominator_scope),
		"top_weighted_risk_stores": derived["store_rows"][:10],
		"incident_share_vs_sales_share": sorted(
			derived["store_rows"],
			key=lambda row: (
				-abs(_to_number(row.get("incident_share_pct")) - _to_number(row.get("sales_share_pct"))),
				-_to_number(row.get("weighted_risk_rate")),
				str(row.get("store_name") or ""),
			),
		)[:10],
		"methodology_notes": [
			"Executive analytics read snapshot-backed benchmark facts only.",
			"Recorded SC and PWD percentages stay separate even when category scope is BOTH.",
			"All-channel gross is allowed only for chain-level cards; store rows automatically fall back to POS original gross.",
		],
	}


def _build_store_benchmark_payload(
	start_day: date,
	end_day: date,
	*,
	denominator_scope: str,
	category_scope: str,
	peer_mode: str,
	store_names: list[str] | None = None,
) -> dict[str, Any]:
	selected_rows, snapshot_meta = _select_snapshot_rows_for_window(start_day, end_day)
	derived = _derive_range_metrics(
		selected_rows,
		denominator_scope=denominator_scope,
		category_scope=category_scope,
		peer_mode=peer_mode,
	)
	rows = derived["store_rows"]
	selected_store_names = store_names or []
	if selected_store_names:
		selected_lookup = {_normalize_text(name) for name in selected_store_names}
		rows = [row for row in rows if _normalize_text(row.get("store_name")) in selected_lookup]
	return {
		"start_date": start_day.isoformat(),
		"end_date": end_day.isoformat(),
		"display_window": _format_display_window(start_day, end_day),
		"category_scope": category_scope,
		"denominator_scope": denominator_scope,
		"peer_mode": _normalize_peer_mode(peer_mode),
		"snapshot_source": snapshot_meta["snapshot_source"],
		"period_granularity": snapshot_meta["period_granularity"],
		"selected_store_names": selected_store_names,
		"rows": rows,
		"top_outlier_store": derived["top_outlier"],
		"methodology_notes": [
			"Weighted risk ranking normalizes same-day store findings against paid order volume.",
			"Peer percentile automatically falls back from store type + legal entity to store type, then chainwide.",
		],
	}


def _build_finance_reconciliation_payload(
	start_day: date,
	end_day: date,
	*,
	denominator_scope: str,
	category_scope: str,
) -> dict[str, Any]:
	selected_rows, snapshot_meta = _select_snapshot_rows_for_window(start_day, end_day)
	derived = _derive_range_metrics(
		selected_rows,
		denominator_scope=denominator_scope,
		category_scope=category_scope,
		peer_mode="auto",
	)
	chain_row = derived["chain_row"] or {}
	original_gross_sales = _to_number(chain_row.get("pos_original_gross_sales"))
	post_discount_gross_sales = _to_number(chain_row.get("pos_post_discount_gross_sales"))
	net_sales_without_vat = _to_number(chain_row.get("pos_net_sales_without_vat"))
	recorded_discount_amount = _to_number(chain_row.get("selected_recorded_discount_amount"))
	statutory_vat_relief = _to_number(chain_row.get("selected_statutory_vat_relief"))
	effective_statutory_benefit = _to_number(chain_row.get("selected_effective_statutory_benefit"))
	gross_gap_before_vat = max(0.0, original_gross_sales - post_discount_gross_sales)
	other_discount_gap = max(0.0, gross_gap_before_vat - effective_statutory_benefit)
	post_discount_vat_component = max(0.0, post_discount_gross_sales - net_sales_without_vat)
	return {
		"start_date": start_day.isoformat(),
		"end_date": end_day.isoformat(),
		"display_window": _format_display_window(start_day, end_day),
		"category_scope": category_scope,
		"denominator_scope": denominator_scope,
		"snapshot_source": snapshot_meta["snapshot_source"],
		"period_granularity": snapshot_meta["period_granularity"],
		"totals": {
			"original_gross_sales": _round_metric(original_gross_sales, 2),
			"post_discount_gross_sales": _round_metric(post_discount_gross_sales, 2),
			"net_sales_without_vat": _round_metric(net_sales_without_vat, 2),
			"recorded_discount_amount": _round_metric(recorded_discount_amount, 2),
			"statutory_vat_relief": _round_metric(statutory_vat_relief, 2),
			"effective_statutory_benefit": _round_metric(effective_statutory_benefit, 2),
			"other_discount_gap": _round_metric(other_discount_gap, 2),
			"post_discount_vat_component": _round_metric(post_discount_vat_component, 2),
			"gross_gap_before_vat": _round_metric(gross_gap_before_vat, 2),
		},
		"ratios": {
			"recorded_discount_pct_of_sales": _round_metric(
				_pct(recorded_discount_amount, _row_denominator_value(chain_row, denominator_scope)),
				2,
			),
			"effective_statutory_benefit_pct_of_sales": _round_metric(
				_pct(effective_statutory_benefit, _row_denominator_value(chain_row, denominator_scope)),
				2,
			),
		},
		"waterfall": [
			{
				"key": "original_gross_sales",
				"label": "Original gross sales",
				"value": _round_metric(original_gross_sales, 2),
			},
			{
				"key": "recorded_discount_amount",
				"label": "Recorded statutory discount",
				"value": _round_metric(-recorded_discount_amount, 2),
			},
			{
				"key": "statutory_vat_relief",
				"label": "Statutory VAT relief",
				"value": _round_metric(-statutory_vat_relief, 2),
			},
			{
				"key": "other_discount_gap",
				"label": "Other discounts and adjustments",
				"value": _round_metric(-other_discount_gap, 2),
			},
			{
				"key": "post_discount_gross_sales",
				"label": "Post-discount gross sales",
				"value": _round_metric(post_discount_gross_sales, 2),
				"is_subtotal": True,
			},
			{
				"key": "post_discount_vat_component",
				"label": "VAT component after discount",
				"value": _round_metric(-post_discount_vat_component, 2),
			},
			{
				"key": "net_sales_without_vat",
				"label": "Net sales without VAT",
				"value": _round_metric(net_sales_without_vat, 2),
				"is_subtotal": True,
			},
		],
		"formula_notes": [
			"Recorded discount amount is the statutory-category-filtered discount value stored on discounted POS item lines.",
			"Statutory VAT relief is allocated per order from the gross gap before VAT after subtracting recorded discounts.",
			"Other discounts and adjustments capture the remaining gap between original gross and post-discount gross after the selected statutory benefit is removed.",
		],
	}


def _replace_store_day_snapshots(target_day: date, rows: list[dict[str, Any]]) -> int:
	_supabase_delete(
		"discount_investigation_store_day",
		{"business_date": f"eq.{target_day.isoformat()}"},
	)
	if rows:
		_supabase_post("discount_investigation_store_day", rows)
	return len(rows)


def _replace_store_month_snapshots(target_month_start: date, rows: list[dict[str, Any]]) -> int:
	_supabase_delete(
		"discount_investigation_store_month",
		{"month_start": f"eq.{target_month_start.isoformat()}"},
	)
	if rows:
		_supabase_post("discount_investigation_store_month", rows)
	return len(rows)


def _refresh_discount_benchmark_snapshots_internal(
	start_day: date,
	end_day: date,
) -> dict[str, Any]:
	if end_day < start_day:
		frappe.throw(frappe._("End date cannot be earlier than start date."))
	day_inserted = 0
	month_inserted = 0
	refreshed_days: list[str] = []
	touched_months: set[date] = set()
	for target_day in _iter_days(start_day, end_day):
		day_rows = _build_store_day_snapshot_rows(target_day)
		day_inserted += _replace_store_day_snapshots(target_day, day_rows)
		refreshed_days.append(target_day.isoformat())
		touched_months.add(_month_start(target_day))

	for target_month_start in sorted(touched_months):
		month_day_rows = _query_store_day_snapshots(target_month_start, _month_end(target_month_start))
		if month_day_rows:
			last_available_day = max(_coerce_date(row.get("business_date")) for row in month_day_rows)
			month_rows = _aggregate_snapshot_rows_for_period(
				month_day_rows,
				start_day=target_month_start,
				end_day=last_available_day,
			)
			for row in month_rows:
				row["month_start"] = target_month_start.isoformat()
				row["business_date"] = last_available_day.isoformat()
		else:
			month_rows = []
		month_inserted += _replace_store_month_snapshots(target_month_start, month_rows)

	return {
		"success": True,
		"start_date": start_day.isoformat(),
		"end_date": end_day.isoformat(),
		"refreshed_days": refreshed_days,
		"refreshed_months": [value.isoformat() for value in sorted(touched_months)],
		"store_day_rows_written": day_inserted,
		"store_month_rows_written": month_inserted,
	}


@frappe.whitelist()
def refresh_discount_benchmark_snapshots(
	start_date: str | None = None,
	end_date: str | None = None,
) -> dict[str, Any]:
	_check_discount_audit_role()
	start_day = _coerce_date(start_date, default=_default_business_date())
	end_day = _coerce_date(end_date, default=start_day)
	return {
		"success": True,
		"data": _refresh_discount_benchmark_snapshots_internal(start_day, end_day),
	}


@frappe.whitelist()
def get_discount_executive_summary(
	start_date: str,
	end_date: str,
	denominator_scope: str | None = None,
	category_scope: str | None = None,
	peer_mode: str | None = None,
) -> dict[str, Any]:
	_check_discount_audit_role()
	start_day = _coerce_date(start_date)
	end_day = _coerce_date(end_date)
	return {
		"success": True,
		"data": _build_executive_summary_payload(
			start_day,
			end_day,
			denominator_scope=_normalize_denominator_scope(denominator_scope),
			category_scope=_normalize_category_scope(category_scope),
			peer_mode=_normalize_peer_mode(peer_mode),
		),
	}


@frappe.whitelist()
def get_discount_store_benchmark(
	start_date: str,
	end_date: str,
	denominator_scope: str | None = None,
	category_scope: str | None = None,
	peer_mode: str | None = None,
	store_names: str | None = None,
) -> dict[str, Any]:
	_check_discount_audit_role()
	start_day = _coerce_date(start_date)
	end_day = _coerce_date(end_date)
	return {
		"success": True,
		"data": _build_store_benchmark_payload(
			start_day,
			end_day,
			denominator_scope=_normalize_denominator_scope(denominator_scope),
			category_scope=_normalize_category_scope(category_scope),
			peer_mode=_normalize_peer_mode(peer_mode),
			store_names=_parse_store_names(store_names),
		),
	}


@frappe.whitelist()
def get_discount_finance_reconciliation(
	start_date: str,
	end_date: str,
	denominator_scope: str | None = None,
	category_scope: str | None = None,
) -> dict[str, Any]:
	_check_discount_audit_role()
	start_day = _coerce_date(start_date)
	end_day = _coerce_date(end_date)
	return {
		"success": True,
		"data": _build_finance_reconciliation_payload(
			start_day,
			end_day,
			denominator_scope=_normalize_denominator_scope(denominator_scope),
			category_scope=_normalize_category_scope(category_scope),
		),
	}


def scheduled_refresh_discount_benchmark_snapshots() -> dict[str, Any]:
	target_day = _default_business_date()
	return _refresh_discount_benchmark_snapshots_internal(target_day, target_day)


def _validate_alert_resolution_payload(payload: dict[str, Any]) -> None:
	required_keys = [
		"event_date",
		"scope",
		"scope_key",
		"discount_bir_category",
		"identity_type",
		"identity_key",
		"detection_type",
	]
	missing = [key for key in required_keys if payload.get(key) in (None, "")]
	if missing:
		frappe.throw(frappe._("Missing alert identity fields: {0}").format(", ".join(missing)))


def _resolve_discount_alert_payload(
	payload: dict[str, Any],
	resolution_code: str,
	resolution_note: str | None = None,
) -> dict[str, Any] | None:
	existing_rows = _supabase_get(
		"discount_abuse_alerts",
		{
			"select": "details",
			**_alert_filters(payload),
			"limit": "1",
		},
	)
	existing_details = _safe_json(existing_rows[0].get("details")) if existing_rows else {}
	resolved_at = _manila_now().isoformat()
	merged_details = {
		**existing_details,
		"resolution_code": resolution_code,
		"resolution_note": resolution_note or "",
		"resolved_by": getattr(getattr(frappe, "session", None), "user", "unknown"),
		"resolved_at_pht": resolved_at,
	}

	updated = _supabase_patch(
		"discount_abuse_alerts",
		_alert_filters(payload),
		{
			"resolved": True,
			"resolved_at": resolved_at,
			"updated_at": resolved_at,
			"details": merged_details,
		},
	)
	return updated[0] if updated else None


@frappe.whitelist()
def resolve_discount_audit_alert(
	alert: str | dict[str, Any],
	resolution_code: str,
	resolution_note: str | None = None,
) -> dict[str, Any]:
	_check_discount_audit_role()
	payload = _safe_json(alert) if not isinstance(alert, dict) else alert
	_validate_alert_resolution_payload(payload)
	updated = _resolve_discount_alert_payload(payload, resolution_code, resolution_note)
	return {
		"success": True,
		"data": updated,
		"message": "Alert resolved",
	}


@frappe.whitelist()
def resolve_discount_audit_incident(
	incident: str | dict[str, Any],
	resolution_code: str,
	resolution_note: str | None = None,
) -> dict[str, Any]:
	_check_discount_audit_role()
	payload = _safe_json(incident) if not isinstance(incident, dict) else incident
	resolution_targets = payload.get("resolution_targets")
	if not isinstance(resolution_targets, list) or not resolution_targets:
		frappe.throw(frappe._("Incident is missing resolution_targets."))
	if payload.get("resolve_scope_policy") not in (None, "", "all_underlying_rows"):
		frappe.throw(frappe._("Unsupported resolve_scope_policy for incident resolution."))

	seen_target_keys: set[str] = set()
	updated_rows: list[dict[str, Any]] = []
	for target in resolution_targets:
		if not isinstance(target, dict):
			continue
		_validate_alert_resolution_payload(target)
		target_key = _resolution_target_key(target)
		if target_key in seen_target_keys:
			continue
		seen_target_keys.add(target_key)
		updated = _resolve_discount_alert_payload(target, resolution_code, resolution_note)
		if updated:
			updated_rows.append(updated)

	return {
		"success": True,
		"data": {
			"cluster_id": payload.get("cluster_id"),
			"resolved_count": len(updated_rows),
			"target_count": len(seen_target_keys),
			"rows": updated_rows,
		},
		"message": "Incident resolved",
	}


@frappe.whitelist()
def generate_daily_discount_audit_report(
	business_date: str | None = None,
	attach_to_user: str | None = None,
) -> dict[str, Any]:
	target_day = _coerce_date(business_date, default=_default_business_date())
	return _generate_daily_discount_audit_report_internal(target_day, attach_to_user, check_permissions=True)


def _generate_daily_discount_audit_report_internal(
	target_day: date,
	attach_to_user: str | None = None,
	check_permissions: bool = False,
) -> dict[str, Any]:
	if check_permissions:
		_check_discount_audit_role()

	same_day_rows = _sort_same_day_rows(_query_same_day_rows(target_day))
	rolling_rows = _sort_rolling_rows(_query_rolling_rows(target_day))
	summary = _make_summary(same_day_rows, rolling_rows)
	workbook_bytes = _build_audit_workbook_bytes(target_day, same_day_rows, rolling_rows, summary)

	from frappe.utils.file_manager import save_file

	target_user = _get_report_owner(attach_to_user or getattr(getattr(frappe, "session", None), "user", None))
	filename = f"{REPORT_FILENAME_PREFIX}{target_day.isoformat()}.xlsx"
	file_doc = save_file(filename, workbook_bytes, "User", target_user, is_private=1)
	return {
		"success": True,
		"data": {
			"business_date": target_day.isoformat(),
			"filename": filename,
			"file_url": file_doc.file_url,
			"attached_to_user": target_user,
			"same_day_count": len(same_day_rows),
			"rolling_count": len(rolling_rows),
			"summary": summary,
		},
	}


@frappe.whitelist()
def list_discount_audit_reports(limit: int | str = 10) -> dict[str, Any]:
	_check_discount_audit_role()
	page_length = min(max(int(limit or 10), 1), 50)
	files = frappe.get_all(
		"File",
		filters=[
			["File", "file_name", "like", f"{REPORT_FILENAME_PREFIX}%"],
			["File", "is_private", "=", 1],
		],
		fields=["name", "file_name", "file_url", "creation", "attached_to_name"],
		order_by="creation desc",
		limit_page_length=page_length,
	)
	return {"success": True, "data": files}


@frappe.whitelist()
def send_critical_discount_alert_notifications(
	business_date: str | None = None,
	force: int | str | bool = False,
) -> dict[str, Any]:
	target_day = _coerce_date(business_date, default=_default_business_date())
	force_send = str(force).lower() in {"1", "true", "yes"}
	return _send_critical_discount_alert_notifications_internal(
		target_day,
		force_send=force_send,
		check_permissions=True,
	)


def _send_critical_discount_alert_notifications_internal(
	target_day: date,
	force_send: bool = False,
	check_permissions: bool = False,
) -> dict[str, Any]:
	if check_permissions:
		_check_discount_audit_role()

	if not force_send and not _notifications_enabled_for_day(target_day):
		return {
			"success": True,
			"skipped": True,
			"reason": "Notifications are still in warm-up window",
			"go_live_date": _get_notification_go_live_date().isoformat(),
		}

	rows = _sort_same_day_rows(_query_same_day_rows(target_day, severity="critical"))
	if not rows:
		return {
			"success": True,
			"skipped": True,
			"reason": "No critical same-day rows",
			"business_date": target_day.isoformat(),
		}

	message = _build_critical_notification_message(target_day, rows)
	chat_sent = False
	email_sent = False

	try:
		from hrms.api.google_chat import send_message_to_space
		from hrms.utils.bei_config import SPACE_ACCOUNTING, get_chat_space

		chat_sent = send_message_to_space(get_chat_space(SPACE_ACCOUNTING), message)
	except Exception as exc:
		frappe.log_error(
			title="Discount Alert Chat Notification Failed",
			message=str(exc),
		)

	recipients = _parse_email_recipients()
	if recipients:
		try:
			frappe.sendmail(
				recipients=recipients,
				subject=f"Critical SC/PWD Alert Digest - {target_day.isoformat()}",
				message=f"<pre>{message}</pre>",
			)
			email_sent = True
		except Exception as exc:
			frappe.log_error(
				title="Discount Alert Email Notification Failed",
				message=str(exc),
			)

	marked_count = _mark_alerts_notified(rows) if (chat_sent or email_sent) else 0
	return {
		"success": True,
		"data": {
			"business_date": target_day.isoformat(),
			"critical_count": len(rows),
			"chat_sent": chat_sent,
			"email_sent": email_sent,
			"marked_notified": marked_count,
		},
	}


def scheduled_generate_daily_discount_audit_report() -> dict[str, Any]:
	target_day = _default_business_date()
	report_owner = _get_report_owner()
	try:
		return _generate_daily_discount_audit_report_internal(
			target_day,
			attach_to_user=report_owner,
		)
	except Exception as exc:
		frappe.log_error(
			title="Scheduled Discount Audit Report Failed",
			message=str(exc),
		)
		return {"success": False, "error": str(exc)}


def scheduled_send_critical_discount_alert_notifications() -> dict[str, Any]:
	target_day = _default_business_date()
	try:
		return _send_critical_discount_alert_notifications_internal(target_day)
	except Exception as exc:
		frappe.log_error(
			title="Scheduled Discount Alert Notification Failed",
			message=str(exc),
		)
		return {"success": False, "error": str(exc)}
