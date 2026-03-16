# Copyright (c) 2026, Frappe Technologies Pvt. Ltd. and contributors
# For license information, please see license.txt
"""
Inventory API
Handles cycle counts, variances, and shelf life extensions
"""

import json
from contextlib import contextmanager

import frappe
from frappe import _
from frappe.utils import add_days, cint, flt, now_datetime, nowdate

from hrms.utils.bei_config import get_company
from hrms.utils.scm_roles import SCM_INVENTORY_ROLES, SCM_STOCK_UPDATE_ROLES
from hrms.utils.scm_roles import check_scm_permission as _check_warehouse_permission

COUNT_TYPE_STORE_MONTHLY = "Store Monthly"
COUNT_TYPE_WAREHOUSE_MONTHLY = "Warehouse Monthly"
COUNT_TYPE_SPOT_CHECK = "Spot Check"
COUNT_TYPE_EXTERNAL_AUDIT = "External Audit"

ALL_CYCLE_COUNT_TYPES = (
	COUNT_TYPE_STORE_MONTHLY,
	COUNT_TYPE_WAREHOUSE_MONTHLY,
	COUNT_TYPE_SPOT_CHECK,
	COUNT_TYPE_EXTERNAL_AUDIT,
)

HQ_STOCK_COUNT_ROLES = {"HQ User", "System Manager", "Administrator"}
SUPERVISOR_STOCK_COUNT_ROLES = {"Store Supervisor", "Area Supervisor"}


@contextmanager
def _run_as_system_user(user: str = "Administrator"):
	"""Temporarily elevate the session user for internal stock adjustments."""
	session = getattr(frappe, "session", None)
	if session is None:
		session = getattr(getattr(frappe, "local", None), "session", None)
	original_user = getattr(session, "user", None)
	try:
		if session and user:
			session.user = user
		yield
	finally:
		if session and original_user:
			session.user = original_user


def _get_allowed_cycle_count_types(user_roles: list[str] | set[str] | None = None) -> list[str]:
	roles = set(user_roles or frappe.get_roles())

	if "External Auditor" in roles:
		return [COUNT_TYPE_EXTERNAL_AUDIT]

	if roles.intersection(HQ_STOCK_COUNT_ROLES):
		return list(ALL_CYCLE_COUNT_TYPES)

	if "Warehouse User" in roles:
		return [COUNT_TYPE_WAREHOUSE_MONTHLY, COUNT_TYPE_SPOT_CHECK]

	if roles.intersection(SUPERVISOR_STOCK_COUNT_ROLES):
		return [COUNT_TYPE_STORE_MONTHLY, COUNT_TYPE_SPOT_CHECK]

	if "Store Staff" in roles:
		return [COUNT_TYPE_STORE_MONTHLY]

	return []


def _validate_cycle_count_type(
	count_type: str | None,
	user_roles: list[str] | set[str] | None = None,
) -> str:
	normalized = (count_type or "").strip() or COUNT_TYPE_STORE_MONTHLY
	if normalized not in ALL_CYCLE_COUNT_TYPES:
		frappe.throw(_("Invalid cycle count type: {0}").format(normalized), frappe.ValidationError)

	allowed_count_types = _get_allowed_cycle_count_types(user_roles=user_roles)
	if normalized not in allowed_count_types:
		frappe.throw(
			_("You do not have permission to create {0} counts").format(normalized),
			frappe.PermissionError,
		)

	return normalized


def _clamp_multiplier(value: float | int, floor: float = 0.70, ceiling: float = 1.50):
	return flt(min(max(flt(value), flt(floor)), flt(ceiling)), 4)


def get_signal_modifiers(
	is_salary_week: bool = False,
	is_holiday: bool = False,
	is_weather_risk: bool = False,
):
	"""Deterministic S019 demand modifiers used by ordering recommendation tuning."""
	salary_week_multiplier = 1.10 if is_salary_week else 1.0
	holiday_multiplier = 1.12 if is_holiday else 1.0
	overlap_multiplier = 1.08 if (is_salary_week and is_holiday) else 1.0
	weather_multiplier = 1.06 if is_weather_risk else 1.0
	composite_multiplier = (
		salary_week_multiplier * holiday_multiplier * overlap_multiplier * weather_multiplier
	)
	return {
		"salary_week_multiplier": flt(salary_week_multiplier, 4),
		"holiday_multiplier": flt(holiday_multiplier, 4),
		"overlap_multiplier": flt(overlap_multiplier, 4),
		"weather_multiplier": flt(weather_multiplier, 4),
		"composite_multiplier": _clamp_multiplier(composite_multiplier),
	}


def apply_adaptive_tuning(
	current_multiplier: float | int,
	adjustment_delta: float | int,
	floor: float = 0.70,
	ceiling: float = 1.50,
):
	"""Clamp weekly tuning updates to agreed S019 guardrails."""
	return _clamp_multiplier(
		flt(current_multiplier) + flt(adjustment_delta),
		floor=floor,
		ceiling=ceiling,
	)


@frappe.whitelist()
def preview_adaptive_multiplier(
	current_multiplier: float | int = 1.0,
	adjustment_delta: float | int = 0.0,
):
	return {
		"multiplier_before": flt(current_multiplier, 4),
		"adjustment_delta": flt(adjustment_delta, 4),
		"multiplier_after": apply_adaptive_tuning(current_multiplier, adjustment_delta),
	}


@frappe.whitelist()
def submit_cycle_count(
	store: str | None = None,
	items: list[dict] | str | None = None,
	count_date: str | None = None,
):
	"""DEPRECATED: Use submit_cycle_count_v2 instead.

	B-7: This v1 endpoint bypasses duplicate check and doesn't set count_type.
	Kept as whitelisted to avoid breaking existing API callers, but throws deprecation error.
	"""
	frappe.throw(_("This endpoint is deprecated. Use submit_cycle_count_v2 instead."))
	# Original implementation below kept for reference only
	try:
		if not store:
			frappe.throw(_("Store is required"))

		if isinstance(items, str):
			items = json.loads(items)

		# Resolve branch name to warehouse name
		warehouse = _resolve_warehouse(store)
		if not warehouse:
			frappe.throw(_("Could not find Store: {0}").format(store))

		# Validate no negative quantities
		for item in items:
			if flt(item.get("counted_qty", 0)) < 0:
				frappe.throw(
					_("Counted quantity cannot be negative for item {0}").format(
						item.get("item_code", "unknown")
					)
				)

		doc = frappe.new_doc("BEI Cycle Count")
		doc.store = warehouse
		doc.count_date = count_date or nowdate()
		doc.counted_by = frappe.session.user

		total_variance = 0
		for item in items:
			# Get system qty from stock ledger - use resolved warehouse
			system_qty = (
				frappe.db.get_value(
					"Bin", {"warehouse": warehouse, "item_code": item["item_code"]}, "actual_qty"
				)
				or 0
			)

			row = doc.append(
				"items",
				{
					"item_code": item["item_code"],
					"system_qty": system_qty,
					"counted_qty": flt(item["counted_qty"]),
					"variance_qty": flt(item["counted_qty"]) - system_qty,
					"remarks": item.get("remarks"),
				},
			)

			# Calculate variance value
			item_price = frappe.db.get_value("Item", item["item_code"], "valuation_rate") or 0
			row.variance_value = row.variance_qty * item_price
			total_variance += row.variance_value

		doc.total_variance_value = total_variance
		doc.status = "Submitted"
		doc.insert(ignore_permissions=True)
		# V-05 fix: submit() to set docstatus=1 — matches submit_cycle_count_v2 behavior.
		# Without this, approve_cycle_count rejects these records (requires docstatus=1).
		doc.flags.ignore_permissions = True
		doc.submit()
		return {"success": True, "name": doc.name, "total_variance": total_variance}

	except Exception as e:
		frappe.log_error(
			f"Cycle Count Error for store {store}: {e!s}\n\n{frappe.get_traceback()}",
			"Cycle Count Submission Error",
		)
		return {
			"success": False,
			"error": _("Failed to submit cycle count. Please try again or contact support."),
		}


@frappe.whitelist()
def get_cycle_counts(
	store: str | None = None,
	date_from: str | None = None,
	date_to: str | None = None,
	status: str | None = None,
	count_type: str | None = None,
	external_auditor: str | None = None,
	limit: int = 20,
):
	"""Get cycle count history with filtering."""
	# External Auditors: verify store access when filtering by store
	if store and "External Auditor" in frappe.get_roles():
		_check_store_access(store)

	limit = min(int(limit or 20), 500)
	filters = {"docstatus": ["in", [0, 1]]}  # Exclude cancelled (docstatus=2)
	if store:
		filters["store"] = store
	if status:
		filters["status"] = status
	if count_type:
		filters["count_type"] = count_type
	if external_auditor:
		filters["external_auditor"] = external_auditor
	if date_from:
		filters["count_date"] = [">=", date_from]
	if date_to:
		if "count_date" in filters:
			filters["count_date"] = ["between", [date_from, date_to]]
		else:
			filters["count_date"] = ["<=", date_to]

	# External Auditors can only see their own counts
	if "External Auditor" in frappe.get_roles() and not external_auditor:
		filters["counted_by"] = frappe.session.user

	counts = frappe.get_all(
		"BEI Cycle Count",
		filters=filters,
		fields=[
			"name",
			"store",
			"count_date",
			"status",
			"total_variance_value",
			"counted_by",
			"count_type",
			"external_auditor",
			"photo_evidence",
			"exported_to_cos_recon",
			"export_date",
			"rejection_reason",
			"resubmission_count",
		],
		order_by="count_date desc",
		limit=int(limit),
	)
	return {"counts": counts}


@frappe.whitelist()
def get_cycle_count(name: str):
	"""Get single cycle count with items."""
	doc = frappe.get_doc("BEI Cycle Count", name)
	doc.check_permission("read")

	# External Auditors can only see their own counts
	if "External Auditor" in frappe.get_roles() and doc.counted_by != frappe.session.user:
		frappe.throw(_("You can only view your own cycle counts"), frappe.PermissionError)

	items = []
	for item in doc.items:
		item_meta = (
			frappe.db.get_value("Item", item.item_code, ["item_name", "item_group"], as_dict=True) or {}
		)
		items.append(
			{
				"name": item.name,
				"item_code": item.item_code,
				"item_name": item_meta.get("item_name", ""),
				"item_group": item_meta.get("item_group", ""),
				"uom": item.uom,
				"system_qty": item.system_qty,
				"counted_qty": item.counted_qty,
				"counted_qty_whole": item.counted_qty_whole,
				"counted_qty_loose": item.counted_qty_loose,
				"conversion_factor": item.conversion_factor,
				"unit_cost": item.unit_cost,
				"variance_qty": item.variance_qty,
				"variance_value": item.variance_value,
				"remarks": item.remarks,
			}
		)

	return {
		"name": doc.name,
		"store": doc.store,
		"count_date": str(doc.count_date),
		"count_type": doc.count_type,
		"status": doc.status,
		"counted_by": doc.counted_by,
		"verified_by": doc.verified_by,
		"external_auditor": doc.external_auditor,
		"photo_evidence": doc.photo_evidence,
		"exported_to_cos_recon": doc.exported_to_cos_recon,
		"export_date": str(doc.export_date) if doc.export_date else None,
		"total_variance_value": doc.total_variance_value,
		"rejection_reason": doc.rejection_reason,
		"resubmission_count": doc.resubmission_count,
		"items": items,
		"docstatus": doc.docstatus,
		"creation": str(doc.creation),
		"modified": str(doc.modified),
	}


@frappe.whitelist()
def approve_cycle_count(cycle_count_name: str, action: str, comment: str | None = None):
	"""Approve or reject a submitted cycle count.

	Args:
	    action: "Approve" or "Reject"
	    comment: Mandatory for rejection
	"""
	if isinstance(action, str) and action not in ("Approve", "Reject"):
		frappe.throw(_("Invalid action. Must be 'Approve' or 'Reject'."))

	doc = frappe.get_doc("BEI Cycle Count", cycle_count_name)
	doc.check_permission("write")

	# C-2 fix: Guard against approving non-submitted docs
	if doc.docstatus != 1:
		frappe.throw(_("Only submitted cycle counts can be approved/rejected."))
	if doc.status not in ("Submitted", "Resubmitted"):
		frappe.throw(_("This cycle count has already been {0}.").format(doc.status))

	# AUDIT-10: Role gate — only supervisors and above can approve/reject
	allowed_roles = {"Store Supervisor", "Area Supervisor", "System Manager", "Administrator"}
	user_roles = set(frappe.get_roles())
	if not user_roles.intersection(allowed_roles):
		frappe.throw(_("You do not have permission to approve/reject cycle counts."), frappe.PermissionError)

	# AUDIT-10: Self-approval prevention
	if doc.counted_by == frappe.session.user:
		frappe.throw(_("You cannot approve your own cycle count."))

	if action == "Reject":
		if not comment:
			frappe.throw(_("A comment is mandatory when rejecting a cycle count."))
		doc.status = "Rejected"
		doc.rejection_reason = comment
		doc.save(ignore_permissions=True)
		doc.add_comment("Comment", comment)
	else:
		doc.status = "Approved"
		doc.approved_by = frappe.session.user
		doc.approved_at = now_datetime()
		if comment:
			doc.add_comment("Comment", comment)
		doc.save(ignore_permissions=True)

	return {"status": doc.status, "approved_by": doc.approved_by}


@frappe.whitelist()
def reject_cycle_count(count_name: str, rejection_reason: str):
	"""Reject a submitted cycle count with a reason.

	DEPRECATED: Use approve_cycle_count(cycle_count_name, action="Reject", comment=rejection_reason) instead.
	Kept for backward compatibility with existing frontend calls.
	"""
	return approve_cycle_count(count_name, action="Reject", comment=rejection_reason)


@frappe.whitelist()
def mark_cycle_count_reconciled(
	count_name: str,
	stock_reconciliation_name: str | None = None,
	notes: str | None = None,
):
	"""Mark a Verified cycle count as Reconciled.

	Called after Finance reconciles the count against the stock reconciliation document.

	Args:
	    count_name: BEI Cycle Count name
	    stock_reconciliation_name: Optional link to Stock Reconciliation doc
	    notes: Optional reconciliation notes

	Returns:
	    dict: {status, reconciled_by, reconciled_at}
	"""
	allowed_roles = {"HQ User", "System Manager", "Administrator"}
	user_roles = set(frappe.get_roles())
	if not user_roles.intersection(allowed_roles):
		frappe.throw(_("Only HQ/Finance can mark cycle counts as reconciled."), frappe.PermissionError)

	doc = frappe.get_doc("BEI Cycle Count", count_name)

	if doc.status not in ("Approved", "Verified"):
		frappe.throw(
			_("Only approved/verified cycle counts can be reconciled. Current status: {0}").format(doc.status)
		)

	doc.status = "Reconciled"
	doc.reconciled_by = frappe.session.user
	doc.reconciled_at = now_datetime()

	if stock_reconciliation_name:
		doc.stock_reconciliation = stock_reconciliation_name

	if notes:
		doc.reconciliation_notes = notes

	doc.save(ignore_permissions=True)

	return {
		"success": True,
		"name": count_name,
		"status": doc.status,
		"reconciled_by": doc.reconciled_by,
		"reconciled_at": str(doc.reconciled_at),
	}


@frappe.whitelist()
def resubmit_cycle_count(count_name: str, items: list[dict] | str):
	"""
	Resubmit a rejected cycle count with corrected counts.
	Creates a new cycle count document linked to the original.
	"""
	if isinstance(items, str):
		items = json.loads(items)

	original = frappe.get_doc("BEI Cycle Count", count_name)

	if original.status != "Rejected":
		frappe.throw(_("Only rejected cycle counts can be resubmitted"))

	# B-2a: Ownership check — only the original counter can resubmit
	if original.counted_by != frappe.session.user:
		if not any(r in frappe.get_roles() for r in ["System Manager", "Administrator"]):
			frappe.throw(_("You can only resubmit your own cycle counts."), frappe.PermissionError)

	# Create new cycle count as resubmission
	doc = frappe.new_doc("BEI Cycle Count")
	doc.store = original.store
	doc.count_date = nowdate()
	doc.count_type = original.count_type  # B-2b: Was missing
	doc.counted_by = frappe.session.user
	doc.original_count = original.name
	doc.resubmission_count = (original.resubmission_count or 0) + 1

	if "External Auditor" in frappe.get_roles():
		doc.external_auditor = frappe.session.user

	for item in items:
		uom = frappe.db.get_value("Item", item["item_code"], "stock_uom") or ""
		doc.append(
			"items",
			{
				"item_code": item["item_code"],
				"uom": uom,
				"counted_qty_whole": item.get("counted_qty_whole", 0),
				"counted_qty_loose": item.get("counted_qty_loose", 0.0),
				"remarks": item.get("remarks"),
			},
		)
	# validate() computes counted_qty, system_qty from Bin, unit_cost, variance automatically

	doc.status = "Resubmitted"

	frappe.db.savepoint("resubmit_cycle_count")
	try:
		doc.insert(ignore_permissions=True)
		doc.flags.ignore_permissions = True
		doc.submit()
		# Mark original as superseded
		original.db_set("status", "Resubmitted")
		original.db_set("notes", f"Resubmitted as {doc.name} on {nowdate()}\n" + (original.notes or ""))
	except Exception:
		frappe.db.rollback(save_point="resubmit_cycle_count")
		raise

	return {
		"success": True,
		"name": doc.name,
		"total_variance": doc.total_variance_value,
		"message": _("Cycle count resubmitted successfully"),
	}


@frappe.whitelist()
def report_variance(
	store: str,
	item_code: str,
	system_qty: float | int = 0,
	actual_qty: float | int = 0,
	variance_type: str | None = None,
	explanation: str | None = None,
	photo: str | None = None,
):
	"""Report inventory variance.

	Bug fixes (C8):
	- Resolve store to warehouse
	- system_qty/actual_qty default to 0
	- variance_type defaults to 'Shortage' if not provided
	- photo converted via save_base64_image
	- insert with ignore_permissions=True
	- Returns user-friendly error
	"""
	try:
		if not store or not item_code:
			frappe.throw(_("Store and item are required"))

		# Resolve branch name to warehouse name
		warehouse = _resolve_warehouse(store)
		if not warehouse:
			frappe.throw(_("Could not find Store: {0}").format(store))

		# Handle photo base64
		photo_url = None
		if photo:
			from hrms.api.store import save_base64_image

			photo_url = save_base64_image(photo, "BEI Inventory Variance", fieldname="photo_evidence")

		doc = frappe.new_doc("BEI Inventory Variance")
		doc.store = warehouse
		doc.variance_date = nowdate()
		doc.item_code = item_code
		doc.system_qty = flt(system_qty)
		doc.actual_qty = flt(actual_qty)
		doc.variance_qty = flt(actual_qty) - flt(system_qty)

		item_price = frappe.db.get_value("Item", item_code, "valuation_rate") or 0
		doc.variance_value = doc.variance_qty * item_price

		doc.variance_type = variance_type or ("Shortage" if doc.variance_qty < 0 else "Overage")
		doc.explanation = explanation or ""
		doc.reported_by = frappe.session.user
		doc.photo_evidence = photo_url
		doc.insert(ignore_permissions=True)
		return {"success": True, "name": doc.name}

	except Exception as e:
		frappe.log_error(
			f"Variance Report Error for store {store}: {e!s}\n\n{frappe.get_traceback()}",
			"Variance Report Submission Error",
		)
		return {
			"success": False,
			"error": _("Failed to submit variance report. Please try again or contact support."),
		}


@frappe.whitelist()
def get_variances(store: str | None = None, status: str | None = None, limit: int = 20):
	"""Get inventory variances."""
	limit = min(int(limit or 20), 500)
	filters = {}
	if store:
		filters["store"] = store
	if status:
		filters["status"] = status

	variances = frappe.get_all(
		"BEI Inventory Variance",
		filters=filters,
		fields=["name", "store", "item_code", "variance_qty", "variance_value", "variance_type", "status"],
		order_by="creation desc",
		limit=int(limit),
	)
	return {"variances": variances}


@frappe.whitelist()
def request_shelf_extension(
	store: str,
	item_code: str,
	original_expiry: str,
	requested_expiry: str,
	quantity: float | int,
	reason: str,
	batch_no: str | None = None,
	photo: str | None = None,
):
	"""Request shelf life extension for an item.

	Bug fix (C9): Store OIC/Staff get permission error because BEI Shelf Life Extension
	DocType only allows System Manager and Stock User roles. Fix: use ignore_permissions=True
	since the API already has @frappe.whitelist() protection, and resolve store to warehouse.
	Also convert photo base64 to file URL.
	"""
	try:
		if not store or not item_code:
			frappe.throw(_("Store and item are required"))

		# Resolve branch name to warehouse name
		warehouse = _resolve_warehouse(store)
		if not warehouse:
			frappe.throw(_("Could not find Store: {0}").format(store))

		# Handle photo base64
		photo_url = None
		if photo:
			from hrms.api.store import save_base64_image

			photo_url = save_base64_image(photo, "BEI Shelf Life Extension", fieldname="photo_evidence")

		doc = frappe.new_doc("BEI Shelf Life Extension")
		doc.store = warehouse
		doc.request_date = nowdate()
		doc.item_code = item_code
		doc.batch_no = batch_no
		doc.original_expiry = original_expiry
		doc.requested_expiry = requested_expiry
		doc.quantity = flt(quantity)
		doc.reason = reason
		doc.requested_by = frappe.session.user
		doc.photo_evidence = photo_url
		# C9 fix: ignore_permissions because DocType only allows Stock User/System Manager
		# but Store Staff/OIC need to submit shelf life checks
		doc.insert(ignore_permissions=True)

		return {"success": True, "name": doc.name}

	except Exception as e:
		frappe.log_error(
			f"Shelf Life Extension Error for store {store}: {e!s}\n\n{frappe.get_traceback()}",
			"Shelf Life Extension Error",
		)
		return {
			"success": False,
			"error": _("Failed to submit shelf life check. Please try again or contact support."),
		}


@frappe.whitelist()
def approve_shelf_extension(
	extension_name: str,
	approved_expiry: str | None = None,
	rejection_reason: str | None = None,
	approve: bool = True,
):
	"""Approve or reject shelf life extension."""
	allowed_roles = ["Area Supervisor", "Store Supervisor", "System Manager"]
	if not any(r in frappe.get_roles(frappe.session.user) for r in allowed_roles):
		frappe.throw(_("Not authorized to approve shelf life extensions"), frappe.PermissionError)

	doc = frappe.get_doc("BEI Shelf Life Extension", extension_name)

	if approve:
		doc.status = "Approved"
		doc.approved_by = frappe.session.user
		doc.approved_expiry = approved_expiry or doc.requested_expiry
	else:
		doc.status = "Rejected"
		doc.approved_by = frappe.session.user
		doc.rejection_reason = rejection_reason

	doc.save()
	return {"success": True, "status": doc.status}


# ============ RETURNS APIs ============


def _resolve_warehouse(store_or_branch: str):
	"""Resolve a branch name to full warehouse name."""
	if not store_or_branch:
		return None

	if frappe.db.exists("Warehouse", store_or_branch):
		return store_or_branch

	warehouse_with_company = f"{store_or_branch} - BEI"
	if frappe.db.exists("Warehouse", warehouse_with_company):
		return warehouse_with_company

	warehouse = frappe.db.get_value(
		"Warehouse", {"warehouse_name": store_or_branch, "company": get_company()}, "name"
	)
	if warehouse:
		return warehouse

	# Fallback without company filter for backward compat
	warehouse = frappe.db.get_value("Warehouse", {"warehouse_name": store_or_branch}, "name")
	if warehouse:
		return warehouse

	return None


def _coerce_flag(value) -> bool:
	return bool(cint(value or 0))


def _get_inventory_warehouse_catalog() -> list[dict]:
	"""Return the canonical warehouse selector catalog for inventory surfaces."""
	mapped_rows = frappe.get_all(
		"BEI Warehouse Department Mapping",
		filters={"is_active": 1},
		fields=["warehouse", "department"],
		order_by="warehouse",
	)

	catalog: list[dict] = []
	seen: set[str] = set()

	for row in mapped_rows:
		warehouse = _resolve_warehouse(row.warehouse) or row.warehouse
		if not warehouse or warehouse in seen or not frappe.db.exists("Warehouse", warehouse):
			continue
		seen.add(warehouse)
		catalog.append(
			{
				"warehouse": warehouse,
				"warehouse_name": frappe.db.get_value("Warehouse", warehouse, "warehouse_name")
				or warehouse,
				"department": row.department,
			}
		)

	if catalog:
		return catalog

	return [
		{
			"warehouse": row.name,
			"warehouse_name": row.warehouse_name or row.name,
			"department": None,
		}
		for row in frappe.get_all(
			"Warehouse",
			filters={"company": get_company(), "is_group": 0, "disabled": 0},
			fields=["name", "warehouse_name"],
			order_by="warehouse_name",
		)
	]


def _get_inventory_scope_context() -> dict:
	"""
	Resolve the current user's warehouse scope for inventory pages.

	Preferred source is the employee branch -> warehouse mapping. Manager/admin
	cohorts get all warehouses. If a warehouse user has no branch mapping yet, we
	fall back to the catalog so the page is still operable instead of shelling out.
	"""
	user = frappe.session.user
	user_roles = set(frappe.get_roles(user))
	catalog = _get_inventory_warehouse_catalog()
	catalog_names = {row["warehouse"] for row in catalog}

	elevated_roles = {
		"HR Manager",
		"Supply Chain Manager",
		"Warehouse Manager",
		"Logistics Coordinator",
		"System Manager",
		"Administrator",
	}
	can_view_all = bool(user_roles.intersection(elevated_roles))

	employee = frappe.db.get_value(
		"Employee",
		{"user_id": user, "status": "Active"},
		["name", "branch", "employee_name"],
		as_dict=True,
	)
	default_warehouse = _resolve_warehouse(employee.branch) if employee and employee.branch else None

	scope_source = "all_warehouses_role" if can_view_all else "catalog_fallback"
	allowed_rows = catalog

	if not can_view_all and default_warehouse:
		scope_source = "employee_branch"
		allowed_rows = [row for row in catalog if row["warehouse"] == default_warehouse]
		if not allowed_rows:
			allowed_rows = [
				{
					"warehouse": default_warehouse,
					"warehouse_name": frappe.db.get_value(
						"Warehouse", default_warehouse, "warehouse_name"
					)
					or default_warehouse,
					"department": None,
				}
			]

	if not default_warehouse and allowed_rows:
		default_warehouse = allowed_rows[0]["warehouse"]

	return {
		"user": user,
		"user_roles": sorted(user_roles),
		"can_view_all": can_view_all,
		"default_warehouse": default_warehouse,
		"scope_source": scope_source,
		"warehouses": allowed_rows,
		"catalog_warehouses": catalog,
		"warehouse_names": {row["warehouse"] for row in allowed_rows} or catalog_names,
	}


def _resolve_inventory_requested_warehouses(scope: dict, warehouse: str | None = None) -> list[str]:
	"""Validate a requested warehouse against the current user's inventory scope."""
	allowed_names = sorted(scope.get("warehouse_names") or [])
	if not allowed_names:
		return []

	if not warehouse:
		return allowed_names

	resolved = _resolve_warehouse(warehouse) or warehouse
	if resolved not in scope["warehouse_names"]:
		frappe.throw(
			_("You do not have inventory access to warehouse {0}").format(warehouse),
			frappe.PermissionError,
		)
	return [resolved]


@frappe.whitelist()
def get_inventory_scope():
	"""Return the warehouse selector contract for inventory and daily update surfaces."""
	_check_warehouse_permission(SCM_INVENTORY_ROLES, "view inventory scope")
	scope = _get_inventory_scope_context()
	return {
		"can_view_all": scope["can_view_all"],
		"default_warehouse": scope["default_warehouse"],
		"scope_source": scope["scope_source"],
		"warehouses": scope["warehouses"],
	}


@frappe.whitelist()
def get_returnable_items(store: str | None = None):
	"""
	Get items that can be returned from a store.
	Returns items with current stock in the store's warehouse.
	"""
	warehouse = _resolve_warehouse(store) if store else None

	if warehouse:
		# Get items with stock in this warehouse
		items = frappe.db.sql(
			"""
            SELECT
                b.item_code as id,
                i.item_name as name,
                i.item_group,
                i.stock_uom as uom,
                b.actual_qty as stock_qty
            FROM `tabBin` b
            JOIN `tabItem` i ON i.name = b.item_code
            WHERE b.warehouse = %s
            AND b.actual_qty > 0
            AND i.disabled = 0
            ORDER BY i.item_group, i.item_name
        """,
			warehouse,
			as_dict=True,
		)
	else:
		# Get all stock items if no store specified
		items = frappe.get_all(
			"Item",
			filters={"is_stock_item": 1, "disabled": 0},
			fields=["name as id", "item_name as name", "item_group", "stock_uom as uom"],
		)

	return {"items": items}


@frappe.whitelist()
def get_return_reasons():
	"""Get list of valid return reasons."""
	# Return from master data if exists, otherwise hardcoded list
	reasons = [
		{"value": "expired", "label": "Expired/Near Expiry"},
		{"value": "damaged", "label": "Damaged Packaging"},
		{"value": "quality", "label": "Quality Issue"},
		{"value": "overstock", "label": "Overstock/Slow Moving"},
		{"value": "recall", "label": "Product Recall"},
		{"value": "other", "label": "Other"},
	]
	return {"reasons": reasons}


@frappe.whitelist()
def submit_return_request(store: str, items: list[dict] | str, photo: str | None = None):
	"""
	Submit a return request to commissary.
	Items should be a list of {item_code, quantity, reason, notes}
	"""
	allowed_roles = ["Store Supervisor", "Store Staff", "Area Supervisor", "System Manager"]
	if not any(r in frappe.get_roles(frappe.session.user) for r in allowed_roles):
		frappe.throw(_("Not authorized to submit return requests"), frappe.PermissionError)

	if not store:
		frappe.throw(_("Store is required"))

	warehouse = _resolve_warehouse(store)
	if not warehouse:
		frappe.throw(_("Invalid store: {0}").format(store))

	if isinstance(items, str):
		items = json.loads(items)

	if not items or len(items) == 0:
		frappe.throw(_("At least one item is required"))

	# Create Stock Entry for Material Transfer (store -> commissary/scrap)
	doc = frappe.new_doc("Stock Entry")
	doc.stock_entry_type = "Material Issue"  # Issue from store
	doc.company = get_company()
	doc.custom_return_request = 1  # Custom flag to identify returns
	doc.custom_return_from_store = warehouse
	doc.custom_return_photo = photo
	doc.set_posting_time = 1
	doc.posting_date = nowdate()
	doc.posting_time = frappe.utils.nowtime()

	# Build remarks from return details
	remarks_parts = []
	for item in items:
		item_name = frappe.db.get_value("Item", item["item_code"], "item_name") or item["item_code"]
		qty = item.get("quantity") or item.get("qty")
		remarks_parts.append(f"{item_name}: {qty} ({item['reason']})")
		if item.get("notes"):
			remarks_parts.append(f"  Notes: {item['notes']}")

	doc.remarks = "Store Return Request:\n" + "\n".join(remarks_parts)

	for item in items:
		qty = item.get("quantity") or item.get("qty")
		doc.append(
			"items",
			{
				"item_code": item["item_code"],
				"qty": flt(qty),
				"s_warehouse": warehouse,
				"custom_return_reason": item["reason"],
				"custom_return_notes": item.get("notes"),
			},
		)

	doc.insert()
	doc.submit()

	total_qty = sum(flt(i.get("quantity") or i.get("qty") or 0) for i in items)
	items_count = len(items)

	return {
		"success": True,
		"name": doc.name,
		"message": _("Return request submitted successfully"),
		"data": {
			"name": doc.name,
			"items_count": items_count,
			"total_qty": total_qty,
			"movement_type": "Material Issue",
			"source_warehouse": warehouse,
		},
	}


@frappe.whitelist()
def get_return_requests(store: str | None = None, status: str | None = None, limit: int = 20):
	"""Get return request history for a store."""
	limit = min(int(limit or 20), 500)
	filters = {}

	# Bug fix C8: Check if custom columns exist before filtering
	if frappe.db.has_column("Stock Entry", "custom_return_request"):
		filters["custom_return_request"] = 1
	if store:
		warehouse = _resolve_warehouse(store)
		if warehouse and frappe.db.has_column("Stock Entry", "custom_return_from_store"):
			filters["custom_return_from_store"] = warehouse

	if status:
		filters["docstatus"] = 1 if status == "submitted" else 0

	# Get safe fields list based on column existence
	fields = ["name", "posting_date", "docstatus", "remarks", "total_outgoing_value as value"]
	if frappe.db.has_column("Stock Entry", "custom_return_from_store"):
		fields.insert(2, "custom_return_from_store as store")

	returns = frappe.get_all(
		"Stock Entry", filters=filters, fields=fields, order_by="posting_date desc", limit=int(limit)
	)

	# Add item count
	for ret in returns:
		ret["item_count"] = frappe.db.count("Stock Entry Detail", {"parent": ret["name"]})
		ret["status"] = "Submitted" if ret["docstatus"] == 1 else "Draft"

	return {"returns": returns}


# ============ STOCK COUNTING APIs (Phase 1) ============


def _map_item_group(item_group: str | None):
	"""Map Frappe Item Group hierarchy to COS RECON categories."""
	mapping = {
		"Finished Goods": "FG",
		"Raw Materials": "RM",
		"Packaging Materials": "PM",
		"Consumables": "CS",
	}
	return mapping.get(item_group, "Other")


def _check_store_access(store: str):
	"""Verify External Auditor has access to this store via BEI External Auditor Store Access."""
	if "External Auditor" not in frappe.get_roles():
		return  # Only check for External Auditors

	warehouse = _resolve_warehouse(store)
	if not warehouse:
		frappe.throw(_("Could not find Store: {0}").format(store))

	access = frappe.db.exists(
		"BEI External Auditor Store Access", {"user": frappe.session.user, "warehouse": warehouse}
	)
	if not access:
		frappe.throw(_("You do not have access to count at store {0}").format(store), frappe.PermissionError)


@frappe.whitelist()
def get_assigned_stores():
	"""Return warehouses the current user can count at.

	- External Auditors: only stores from BEI External Auditor Store Access
	- Store Staff: their branch warehouse only
	- Supervisors/HQ: all warehouses
	"""
	user = frappe.session.user
	roles = set(frappe.get_roles())
	allowed_count_types = _get_allowed_cycle_count_types(user_roles=roles)
	default_count_type = allowed_count_types[0] if allowed_count_types else None

	# External Auditor branch checked FIRST — intentional. Dual-role users (Store Staff +
	# External Auditor) see only audit-assigned stores, not their own branch. This preserves
	# audit independence. To let them count their own store, add it to Store Access explicitly.
	if "External Auditor" in roles:
		return {
			"stores": frappe.get_all(
				"BEI External Auditor Store Access",
				filters={"user": user},
				fields=["warehouse as store", "warehouse_name as store_name"],
			),
			"allowed_count_types": allowed_count_types,
			"default_count_type": default_count_type,
		}

	if "Store Staff" in roles and not roles.intersection(
		{"Store Supervisor", "Area Supervisor", "HQ User", "System Manager", "Administrator"}
	):
		emp_branch = frappe.db.get_value("Employee", {"user_id": user}, "branch")
		if emp_branch:
			warehouse = _resolve_warehouse(emp_branch)
			if warehouse:
				return {
					"stores": [{"store": warehouse, "store_name": emp_branch}],
					"allowed_count_types": allowed_count_types,
					"default_count_type": default_count_type,
				}
		return {
			"stores": [],
			"allowed_count_types": allowed_count_types,
			"default_count_type": default_count_type,
		}

	# Supervisors, HQ, Admin — all warehouses
	return {
		"stores": frappe.get_all(
			"Warehouse",
			filters={"company": get_company(), "is_group": 0},
			fields=["name as store", "warehouse_name as store_name"],
			order_by="warehouse_name",
		),
		"allowed_count_types": allowed_count_types,
		"default_count_type": default_count_type,
	}


@frappe.whitelist()
def get_items_for_count(store: str, count_type: str | None = None):
	"""Return ALL stock items for a warehouse, grouped by category.

	Design Decision (2026-02-20):
	- Loads ALL items from Item Master (not just 90-day SLE)
	- system_qty is always 0 (blind count — counter must not see expected qty)
	- Real system_qty is populated server-side in validate() at submit time
	- This prevents cheating and ensures auditors count items the system says don't exist
	"""
	try:
		# External Auditors: verify store access
		_check_store_access(store)
		validated_count_type = _validate_cycle_count_type(count_type)

		warehouse = _resolve_warehouse(store)
		if not warehouse:
			frappe.throw(_("No warehouse found for store {0}").format(store))

		# Load ALL stock items — blind count (system_qty always 0)
		items = frappe.db.sql(
			"""
            SELECT DISTINCT
                i.name as item_code, i.item_name, i.item_group,
                i.stock_uom as uom,
                0 as system_qty,
                COALESCE(ucd.conversion_factor, 1.0) as conversion_factor
            FROM `tabItem` i
            LEFT JOIN `tabUOM Conversion Detail` ucd
                ON ucd.parent = i.name AND ucd.uom = i.stock_uom
            WHERE i.is_stock_item = 1
              AND i.disabled = 0
            ORDER BY i.item_group, i.item_name
        """,
			as_dict=1,
		)

		# Group by category for frontend display
		grouped = {}
		for item in items:
			group = _map_item_group(item.item_group)
			grouped.setdefault(group, []).append(item)

		return {
			"items": items,
			"grouped": grouped,
			"warehouse": warehouse,
			"count": len(items),
			"count_type": validated_count_type,
		}
	except frappe.PermissionError:
		raise
	except frappe.ValidationError:
		raise
	except Exception as e:
		frappe.log_error(f"get_items_for_count failed for store={store}: {e}")
		frappe.throw(_("Failed to load items for counting. Please try again."))


@frappe.whitelist()
def submit_cycle_count_v2(
	store: str,
	count_date: str,
	items: list[dict] | str,
	count_type: str = COUNT_TYPE_STORE_MONTHLY,
	photo_url: str | None = None,
	name: str | None = None,
):
	"""Create or submit cycle count with WHOLE + LOOSE quantities.

	Args:
	    items: list of {item_code, counted_qty_whole, counted_qty_loose}
	    photo_url: File URL from prior upload via /api/method/upload_file (AUDIT-6)
	    name: Optional — if provided, submits an existing Draft instead of creating new
	"""
	# Permission check
	if "External Auditor" in frappe.get_roles():
		_check_store_access(store)

	if isinstance(items, str):
		items = json.loads(items)

	count_type = _validate_cycle_count_type(count_type)

	# Resolve warehouse
	warehouse = _resolve_warehouse(store)
	if not warehouse:
		frappe.throw(_("Could not find Store: {0}").format(store))

	try:
		# AUDIT-3: DM-2 compliance — savepoint around insert+submit
		frappe.db.savepoint("cycle_count_submit")

		if name:
			# B-8: Submit existing draft
			doc = frappe.get_doc("BEI Cycle Count", name)
			if doc.docstatus != 0:
				frappe.throw(_("This count has already been submitted."))
			if doc.counted_by != frappe.session.user:
				frappe.throw(_("You can only submit your own draft."))
			# Clear existing items and replace with new data
			doc.items = []
			for item_data in items:
				uom = frappe.db.get_value("Item", item_data["item_code"], "stock_uom") or ""
				doc.append(
					"items",
					{
						"item_code": item_data["item_code"],
						"uom": uom,
						"counted_qty_whole": item_data.get("counted_qty_whole", 0),
						"counted_qty_loose": item_data.get("counted_qty_loose", 0.0),
					},
				)
			doc.store = warehouse
			doc.count_date = count_date
			doc.count_type = count_type
			if photo_url:
				doc.photo_evidence = photo_url
			doc.save(ignore_permissions=True)
			doc.flags.ignore_permissions = True
			doc.submit()
		else:
			# AUDIT-8: Duplicate detection — exclude Rejected/Resubmitted (resubmissions allowed)
			existing = frappe.db.sql(
				"""
                SELECT name FROM `tabBEI Cycle Count`
                WHERE store = %s AND count_date = %s AND count_type = %s
                  AND docstatus = 1
                  AND status NOT IN ('Rejected', 'Resubmitted')
                LIMIT 1
            """,
				(warehouse, count_date, count_type),
			)
			if existing:
				frappe.throw(
					f"A cycle count already exists for {store} on {count_date} ({count_type}). "
					f"Please use the existing record: {existing[0][0]}"
				)

			doc = frappe.new_doc("BEI Cycle Count")
			doc.store = warehouse
			doc.count_date = count_date
			doc.count_type = count_type
			doc.counted_by = frappe.session.user

			if "External Auditor" in frappe.get_roles():
				doc.external_auditor = frappe.session.user

			# AUDIT-6: Photo is referenced by URL (uploaded separately via multipart/form-data)
			if photo_url:
				doc.photo_evidence = photo_url

			for item_data in items:
				uom = frappe.db.get_value("Item", item_data["item_code"], "stock_uom") or ""
				doc.append(
					"items",
					{
						"item_code": item_data["item_code"],
						"uom": uom,
						"counted_qty_whole": item_data.get("counted_qty_whole", 0),
						"counted_qty_loose": item_data.get("counted_qty_loose", 0.0),
					},
				)
			# validate() computes counted_qty, system_qty from Bin, unit_cost, variance automatically

			doc.insert(ignore_permissions=True)
			doc.flags.ignore_permissions = True
			doc.submit()

			# Link orphan photo file to newly created doc (two-step upload pattern)
			if photo_url:
				frappe.db.set_value(
					"File",
					{"file_url": photo_url, "attached_to_name": ["is", "not set"]},
					{"attached_to_doctype": "BEI Cycle Count", "attached_to_name": doc.name},
				)
	except Exception:
		frappe.db.rollback(save_point="cycle_count_submit")
		raise

	return {"name": doc.name, "status": "Submitted", "items_count": len(doc.items)}


@frappe.whitelist()
def save_cycle_count_draft(
	store: str,
	count_date: str,
	items: list[dict] | str,
	count_type: str = COUNT_TYPE_STORE_MONTHLY,
	photo_url: str | None = None,
	name: str | None = None,
):
	"""Save cycle count as draft (not submitted). Supports create and update."""
	if "External Auditor" in frappe.get_roles():
		_check_store_access(store)

	if isinstance(items, str):
		items = json.loads(items)

	count_type = _validate_cycle_count_type(count_type)

	warehouse = _resolve_warehouse(store)
	if not warehouse:
		frappe.throw(_("Could not find Store: {0}").format(store))

	if name:
		# Update existing draft
		doc = frappe.get_doc("BEI Cycle Count", name)
		if doc.docstatus != 0:
			frappe.throw(_("Cannot update a submitted cycle count"))
		doc.items = []
	else:
		doc = frappe.new_doc("BEI Cycle Count")

	doc.store = warehouse
	doc.count_date = count_date
	doc.count_type = count_type
	doc.counted_by = frappe.session.user
	doc.status = "Draft"

	if "External Auditor" in frappe.get_roles():
		doc.external_auditor = frappe.session.user

	if photo_url:
		doc.photo_evidence = photo_url

	for item_data in items:
		doc.append(
			"items",
			{
				"item_code": item_data["item_code"],
				"counted_qty_whole": item_data.get("counted_qty_whole", 0),
				"counted_qty_loose": item_data.get("counted_qty_loose", 0.0),
			},
		)

	doc.save()
	return {"name": doc.name, "status": "Draft", "items_count": len(doc.items)}


@frappe.whitelist()
def export_count_to_cos_recon(cycle_count_name: str):
	"""Generate COS RECON Excel and attach to cycle count record."""
	import openpyxl

	from frappe.utils.file_manager import save_file

	from hrms.utils.cos_recon_schema import ACTIVE_VERSION

	doc = frappe.get_doc("BEI Cycle Count", cycle_count_name)
	doc.check_permission("read")

	# B-4a: Status gate — only export approved/verified/reconciled counts
	if doc.status not in ("Approved", "Verified", "Reconciled"):
		frappe.throw(_("Only approved counts can be exported. Current status: {0}").format(doc.status))

	# B-4b: Role gate — HQ/Finance only
	allowed_roles = {"HQ User", "System Manager", "Administrator"}
	user_roles = set(frappe.get_roles())
	if not user_roles.intersection(allowed_roles):
		frappe.throw(_("Only HQ/Finance can export cycle counts."), frappe.PermissionError)

	filename = f"COS_RECON_{doc.store}_{doc.count_date}.xlsx"

	# Idempotency: return existing export if already done
	if doc.exported_to_cos_recon and doc.export_date:
		existing = frappe.get_all(
			"File", {"attached_to_name": doc.name, "file_name": ["like", "COS_RECON%"]}, ["file_url"]
		)
		if existing:
			return {"file_url": existing[0].file_url, "filename": filename, "already_exported": True}

	# Prefetch all item data in one query (fixes N+1)
	item_codes = [item.item_code for item in doc.items]
	item_data = {}
	if item_codes:
		for d in frappe.get_all(
			"Item",
			filters={"name": ["in", item_codes]},
			fields=["name", "item_name", "item_group", "description"],
		):
			item_data[d.name] = d

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = doc.store

	# Write headers from versioned schema
	col_map = {}
	for col_def in ACTIVE_VERSION["columns"]:
		col_idx = ord(col_def["col"]) - ord("A") + 1  # C=3, D=4, etc.
		ws.cell(row=1, column=col_idx, value=col_def["header"])
		col_map[col_def["field"]] = col_idx

	# Write items sorted by category
	row = 2
	for item in sorted(
		doc.items, key=lambda x: _map_item_group((item_data.get(x.item_code) or {}).get("item_group", ""))
	):
		meta = item_data.get(item.item_code) or {}
		ws.cell(row=row, column=col_map["item_code"], value=item.item_code)
		ws.cell(row=row, column=col_map["item_name"], value=meta.get("item_name", ""))
		ws.cell(row=row, column=col_map["description"], value=meta.get("description", ""))
		ws.cell(
			row=row, column=col_map["grams"], value=0
		)  # TODO: populate from item properties when available
		ws.cell(row=row, column=col_map["uom"], value=item.uom)
		ws.cell(row=row, column=col_map["counted_qty_whole"], value=item.counted_qty_whole or 0)
		ws.cell(row=row, column=col_map["counted_qty_loose"], value=item.counted_qty_loose or 0.0)
		ws.cell(row=row, column=col_map["unit_cost"], value=item.unit_cost or 0.0)
		ws.cell(row=row, column=col_map["total_cost"], value=(item.counted_qty or 0) * (item.unit_cost or 0))
		row += 1

	# Save to bytes and attach via File DocType
	from io import BytesIO

	buffer = BytesIO()
	wb.save(buffer)
	buffer.seek(0)

	file_doc = save_file(filename, buffer.read(), "BEI Cycle Count", doc.name, is_private=1)

	# Update flags atomically
	frappe.db.set_value(
		"BEI Cycle Count", doc.name, {"exported_to_cos_recon": 1, "export_date": frappe.utils.today()}
	)

	return {"file_url": file_doc.file_url, "filename": filename}


# ============ PHASE 2A: REAL-TIME WAREHOUSE STOCK APIs ============
# Gives Ian real-time stock visibility across 6 BEI warehouses:
# 3MD Cold, 3MD Dry, JENTEC, RCS, PINNACLE, SHAW


@frappe.whitelist()
def get_warehouse_stock(
	warehouse: str | None = None,
	item_group: str | None = None,
	include_zero_stock: int | str = 0,
):
	"""
	GET stock levels across warehouses.
	If warehouse is specified, filter to that warehouse only.

	Returns: [{item_code, item_name, item_group, warehouse, actual_qty,
	            reserved_qty, available_qty, reorder_point, is_low_stock}]
	"""
	_check_warehouse_permission(SCM_INVENTORY_ROLES, "view warehouse stock")
	scope = _get_inventory_scope_context()
	warehouse_names = _resolve_inventory_requested_warehouses(scope, warehouse)
	if not warehouse_names:
		return []

	conditions = ["b.warehouse IN %(warehouses)s"]
	values = {"warehouses": tuple(warehouse_names)}

	if not _coerce_flag(include_zero_stock):
		conditions.append("(b.actual_qty > 0 OR b.reserved_qty > 0)")

	if warehouse:
		values["warehouse"] = warehouse_names[0]

	if item_group:
		conditions.append("i.item_group = %(item_group)s")
		values["item_group"] = item_group

	where_clause = " AND ".join(conditions)

	query = (
		"""
        SELECT
            b.item_code,
            i.item_name,
            i.item_group,
            b.warehouse,
            b.actual_qty,
            b.reserved_qty,
            (b.actual_qty - b.reserved_qty) AS available_qty,
            COALESCE(ir.warehouse_reorder_level, 0) AS reorder_point,
            CASE
                WHEN ir.warehouse_reorder_level > 0
                     AND b.actual_qty <= ir.warehouse_reorder_level THEN 1
                ELSE 0
            END AS is_low_stock
        FROM `tabBin` b
        INNER JOIN `tabItem` i ON i.name = b.item_code
        LEFT JOIN `tabItem Reorder` ir
            ON ir.parent = b.item_code
            AND ir.warehouse = b.warehouse
        WHERE """
		+ where_clause
		+ """
        ORDER BY b.warehouse, i.item_name
    """
	)
	stock = frappe.db.sql(query, values, as_dict=True)

	return stock


@frappe.whitelist()
def daily_stock_update(warehouse: str, items: list[dict] | str):
	"""
	POST daily SOH (Stock on Hand) update for a warehouse.
	Creates a Stock Reconciliation entry in Frappe.

	Args:
	    warehouse (str): Target warehouse name
	    items (list|str): [{item_code, qty}]

	Returns: {name, status, warehouse, item_count}
	"""
	_check_warehouse_permission(SCM_STOCK_UPDATE_ROLES, "submit daily stock update")
	scope = _get_inventory_scope_context()

	if isinstance(items, str):
		items = json.loads(items)

	if not items:
		frappe.throw(_("No items provided for stock update"))

	if not warehouse:
		frappe.throw(_("Warehouse is required"))

	allowed_warehouses = _resolve_inventory_requested_warehouses(scope, warehouse)
	if not allowed_warehouses:
		frappe.throw(_("No warehouse scope is assigned to this account"), frappe.PermissionError)
	warehouse = allowed_warehouses[0]

	frappe.db.savepoint("daily_stock_update_start")

	try:
		sr = frappe.new_doc("Stock Reconciliation")
		sr.purpose = "Stock Reconciliation"
		sr.posting_date = nowdate()
		sr.posting_time = now_datetime().strftime("%H:%M:%S")
		sr.company = frappe.defaults.get_global_default("company")
		sr.remarks = f"BEI Daily SOH Update — Warehouse: {warehouse} — Submitted via my.bebang.ph"

		for row in items:
			item_code = row.get("item_code")
			qty = flt(row.get("qty", 0))
			if not item_code:
				continue
			sr.append(
				"items",
				{
					"item_code": item_code,
					"warehouse": warehouse,
					"qty": qty,
				},
			)

		if not sr.items:
			frappe.throw(_("No valid items to reconcile"))

		sr.insert(ignore_permissions=True)
		sr.submit()

		return {"name": sr.name, "status": "Submitted", "warehouse": warehouse, "item_count": len(sr.items)}

	except Exception:
		frappe.db.rollback(save_point="daily_stock_update_start")
		raise


@frappe.whitelist()
def get_low_stock_alerts(warehouse: str | None = None):
	"""
	GET items currently below their reorder point.

	Returns: [{item_code, item_name, item_group, warehouse, actual_qty,
	            reorder_point, shortage, suggested_order_qty}]
	"""
	_check_warehouse_permission(SCM_INVENTORY_ROLES, "view low stock alerts")
	scope = _get_inventory_scope_context()
	warehouse_names = _resolve_inventory_requested_warehouses(scope, warehouse)
	if not warehouse_names:
		return []

	conditions = [
		"b.warehouse IN %(warehouses)s",
		"ir.warehouse_reorder_level > 0",
		"b.actual_qty <= ir.warehouse_reorder_level",
	]
	values = {"warehouses": tuple(warehouse_names)}

	if warehouse:
		values["warehouse"] = warehouse_names[0]

	where_clause = " AND ".join(conditions)

	query = (
		"""
        SELECT
            b.item_code,
            i.item_name,
            i.item_group,
            b.warehouse,
            b.actual_qty,
            ir.warehouse_reorder_level AS reorder_point,
            (ir.warehouse_reorder_level - b.actual_qty) AS shortage,
            ir.warehouse_reorder_qty AS suggested_order_qty
        FROM `tabBin` b
        INNER JOIN `tabItem` i ON i.name = b.item_code
        INNER JOIN `tabItem Reorder` ir
            ON ir.parent = b.item_code
            AND ir.warehouse = b.warehouse
        WHERE """
		+ where_clause
		+ """
        ORDER BY shortage DESC, b.warehouse, i.item_name
    """
	)
	alerts = frappe.db.sql(query, values, as_dict=True)

	return alerts


@frappe.whitelist()
def get_multi_warehouse_summary():
	"""
	GET summary across all 6 BEI warehouses for the inventory dashboard.

	Returns: [{warehouse, total_items, low_stock_count, last_updated}]
	"""
	_check_warehouse_permission(SCM_INVENTORY_ROLES, "view warehouse summary")
	scope = _get_inventory_scope_context()
	warehouse_rows = scope["warehouses"]
	warehouse_names = sorted(scope["warehouse_names"])
	if not warehouse_names:
		return []

	summary_rows = frappe.db.sql(
		"""
        SELECT
            b.warehouse,
            COUNT(
                DISTINCT CASE
                    WHEN (b.actual_qty > 0 OR b.reserved_qty > 0) THEN b.item_code
                    ELSE NULL
                END
            ) AS total_items,
            COALESCE(SUM(
                CASE
                    WHEN ir.warehouse_reorder_level > 0
                         AND b.actual_qty <= ir.warehouse_reorder_level THEN 1
                    ELSE 0
                END
            ), 0) AS low_stock_count,
            MAX(sle.posting_datetime) AS last_updated
        FROM `tabBin` b
        LEFT JOIN `tabItem Reorder` ir
            ON ir.parent = b.item_code
            AND ir.warehouse = b.warehouse
        LEFT JOIN `tabStock Ledger Entry` sle
            ON sle.item_code = b.item_code
            AND sle.warehouse = b.warehouse
            AND sle.is_cancelled = 0
        WHERE b.warehouse IN %(warehouses)s
        GROUP BY b.warehouse
        ORDER BY b.warehouse
    """,
		{"warehouses": tuple(warehouse_names)},
		as_dict=True,
	)

	summary_map = {row.warehouse: row for row in summary_rows}
	return [
		{
			"warehouse": row["warehouse"],
			"warehouse_name": row["warehouse_name"],
			"total_items": cint(getattr(summary_map.get(row["warehouse"]), "total_items", 0) or 0),
			"low_stock_count": cint(
				getattr(summary_map.get(row["warehouse"]), "low_stock_count", 0) or 0
			),
			"last_updated": getattr(summary_map.get(row["warehouse"]), "last_updated", None),
		}
		for row in warehouse_rows
	]


@frappe.whitelist()
def get_network_inventory_overview(
	warehouse: str | None = None,
	item_group: str | None = None,
	query: str | None = None,
	limit: int | str = 100,
	include_zero_stock: int | str = 0,
):
	"""Return a bird's-eye item x warehouse inventory view for the current scope."""
	_check_warehouse_permission(SCM_INVENTORY_ROLES, "view network inventory overview")
	scope = _get_inventory_scope_context()
	warehouse_names = _resolve_inventory_requested_warehouses(scope, warehouse)
	if not warehouse_names:
		return {
			"summary": {
				"total_on_hand": 0,
				"total_reserved": 0,
				"total_available": 0,
				"warehouse_count": 0,
				"item_count": 0,
				"low_stock_items": 0,
				"stockout_items": 0,
			},
			"items": [],
		}

	limit = min(max(cint(limit or 100), 1), 250)
	conditions = ["b.warehouse IN %(warehouses)s"]
	values: dict[str, object] = {"warehouses": tuple(warehouse_names)}

	if item_group:
		conditions.append("i.item_group = %(item_group)s")
		values["item_group"] = item_group

	if query:
		values["query"] = f"%{query.strip()}%"
		conditions.append("(b.item_code LIKE %(query)s OR i.item_name LIKE %(query)s)")

	if not _coerce_flag(include_zero_stock):
		conditions.append("(b.actual_qty > 0 OR b.reserved_qty > 0 OR COALESCE(ir.warehouse_reorder_level, 0) > 0)")

	where_clause = " AND ".join(conditions)

	summary = frappe.db.sql(
		"""
        SELECT
            COALESCE(SUM(b.actual_qty), 0) AS total_on_hand,
            COALESCE(SUM(b.reserved_qty), 0) AS total_reserved,
            COALESCE(SUM(b.actual_qty - b.reserved_qty), 0) AS total_available,
            COUNT(DISTINCT b.warehouse) AS warehouse_count,
            COUNT(DISTINCT b.item_code) AS item_count,
            COUNT(
                DISTINCT CASE
                    WHEN ir.warehouse_reorder_level > 0
                         AND b.actual_qty <= ir.warehouse_reorder_level THEN b.item_code
                    ELSE NULL
                END
            ) AS low_stock_items,
            COUNT(
                DISTINCT CASE
                    WHEN ir.warehouse_reorder_level > 0
                         AND b.actual_qty <= 0 THEN b.item_code
                    ELSE NULL
                END
            ) AS stockout_items
        FROM `tabBin` b
        INNER JOIN `tabItem` i ON i.name = b.item_code
        LEFT JOIN `tabItem Reorder` ir
            ON ir.parent = b.item_code
            AND ir.warehouse = b.warehouse
        WHERE """
		+ where_clause,
		values,
		as_dict=True,
	)[0]

	rows = frappe.db.sql(
		"""
        SELECT
            b.item_code,
            i.item_name,
            i.item_group,
            b.warehouse,
            b.actual_qty,
            b.reserved_qty,
            (b.actual_qty - b.reserved_qty) AS available_qty,
            COALESCE(ir.warehouse_reorder_level, 0) AS reorder_point,
            CASE
                WHEN ir.warehouse_reorder_level > 0
                     AND b.actual_qty <= ir.warehouse_reorder_level THEN 1
                ELSE 0
            END AS is_low_stock
        FROM `tabBin` b
        INNER JOIN `tabItem` i ON i.name = b.item_code
        LEFT JOIN `tabItem Reorder` ir
            ON ir.parent = b.item_code
            AND ir.warehouse = b.warehouse
        WHERE """
		+ where_clause
		+ """
        ORDER BY i.item_name, b.warehouse
    """,
		values,
		as_dict=True,
	)

	items_map: dict[str, dict] = {}
	for row in rows:
		item = items_map.setdefault(
			row.item_code,
			{
				"item_code": row.item_code,
				"item_name": row.item_name,
				"item_group": row.item_group,
				"total_on_hand": 0.0,
				"total_reserved": 0.0,
				"total_available": 0.0,
				"warehouse_count": 0,
				"low_stock_warehouse_count": 0,
				"stockout_warehouse_count": 0,
				"warehouses": [],
			},
		)
		item["total_on_hand"] += flt(row.actual_qty)
		item["total_reserved"] += flt(row.reserved_qty)
		item["total_available"] += flt(row.available_qty)
		item["warehouse_count"] += 1
		item["low_stock_warehouse_count"] += cint(row.is_low_stock or 0)
		if flt(row.actual_qty) <= 0:
			item["stockout_warehouse_count"] += 1
		item["warehouses"].append(
			{
				"warehouse": row.warehouse,
				"actual_qty": flt(row.actual_qty),
				"reserved_qty": flt(row.reserved_qty),
				"available_qty": flt(row.available_qty),
				"reorder_point": flt(row.reorder_point),
				"is_low_stock": cint(row.is_low_stock or 0),
			}
		)

	items = sorted(
		items_map.values(),
		key=lambda row: (
			-row["stockout_warehouse_count"],
			-row["low_stock_warehouse_count"],
			row["item_name"],
		),
	)[:limit]

	return {
		"summary": {
			"total_on_hand": flt(summary.total_on_hand),
			"total_reserved": flt(summary.total_reserved),
			"total_available": flt(summary.total_available),
			"warehouse_count": cint(summary.warehouse_count),
			"item_count": cint(summary.item_count),
			"low_stock_items": cint(summary.low_stock_items),
			"stockout_items": cint(summary.stockout_items),
		},
		"items": items,
	}


@frappe.whitelist()
def get_item_stock_history(item_code: str, warehouse: str | None = None, days: int = 30):
	"""
	GET stock movement history for an item from Stock Ledger Entry.

	Args:
	    item_code (str): Item code to query
	    warehouse (str, optional): Filter to a specific warehouse
	    days (int): Days back to query (default: 30, max: 365)

	Returns: [{date, posting_datetime, warehouse, voucher_type, voucher_no,
	            actual_qty, qty_after_transaction, stock_value_difference}]
	"""
	_check_warehouse_permission(SCM_INVENTORY_ROLES, "view stock history")
	scope = _get_inventory_scope_context()
	warehouse_names = _resolve_inventory_requested_warehouses(scope, warehouse)
	if not warehouse_names:
		return []

	days = min(int(days), 365)
	from_date = add_days(nowdate(), -days)

	conditions = [
		"sle.item_code = %(item_code)s",
		"sle.posting_date >= %(from_date)s",
		"sle.is_cancelled = 0",
		"sle.warehouse IN %(warehouses)s",
	]
	values = {"item_code": item_code, "from_date": from_date, "warehouses": tuple(warehouse_names)}

	if warehouse:
		values["warehouse"] = warehouse_names[0]

	where_clause = " AND ".join(conditions)

	query = (
		"""
        SELECT
            sle.posting_date AS `date`,
            sle.posting_datetime,
            sle.warehouse,
            sle.voucher_type,
            sle.voucher_no,
            sle.actual_qty,
            sle.qty_after_transaction,
            sle.stock_value_difference
        FROM `tabStock Ledger Entry` sle
        WHERE """
		+ where_clause
		+ """
        ORDER BY sle.posting_datetime DESC
        LIMIT 500
    """
	)
	history = frappe.db.sql(query, values, as_dict=True)

	return history


@frappe.whitelist()
def notify_gr_completion(gr_name: str):
	"""
	POST hook: called after a BEI Goods Receipt is submitted.
	Sends a Google Chat notification to Ian's and Jay's spaces.
	"""
	_check_warehouse_permission(SCM_INVENTORY_ROLES, "trigger GR notifications")

	gr = frappe.get_doc("BEI Goods Receipt", gr_name)

	lines = [
		"*Goods Receipt Submitted* :white_check_mark:",
		"",
		f"*GR No:* {gr.gr_no or gr.name}",
		f"*Supplier:* {gr.supplier_name or gr.supplier or 'N/A'}",
		f"*Warehouse:* {gr.warehouse or 'N/A'}",
		f"*Receipt Date:* {gr.receipt_date}",
		f"*PO Reference:* {gr.purchase_order or 'N/A'}",
		f"*Total Amount:* \u20b1{flt(gr.total_amount):,.2f}",
		f"*Total Received Qty:* {flt(gr.total_received_qty)}",
		"",
		f"Submitted by: {frappe.session.user}",
	]

	_send_warehouse_gchat_notification("\n".join(lines))

	return {"status": "notified", "gr": gr_name}


def _send_warehouse_gchat_notification(text: str):
	"""
	Send a Google Chat message for warehouse/inventory alerts.
	Delegates to shared send_message_to_space() utility.
	"""
	from hrms.api.google_chat import send_message_to_space
	from hrms.utils.bei_config import SPACE_NOTIFICATIONS, get_chat_space

	send_message_to_space(get_chat_space(SPACE_NOTIFICATIONS), text)


def send_low_stock_daily_alert():
	"""
	Daily scheduler job: check all warehouses for items below reorder point
	and send a GChat summary alert if any are found.

	Registered in hooks.py under scheduler_events["daily"].
	"""
	try:
		alerts = frappe.db.sql(
			"""
            SELECT
                b.item_code,
                i.item_name,
                b.warehouse,
                b.actual_qty,
                ir.warehouse_reorder_level AS reorder_point,
                (ir.warehouse_reorder_level - b.actual_qty) AS shortage
            FROM `tabBin` b
            INNER JOIN `tabItem` i ON i.name = b.item_code
            INNER JOIN `tabItem Reorder` ir
                ON ir.parent = b.item_code
                AND ir.warehouse = b.warehouse
            WHERE ir.warehouse_reorder_level > 0
              AND b.actual_qty <= ir.warehouse_reorder_level
            ORDER BY shortage DESC
            LIMIT 50
        """,
			as_dict=True,
		)

		if not alerts:
			return

		lines = [
			"*Daily Low Stock Alert* :warning:",
			f"*Date:* {nowdate()}",
			f"*Items below reorder point:* {len(alerts)}",
			"",
		]

		for a in alerts[:20]:
			lines.append(
				"\u2022 *{item_name}* ({item_code}) @ {warehouse} \u2014 "
				"Stock: {actual_qty}, Reorder at: {reorder_point}, "
				"Short by: {shortage}".format(**a)
			)

		if len(alerts) > 20:
			lines.append(f"... and {len(alerts) - 20} more items")

		_send_warehouse_gchat_notification("\n".join(lines))

	except Exception:
		frappe.log_error(frappe.get_traceback(), "Low Stock Daily Alert Failed")


# ================================
# VARIANCE INVESTIGATION WORKFLOW
# ================================


@frappe.whitelist()
def get_open_variances(store: str | None = None):
	"""Get open inventory variances for a store (or all stores).

	Used by store dashboard to show count of unresolved variances.

	Args:
	    store: Optional warehouse name. If None, returns all open variances.

	Returns:
	    dict: {variances, count}
	"""
	filters = {"status": ["in", ["Open", "Investigating"]]}
	if store:
		filters["store"] = store

	variances = frappe.get_all(
		"BEI Inventory Variance",
		filters=filters,
		fields=[
			"name",
			"store",
			"variance_date",
			"item_code",
			"variance_type",
			"status",
			"variance_qty",
			"variance_value",
			"reported_by",
		],
		order_by="variance_date desc",
	)

	return {
		"variances": variances,
		"count": len(variances),
		"store": store,
	}


@frappe.whitelist()
def start_variance_investigation(variance_name: str):
	"""Transition a variance from Open to Investigating.

	Args:
	    variance_name: BEI Inventory Variance name

	Returns:
	    dict: {success, status}
	"""
	allowed_roles = {"Store Supervisor", "Area Supervisor", "Warehouse User", "System Manager"}
	if not set(frappe.get_roles()).intersection(allowed_roles):
		frappe.throw(
			_("You do not have permission to manage variance investigations"), frappe.PermissionError
		)

	doc = frappe.get_doc("BEI Inventory Variance", variance_name)

	if doc.status != "Open":
		frappe.throw(_("Only 'Open' variances can be moved to Investigating"))

	doc.status = "Investigating"
	doc.save(ignore_permissions=True)

	return {"success": True, "name": variance_name, "status": doc.status}


@frappe.whitelist()
def resolve_variance(
	variance_name: str,
	resolution_type: str,
	resolution_notes: str,
	adjustment_qty: float | int | None = None,
):
	"""Resolve a variance and optionally create a Stock Entry for write-offs.

	Resolution types:
	  - "Write-Off": Creates Stock Entry (Material Issue) to remove stock
	  - "Recount Corrected": Creates Stock Reconciliation
	  - "Theft", "Damage", "System Error": Records resolution, no stock entry

	Args:
	    variance_name: BEI Inventory Variance name
	    resolution_type: One of Write-Off / Recount Corrected / Theft / Damage / System Error
	    resolution_notes: Required explanation
	    adjustment_qty: Override quantity for stock adjustment (defaults to variance_qty)

	Returns:
	    dict: {success, status, stock_entry} (stock_entry may be None)
	"""
	allowed_roles = {"Store Supervisor", "Area Supervisor", "Warehouse User", "System Manager"}
	if not set(frappe.get_roles()).intersection(allowed_roles):
		frappe.throw(
			_("You do not have permission to manage variance investigations"), frappe.PermissionError
		)

	valid_types = {"Write-Off", "Recount Corrected", "Theft", "Damage", "System Error"}
	if resolution_type not in valid_types:
		frappe.throw(_("Invalid resolution type. Must be one of: {0}").format(", ".join(valid_types)))

	if not resolution_notes:
		frappe.throw(_("Resolution notes are required"))

	doc = frappe.get_doc("BEI Inventory Variance", variance_name)

	if doc.status not in ("Open", "Investigating"):
		frappe.throw(_("Cannot resolve a variance with status '{0}'").format(doc.status))

	if not doc.item_code:
		frappe.throw(_("Variance has no item code — cannot create stock entry"))

	stock_entry_name = None
	qty = flt(adjustment_qty) if adjustment_qty else flt(doc.variance_qty)

	sp = frappe.db.savepoint(f"resolve_variance_{variance_name.replace('-', '_')}")
	try:
		# Create Stock Entry for Write-Off (removes stock from store)
		if resolution_type == "Write-Off" and qty and qty > 0:
			company = get_company()
			se = frappe.new_doc("Stock Entry")
			se.stock_entry_type = "Material Issue"
			se.company = company
			se.posting_date = nowdate()
			se.remarks = f"Variance write-off: {variance_name} — {resolution_notes}"

			se.append(
				"items",
				{
					"item_code": doc.item_code,
					"s_warehouse": doc.store,
					"qty": qty,
					"uom": frappe.db.get_value("Item", doc.item_code, "stock_uom") or "Nos",
				},
			)

			with _run_as_system_user():
				se.insert(ignore_permissions=True)
				se.flags.ignore_permissions = True
				se.submit()
			stock_entry_name = se.name
			doc.status = "Written Off"

		elif resolution_type == "Recount Corrected" and qty is not None:
			# Stock Reconciliation to set system qty to actual
			company = get_company()
			sr = frappe.new_doc("Stock Reconciliation")
			sr.company = company
			sr.posting_date = nowdate()
			sr.purpose = "Stock Reconciliation"
			sr.remarks = f"Variance recount correction: {variance_name} — {resolution_notes}"

			sr.append(
				"items",
				{
					"item_code": doc.item_code,
					"warehouse": doc.store,
					"qty": flt(doc.actual_qty) if doc.actual_qty else (flt(doc.system_qty) - qty),
				},
			)

			with _run_as_system_user():
				sr.insert(ignore_permissions=True)
				sr.flags.ignore_permissions = True
				sr.submit()
			stock_entry_name = sr.name
			doc.status = "Resolved"

		elif resolution_type in ("Theft", "Damage"):
			# NG-01 fix: Theft/Damage must create Material Issue to remove phantom inventory.
			# Without this, stock ledger retains quantity that no longer physically exists.
			company = get_company()
			se = frappe.new_doc("Stock Entry")
			se.stock_entry_type = "Material Issue"
			se.company = company
			se.posting_date = nowdate()
			se.remarks = f"Variance {resolution_type.lower()}: {variance_name} — {resolution_notes}"

			se.append(
				"items",
				{
					"item_code": doc.item_code,
					"s_warehouse": doc.store,
					"qty": qty if qty and qty > 0 else abs(flt(doc.variance_qty)),
					"uom": frappe.db.get_value("Item", doc.item_code, "stock_uom") or "Nos",
				},
			)

			with _run_as_system_user():
				se.insert(ignore_permissions=True)
				se.flags.ignore_permissions = True
				se.submit()
			stock_entry_name = se.name
			doc.status = "Resolved"

		else:
			# System Error — no stock adjustment, just mark resolved
			doc.status = "Resolved"

		# Record resolution on the variance doc
		doc.resolution = f"[{resolution_type}] {resolution_notes}"
		doc.save(ignore_permissions=True)

	except Exception as e:
		frappe.db.rollback(save_point=sp)
		frappe.log_error(frappe.get_traceback(), f"resolve_variance failed for {variance_name}")
		frappe.throw(_("Failed to resolve variance: {0}").format(str(e)))

	return {
		"success": True,
		"name": variance_name,
		"status": doc.status,
		"resolution_type": resolution_type,
		"stock_entry": stock_entry_name,
	}


@frappe.whitelist()
def get_variance_resolution_contract():
	"""Expose supported variance-resolution contract for frontend wiring."""
	return {
		"success": True,
		"action": "resolve_variance",
		"required_fields": ["variance_name", "resolution_type", "resolution_notes"],
		"optional_fields": ["adjustment_qty"],
		"resolution_types": ["Write-off", "Recount Corrected", "Theft", "Damage", "System Error"],
	}
