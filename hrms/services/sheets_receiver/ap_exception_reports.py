"""
AP exception report generation for the live SUPPLIERS SOA workbook.

Builds a daily/actionable accounting workbook from the authoritative sheet
instead of relying on a hand-maintained local copy.
"""

from __future__ import annotations

import csv
import json
import logging
import shutil
from collections import Counter
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import get_config, get_sheet_config
from .models import SyncLog, get_db
from .sheets_client import get_sheets_client

logger = logging.getLogger(__name__)

FIELDNAMES = [
	"sheet_row",
	"date_entry",
	"entered_by",
	"billed_to",
	"acctg_status",
	"fin_status",
	"invoice_no",
	"invoice_date",
	"terms",
	"supplier_name",
	"category",
	"particulars",
	"rfp_id",
	"status",
	"payment_date",
	"amount",
	"payment",
	"outstanding_balance",
	"due_date",
	"aging_days",
	"missing_supplier",
	"missing_invoice_no",
	"recommended_action",
]

HEADER_ALIASES = {
	"date_entry": ("date_entry", "date"),
	"entered_by": ("entered_by",),
	"billed_to": ("billed_to",),
	"acctg_status": ("acctg_status",),
	"fin_status": ("fin_status",),
	"invoice_no": ("invoice_no", "invoice_no.", "invoice_number"),
	"invoice_date": ("invoice_date",),
	"terms": ("terms",),
	"supplier_name": ("supplier", "supplier_name"),
	"category": ("category",),
	"particulars": ("particulars",),
	"rfp_id": ("rfp_id", "rfp"),
	"status": ("status",),
	"payment_date": ("payment_date",),
	"amount": ("amount",),
	"payment": ("payment",),
	"outstanding_balance": ("outstanding_balance", "outstanding"),
	"due_date": ("due_date",),
	"aging_days": ("aging_days", "aging"),
}

HEADER_FALLBACK_INDEX = {
	"date_entry": 0,
	"entered_by": 1,
	"billed_to": 2,
	"acctg_status": 3,
	"fin_status": 4,
	"invoice_no": 6,
	"invoice_date": 7,
	"terms": 8,
	"supplier_name": 9,
	"category": 10,
	"particulars": 11,
	"rfp_id": 12,
	"status": 13,
	"payment_date": 14,
	"amount": 15,
	"payment": 16,
	"outstanding_balance": 17,
	"due_date": 18,
	"aging_days": 19,
}

BEI_GREEN = "04400A"
BEI_GOLD = "C8900A"
BEI_CREAM = "F9F5EB"
BEI_GREEN_TINT = "E6ECE7"
BEI_BORDER = "D4D0C8"
TEXT_DARK = "1A1A1A"
TEXT_WHITE = "FFFFFF"
PHP_ACCOUNTING = '₱#,##0.00_);[Red](₱#,##0.00);₱"-"??_)'
THIN_BORDER = Border(
	left=Side(style="thin", color=BEI_BORDER),
	right=Side(style="thin", color=BEI_BORDER),
	top=Side(style="thin", color=BEI_BORDER),
	bottom=Side(style="thin", color=BEI_BORDER),
)


def _normalize_header(value: Any) -> str:
	return str(value or "").strip().lower().replace(".", "").replace(" ", "_")


def _clean_text(value: Any) -> str:
	return str(value or "").strip()


def _parse_decimal(value: Any) -> Decimal:
	text = _clean_text(value).replace(",", "")
	if text in {"", "-", "-   ", "-   -"}:
		return Decimal("0")
	text = text.replace("-   ", "").replace("₱", "")
	try:
		return Decimal(text)
	except InvalidOperation:
		return Decimal("0")


def _parse_int(value: Any) -> int | None:
	text = _clean_text(value).replace(",", "")
	if not text:
		return None
	try:
		return int(float(text))
	except ValueError:
		return None


def _decimal_to_number(value: Decimal) -> int | float:
	if value == value.to_integral():
		return int(value)
	return float(value)


def _resolve_value(row: list[Any], header_indexes: dict[str, int], canonical_name: str) -> Any:
	index = header_indexes.get(canonical_name)
	if index is None:
		index = HEADER_FALLBACK_INDEX[canonical_name]
	if index >= len(row):
		return ""
	return row[index]


def extract_missing_invoice_rows(raw_values: list[list[Any]]) -> list[dict[str, Any]]:
	"""Extract rows with blank invoice numbers from raw SUPPLIERS SOA values."""
	if len(raw_values) < 2:
		return []

	headers = [_normalize_header(value) for value in raw_values[1]]
	header_indexes: dict[str, int] = {}
	for canonical_name, aliases in HEADER_ALIASES.items():
		for alias in aliases:
			normalized = _normalize_header(alias)
			if normalized in headers:
				header_indexes[canonical_name] = headers.index(normalized)
				break

	rows: list[dict[str, Any]] = []
	for sheet_row_number, raw_row in enumerate(raw_values[2:], start=3):
		invoice_no = _clean_text(_resolve_value(raw_row, header_indexes, "invoice_no"))
		if invoice_no:
			continue

		supplier_name = _clean_text(_resolve_value(raw_row, header_indexes, "supplier_name"))
		amount = _parse_decimal(_resolve_value(raw_row, header_indexes, "amount"))
		payment = _parse_decimal(_resolve_value(raw_row, header_indexes, "payment"))
		outstanding_balance = _parse_decimal(_resolve_value(raw_row, header_indexes, "outstanding_balance"))

		rows.append(
			{
				"sheet_row": sheet_row_number,
				"date_entry": _clean_text(_resolve_value(raw_row, header_indexes, "date_entry")),
				"entered_by": _clean_text(_resolve_value(raw_row, header_indexes, "entered_by")),
				"billed_to": _clean_text(_resolve_value(raw_row, header_indexes, "billed_to")),
				"acctg_status": _clean_text(_resolve_value(raw_row, header_indexes, "acctg_status")),
				"fin_status": _clean_text(_resolve_value(raw_row, header_indexes, "fin_status")),
				"invoice_no": "",
				"invoice_date": _clean_text(_resolve_value(raw_row, header_indexes, "invoice_date")),
				"terms": _parse_int(_resolve_value(raw_row, header_indexes, "terms")),
				"supplier_name": supplier_name,
				"category": _clean_text(_resolve_value(raw_row, header_indexes, "category")),
				"particulars": _clean_text(_resolve_value(raw_row, header_indexes, "particulars")),
				"rfp_id": _clean_text(_resolve_value(raw_row, header_indexes, "rfp_id")),
				"status": _clean_text(_resolve_value(raw_row, header_indexes, "status")),
				"payment_date": _clean_text(_resolve_value(raw_row, header_indexes, "payment_date")),
				"amount": _decimal_to_number(amount),
				"payment": _decimal_to_number(payment),
				"outstanding_balance": _decimal_to_number(outstanding_balance),
				"due_date": _clean_text(_resolve_value(raw_row, header_indexes, "due_date")),
				"aging_days": _parse_int(_resolve_value(raw_row, header_indexes, "aging_days")),
				"missing_supplier": not supplier_name,
				"missing_invoice_no": True,
				"recommended_action": "Fill in invoice number before AP sync rerun",
			}
		)

	return rows


def _build_summary(
	rows: list[dict[str, Any]], latest_sync: SyncLog | None, generated_at: datetime
) -> dict[str, Any]:
	total_amount = sum(Decimal(str(row["amount"] or 0)) for row in rows)
	total_payment = sum(Decimal(str(row["payment"] or 0)) for row in rows)
	total_outstanding = sum(Decimal(str(row["outstanding_balance"] or 0)) for row in rows)
	supplier_counts = Counter(row["supplier_name"] or "(missing supplier)" for row in rows)
	receiver_rows_failed = latest_sync.rows_failed if latest_sync else None
	return {
		"generated_at_utc": generated_at.astimezone(UTC).isoformat(),
		"live_blocked_rows": len(rows),
		"affected_suppliers": len(supplier_counts),
		"rows_missing_supplier": sum(1 for row in rows if row["missing_supplier"]),
		"rows_missing_invoice_no": sum(1 for row in rows if row["missing_invoice_no"]),
		"total_amount": _decimal_to_number(total_amount),
		"total_payment": _decimal_to_number(total_payment),
		"total_outstanding_balance": _decimal_to_number(total_outstanding),
		"receiver_rows_failed": receiver_rows_failed,
		"receiver_discrepancy": (
			abs(receiver_rows_failed - len(rows)) if receiver_rows_failed is not None else None
		),
		"latest_sync_log_id": latest_sync.id if latest_sync else None,
		"latest_sync_created_at": latest_sync.created_at.isoformat() if latest_sync else None,
		"suppliers": dict(sorted(supplier_counts.items())),
	}


def build_team_message(summary: dict[str, Any]) -> list[str]:
	"""Return the copy-ready accounting message lines for the current exception list."""
	return [
		"Team, we validated the live AP workbook today and found rows in SUPPLIERS SOA that cannot sync into Frappe yet.",
		"",
		f"There are {summary['live_blocked_rows']} rows that need invoice numbers completed across {summary['affected_suppliers']} suppliers.",
		(
			f"The latest AP sync failed on {summary['receiver_rows_failed']} rows, and the current live workbook now shows {summary['live_blocked_rows']} rows missing INVOICE NO."
			if summary["receiver_rows_failed"] is not None
			else f"The current live workbook now shows {summary['live_blocked_rows']} rows missing INVOICE NO."
		),
		"",
		"Please update the INVOICE NO. field directly in the SUPPLIERS SOA tab for the rows listed in this workbook.",
		"Do not change amount, payment, outstanding balance, supplier name, or dates unless the source document itself is wrong.",
		"If a row should not remain in SUPPLIERS SOA because it is already fully paid or entered by mistake, mark it for review before deleting anything.",
		"",
		"Once all invoice numbers are filled, reply in the group so we can rerun AP sync and verify that the blocked rows clear in Frappe.",
	]


def _write_csv(rows: list[dict[str, Any]], destination: Path) -> None:
	with destination.open("w", encoding="utf-8", newline="") as handle:
		writer = csv.DictWriter(handle, fieldnames=FIELDNAMES)
		writer.writeheader()
		for row in rows:
			writer.writerow(row)


def _write_json(payload: dict[str, Any], destination: Path) -> None:
	destination.write_text(json.dumps(payload, indent=2, sort_keys=True, default=str), encoding="utf-8")


def _write_markdown(
	*,
	destination: Path,
	summary: dict[str, Any],
	message_lines: list[str],
	rows: list[dict[str, Any]],
) -> None:
	lines = [
		"# AP Missing Invoice Exception Report",
		"",
		f"- Generated (UTC): `{summary['generated_at_utc']}`",
		f"- Live blocked rows: `{summary['live_blocked_rows']}`",
		f"- Affected suppliers: `{summary['affected_suppliers']}`",
		f"- Total amount: `{summary['total_amount']}`",
		f"- Total payment: `{summary['total_payment']}`",
		f"- Total outstanding balance: `{summary['total_outstanding_balance']}`",
	]
	if summary["receiver_rows_failed"] is not None:
		lines.extend(
			[
				f"- Latest receiver `rows_failed`: `{summary['receiver_rows_failed']}`",
				f"- Workbook vs receiver discrepancy: `{summary['receiver_discrepancy']}`",
			]
		)
	lines.extend(["", "## Accounting Message", ""])
	lines.extend([f"- {line}" if line else "" for line in message_lines])
	lines.extend(["", "## Supplier Breakdown", ""])
	for supplier_name, count in summary["suppliers"].items():
		lines.append(f"- `{supplier_name}`: `{count}`")
	lines.extend(["", "## Sample Rows", ""])
	for row in rows[:10]:
		lines.append(
			f"- Row `{row['sheet_row']}` | Supplier `{row['supplier_name'] or '(missing supplier)'}` | "
			f"Amount `{row['amount']}` | Outstanding `{row['outstanding_balance']}`"
		)
	destination.write_text("\n".join(lines).strip() + "\n", encoding="utf-8")


def _style_cell(cell, *, bold: bool = False, fill_color: str = BEI_CREAM, wrap: bool = False) -> None:
	cell.font = Font(
		name="Calibri", size=10, bold=bold, color=TEXT_WHITE if fill_color == BEI_GREEN else TEXT_DARK
	)
	cell.fill = PatternFill("solid", fgColor=fill_color)
	cell.alignment = Alignment(vertical="center", wrap_text=wrap)
	cell.border = THIN_BORDER


def _autosize(ws) -> None:
	for column_cells in ws.columns:
		column_letter = get_column_letter(column_cells[0].column)
		width = 10
		for cell in column_cells:
			width = max(width, len(str(cell.value or "")) + 2)
		ws.column_dimensions[column_letter].width = min(width, 42)


def _write_workbook(
	*,
	destination: Path,
	summary: dict[str, Any],
	message_lines: list[str],
	rows: list[dict[str, Any]],
) -> None:
	workbook = Workbook()
	summary_ws = workbook.active
	summary_ws.title = "Summary"
	message_ws = workbook.create_sheet("Team Message")
	rows_ws = workbook.create_sheet("Missing Invoice Rows")

	summary_ws["A1"] = "AP Exception Summary"
	_style_cell(summary_ws["A1"], bold=True, fill_color=BEI_GREEN)
	summary_ws.merge_cells("A1:C1")
	summary_ws["A2"] = "Live SUPPLIERS SOA rows currently blocked from AP sync."
	_style_cell(summary_ws["A2"], fill_color=BEI_CREAM)
	summary_ws.merge_cells("A2:C2")

	summary_metrics = [
		("Generated (UTC)", summary["generated_at_utc"]),
		("Live blocked rows", summary["live_blocked_rows"]),
		("Affected suppliers", summary["affected_suppliers"]),
		("Rows missing supplier", summary["rows_missing_supplier"]),
		("Rows missing invoice number", summary["rows_missing_invoice_no"]),
		("Latest receiver rows_failed", summary["receiver_rows_failed"]),
		("Workbook vs receiver discrepancy", summary["receiver_discrepancy"]),
		("Total amount", summary["total_amount"]),
		("Total payment", summary["total_payment"]),
		("Total outstanding balance", summary["total_outstanding_balance"]),
	]
	for column, header in enumerate(("Metric", "Value", "Notes"), start=1):
		cell = summary_ws.cell(row=4, column=column, value=header)
		_style_cell(cell, bold=True, fill_color=BEI_GOLD)
	for row_index, (label, value) in enumerate(summary_metrics, start=5):
		summary_ws.cell(row=row_index, column=1, value=label)
		summary_ws.cell(row=row_index, column=2, value=value)
		summary_ws.cell(
			row=row_index,
			column=3,
			value="Receiver comparison" if "receiver" in label.lower() else "",
		)
		for column in range(1, 4):
			_style_cell(
				summary_ws.cell(row=row_index, column=column),
				fill_color=BEI_GREEN_TINT if row_index % 2 == 0 else BEI_CREAM,
			)
	for row_index in (12, 13, 14):
		summary_ws.cell(row=row_index, column=2).number_format = PHP_ACCOUNTING

	message_ws["A1"] = "Accounting Team Message"
	_style_cell(message_ws["A1"], bold=True, fill_color=BEI_GREEN)
	message_ws.merge_cells("A1:C1")
	for row_index, line in enumerate(message_lines, start=3):
		message_ws.cell(row=row_index, column=1, value=line)
		message_ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
		_style_cell(message_ws.cell(row=row_index, column=1), fill_color=BEI_CREAM, wrap=True)

	for column_index, header in enumerate(FIELDNAMES, start=1):
		cell = rows_ws.cell(row=1, column=column_index, value=header)
		_style_cell(cell, bold=True, fill_color=BEI_GOLD)
	for row_index, row in enumerate(rows, start=2):
		for column_index, field_name in enumerate(FIELDNAMES, start=1):
			cell = rows_ws.cell(row=row_index, column=column_index, value=row.get(field_name))
			_style_cell(cell, fill_color=BEI_GREEN_TINT if row_index % 2 == 0 else BEI_CREAM)
			if field_name in {"amount", "payment", "outstanding_balance"}:
				cell.number_format = PHP_ACCOUNTING
			elif field_name in {"missing_supplier", "missing_invoice_no"}:
				cell.value = "True" if row.get(field_name) else "False"

	summary_ws.freeze_panes = "A5"
	rows_ws.freeze_panes = "A2"
	_autosize(summary_ws)
	_autosize(message_ws)
	_autosize(rows_ws)
	workbook.save(destination)


def _update_latest_copies(timestamped_paths: dict[str, Path], output_dir: Path) -> dict[str, Path]:
	latest_paths: dict[str, Path] = {}
	for suffix, timestamped_path in timestamped_paths.items():
		latest_path = output_dir / f"ap_missing_invoice_report_latest.{suffix}"
		shutil.copyfile(timestamped_path, latest_path)
		latest_paths[suffix] = latest_path
	return latest_paths


def generate_ap_exception_report(
	*,
	sheets_client=None,
	db=None,
	output_dir: str | Path | None = None,
	generated_at: datetime | None = None,
) -> dict[str, Any]:
	"""Generate current AP missing-invoice exception artifacts from live Google Sheets data."""
	config = get_config()
	sheets_client = sheets_client or get_sheets_client()
	db = db or get_db()
	generated_at = (generated_at or datetime.now(UTC)).astimezone(UTC)
	sheet_config = get_sheet_config("ap_opening_balance")
	if sheet_config is None:
		raise RuntimeError("Missing sheet config for ap_opening_balance")

	range_name = f"{sheet_config.sheet_name}!A:Z"
	raw_values = (
		sheets_client.sheets.spreadsheets()
		.values()
		.get(
			spreadsheetId=sheet_config.spreadsheet_id,
			range=range_name,
			dateTimeRenderOption="FORMATTED_STRING",
		)
		.execute()
		.get("values", [])
	)
	rows = extract_missing_invoice_rows(raw_values)
	latest_sync = db.get_latest_sync(
		sheet_config.spreadsheet_id,
		sheet_config.sheet_name,
		status="success",
	)
	summary = _build_summary(rows, latest_sync, generated_at)
	message_lines = build_team_message(summary)

	report_dir = Path(output_dir or config.ap_exception_report_dir)
	report_dir.mkdir(parents=True, exist_ok=True)
	timestamp = generated_at.strftime("%Y%m%d_%H%M%S")
	base_path = report_dir / f"ap_missing_invoice_report_{timestamp}"
	timestamped_paths = {
		"csv": base_path.with_suffix(".csv"),
		"json": base_path.with_suffix(".json"),
		"md": base_path.with_suffix(".md"),
		"xlsx": base_path.with_suffix(".xlsx"),
	}

	_write_csv(rows, timestamped_paths["csv"])
	_write_json(
		{
			"summary": summary,
			"team_message_lines": message_lines,
			"rows": rows,
		},
		timestamped_paths["json"],
	)
	_write_markdown(
		destination=timestamped_paths["md"],
		summary=summary,
		message_lines=message_lines,
		rows=rows,
	)
	_write_workbook(
		destination=timestamped_paths["xlsx"],
		summary=summary,
		message_lines=message_lines,
		rows=rows,
	)
	latest_paths = _update_latest_copies(timestamped_paths, report_dir)

	result = {
		"summary": summary,
		"team_message_lines": message_lines,
		"timestamped_files": {key: str(path) for key, path in timestamped_paths.items()},
		"latest_files": {key: str(path) for key, path in latest_paths.items()},
	}
	logger.info(
		"Generated AP exception report with %s blocked rows at %s",
		summary["live_blocked_rows"],
		timestamped_paths["xlsx"],
	)
	return result
