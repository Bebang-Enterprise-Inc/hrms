import importlib.util
import sys
import tempfile
import types
import unittest
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
	sys.path.insert(0, str(ROOT))


def _install_fake_dependencies():
	if "hrms" not in sys.modules:
		hrms_pkg = types.ModuleType("hrms")
		hrms_pkg.__path__ = []
		sys.modules["hrms"] = hrms_pkg

	if "hrms.services" not in sys.modules:
		services_pkg = types.ModuleType("hrms.services")
		services_pkg.__path__ = []
		sys.modules["hrms.services"] = services_pkg

	if "hrms.services.sheets_receiver" not in sys.modules:
		sheets_pkg = types.ModuleType("hrms.services.sheets_receiver")
		sheets_pkg.__path__ = []
		sys.modules["hrms.services.sheets_receiver"] = sheets_pkg

	config_mod = types.ModuleType("hrms.services.sheets_receiver.config")
	config_mod.get_config = lambda: types.SimpleNamespace(ap_exception_report_dir=".")
	config_mod.get_sheet_config = lambda _key: None
	sys.modules["hrms.services.sheets_receiver.config"] = config_mod

	models_mod = types.ModuleType("hrms.services.sheets_receiver.models")
	models_mod.SyncLog = types.SimpleNamespace
	models_mod.get_db = lambda: types.SimpleNamespace()
	sys.modules["hrms.services.sheets_receiver.models"] = models_mod

	sheets_client_mod = types.ModuleType("hrms.services.sheets_receiver.sheets_client")
	sheets_client_mod.get_sheets_client = lambda: types.SimpleNamespace()
	sys.modules["hrms.services.sheets_receiver.sheets_client"] = sheets_client_mod


_install_fake_dependencies()
report_spec = importlib.util.spec_from_file_location(
	"hrms.services.sheets_receiver.ap_exception_reports_under_test",
	ROOT / "hrms" / "services" / "sheets_receiver" / "ap_exception_reports.py",
)
report_mod = importlib.util.module_from_spec(report_spec)
report_spec.loader.exec_module(report_mod)


class TestSheetsReceiverApExceptionReports(unittest.TestCase):
	def test_extract_missing_invoice_rows_uses_header_row_two(self):
		raw_values = [
			["Summary block", "", "", ""],
			[
				"DATE ENTRY",
				"ENTERED BY",
				"BILLED TO",
				"ACCTG STATUS",
				"FIN STATUS",
				"NO.",
				"INVOICE NO.",
				"INVOICE DATE",
				"TERMS",
				"SUPPLIER",
				"CATEGORY",
				"PARTICULARS",
				"RFP ID",
				"STATUS",
				"PAYMENT DATE",
				"AMOUNT",
				"PAYMENT",
				"OUTSTANDING BALANCE",
				"DUE DATE",
				"AGING DAYS",
			],
			[
				"2026-03-10",
				"Alyssa",
				"Stores - BEI",
				"OPEN",
				"OPEN",
				"1",
				"",
				"2026-03-05",
				"30",
				"ALYANA CHUA",
				"Supplies",
				"Office supplies",
				"RFP-1",
				"UNPAID",
				"",
				"1500",
				"0",
				"1500",
				"2026-04-04",
				"5",
			],
			[
				"2026-03-10",
				"Alyssa",
				"Stores - BEI",
				"OPEN",
				"OPEN",
				"2",
				"INV-001",
				"2026-03-05",
				"30",
				"MAC SIGNS",
				"Signage",
				"Poster",
				"RFP-2",
				"UNPAID",
				"",
				"2500",
				"500",
				"2000",
				"2026-04-04",
				"5",
			],
		]

		rows = report_mod.extract_missing_invoice_rows(raw_values)

		self.assertEqual(len(rows), 1)
		self.assertEqual(rows[0]["sheet_row"], 3)
		self.assertEqual(rows[0]["supplier_name"], "ALYANA CHUA")
		self.assertEqual(rows[0]["amount"], 1500)
		self.assertTrue(rows[0]["missing_invoice_no"])

	def test_generate_ap_exception_report_writes_expected_artifacts(self):
		raw_values = [
			["Summary block", "", "", ""],
			[
				"DATE ENTRY",
				"ENTERED BY",
				"BILLED TO",
				"ACCTG STATUS",
				"FIN STATUS",
				"NO.",
				"INVOICE NO.",
				"INVOICE DATE",
				"TERMS",
				"SUPPLIER",
				"CATEGORY",
				"PARTICULARS",
				"RFP ID",
				"STATUS",
				"PAYMENT DATE",
				"AMOUNT",
				"PAYMENT",
				"OUTSTANDING BALANCE",
				"DUE DATE",
				"AGING DAYS",
			],
			[
				"2026-03-10",
				"Alyssa",
				"Stores - BEI",
				"OPEN",
				"OPEN",
				"1",
				"",
				"2026-03-05",
				"30",
				"ALYANA CHUA",
				"Supplies",
				"Office supplies",
				"RFP-1",
				"UNPAID",
				"",
				"1500",
				"0",
				"1500",
				"2026-04-04",
				"5",
			],
			[
				"2026-03-10",
				"Alyssa",
				"Stores - BEI",
				"OPEN",
				"OPEN",
				"2",
				"",
				"2026-03-06",
				"15",
				"MAC SIGNS",
				"Signage",
				"Poster",
				"RFP-2",
				"UNPAID",
				"",
				"2500",
				"500",
				"2000",
				"2026-03-21",
				"4",
			],
		]

		class _FakeGetRequest:
			def execute(self):
				return {"values": raw_values}

		class _FakeValuesService:
			def get(self, **_kwargs):
				return _FakeGetRequest()

		class _FakeSpreadsheetsService:
			def values(self):
				return _FakeValuesService()

		class _FakeSheetsService:
			def spreadsheets(self):
				return _FakeSpreadsheetsService()

		fake_client = types.SimpleNamespace(sheets=_FakeSheetsService())
		fake_db = types.SimpleNamespace(
			get_latest_sync=lambda *_args, **_kwargs: types.SimpleNamespace(
				id=77,
				rows_failed=3,
				created_at=datetime(2026, 3, 10, 0, 0, tzinfo=UTC),
			)
		)
		report_mod.get_sheet_config = lambda _key: types.SimpleNamespace(
			spreadsheet_id="sheet-123",
			sheet_name="SUPPLIERS SOA",
		)
		report_mod.get_config = lambda: types.SimpleNamespace(ap_exception_report_dir=".")

		with tempfile.TemporaryDirectory() as tmpdir:
			result = report_mod.generate_ap_exception_report(
				sheets_client=fake_client,
				db=fake_db,
				output_dir=tmpdir,
				generated_at=datetime(2026, 3, 10, 1, 2, 3, tzinfo=UTC),
			)

			self.assertEqual(result["summary"]["live_blocked_rows"], 2)
			self.assertEqual(result["summary"]["affected_suppliers"], 2)
			self.assertEqual(result["summary"]["receiver_rows_failed"], 3)
			for file_path in result["timestamped_files"].values():
				self.assertTrue(Path(file_path).exists(), file_path)
			for file_path in result["latest_files"].values():
				self.assertTrue(Path(file_path).exists(), file_path)

			workbook = load_workbook(result["latest_files"]["xlsx"])
			self.assertEqual(
				workbook.sheetnames,
				["Summary", "Team Message", "Missing Invoice Rows"],
			)
			rows_ws = workbook["Missing Invoice Rows"]
			self.assertEqual(rows_ws["A2"].value, 3)
			self.assertEqual(rows_ws["J2"].value, "ALYANA CHUA")


if __name__ == "__main__":
	unittest.main()
