import importlib.util
import sys
import tempfile
import types
import unittest
from unittest.mock import MagicMock, patch
from datetime import date
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
	sys.path.insert(0, str(ROOT))


def _install_fake_frappe():
	class _FakeFrappe(types.ModuleType):
		@property
		def db(self):
			return self.local.db

	frappe = _FakeFrappe("frappe")
	frappe.get_site_path = lambda *parts: str(ROOT / "tmp_test_site" / Path(*parts))
	frappe.local = types.SimpleNamespace(
		db=types.SimpleNamespace(exists=lambda *args, **kwargs: None, commit=lambda: None)
	)
	frappe.get_doc = lambda *args, **kwargs: None
	sys.modules["frappe"] = frappe


_install_fake_frappe()
spec = importlib.util.spec_from_file_location(
	"store_inventory_shadow_sync_under_test",
	ROOT / "hrms" / "utils" / "store_inventory_shadow_sync.py",
)
shadow_sync = importlib.util.module_from_spec(spec)
assert spec and spec.loader
sys.modules["store_inventory_shadow_sync_under_test"] = shadow_sync
spec.loader.exec_module(shadow_sync)


class TestStoreInventoryShadowSync(unittest.TestCase):
	def test_ensure_required_master_data_reuses_existing_warehouse_name_match(self):
		store = shadow_sync.StoreSyncConfig(
			store_code="NAIA",
			store_name="NAIA",
			spreadsheet_id="sheet-1",
			warehouse_name="NAIA T3",
			warehouse_docname="NAIA T3 - Bebang Enterprise Inc.",
		)

		def db_exists(doctype, name=None):
			if doctype == "Warehouse":
				return name == "Stores - BEI"
			if doctype == "Item":
				return True
			if doctype == "Company":
				return name == "Bebang Enterprise Inc."
			return None

		def db_get_value(doctype, filters=None, fieldname=None):
			if doctype == "Warehouse" and filters == {"warehouse_name": "NAIA T3"}:
				return "NAIA T3 - BKI"
			return None

		shadow_sync.frappe.local.db.exists = db_exists
		shadow_sync.frappe.local.db.get_value = db_get_value

		import hrms

		fake_erp_sync = types.ModuleType("hrms.api.erp_sync")
		fake_erp_sync._normalize_company = MagicMock(return_value="Bebang Enterprise Inc.")
		fake_api = types.ModuleType("hrms.api")
		fake_api.erp_sync = fake_erp_sync
		original_api = getattr(hrms, "api", None)
		original_api_module = sys.modules.get("hrms.api")
		original_erp_sync_module = sys.modules.get("hrms.api.erp_sync")
		hrms.api = fake_api
		sys.modules["hrms.api"] = fake_api
		sys.modules["hrms.api.erp_sync"] = fake_erp_sync

		try:
			result = shadow_sync.ensure_required_master_data([store], {})
		finally:
			if original_api is None:
				delattr(hrms, "api")
			else:
				hrms.api = original_api
			if original_api_module is None:
				sys.modules.pop("hrms.api", None)
			else:
				sys.modules["hrms.api"] = original_api_module
			if original_erp_sync_module is None:
				sys.modules.pop("hrms.api.erp_sync", None)
			else:
				sys.modules["hrms.api.erp_sync"] = original_erp_sync_module

		self.assertEqual(store.warehouse_docname, "NAIA T3 - BKI")
		self.assertEqual(result["warehouses_created"], [])
		self.assertEqual(result["items_created"], [])

	def test_resolve_current_qty_prefers_encode_then_history_then_blank_zero(self):
		qty, source, inventory_date, detail = shadow_sync._resolve_current_qty(
			{"encode": 7, "total": "", "whole": "", "loose": "", "wt": 1},
			[],
			date(2026, 3, 10),
		)
		self.assertEqual(qty, 7)
		self.assertEqual(source, "encode")
		self.assertEqual(inventory_date, "2026-03-10")
		self.assertEqual(detail, "")

		qty, source, inventory_date, _ = shadow_sync._resolve_current_qty(
			{"encode": "", "total": "", "whole": "", "loose": "", "wt": 1},
			[
				{"inventory_date": "2026-03-08", "end": 4},
				{"inventory_date": "2026-03-09", "end": 6},
				{"inventory_date": "2026-03-11", "end": 8},
			],
			date(2026, 3, 10),
		)
		self.assertEqual(qty, 6)
		self.assertEqual(source, "historical_end")
		self.assertEqual(inventory_date, "2026-03-09")

		qty, source, inventory_date, detail = shadow_sync._resolve_current_qty(
			{"encode": "", "total": "", "whole": "", "loose": "", "wt": 1},
			[],
			date(2026, 3, 10),
		)
		self.assertEqual(qty, 0.0)
		self.assertEqual(source, "blank_zero_policy")
		self.assertEqual(inventory_date, "2026-03-10")
		self.assertIn("zeroed", detail)

	def test_extract_store_inventory_payload_generates_payload_and_exceptions(self):
		workbook = Workbook()
		ws = workbook.active
		ws.title = "3. INVENTORY"
		ws["A1"] = "WT."
		ws["B1"] = "WHOLE"
		ws["C1"] = "LOOSE"
		ws["D1"] = "TOTAL"
		ws["E1"] = "ENCODE"
		ws["F1"] = "2026-03-09"
		ws["F2"] = "BEG"
		ws["G2"] = "IN"
		ws["H2"] = "OUT"
		ws["I2"] = "END"
		ws["J2"] = "CODE"
		ws["K2"] = "ITEMS"
		ws["L2"] = "DESCRIPTION"
		ws["M2"] = "UOM"
		ws["J3"] = "RM-ENC"
		ws["K3"] = "ENC ITEM"
		ws["M3"] = "PIECE"
		ws["A3"] = 1
		ws["E3"] = 3
		ws["J4"] = "RM-HIST"
		ws["K4"] = "HIST ITEM"
		ws["M4"] = "PIECE"
		ws["A4"] = 1
		ws["I4"] = 5
		ws["J5"] = "RM-BLANK"
		ws["K5"] = "BLANK ITEM"
		ws["M5"] = "PIECE"
		ws["A5"] = 1
		ws["J6"] = "RM-ERR"
		ws["K6"] = "ERR ITEM"
		ws["M6"] = "PIECE"
		ws["A6"] = 1
		ws["E6"] = "#REF!"

		with tempfile.TemporaryDirectory() as tmp_dir:
			workbook_path = Path(tmp_dir) / "sample.xlsx"
			workbook.save(workbook_path)

			mapping = {
				"RM-ENC": shadow_sync.ItemMapping(
					inventory_code="RM-ENC",
					inventory_items="ENC ITEM",
					inventory_description="",
					inventory_uom="PIECE",
					resolution_type="exact_live_item",
					target_item_code="RM-ENC",
					target_item_name="ENC ITEM",
					target_stock_uom="PIECE",
					target_item_group="Raw Materials",
					import_policy="import",
					notes="",
				),
				"RM-HIST": shadow_sync.ItemMapping(
					inventory_code="RM-HIST",
					inventory_items="HIST ITEM",
					inventory_description="",
					inventory_uom="PIECE",
					resolution_type="exact_live_item",
					target_item_code="RM-HIST",
					target_item_name="HIST ITEM",
					target_stock_uom="PIECE",
					target_item_group="Raw Materials",
					import_policy="import",
					notes="",
				),
				"RM-BLANK": shadow_sync.ItemMapping(
					inventory_code="RM-BLANK",
					inventory_items="BLANK ITEM",
					inventory_description="",
					inventory_uom="PIECE",
					resolution_type="exact_live_item",
					target_item_code="RM-BLANK",
					target_item_name="BLANK ITEM",
					target_stock_uom="PIECE",
					target_item_group="Raw Materials",
					import_policy="import",
					notes="",
				),
				"RM-ERR": shadow_sync.ItemMapping(
					inventory_code="RM-ERR",
					inventory_items="ERR ITEM",
					inventory_description="",
					inventory_uom="PIECE",
					resolution_type="exact_live_item",
					target_item_code="RM-ERR",
					target_item_name="ERR ITEM",
					target_stock_uom="PIECE",
					target_item_group="Raw Materials",
					import_policy="import",
					notes="",
				),
			}
			config = shadow_sync.StoreSyncConfig(
				store_code="TST",
				store_name="Test Store",
				spreadsheet_id="sheet-1",
				warehouse_name="Test Store",
				warehouse_docname="Test Store - BEI",
			)

			result = shadow_sync.extract_store_inventory_payload(
				config,
				workbook_path,
				mapping,
				date(2026, 3, 10),
			)

		self.assertEqual(len(result["payload_rows"]), 3)
		self.assertEqual(len(result["exception_rows"]), 1)
		by_code = {row["item_code"]: row for row in result["payload_rows"]}
		self.assertEqual(by_code["RM-ENC"]["qty"], 3)
		self.assertEqual(by_code["RM-ENC"]["qty_source"], "encode")
		self.assertEqual(by_code["RM-HIST"]["qty"], 5)
		self.assertEqual(by_code["RM-HIST"]["qty_source"], "historical_end")
		self.assertEqual(by_code["RM-BLANK"]["qty"], 0)
		self.assertEqual(by_code["RM-BLANK"]["qty_source"], "blank_zero_policy")
		self.assertEqual(result["exception_rows"][0]["inventory_code"], "RM-ERR")
		self.assertEqual(result["exception_rows"][0]["classification"], "formula_error")

	def test_export_store_workbook_uses_temp_file_before_promoting(self):
		config = shadow_sync.StoreSyncConfig(
			store_code="TST",
			store_name="Test Store",
			spreadsheet_id="sheet-1",
			warehouse_name="Test Store",
			warehouse_docname="Test Store - BEI",
		)
		seen_destinations: list[Path] = []

		def fake_drive_export(_drive, _spreadsheet_id, dest: Path):
			seen_destinations.append(dest)
			dest.write_bytes(b"workbook-bytes")

		with (
			tempfile.TemporaryDirectory() as tmp_dir,
			patch.object(
				shadow_sync,
				"_export_workbook_via_drive",
				side_effect=fake_drive_export,
			),
		):
			result = shadow_sync.export_store_workbook(config, Path(tmp_dir), drive=None, sheets=None)
			self.assertEqual(seen_destinations[0].suffixes[-2:], [".xlsx", ".part"])
			self.assertEqual(result.read_bytes(), b"workbook-bytes")
			self.assertFalse(result.with_suffix(".xlsx.part").exists())

	def test_run_store_inventory_shadow_sync_persists_progress_after_each_store(self):
		store_one = shadow_sync.StoreSyncConfig(
			store_code="AFT",
			store_name="Store One",
			spreadsheet_id="sheet-1",
			warehouse_name="Store One",
			warehouse_docname="Store One - BEI",
		)
		store_two = shadow_sync.StoreSyncConfig(
			store_code="AMM",
			store_name="Store Two",
			spreadsheet_id="sheet-2",
			warehouse_name="Store Two",
			warehouse_docname="Store Two - BEI",
		)
		fake_erp_sync = types.ModuleType("hrms.api.erp_sync")
		fake_erp_sync._sync_inventory_rows = MagicMock(
			return_value={"rows_created": 1, "rows_updated": 0, "rows_failed": 0, "errors": []}
		)
		fake_api = types.ModuleType("hrms.api")
		fake_api.erp_sync = fake_erp_sync

		import hrms

		original_api = getattr(hrms, "api", None)
		original_api_module = sys.modules.get("hrms.api")
		original_erp_sync_module = sys.modules.get("hrms.api.erp_sync")
		hrms.api = fake_api
		sys.modules["hrms.api"] = fake_api
		sys.modules["hrms.api.erp_sync"] = fake_erp_sync

		try:
			with tempfile.TemporaryDirectory() as tmp_dir:
				tmp_path = Path(tmp_dir)
				with (
					patch.object(shadow_sync, "load_store_registry", return_value=[store_one, store_two]),
					patch.object(shadow_sync, "load_item_mapping", return_value={}),
					patch.object(shadow_sync, "ensure_required_master_data", return_value={}),
					patch.object(shadow_sync, "_build_google_services", return_value=(None, None)),
					patch.object(
						shadow_sync,
						"export_store_workbook",
						side_effect=[tmp_path / "one.xlsx", tmp_path / "two.xlsx"],
					),
					patch.object(
						shadow_sync,
						"extract_store_inventory_payload",
						side_effect=[
							{
								"payload_rows": [
									{
										"store_code": "AFT",
										"store_name": "Store One",
										"warehouse": "Store One - BEI",
										"item_code": "ITEM-1",
										"qty": 1,
										"qty_source": "encode",
										"inventory_date": "2026-03-10",
										"spreadsheet_id": "sheet-1",
										"source_row": 3,
										"workbook_path": str(tmp_path / "one.xlsx"),
									}
								],
								"exception_rows": [],
								"audit_rows": [],
								"checksum": "checksum-1",
							},
							{
								"payload_rows": [
									{
										"store_code": "AMM",
										"store_name": "Store Two",
										"warehouse": "Store Two - BEI",
										"item_code": "ITEM-2",
										"qty": 2,
										"qty_source": "encode",
										"inventory_date": "2026-03-10",
										"spreadsheet_id": "sheet-2",
										"source_row": 4,
										"workbook_path": str(tmp_path / "two.xlsx"),
									}
								],
								"exception_rows": [],
								"audit_rows": [],
								"checksum": "checksum-2",
							},
						],
					),
					patch.object(shadow_sync, "_persist_shadow_sync_progress") as persist_mock,
				):
					result = shadow_sync.run_store_inventory_shadow_sync(
						run_date="2026-03-10",
						force=True,
						output_dir=str(tmp_path / "run"),
						registry_path=str(tmp_path / "registry.csv"),
						state_path=str(tmp_path / "state.json"),
					)
		finally:
			if original_api is None:
				delattr(hrms, "api")
			else:
				hrms.api = original_api
			if original_api_module is None:
				sys.modules.pop("hrms.api", None)
			else:
				sys.modules["hrms.api"] = original_api_module
			if original_erp_sync_module is None:
				sys.modules.pop("hrms.api.erp_sync", None)
			else:
				sys.modules["hrms.api.erp_sync"] = original_erp_sync_module

		self.assertEqual(result["status"], "completed")
		self.assertEqual(result["imported_stores"], 2)
		self.assertEqual(fake_erp_sync._sync_inventory_rows.call_count, 2)
		self.assertEqual(persist_mock.call_count, 4)
		self.assertEqual(persist_mock.call_args_list[0].kwargs["summary"]["status"], "in_progress")
		self.assertEqual(persist_mock.call_args_list[-1].kwargs["summary"]["status"], "completed")
		self.assertEqual(
			persist_mock.call_args_list[0].kwargs["runtime_state"]["last_run"]["recovery_enqueued_at"], ""
		)
		self.assertIn("updated_at", persist_mock.call_args_list[-1].kwargs["runtime_state"]["last_run"])


if __name__ == "__main__":
	unittest.main()
