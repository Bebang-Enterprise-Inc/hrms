from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import re
import shutil
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
	import frappe  # type: ignore
except Exception:  # pragma: no cover
	frappe = None


REPO_ROOT = Path(__file__).resolve().parents[2]
FIXTURE_DIR = REPO_ROOT / "hrms" / "fixtures" / "store_inventory_shadow_sync"
DEFAULT_REGISTRY_FIXTURE = FIXTURE_DIR / "store_inventory_shadow_sync_registry.csv"
ITEM_MAPPING_FIXTURE = FIXTURE_DIR / "store_inventory_item_mapping.csv"
DEFAULT_RUNTIME_REGISTRY_NAME = "store_inventory_shadow_sync_registry.csv"
DEFAULT_RUNTIME_STATE_NAME = "store_inventory_shadow_sync_state.json"
DEFAULT_RUNS_DIRNAME = "store_inventory_shadow_sync_runs"
INVENTORY_SHEET_NAME = "3. INVENTORY"
GOOGLE_IMPERSONATE_USER = "sam@bebang.ph"
GOOGLE_SCOPES = [
	"https://www.googleapis.com/auth/drive.readonly",
	"https://www.googleapis.com/auth/spreadsheets.readonly",
]
EXPORT_TIMEOUT_SECONDS = 600
SECTION_NAMES = {
	"RAW MATERIALS",
	"COMMISSARY",
	"PACKAGING MATERIALS",
	"CLEANING SUPPLIES",
	"OFFICE SUPPLIES",
	"KALINISAN",
}
IMPORTABLE_STATES = {"shadow_sync"}
FORMULA_ERROR_PREFIX = "#"


@dataclass
class StoreSyncConfig:
	store_code: str
	store_name: str
	spreadsheet_id: str
	warehouse_name: str
	warehouse_docname: str
	state: str = "shadow_sync"
	sheet_sync_enabled: bool = True
	last_checksum: str = ""
	last_success_at: str = ""
	last_error: str = ""
	last_import_rows: int = 0
	last_inventory_date: str = ""
	notes: str = ""


@dataclass
class ItemMapping:
	inventory_code: str
	inventory_items: str
	inventory_description: str
	inventory_uom: str
	resolution_type: str
	target_item_code: str
	target_item_name: str
	target_stock_uom: str
	target_item_group: str
	import_policy: str
	notes: str


@dataclass
class InventoryLayout:
	header_row: int
	labels_row: int
	metadata_cols: dict[str, int]
	current_cols: dict[str, int]
	group_starts: list[int]
	group_dates: list[str | None]
	max_row: int
	max_col: int


def _now_ts() -> str:
	return datetime.now().strftime("%Y-%m-%dT%H:%M:%S")


def _bool_from_csv(value: str | bool | None) -> bool:
	if isinstance(value, bool):
		return value
	text = str(value or "").strip().lower()
	return text in {"1", "true", "yes", "y", "on"}


def _coerce_float(value: Any) -> float | None:
	if value in (None, ""):
		return None
	if isinstance(value, int | float):
		if isinstance(value, float) and math.isnan(value):
			return None
		return float(value)
	text = str(value).strip()
	if not text:
		return None
	if text.startswith(FORMULA_ERROR_PREFIX):
		return None
	try:
		return float(text.replace(",", ""))
	except Exception:
		return None


def _is_formula_error(value: Any) -> bool:
	return isinstance(value, str) and value.strip().startswith(FORMULA_ERROR_PREFIX)


def _normalize_text(value: Any) -> str:
	return str(value).strip() if value is not None else ""


def _normalize_header(value: Any) -> str:
	return _normalize_text(value).upper()


def _parse_date_cell(value: Any) -> str | None:
	if value in (None, ""):
		return None
	if isinstance(value, datetime):
		return value.date().isoformat()
	if isinstance(value, date):
		return value.isoformat()
	text = _normalize_text(value)
	if not text:
		return None
	for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
		try:
			return datetime.strptime(text, fmt).date().isoformat()
		except ValueError:
			continue
	return None


def _read_cell(row_values: tuple[Any, ...], col_idx: int | None) -> Any:
	if not col_idx or col_idx <= 0:
		return None
	offset = col_idx - 1
	if offset >= len(row_values):
		return None
	return row_values[offset]


def _write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
	path.parent.mkdir(parents=True, exist_ok=True)
	with path.open("w", newline="", encoding="utf-8") as handle:
		writer = csv.DictWriter(handle, fieldnames=fieldnames)
		writer.writeheader()
		writer.writerows(rows)


def _write_json(path: Path, payload: Any) -> None:
	path.parent.mkdir(parents=True, exist_ok=True)
	path.write_text(json.dumps(payload, indent=2, ensure_ascii=False, default=str), encoding="utf-8")


def _runtime_root() -> Path:
	if frappe and hasattr(frappe, "get_site_path"):
		return Path(frappe.get_site_path("private", "files"))
	return REPO_ROOT / "output" / "store_inventory_shadow_sync"


def get_runtime_registry_path() -> Path:
	return _runtime_root() / DEFAULT_RUNTIME_REGISTRY_NAME


def get_runtime_state_path() -> Path:
	return _runtime_root() / DEFAULT_RUNTIME_STATE_NAME


def _runs_root() -> Path:
	return _runtime_root() / DEFAULT_RUNS_DIRNAME


def _default_run_dir(run_date: str) -> Path:
	return _runs_root() / f"{run_date}_store_inventory_shadow_sync"


def load_runtime_state(state_path: Path | None = None) -> dict[str, Any]:
	path = state_path or get_runtime_state_path()
	if not path.exists():
		return {"stores": {}, "last_run": {}}
	try:
		return json.loads(path.read_text(encoding="utf-8"))
	except Exception:
		return {"stores": {}, "last_run": {}}


def save_runtime_state(payload: dict[str, Any], state_path: Path | None = None) -> None:
	path = state_path or get_runtime_state_path()
	_write_json(path, payload)


def ensure_runtime_registry_exists(
	registry_path: Path | None = None,
	default_fixture_path: Path | None = None,
) -> Path:
	path = registry_path or get_runtime_registry_path()
	fixture = default_fixture_path or DEFAULT_REGISTRY_FIXTURE
	if path.exists():
		return path
	path.parent.mkdir(parents=True, exist_ok=True)
	shutil.copyfile(fixture, path)
	return path


def load_store_registry(registry_path: Path | None = None) -> list[StoreSyncConfig]:
	path = ensure_runtime_registry_exists(registry_path)
	rows: list[StoreSyncConfig] = []
	with path.open("r", encoding="utf-8-sig", newline="") as handle:
		for row in csv.DictReader(handle):
			rows.append(
				StoreSyncConfig(
					store_code=_normalize_text(row.get("store_code")),
					store_name=_normalize_text(row.get("store_name")),
					spreadsheet_id=_normalize_text(row.get("spreadsheet_id")),
					warehouse_name=_normalize_text(row.get("warehouse_name")),
					warehouse_docname=_normalize_text(row.get("warehouse_docname")),
					state=_normalize_text(row.get("state")) or "shadow_sync",
					sheet_sync_enabled=_bool_from_csv(row.get("sheet_sync_enabled")),
					last_checksum=_normalize_text(row.get("last_checksum")),
					last_success_at=_normalize_text(row.get("last_success_at")),
					last_error=_normalize_text(row.get("last_error")),
					last_import_rows=int(float(row.get("last_import_rows") or 0)),
					last_inventory_date=_normalize_text(row.get("last_inventory_date")),
					notes=_normalize_text(row.get("notes")),
				)
			)
	return rows


def save_store_registry(rows: list[StoreSyncConfig], registry_path: Path | None = None) -> None:
	path = registry_path or get_runtime_registry_path()
	fieldnames = [
		"store_code",
		"store_name",
		"spreadsheet_id",
		"warehouse_name",
		"warehouse_docname",
		"state",
		"sheet_sync_enabled",
		"last_checksum",
		"last_success_at",
		"last_error",
		"last_import_rows",
		"last_inventory_date",
		"notes",
	]
	_write_csv(
		path,
		fieldnames,
		[
			{
				"store_code": row.store_code,
				"store_name": row.store_name,
				"spreadsheet_id": row.spreadsheet_id,
				"warehouse_name": row.warehouse_name,
				"warehouse_docname": row.warehouse_docname,
				"state": row.state,
				"sheet_sync_enabled": "true" if row.sheet_sync_enabled else "false",
				"last_checksum": row.last_checksum,
				"last_success_at": row.last_success_at,
				"last_error": row.last_error,
				"last_import_rows": row.last_import_rows,
				"last_inventory_date": row.last_inventory_date,
				"notes": row.notes,
			}
			for row in rows
		],
	)


def load_item_mapping(mapping_path: Path | None = None) -> dict[str, ItemMapping]:
	path = mapping_path or ITEM_MAPPING_FIXTURE
	mapping: dict[str, ItemMapping] = {}
	with path.open("r", encoding="utf-8-sig", newline="") as handle:
		for row in csv.DictReader(handle):
			key = _normalize_text(row.get("inventory_code"))
			if not key:
				continue
			mapping[key] = ItemMapping(
				inventory_code=key,
				inventory_items=_normalize_text(row.get("inventory_items")),
				inventory_description=_normalize_text(row.get("inventory_description")),
				inventory_uom=_normalize_text(row.get("inventory_uom")),
				resolution_type=_normalize_text(row.get("resolution_type")),
				target_item_code=_normalize_text(row.get("target_item_code")),
				target_item_name=_normalize_text(row.get("target_item_name")),
				target_stock_uom=_normalize_text(row.get("target_stock_uom")),
				target_item_group=_normalize_text(row.get("target_item_group")),
				import_policy=_normalize_text(row.get("import_policy")) or "import",
				notes=_normalize_text(row.get("notes")),
			)
	return mapping


def _get_service_account_path() -> str:
	if frappe:
		try:
			from hrms.utils.bei_config import get_service_account_path

			candidate = get_service_account_path()
			if candidate:
				return candidate
		except Exception:
			pass
	return str(REPO_ROOT / "credentials" / "task-manager-service.json")


def _build_google_services() -> tuple[Any, Any]:
	import httplib2  # type: ignore
	from google.oauth2 import service_account  # type: ignore
	from google_auth_httplib2 import AuthorizedHttp  # type: ignore
	from googleapiclient.discovery import build  # type: ignore

	credentials = service_account.Credentials.from_service_account_file(
		_get_service_account_path(),
		scopes=GOOGLE_SCOPES,
	).with_subject(GOOGLE_IMPERSONATE_USER)
	drive_http = AuthorizedHttp(credentials, http=httplib2.Http(timeout=EXPORT_TIMEOUT_SECONDS))
	sheets_http = AuthorizedHttp(credentials, http=httplib2.Http(timeout=EXPORT_TIMEOUT_SECONDS))
	drive = build("drive", "v3", http=drive_http, cache_discovery=False)
	sheets = build("sheets", "v4", http=sheets_http, cache_discovery=False)
	return drive, sheets


def _retryable_http_error(exc: Exception) -> bool:
	try:
		from googleapiclient.errors import HttpError  # type: ignore
	except Exception:  # pragma: no cover
		return False
	if not isinstance(exc, HttpError):
		return False
	return bool(getattr(exc, "resp", None)) and exc.resp.status in {429, 500, 502, 503, 504}


def _retry_google_call(fn, max_retries: int = 4):
	for attempt in range(max_retries):
		try:
			return fn()
		except Exception as exc:
			if attempt == max_retries - 1 or not _retryable_http_error(exc):
				raise
			import time

			time.sleep(2**attempt)


def _export_too_large(exc: Exception) -> bool:
	message = str(exc)
	return "exportSizeLimitExceeded" in message or "too large to be exported" in message.lower()


def _safe_filename(name: str) -> str:
	name = name.replace("/", "_").replace("\\", "_").replace(":", " -")
	return re.sub(r"\s+", " ", name).strip()


def _export_workbook_via_drive(drive: Any, spreadsheet_id: str, dest: Path) -> None:
	from googleapiclient.http import MediaIoBaseDownload  # type: ignore

	def _download() -> None:
		request = drive.files().export_media(
			fileId=spreadsheet_id,
			mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
		)
		with io.FileIO(dest, "wb") as handle:
			downloader = MediaIoBaseDownload(handle, request)
			done = False
			while not done:
				_, done = downloader.next_chunk()

	_retry_google_call(_download)


def _make_safe_sheet_title(title: str, existing: set[str]) -> str:
	cleaned = re.sub(r"[\\/*?:\[\]]", "_", title).strip() or "Sheet"
	candidate = cleaned[:31]
	index = 1
	while candidate in existing:
		suffix = f"_{index}"
		candidate = f"{cleaned[: 31 - len(suffix)]}{suffix}"
		index += 1
	existing.add(candidate)
	return candidate


def _export_workbook_via_sheets_api(sheets: Any, spreadsheet_id: str, dest: Path) -> None:
	from openpyxl import Workbook  # type: ignore

	meta = _retry_google_call(lambda: sheets.spreadsheets().get(spreadsheetId=spreadsheet_id).execute())
	titles = [sheet["properties"]["title"] for sheet in meta.get("sheets", [])]
	workbook = Workbook()
	workbook.remove(workbook.active)
	used_titles: set[str] = set()

	for title in titles:
		values = _retry_google_call(
			lambda title=title: (
				sheets.spreadsheets()
				.values()
				.get(
					spreadsheetId=spreadsheet_id,
					range=f"'{title}'",
					valueRenderOption="UNFORMATTED_VALUE",
					dateTimeRenderOption="FORMATTED_STRING",
				)
				.execute()
				.get("values", [])
			)
		)
		ws = workbook.create_sheet(title=_make_safe_sheet_title(title, used_titles))
		for row_idx, row in enumerate(values, start=1):
			for col_idx, value in enumerate(row, start=1):
				ws.cell(row=row_idx, column=col_idx, value=value)

	workbook.save(dest)
	workbook.close()


def export_store_workbook(config: StoreSyncConfig, dest_dir: Path, drive: Any, sheets: Any) -> Path:
	dest_dir.mkdir(parents=True, exist_ok=True)
	dest = dest_dir / f"{_safe_filename(config.store_code + ' - ' + config.store_name)}.xlsx"
	if dest.exists():
		dest.unlink()
	try:
		_export_workbook_via_drive(drive, config.spreadsheet_id, dest)
	except Exception as exc:
		if not _export_too_large(exc):
			raise
		_export_workbook_via_sheets_api(sheets, config.spreadsheet_id, dest)
	return dest


def detect_inventory_layout(ws) -> InventoryLayout:
	header_row = None
	metadata_cols: dict[str, int] = {}

	for row_idx in range(1, min(ws.max_row, 20) + 1):
		normalized = [
			_normalize_header(ws.cell(row_idx, col_idx).value)
			for col_idx in range(1, min(ws.max_column, 40) + 1)
		]
		if {"CODE", "ITEMS", "DESCRIPTION", "UOM"}.issubset(set(normalized)):
			header_row = row_idx
			for col_idx, value in enumerate(normalized, start=1):
				if value in {"CODE", "ITEMS", "DESCRIPTION", "UOM"}:
					metadata_cols[value] = col_idx
			break

	if header_row is None:
		raise ValueError("Could not detect 3. INVENTORY header row")

	labels_row = header_row - 1
	current_cols: dict[str, int] = {}
	for col_idx in range(1, min(ws.max_column, 20) + 1):
		label = _normalize_header(ws.cell(labels_row, col_idx).value)
		if label in {"WT.", "WHOLE", "LOOSE", "TOTAL", "ENCODE"}:
			current_cols[label] = col_idx

	group_starts: list[int] = []
	group_dates: list[str | None] = []
	col_idx = 1
	while col_idx <= ws.max_column - 3:
		values = [_normalize_header(ws.cell(header_row, col_idx + offset).value) for offset in range(4)]
		if values == ["BEG", "IN", "OUT", "END"]:
			group_starts.append(col_idx)
			group_dates.append(_parse_date_cell(ws.cell(labels_row, col_idx).value))
			col_idx += 4
		else:
			col_idx += 1

	if not group_starts:
		raise ValueError("Could not detect BEG/IN/OUT/END groups")

	return InventoryLayout(
		header_row=header_row,
		labels_row=labels_row,
		metadata_cols=metadata_cols,
		current_cols=current_cols,
		group_starts=group_starts,
		group_dates=group_dates,
		max_row=ws.max_row,
		max_col=ws.max_column,
	)


def _classify_section_row(code: str, items: str, description: str, uom: str) -> bool:
	return bool(code and not items and not description and not uom)


def _resolve_current_qty(
	row_payload: dict[str, Any],
	daily_records: list[dict[str, Any]],
	run_date: date,
) -> tuple[float | None, str, str | None, str]:
	current_values = {
		"encode": row_payload["encode"],
		"total": row_payload["total"],
		"whole": row_payload["whole"],
		"loose": row_payload["loose"],
		"wt": row_payload["wt"],
	}
	if any(_is_formula_error(value) for value in current_values.values()):
		return None, "formula_error", None, "Current balance field contains formula error"

	encode_qty = _coerce_float(row_payload["encode"])
	if encode_qty is not None:
		return encode_qty, "encode", run_date.isoformat(), ""

	total_qty = _coerce_float(row_payload["total"])
	if total_qty is not None:
		return total_qty, "total", run_date.isoformat(), ""

	best_daily: tuple[date, float] | None = None
	for daily in daily_records:
		inventory_date = daily.get("inventory_date")
		end_value = daily.get("end")
		if not inventory_date:
			continue
		try:
			parsed_date = datetime.strptime(str(inventory_date), "%Y-%m-%d").date()
		except ValueError:
			continue
		if parsed_date > run_date:
			continue
		if _is_formula_error(end_value):
			continue
		end_qty = _coerce_float(end_value)
		if end_qty is None:
			continue
		if not best_daily or parsed_date > best_daily[0]:
			best_daily = (parsed_date, end_qty)

	if best_daily:
		return best_daily[1], "historical_end", best_daily[0].isoformat(), ""

	whole_qty = _coerce_float(row_payload["whole"])
	loose_qty = _coerce_float(row_payload["loose"])
	if whole_qty is not None or loose_qty is not None:
		return None, "needs_conversion_rule", None, "WHOLE/LOOSE present but no trusted conversion rule"

	return (
		0.0,
		"blank_zero_policy",
		run_date.isoformat(),
		"No current or historical stock signal; zeroed for deterministic mirror",
	)


def ensure_required_master_data(
	store_registry: list[StoreSyncConfig],
	item_mapping: dict[str, ItemMapping],
) -> dict[str, list[str]]:
	if not frappe:
		return {"warehouses_created": [], "items_created": []}

	created = {"warehouses_created": [], "items_created": []}
	from hrms.api.erp_sync import _normalize_company

	company = _normalize_company()
	parent_warehouse = "Stores - BEI"
	if not frappe.db.exists("Warehouse", parent_warehouse):
		raise RuntimeError(f"Required parent warehouse not found: {parent_warehouse}")

	for store in store_registry:
		if not store.warehouse_docname:
			continue
		if frappe.db.exists("Warehouse", store.warehouse_docname):
			continue
		doc = frappe.get_doc(
			{
				"doctype": "Warehouse",
				"warehouse_name": store.warehouse_name,
				"company": company,
				"parent_warehouse": parent_warehouse,
				"is_group": 0,
			}
		)
		doc.insert(ignore_permissions=True)
		created["warehouses_created"].append(store.warehouse_docname)

	for mapping in item_mapping.values():
		if mapping.resolution_type != "create_item":
			continue
		if not mapping.target_item_code:
			continue
		if frappe.db.exists("Item", mapping.target_item_code):
			continue
		doc = frappe.get_doc(
			{
				"doctype": "Item",
				"item_code": mapping.target_item_code,
				"item_name": mapping.target_item_name or mapping.inventory_items,
				"item_group": mapping.target_item_group or "Raw Materials",
				"stock_uom": mapping.target_stock_uom or mapping.inventory_uom or "Nos",
				"is_stock_item": 1,
			}
		)
		doc.insert(ignore_permissions=True)
		created["items_created"].append(mapping.target_item_code)

	return created


def extract_store_inventory_payload(
	config: StoreSyncConfig,
	workbook_path: Path,
	item_mapping: dict[str, ItemMapping],
	run_date: date,
) -> dict[str, Any]:
	from openpyxl import load_workbook  # type: ignore

	workbook = load_workbook(workbook_path, data_only=True, read_only=False)
	if INVENTORY_SHEET_NAME not in workbook.sheetnames:
		workbook.close()
		raise ValueError(f"Workbook missing {INVENTORY_SHEET_NAME} tab")

	ws = workbook[INVENTORY_SHEET_NAME]
	layout = detect_inventory_layout(ws)
	payload_rows: list[dict[str, Any]] = []
	exception_rows: list[dict[str, Any]] = []
	current_rows_audit: list[dict[str, Any]] = []
	section_name = ""

	for excel_row_num, row_values in enumerate(
		ws.iter_rows(
			min_row=layout.header_row + 1,
			max_row=layout.max_row,
			min_col=1,
			max_col=layout.max_col,
			values_only=True,
		),
		start=layout.header_row + 1,
	):
		code = _normalize_text(_read_cell(row_values, layout.metadata_cols.get("CODE")))
		items = _normalize_text(_read_cell(row_values, layout.metadata_cols.get("ITEMS")))
		description = _normalize_text(_read_cell(row_values, layout.metadata_cols.get("DESCRIPTION")))
		uom = _normalize_text(_read_cell(row_values, layout.metadata_cols.get("UOM")))

		if not any([code, items, description, uom]):
			continue

		if _classify_section_row(code, items, description, uom):
			section_name = code if _normalize_header(code) in SECTION_NAMES else code
			continue

		current_row = {
			"store_code": config.store_code,
			"store_name": config.store_name,
			"warehouse_name": config.warehouse_name,
			"warehouse_docname": config.warehouse_docname,
			"workbook_path": str(workbook_path),
			"source_row": excel_row_num,
			"section": section_name,
			"code": code,
			"items": items,
			"description": description,
			"uom": uom,
			"wt": _read_cell(row_values, layout.current_cols.get("WT.")),
			"whole": _read_cell(row_values, layout.current_cols.get("WHOLE")),
			"loose": _read_cell(row_values, layout.current_cols.get("LOOSE")),
			"total": _read_cell(row_values, layout.current_cols.get("TOTAL")),
			"encode": _read_cell(row_values, layout.current_cols.get("ENCODE")),
		}

		daily_records: list[dict[str, Any]] = []
		for group_start, inventory_date in zip(layout.group_starts, layout.group_dates, strict=False):
			daily_records.append(
				{
					"inventory_date": inventory_date,
					"beg": _read_cell(row_values, group_start),
					"in": _read_cell(row_values, group_start + 1),
					"out": _read_cell(row_values, group_start + 2),
					"end": _read_cell(row_values, group_start + 3),
				}
			)

		resolved_qty, qty_source, inventory_date, detail = _resolve_current_qty(
			current_row, daily_records, run_date
		)
		mapping = item_mapping.get(code)
		target_item_code = mapping.target_item_code if mapping else ""
		classification = "ready_to_import"

		if not config.warehouse_docname:
			classification = "warehouse_unmapped"
			detail = "Missing target warehouse"
		elif qty_source == "formula_error":
			classification = "formula_error"
		elif qty_source == "needs_conversion_rule":
			classification = "needs_conversion_rule"
		elif not mapping:
			classification = "item_unmapped"
			detail = "No mapping found for inventory code"
		elif mapping.import_policy != "import":
			classification = "skip_non_stock_or_zero_policy"
			detail = mapping.notes or "Row marked non-importable"
		elif not target_item_code:
			classification = "item_unmapped"
			detail = "Mapping has no target item code"

		audit_row = {
			"store_code": config.store_code,
			"store_name": config.store_name,
			"warehouse_name": config.warehouse_name,
			"warehouse_docname": config.warehouse_docname,
			"source_row": excel_row_num,
			"section": section_name,
			"inventory_code": code,
			"inventory_items": items,
			"inventory_description": description,
			"inventory_uom": uom,
			"classification": classification,
			"target_item_code": target_item_code,
			"resolved_qty": resolved_qty if resolved_qty is not None else "",
			"qty_source": qty_source,
			"inventory_date": inventory_date or "",
			"detail": detail,
		}
		current_rows_audit.append(audit_row)

		if classification != "ready_to_import":
			exception_rows.append(audit_row)
			continue

		payload_rows.append(
			{
				"store_code": config.store_code,
				"store_name": config.store_name,
				"warehouse": config.warehouse_docname,
				"item_code": target_item_code,
				"qty": resolved_qty,
				"qty_source": qty_source,
				"inventory_date": inventory_date or run_date.isoformat(),
				"spreadsheet_id": config.spreadsheet_id,
				"source_row": excel_row_num,
				"workbook_path": str(workbook_path),
			}
		)

	workbook.close()
	payload_rows.sort(key=lambda row: (row["warehouse"], row["item_code"]))
	checksum = hashlib.sha1(
		json.dumps(
			[
				{
					"warehouse": row["warehouse"],
					"item_code": row["item_code"],
					"qty": row["qty"],
					"inventory_date": row["inventory_date"],
				}
				for row in payload_rows
			],
			sort_keys=True,
		).encode("utf-8")
	).hexdigest()
	return {
		"store_code": config.store_code,
		"store_name": config.store_name,
		"warehouse_name": config.warehouse_name,
		"warehouse_docname": config.warehouse_docname,
		"workbook_path": str(workbook_path),
		"payload_rows": payload_rows,
		"exception_rows": exception_rows,
		"audit_rows": current_rows_audit,
		"checksum": checksum,
	}


def _summary_markdown(summary: dict[str, Any]) -> str:
	lines = [
		"# Store Inventory Shadow Sync Summary",
		"",
		f"- run_date: `{summary['run_date']}`",
		f"- generated_at: `{summary['generated_at']}`",
		f"- enabled_stores: `{summary['enabled_stores']}`",
		f"- imported_stores: `{summary['imported_stores']}`",
		f"- skipped_unchanged: `{summary['skipped_unchanged']}`",
		f"- skipped_non_shadow: `{summary['skipped_non_shadow']}`",
		f"- skipped_disabled: `{summary['skipped_disabled']}`",
		f"- payload_rows: `{summary['payload_rows']}`",
		f"- exception_rows: `{summary['exception_rows']}`",
		f"- rows_created: `{summary['rows_created']}`",
		f"- rows_updated: `{summary['rows_updated']}`",
		f"- rows_failed: `{summary['rows_failed']}`",
	]
	if summary.get("master_data_bootstrap"):
		bootstrap = summary["master_data_bootstrap"]
		lines.extend(
			[
				"",
				"## Master Data Bootstrap",
				"",
				f"- warehouses_created: `{len(bootstrap.get('warehouses_created', []))}`",
				f"- items_created: `{len(bootstrap.get('items_created', []))}`",
			]
		)
	if summary.get("failed_stores"):
		lines.extend(["", "## Failed Stores", ""])
		for row in summary["failed_stores"]:
			lines.append(f"- `{row['store_code']}`: {row['error']}")
	return "\n".join(lines) + "\n"


def run_store_inventory_shadow_sync(
	*,
	run_date: str | None = None,
	force: bool = False,
	store_codes: list[str] | None = None,
	output_dir: str | None = None,
	registry_path: str | None = None,
	state_path: str | None = None,
) -> dict[str, Any]:
	if not frappe:
		raise RuntimeError("run_store_inventory_shadow_sync requires Frappe context")

	run_date_value = datetime.strptime(run_date or frappe.utils.nowdate(), "%Y-%m-%d").date()
	registry_file = Path(registry_path) if registry_path else get_runtime_registry_path()
	state_file = Path(state_path) if state_path else get_runtime_state_path()
	run_root = Path(output_dir) if output_dir else _default_run_dir(run_date_value.isoformat())
	run_root.mkdir(parents=True, exist_ok=True)
	downloads_dir = run_root / "downloads"

	store_registry = load_store_registry(registry_file)
	item_mapping = load_item_mapping()
	runtime_state = load_runtime_state(state_file)
	bootstrap = ensure_required_master_data(store_registry, item_mapping)
	drive, sheets = _build_google_services()

	payload_rows_all: list[dict[str, Any]] = []
	exception_rows_all: list[dict[str, Any]] = []
	audit_rows_all: list[dict[str, Any]] = []
	failed_stores: list[dict[str, str]] = []
	imported_stores = 0
	skipped_unchanged = 0
	skipped_non_shadow = 0
	skipped_disabled = 0
	rows_created = 0
	rows_updated = 0
	rows_failed = 0
	enabled_stores = 0

	from hrms.api import erp_sync

	selected_codes = {code.strip().upper() for code in (store_codes or []) if code and code.strip()}

	for store in store_registry:
		if selected_codes and store.store_code.upper() not in selected_codes:
			continue
		if not store.sheet_sync_enabled:
			skipped_disabled += 1
			continue
		if store.state not in IMPORTABLE_STATES:
			skipped_non_shadow += 1
			continue

		enabled_stores += 1
		try:
			workbook_path = export_store_workbook(store, downloads_dir, drive, sheets)
			store_result = extract_store_inventory_payload(store, workbook_path, item_mapping, run_date_value)
			payload_rows = store_result["payload_rows"]
			exception_rows = store_result["exception_rows"]
			audit_rows = store_result["audit_rows"]
			payload_rows_all.extend(payload_rows)
			exception_rows_all.extend(exception_rows)
			audit_rows_all.extend(audit_rows)

			if not force and store.last_checksum and store.last_checksum == store_result["checksum"]:
				skipped_unchanged += 1
				continue

			if payload_rows:
				sync_result = erp_sync._sync_inventory_rows(
					sheet_name=f"Store Inventory Shadow Sync {store.store_code}",
					data=payload_rows,
					checksum=store_result["checksum"],
					require_auth=False,
				)
				rows_created += sync_result["rows_created"]
				rows_updated += sync_result["rows_updated"]
				rows_failed += sync_result["rows_failed"]
				if sync_result["rows_failed"]:
					store.last_error = "; ".join(sync_result["errors"])
				else:
					store.last_error = ""
				imported_stores += 1
			else:
				store.last_error = f"No payload rows generated; {len(exception_rows)} exception rows"

			store.last_checksum = store_result["checksum"]
			store.last_success_at = _now_ts()
			store.last_import_rows = len(payload_rows)
			store.last_inventory_date = run_date_value.isoformat()
			runtime_state.setdefault("stores", {})[store.store_code] = {
				"last_checksum": store.last_checksum,
				"last_success_at": store.last_success_at,
				"last_import_rows": store.last_import_rows,
				"last_inventory_date": store.last_inventory_date,
				"last_error": store.last_error,
			}

		except Exception as exc:
			store.last_error = str(exc)
			failed_stores.append({"store_code": store.store_code, "error": str(exc)})

	save_store_registry(store_registry, registry_file)
	runtime_state["last_run"] = {
		"run_date": run_date_value.isoformat(),
		"generated_at": _now_ts(),
		"imported_stores": imported_stores,
		"skipped_unchanged": skipped_unchanged,
		"failed_stores": len(failed_stores),
	}
	save_runtime_state(runtime_state, state_file)

	_write_csv(
		run_root / "store_inventory_payload.csv",
		[
			"store_code",
			"store_name",
			"warehouse",
			"item_code",
			"qty",
			"qty_source",
			"inventory_date",
			"spreadsheet_id",
			"source_row",
			"workbook_path",
		],
		payload_rows_all,
	)
	_write_csv(
		run_root / "store_inventory_exceptions.csv",
		[
			"store_code",
			"store_name",
			"warehouse_name",
			"warehouse_docname",
			"source_row",
			"section",
			"inventory_code",
			"inventory_items",
			"inventory_description",
			"inventory_uom",
			"classification",
			"target_item_code",
			"resolved_qty",
			"qty_source",
			"inventory_date",
			"detail",
		],
		exception_rows_all,
	)
	_write_csv(
		run_root / "store_inventory_resolution_audit.csv",
		[
			"store_code",
			"store_name",
			"warehouse_name",
			"warehouse_docname",
			"source_row",
			"section",
			"inventory_code",
			"inventory_items",
			"inventory_description",
			"inventory_uom",
			"classification",
			"target_item_code",
			"resolved_qty",
			"qty_source",
			"inventory_date",
			"detail",
		],
		audit_rows_all,
	)

	summary = {
		"run_date": run_date_value.isoformat(),
		"generated_at": _now_ts(),
		"registry_path": str(registry_file),
		"state_path": str(state_file),
		"run_root": str(run_root),
		"enabled_stores": enabled_stores,
		"imported_stores": imported_stores,
		"skipped_unchanged": skipped_unchanged,
		"skipped_non_shadow": skipped_non_shadow,
		"skipped_disabled": skipped_disabled,
		"payload_rows": len(payload_rows_all),
		"exception_rows": len(exception_rows_all),
		"rows_created": rows_created,
		"rows_updated": rows_updated,
		"rows_failed": rows_failed,
		"failed_stores": failed_stores,
		"master_data_bootstrap": bootstrap,
	}
	_write_json(run_root / "summary.json", summary)
	(run_root / "summary.md").write_text(_summary_markdown(summary), encoding="utf-8")
	return summary


def cli_run(
	run_date: str | None = None,
	force: bool = False,
	store_codes: list[str] | None = None,
	output_dir: str | None = None,
) -> int:
	if frappe:
		result = run_store_inventory_shadow_sync(
			run_date=run_date,
			force=force,
			store_codes=store_codes,
			output_dir=output_dir,
		)
		print(json.dumps(result, indent=2))
		return 0 if not result["failed_stores"] else 10

	raise RuntimeError("CLI mode without Frappe is not supported for this runner")
