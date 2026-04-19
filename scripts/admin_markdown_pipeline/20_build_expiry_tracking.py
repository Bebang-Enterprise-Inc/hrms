"""Build _EXPIRING_SOON.csv + _EXPIRING_SOON.xlsx with conditional formatting."""

from __future__ import annotations

import csv
import json
import re
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
MD_ROOT = REPO_ROOT / "data" / "admin_markdown"
CSV_OUT = MD_ROOT / "_EXPIRING_SOON.csv"
XLSX_OUT = MD_ROOT / "_EXPIRING_SOON.xlsx"


def _parse_yaml_value(raw: str):
    raw = raw.strip()
    if raw in {"", "null", "~"}: return None
    if raw.startswith(('"', "[", "{")):
        try: return json.loads(raw)
        except json.JSONDecodeError: return raw.strip('"')
    try: return float(raw) if "." in raw else int(raw)
    except ValueError: return raw


def _parse_frontmatter(text: str) -> dict:
    if not text.startswith("---"): return {}
    parts = text.split("---", 2)
    if len(parts) < 3: return {}
    out: dict = {}
    for line in parts[1].splitlines():
        if not line.strip() or line.startswith("#"): continue
        m = re.match(r"^([A-Za-z0-9_]+):\s?(.*)$", line)
        if m: out[m.group(1)] = _parse_yaml_value(m.group(2))
    return out


def _parse_date(s) -> Optional[date]:
    if not s: return None
    s = str(s).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y", "%B %d, %Y", "%B %d %Y"):
        try: return datetime.strptime(s, fmt).date()
        except ValueError: continue
    try: return date.fromisoformat(s[:10])
    except ValueError: return None


def status_for(days):
    if days is None: return "UNKNOWN"
    if days < 0: return "EXPIRED"
    if days <= 30: return "EXPIRING_30_DAYS"
    if days <= 90: return "EXPIRING_90_DAYS"
    return "VALID"


def action_for(status, permit):
    if status == "EXPIRED": return f"RENEW NOW: {permit} is past expiry."
    if status == "EXPIRING_30_DAYS": return f"URGENT: renew {permit} within 30 days."
    if status == "EXPIRING_90_DAYS": return f"Renew {permit} this quarter."
    return ""


def main():
    today = date.today()
    rows: list[dict] = []
    for p in MD_ROOT.rglob("*.md"):
        if p.name.startswith("_") or "_staging" in p.parts or "_validation" in p.parts or "_pages" in p.parts:
            continue
        fm = _parse_frontmatter(p.read_text(encoding="utf-8"))
        if not fm: continue
        expiry_dt = _parse_date(str(fm.get("canonical_expiry_date") or ""))
        issue_dt = _parse_date(str(fm.get("canonical_issue_date") or ""))
        days = (expiry_dt - today).days if expiry_dt else None
        status = status_for(days)
        rows.append({
            "entity_code": fm.get("entity_code") or "",
            "entity_name": fm.get("entity_legal_name") or "",
            "store_mapping": fm.get("entity_store_mapping") or "",
            "permit_code": fm.get("permit_code") or "",
            "document_type": fm.get("document_type") or "",
            "permit_number": fm.get("canonical_permit_number") or "",
            "issuing_authority": fm.get("canonical_issuing_authority") or "",
            "issue_date": issue_dt.isoformat() if issue_dt else (str(fm.get("canonical_issue_date") or "")),
            "expiry_date": expiry_dt.isoformat() if expiry_dt else (str(fm.get("canonical_expiry_date") or "")),
            "days_until_expiry": days if days is not None else "",
            "status": status,
            "owner_action_needed": action_for(status, fm.get("permit_code") or ""),
            "source_drive_url": fm.get("source_drive_url") or "",
            "md_file_path": str(p.relative_to(MD_ROOT)).replace("\\", "/"),
            "validation_method": fm.get("validation_method") or "",
            "dd_ready": fm.get("dd_ready"),
        })

    def sort_key(r):
        d = r["days_until_expiry"]
        return (0 if d == "" else (1 if isinstance(d, int) and d < 0 else 2), d if d != "" else 10**9)
    rows.sort(key=sort_key)

    cols = list(rows[0].keys()) if rows else []
    with CSV_OUT.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for r in rows: w.writerow(r)
    print(f"wrote {CSV_OUT.relative_to(REPO_ROOT)} ({len(rows)} rows)")

    wb = Workbook()
    ws = wb.active
    ws.title = "Expiring Permits"
    ws.append(cols)
    for r in rows: ws.append([r.get(c) for c in cols])

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="04400A")
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill

    red = PatternFill("solid", fgColor="F4CCCC")
    amber = PatternFill("solid", fgColor="FFE5B4")
    yellow = PatternFill("solid", fgColor="FFF2CC")
    green = PatternFill("solid", fgColor="D9EAD3")

    last_col = get_column_letter(len(cols))
    status_col = get_column_letter(cols.index("status") + 1)
    last_row = ws.max_row
    rng = f"A2:{last_col}{last_row}"
    for status_val, fill in [("EXPIRED", red), ("EXPIRING_30_DAYS", amber), ("EXPIRING_90_DAYS", yellow), ("VALID", green)]:
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'${status_col}2="{status_val}"'], fill=fill))

    ws.freeze_panes = "A2"
    for i, c in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(len(c) + 2, 12), 60)

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Status"; ws2["B1"] = "Count"
    for cell in [ws2["A1"], ws2["B1"]]:
        cell.font = header_font; cell.fill = header_fill
    c = Counter(r["status"] for r in rows)
    row_cursor = 2
    for s in ["EXPIRED", "EXPIRING_30_DAYS", "EXPIRING_90_DAYS", "VALID", "UNKNOWN"]:
        ws2.cell(row=row_cursor, column=1, value=s)
        ws2.cell(row=row_cursor, column=2, value=c.get(s, 0))
        row_cursor += 1

    wb.save(XLSX_OUT)
    print(f"wrote {XLSX_OUT.relative_to(REPO_ROOT)}")
    for s, n in c.most_common():
        print(f"  {s:20s} {n}")


if __name__ == "__main__":
    main()
