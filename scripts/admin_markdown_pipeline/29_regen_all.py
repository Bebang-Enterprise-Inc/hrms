"""Regenerate _MASTER_MANIFEST.csv, _INDEX.md, and _DD_COVERAGE_MATRIX.xlsx
reading from the new CEO/Valuation/admin_compliance_dd/ location with ai_* fields.
"""
from __future__ import annotations
import csv, json, re, sys
from collections import defaultdict
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
DD = ROOT / "CEO" / "Valuation" / "admin_compliance_dd"
REGISTRY = ROOT / "data" / "admin_drive_audit" / "company_registry.csv"

MANIFEST = DD / "_MASTER_MANIFEST.csv"
INDEX_MD = DD / "_INDEX.md"
XLSX = DD / "_DD_COVERAGE_MATRIX.xlsx"


def parse_fm(text: str) -> dict:
    if not text.startswith("---"):
        return {}
    parts = text.split("---", 2)
    if len(parts) < 3:
        return {}
    out = {}
    for line in parts[1].splitlines():
        if not line.strip() or line.startswith("#"):
            continue
        m = re.match(r"^([A-Za-z0-9_]+):\s?(.*)$", line)
        if not m:
            continue
        k, v = m.group(1), m.group(2).strip()
        if v.startswith('"') and v.endswith('"'):
            v = v[1:-1]
        elif v in {"null", "~", ""}:
            v = None
        elif v.startswith("[") or v.startswith("{"):
            try:
                v = json.loads(v)
            except Exception:
                pass
        out[k] = v
    return out


def collect_mds() -> list[dict]:
    rows = []
    for p in sorted(DD.rglob("*.md")):
        if p.name.startswith("_"):
            continue
        if any(part.startswith("_") for part in p.relative_to(DD).parts[:-1]):
            continue
        try:
            fm = parse_fm(p.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not fm:
            continue
        fm["_md_relative"] = str(p.relative_to(DD)).replace("\\", "/")
        rows.append(fm)
    return rows


def write_manifest(rows: list[dict]) -> None:
    preferred = [
        "_md_relative", "entity_code", "entity_legal_name", "entity_store_mapping",
        "permit_code", "document_type", "ai_label", "ai_category", "short_description",
        "canonical_business_name", "canonical_trade_name", "canonical_permit_number",
        "canonical_tin", "canonical_ocn", "canonical_psic_code",
        "canonical_issuing_authority", "canonical_issue_date", "canonical_expiry_date",
        "canonical_location_address", "canonical_registered_address", "canonical_signatories",
        "validation_method", "disagreements", "dd_ready",
        "mistral_document_type", "mistral_tin", "mistral_ocn", "mistral_signatories",
        "gemini_pro_document_type", "gemini_pro_tin", "gemini_pro_ocn", "gemini_pro_signatories",
        "opus_arbitration_reasoning", "opus_source_of_truth",
        "source_drive_id", "source_drive_url", "source_path", "source_md5",
        "source_modified", "size_bytes",
        "extraction_date", "extraction_models", "extraction_cost_usd",
    ]
    keys, seen = [], set()
    for k in preferred:
        keys.append(k); seen.add(k)
    for r in rows:
        for k in r:
            if k not in seen:
                keys.append(k); seen.add(k)

    def _val(v):
        if v is None: return ""
        if isinstance(v, (list, dict)): return json.dumps(v, ensure_ascii=False)
        return v

    with MANIFEST.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=keys, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: _val(r.get(k)) for k in keys})
    print(f"wrote {MANIFEST} ({len(rows)} rows)")


def write_index(rows: list[dict]) -> None:
    dd_count = sum(1 for r in rows if r.get("dd_ready") is True)
    arb = sum(1 for r in rows if r.get("validation_method") == "opus_arbitrated")
    dual = sum(1 for r in rows if r.get("validation_method") == "dual_match")
    lines = [
        "# BEI Admin Compliance — Master Index",
        "",
        f"**Total files:** {len(rows)} · DD ready: {dd_count} · Opus-arbitrated: {arb} · Dual-match: {dual}",
        "",
    ]
    by_entity = defaultdict(list)
    for r in rows:
        by_entity[str(r.get("entity_code") or "UNKNOWN")].append(r)
    for ent in sorted(by_entity):
        ent_rows = by_entity[ent]
        legal = ent_rows[0].get("entity_legal_name") or ""
        store = ent_rows[0].get("entity_store_mapping") or ""
        lines.append(f"## {ent}" + (f" — {legal}" if legal else "") + (f"  *(store: {store})*" if store else ""))
        lines.append("")
        lines.append("| AI Label | Category | Doc Type | Issue | Expiry | DD | File |")
        lines.append("|---|---|---|---|---|---|---|")
        ent_rows.sort(key=lambda r: (str(r.get("ai_category") or ""), str(r.get("document_type") or ""), str(r.get("canonical_issue_date") or "")))
        for r in ent_rows:
            lbl = r.get("ai_label") or "(unlabeled)"
            cat = r.get("ai_category") or ""
            dt = r.get("document_type") or ""
            issue = r.get("canonical_issue_date") or ""
            expiry = r.get("canonical_expiry_date") or ""
            dd = "✅" if r.get("dd_ready") is True else "⚠️"
            md_rel = r.get("_md_relative") or ""
            url = r.get("source_drive_url") or ""
            file_md = f"[{Path(md_rel).name}]({md_rel})" if md_rel else ""
            if url:
                file_md = f"{file_md} · [Drive]({url})"
            lines.append(f"| {lbl} | {cat} | {dt} | {issue} | {expiry} | {dd} | {file_md} |")
        lines.append("")
    INDEX_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"wrote {INDEX_MD}")


CORP_DOC_TYPES = [
    ("SEC_COI", "SEC Certificate of Incorporation"),
    ("SEC_AOI", "Articles of Incorporation"),
    ("SEC_BYLAWS", "By-Laws"),
    ("SEC_GIS", "General Information Sheet (latest)"),
    ("SEC_BOARD_RES", "Board Resolution"),
    ("BIR_2303", "BIR 2303 (Certificate of Registration)"),
]

STORE_DOC_TYPES = [
    ("BIR_2303", "BIR 2303 (branch)"),
    ("BIR_PTU_POS", "BIR Permit to Use POS"),
    ("MAYORS_PERMIT", "Mayor's Permit"),
    ("FSIC", "Fire Safety Cert (FSIC)"),
    ("SANITARY", "Sanitary Permit"),
    ("BARANGAY_CLEARANCE", "Barangay Clearance"),
    ("BUILDING_PERMIT", "Building Permit"),
    ("CERT_OCCUPANCY", "Cert of Occupancy"),
    ("LEASE", "Lease Contract"),
    ("INSURANCE", "CGL / Liability Insurance"),
]

CORP_ALIAS = {
    "BB ESTANCIA FOOD CORP.": "CORP_BB_ESTANCIA", "BEBANG BF HOMES INC.": "CORP_BF_HOMES",
    "BEBANG ENTERPRISE INC.": "CORP_BEI", "BEBANG FESTIVAL INC.": "CORP_FESTIVAL",
    "BEBANG FT INC.": "CORP_FAIRVIEW_TERRACES", "BEBANG FRANCHISE CORP.": "CORP_FRANCHISE",
    "BEBANG FRANCHISE OPC": "CORP_FRANCHISE_OPC", "BEIFRANCHISE FOOD OPC": "CORP_FRANCHISE_OPC",
    "BEBANG GRAND CENTRAL INC.": "CORP_GRAND_CENTRAL", "BEBANG LCT INC.": "CORP_LCT",
    "BEBANG MARILAO INC.": "CORP_MARILAO", "BEBANG MARKET MARKET INC.": "CORP_MARKET_MARKET",
    "BEBANG MEGA INC.": "CORP_MEGA", "BEBANG NORTH EDSA INC.": "CORP_NORTH_EDSA",
    "BEBANG PASEO INC.": "CORP_PASEO", "BEBANG PITX INC.": "CORP_PITX",
    "BEBANG SM BICUTAN INC.": "CORP_BICUTAN", "BEBANG SM MARIKINA INC.": "CORP_SM_MARIKINA",
    "BEBANG SMEO INC.": "CORP_SMEO", "BEBANG SMM INC.": "CORP_SMM",
    "BEBANG SMOA INC.": "CORP_SMOA", "BEBANG SMV INC.": "CORP_SMV",
    "BEBANG STARMALL ALABANG INC.": "CORP_STARMALL_ALABANG", "BEBANG UP TOWN CENTER INC.": "CORP_UPTC",
    "BEBANG VENICE GRAND CANAL INC.": "CORP_VENICE", "Bebang Kitchen Inc.": "CORP_BKI",
    "DAY ONES FOOD AND DRINK ESTABLISHMENTS CORP.": "CORP_DAY_ONES", "DLS DESSERT CRAFT INC.": "CORP_DLS",
    "DMD HOLDINGS INC.": "CORP_UPTOWN_BGC", "FREEZE DELIGHT INC.": "CORP_FREEZE_DELIGHT",
    "HALO-HALO TERMINAL FOOD CORP.": "CORP_HALO_TERMINAL", "HFFM SOLENAD FOOD SERVICES INC.": "STORE_SOLENAD",
    "Irresistible Infusions Inc.": "CORP_UPTOWN_BGC", "JL TRADE OPC": "CORP_SJDM",
    "LEGACY77 FOOD CORP.": "CORP_LEGACY77", "PERPETUAL FOOD CORP.": "CORP_PERPETUAL_MONTALBAN",
    "RED TALDAWA FOODS OPC": "CORP_RED_TALDAWA", "SWEET HARMONY FOOD CORP.": "CORP_SWEET_HARMONY",
    "TAJ FOOD CORP.": "CORP_TAJ", "TASTECARTEL CORP.": "CORP_TASTECARTEL",
    "TRICERN FOOD CORP.": "CORP_TUNGSTEN", "TUNGSTEN CAPITAL HOLDINGS OPC": "CORP_TUNGSTEN",
    "B CUBED VENTURES CORP.": "CORP_B_CUBED",
}


GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
HEADER_FILL = PatternFill("solid", fgColor="04400A")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BORDER = Border(left=Side(style="thin", color="BFBFBF"), right=Side(style="thin", color="BFBFBF"),
                top=Side(style="thin", color="BFBFBF"), bottom=Side(style="thin", color="BFBFBF"))


def write_xlsx(rows: list[dict]) -> None:
    # Build availability index
    by_ec_dt: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in rows:
        ec = (r.get("entity_code") or "").upper()
        dt = (r.get("document_type") or "").upper()
        by_ec_dt[ec][dt] += 1

    # Load stores
    stores = []
    with REGISTRY.open(encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            if (r.get("entity_category") or "").strip() != "Store": continue
            if (r.get("operational_status") or "").strip() != "Active": continue
            stores.append({
                "name": (r.get("full_company_name") or "").strip(),
                "store_prefix": (r.get("store_prefix") or "").strip(),
                "abbr": (r.get("abbr") or "").strip(),
                "corp": (r.get("corp_suffix") or "").strip(),
                "ownership": (r.get("store_ownership_type") or "").strip(),
            })

    wb = Workbook()

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "BEI Admin Compliance — DD Coverage Summary"
    ws["A1"].font = Font(bold=True, size=14, color="04400A")
    ws.merge_cells("A1:E1")
    ws["A3"] = "Generated from 1,000 extracted admin compliance PDFs (Mistral OCR 3 + Gemini 3.1 Pro v2 + Claude Opus 4.7 forensic arbitration + AI reclassification for 166 ambiguous files)."
    ws["A3"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A3:E3")

    ws["A5"] = "Scope"; ws["B5"] = "Count"; ws["C5"] = "Notes"
    for c in (ws["A5"], ws["B5"], ws["C5"]):
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.border = BORDER
    summary_rows = [
        ("Total extracted files", len(rows), "with full triple-validation"),
        ("Active stores in registry", len(stores), "47 active stores"),
        ("Managed Franchise stores", sum(1 for s in stores if s["ownership"] == "Managed Franchise"), "requires FA"),
        ("JV stores", sum(1 for s in stores if s["ownership"] == "JV"), "requires JV agreement"),
        ("Full Franchise stores", sum(1 for s in stores if s["ownership"] == "Full Franchise"), "requires FA"),
        ("Company Owned stores", sum(1 for s in stores if s["ownership"] == "Company Owned"), ""),
    ]
    for i, (a, b, c) in enumerate(summary_rows, start=6):
        ws.cell(row=i, column=1, value=a).border = BORDER
        ws.cell(row=i, column=2, value=b).border = BORDER
        ws.cell(row=i, column=3, value=c).border = BORDER

    # Sheet 2: Corps
    ws2 = wb.create_sheet("Corps")
    ws2["A1"] = "Corporation"; ws2["B1"] = "Alias Code"; ws2["C1"] = "Total Files"
    for i, (dt, label) in enumerate(CORP_DOC_TYPES, start=4):
        ws2.cell(row=1, column=i, value=label)
    for c in ws2[1]:
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center"); c.border = BORDER
    ws2.row_dimensions[1].height = 45

    all_corps, seen = [], set()
    corps_from_registry = {s["corp"] for s in stores if s["corp"]}
    for c in sorted(corps_from_registry):
        ec = CORP_ALIAS.get(c)
        if ec and ec not in seen:
            all_corps.append((c, ec)); seen.add(ec)
    for extra in [("Bebang Kitchen Inc.", "CORP_BKI"), ("Resto Tech Inc.", "CORP_RESTO_TECH"),
                  ("Bebang Shaw Inc.", "CORP_SHAW"), ("Bebang Comintanes Food Corp", "CORP_COMINTANES"),
                  ("Halo-Halo Alabang Food Corp", "CORP_HALO_ALABANG")]:
        if extra[1] not in seen:
            all_corps.append(extra); seen.add(extra[1])
    # S208: append FRANCHISE_*/JV_* partner entities discovered in extracted docs
    partner_codes = sorted(
        ec for ec in by_ec_dt.keys()
        if (ec.startswith("FRANCHISE_") or ec.startswith("JV_")) and ec not in seen
    )
    for ec in partner_codes:
        display = ec.replace("FRANCHISE_", "Franchise: ").replace("JV_", "JV: ").replace("_", " ").title()
        all_corps.append((display, ec))
        seen.add(ec)

    for row_i, (corp_name, ec) in enumerate(all_corps, start=2):
        ws2.cell(row=row_i, column=1, value=corp_name).border = BORDER
        ws2.cell(row=row_i, column=2, value=ec).border = BORDER
        total = sum(by_ec_dt[ec].values())
        ws2.cell(row=row_i, column=3, value=total).border = BORDER
        for i, (dt, _) in enumerate(CORP_DOC_TYPES, start=4):
            n = by_ec_dt[ec].get(dt, 0)
            if dt == "SEC_COI" and n == 0:
                n = by_ec_dt[ec].get("SEC_CERT", 0)
            cell = ws2.cell(row=row_i, column=i, value=n if n else "—")
            cell.alignment = Alignment(horizontal="center"); cell.border = BORDER
            cell.fill = GREEN if n >= 1 else RED

    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 26
    ws2.column_dimensions["C"].width = 12
    for i in range(4, 4 + len(CORP_DOC_TYPES)):
        ws2.column_dimensions[get_column_letter(i)].width = 18
    ws2.freeze_panes = "D2"

    # Sheet 3: Stores
    ws3 = wb.create_sheet("Stores")
    headers3 = ["Store", "Corp", "Ownership", "Total Files"] + [lbl for _, lbl in STORE_DOC_TYPES] + ["Franchise / JV Agreement"]
    for i, h in enumerate(headers3, start=1):
        c = ws3.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center"); c.border = BORDER
    ws3.row_dimensions[1].height = 45

    # S208: stores whose FA/JV/Management contract was ingested under a
    # partner-side entity code instead of (or in addition to) the BEI CORP_*
    # code. Listing them here lets the Stores sheet count those docs as
    # covered instead of showing a misleading "—" (missing) dash.
    STORE_PARTNER_CODES = {
        "MEGAWIDE PITX": ["JV_EMPIRE77", "JV_TOPLEVEL"],
        "MEGAWORLD PASEO CENTER": ["JV_PERPETUALLY_CANDID", "JV_TOPLEVEL"],
        "SM EAST ORTIGAS": ["JV_PERPETUALLY_CANDID"],
        "BF HOMES": ["JV_EDWARD_SY_RALPH_TY"],
        "UP TOWN MALL BGC": ["JV_IMELDA"],
        "AYALA VERMOSA": ["JV_TOPLEVEL"],
        "AYALA EVO CITY": ["JV_TOPLEVEL"],
        "ROBINSONS IMUS": ["JV_TOPLEVEL"],
        "SM TANZA": ["JV_TOPLEVEL"],
        "THE GRID ROCKWELL": ["FRANCHISE_TASTECARTEL"],
    }

    def store_entity_codes(s):
        cands = []
        clean = re.sub(r"[^A-Z0-9]", "", s["store_prefix"].upper())
        if clean: cands.append(f"STORE_{clean}")
        if s["abbr"]: cands.append(f"STORE_{s['abbr'].upper()}")
        corp_ec = CORP_ALIAS.get(s["corp"])
        if corp_ec: cands.append(corp_ec)
        # Add partner-side entity codes when the store has cross-mapped docs
        store_name_upper = s["name"].upper()
        for key, partners in STORE_PARTNER_CODES.items():
            if key in store_name_upper:
                cands.extend(partners)
                break
        return cands

    for row_i, s in enumerate(stores, start=2):
        ecs = store_entity_codes(s)
        merged = defaultdict(int)
        for ec in ecs:
            for dt, n in by_ec_dt[ec].items():
                merged[dt] += n
        total = sum(merged.values())

        ws3.cell(row=row_i, column=1, value=s["name"]).border = BORDER
        ws3.cell(row=row_i, column=2, value=s["corp"]).border = BORDER
        ws3.cell(row=row_i, column=3, value=s["ownership"]).border = BORDER
        ws3.cell(row=row_i, column=4, value=total).border = BORDER
        for i, (dt, _) in enumerate(STORE_DOC_TYPES, start=5):
            n = merged.get(dt, 0)
            if dt == "MAYORS_PERMIT" and n == 0:
                n = merged.get("BUSINESS_PERMIT", 0)
            cell = ws3.cell(row=row_i, column=i, value=n if n else "—")
            cell.alignment = Alignment(horizontal="center"); cell.border = BORDER
            cell.fill = GREEN if n >= 1 else (YELLOW if dt in {"BUILDING_PERMIT", "CERT_OCCUPANCY"} else RED)

        last_col = len(headers3)
        if s["ownership"] == "Company Owned":
            cell = ws3.cell(row=row_i, column=last_col, value="N/A"); cell.fill = YELLOW
        else:
            n_fa = merged.get("FRANCHISE_AGREEMENT", 0) + merged.get("JV_AGREEMENT", 0) + merged.get("CONTRACT", 0)
            cell = ws3.cell(row=row_i, column=last_col, value=n_fa if n_fa else "—")
            cell.fill = GREEN if n_fa else RED
        cell.alignment = Alignment(horizontal="center"); cell.border = BORDER

    ws3.column_dimensions["A"].width = 45
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 12
    for i in range(5, 5 + len(STORE_DOC_TYPES) + 1):
        ws3.column_dimensions[get_column_letter(i)].width = 14
    ws3.freeze_panes = "E2"

    # Sheet 4: All_Docs with ai_* fields
    ws4 = wb.create_sheet("All_Docs")
    hdr = ["entity_code", "ai_category", "document_type", "ai_label", "short_description",
           "canonical_business_name", "canonical_trade_name", "canonical_issue_date",
           "canonical_expiry_date", "validation_method", "dd_ready",
           "_md_relative", "source_drive_url"]
    for i, h in enumerate(hdr, start=1):
        c = ws4.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.border = BORDER
    ws4.row_dimensions[1].height = 30
    for ridx, r in enumerate(rows, start=2):
        for cidx, k in enumerate(hdr, start=1):
            v = r.get(k, "")
            if isinstance(v, (list, dict)):
                v = json.dumps(v, ensure_ascii=False)
            ws4.cell(row=ridx, column=cidx, value=v).border = BORDER
    for i, h in enumerate(hdr, start=1):
        width = 22
        if h in {"ai_label", "short_description", "canonical_business_name", "canonical_trade_name", "_md_relative", "source_drive_url"}:
            width = 45
        ws4.column_dimensions[get_column_letter(i)].width = width
    ws4.freeze_panes = "A2"
    ws4.auto_filter.ref = ws4.dimensions

    # Sheet 5: By_Category (pivot)
    ws5 = wb.create_sheet("By_Category")
    ws5["A1"] = "AI Category"; ws5["B1"] = "Doc Type"; ws5["C1"] = "Count"
    for c in ws5[1]:
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.border = BORDER
    piv = defaultdict(int)
    for r in rows:
        piv[((r.get("ai_category") or "?"), (r.get("document_type") or "?"))] += 1
    ordered = sorted(piv.items(), key=lambda x: (-x[1], x[0]))
    for i, ((cat, dt), n) in enumerate(ordered, start=2):
        ws5.cell(row=i, column=1, value=cat).border = BORDER
        ws5.cell(row=i, column=2, value=dt).border = BORDER
        ws5.cell(row=i, column=3, value=n).border = BORDER
    ws5.column_dimensions["A"].width = 20
    ws5.column_dimensions["B"].width = 30
    ws5.column_dimensions["C"].width = 10

    wb.save(XLSX)
    print(f"wrote {XLSX}")


def main() -> None:
    rows = collect_mds()
    print(f"collected {len(rows)} MDs from {DD}")
    write_manifest(rows)
    write_index(rows)
    write_xlsx(rows)

    # Stats
    has_ai_label = sum(1 for r in rows if r.get("ai_label"))
    cats = defaultdict(int)
    for r in rows:
        cats[r.get("ai_category") or "(none)"] += 1
    print(f"\nAI-labeled files: {has_ai_label}/{len(rows)}")
    print("By ai_category:")
    for cat, n in sorted(cats.items(), key=lambda x: -x[1]):
        print(f"  {cat:20} {n}")


if __name__ == "__main__":
    main()
