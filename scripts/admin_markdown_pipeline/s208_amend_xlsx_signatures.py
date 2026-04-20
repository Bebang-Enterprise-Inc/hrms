"""Amend _DD_COVERAGE_MATRIX.xlsx with signature-status info.

Adds:
- New sheet `Signature_Gaps` listing every FA/MA/JV/Contract doc that is NOT
  dual-signed by BEI CEO (or delegate) + counterparty.
- New column on `Stores` sheet `Signature Status` with a note for stores that
  need signature action. Uses explicit file paths for cross-mapped stores so
  the team doesn't have to hunt.
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
DD = ROOT / "CEO" / "Valuation" / "admin_compliance_dd"
XLSX = DD / "_DD_COVERAGE_MATRIX.xlsx"

CONTRACT_TYPES = {"FRANCHISE_AGREEMENT", "JV_AGREEMENT", "CONTRACT"}
BEI_CEO_TERMS = ["sam karazi", "samer karazi", "karazi"]
BEI_DELEGATE_TERMS = ["manansala"]
CONTRACT_KEYWORDS = ["franchise", "jv", "management agreement",
                     "joint venture", "memorandum", "agreement", "contract"]

GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
HEADER_FILL = PatternFill("solid", fgColor="04400A")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BORDER = Border(left=Side(style="thin", color="BFBFBF"), right=Side(style="thin", color="BFBFBF"),
                top=Side(style="thin", color="BFBFBF"), bottom=Side(style="thin", color="BFBFBF"))

_field_re = {k: re.compile(rf"^{k}:\s?(.*)$", re.MULTILINE) for k in
             ("document_type", "ai_label", "canonical_business_name",
              "entity_code", "entity_store_mapping", "permit_code",
              "source_drive_url", "canonical_issue_date")}
_sigs_re = re.compile(r"canonical_signatories:\s*(\[.*?\])(?=\n[a-z_]+:|\n---)", re.DOTALL)


def _unquote(s: str) -> str:
    s = s.strip()
    if s.startswith('"') and s.endswith('"'):
        return s[1:-1]
    return s


def scan_contracts() -> list[dict]:
    out: list[dict] = []
    for md in DD.rglob("*.md"):
        rel = md.relative_to(DD)
        if rel.parts and rel.parts[0].startswith("_"):
            continue
        if md.name.startswith("_"):
            continue
        try:
            text = md.read_text(encoding="utf-8")
        except Exception:
            continue
        if not text.startswith("---"):
            continue
        parts = text.split("---", 2)
        if len(parts) < 3:
            continue
        fm = parts[1]
        fields = {k: (_unquote(rx.search(fm).group(1)) if rx.search(fm) else "")
                  for k, rx in _field_re.items()}
        dt = fields["document_type"]
        if dt not in CONTRACT_TYPES:
            continue
        if not any(k in md.name.lower() for k in CONTRACT_KEYWORDS):
            continue
        sigs_m = _sigs_re.search(fm)
        sigs: list = []
        if sigs_m:
            try:
                sigs = json.loads(sigs_m.group(1))
            except Exception:
                sigs = []
        sig_names = [(s.get("name", "") if isinstance(s, dict) else str(s)) for s in sigs]
        sig_titles = [(s.get("title", "") if isinstance(s, dict) else "") for s in sigs]
        out.append({
            "path": str(rel),
            **fields,
            "sig_names": sig_names,
            "sig_titles": sig_titles,
        })
    return out


def classify(a: dict) -> tuple[str, str]:
    sigs_lower = [s.lower() for s in a["sig_names"]]
    has_ceo = any(any(t in s for t in BEI_CEO_TERMS) for s in sigs_lower)
    has_delegate = any(any(t in s for t in BEI_DELEGATE_TERMS) for s in sigs_lower)
    n = len(a["sig_names"])
    if n == 0:
        return ("NO_SIGNATORIES", "OCR captured zero signatures — manually verify PDF.")
    if n == 1 and has_ceo:
        return ("BLANK_TEMPLATE", "Only BEI CEO signed — likely a blank/draft template; counterparty pending.")
    if n == 1:
        return ("SINGLE_SIG", "Only one signature captured — not fully executed.")
    if has_ceo:
        return ("DUAL_SIGNED_CEO", "Signed by BEI CEO + counterparty.")
    if has_delegate:
        return ("DUAL_SIGNED_DELEGATE", "Signed by BEI delegate (not Sam). Buyer's counsel may request CEO re-execution.")
    return ("BEI_SIG_MISSING", "No BEI signature captured — manual PDF re-review required.")


# Stores that need action. Each entry: store-name-substring → (status, detailed note)
STORE_NOTES: dict[str, tuple[str, str]] = {
    # Truly missing — no FA/JV/Management contract anywhere
    "AYALA MARKET MARKET": ("MISSING", "❗ MISSING JV Agreement — collect signed copy from partner."),
    "SM GRAND CENTRAL": ("MISSING", "❗ MISSING JV Agreement — collect signed copy from partner."),
    "NAIA T3": ("MISSING", "❗ MISSING Franchise/Management Agreement — collect signed copy from Halo-Halo Terminal Food Corp."),
    "SM PULILAN": ("MISSING", "❗ MISSING Franchise/Management Agreement — collect signed copy from franchisee."),
    "SM TAYTAY": ("MISSING", "❗ MISSING Franchise/Management Agreement — collect signed copy from Day Ones Food Corp."),
    "XENTROMALL MONTALBAN": ("MISSING", "❗ MISSING Franchise/Management Agreement — collect signed copy from Perpetual Food Corp."),
    # Lease present but FA/MA missing (HFFM Solenad + BB Estancia)
    "AYALA SOLENAD": ("MISSING_FA", "❗ Lease on file (STORE_SOLENAD/LEASE/...) but MISSING signed Franchise/Management Agreement with HFFM Solenad Food Services Inc."),
    "ORTIGAS ESTANCIA": ("MISSING_FA", "❗ Lease on file (CORP_BB_ESTANCIA/LEASE/...) but MISSING signed Franchise/Management Agreement with BB Estancia Food Corp."),
    # Cross-mapped — show exact file paths so team doesn't hunt
    "MEGAWIDE PITX": ("CROSS_MAPPED", "✓ Signed PITX JV Contract on file at JV_EMPIRE77/JV_AGREEMENT/PITX JV Contract.md + JV_TOPLEVEL/JV_AGREEMENT/PITX JV Contract.md"),
    "MEGAWORLD PASEO CENTER": ("CROSS_MAPPED", "✓ Signed JV Contracts on file at JV_PERPETUALLY_CANDID/UNKNOWN/JV_Contract_Paseo_Center.md + JV_TOPLEVEL/JV_AGREEMENT/Paseo JV Agreement.md"),
}


def main() -> None:
    wb = load_workbook(XLSX)
    contracts = scan_contracts()
    print(f"scanned {len(contracts)} FA/MA/JV/Contract docs")

    by_status: dict[str, list[dict]] = {}
    for a in contracts:
        status, note = classify(a)
        a["status"] = status
        a["note"] = note
        by_status.setdefault(status, []).append(a)

    print(f"\nBy signature status:")
    for s, items in by_status.items():
        print(f"  {s}: {len(items)}")

    # --- Sheet: Signature_Gaps ---
    if "Signature_Gaps" in wb.sheetnames:
        del wb["Signature_Gaps"]
    ws = wb.create_sheet("Signature_Gaps")
    headers = ["Status", "Entity", "Document Type", "Label", "Path", "# Sigs",
               "Signatories (name — title)", "Action Note", "Drive URL"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center"); c.border = BORDER
    ws.row_dimensions[1].height = 32

    gap_rows = []
    for status, items in by_status.items():
        if status == "DUAL_SIGNED_CEO":
            continue
        gap_rows.extend(items)

    priority = {
        "BEI_SIG_MISSING": 0, "NO_SIGNATORIES": 1, "SINGLE_SIG": 2,
        "DUAL_SIGNED_DELEGATE": 3, "BLANK_TEMPLATE": 4,
    }
    gap_rows.sort(key=lambda x: (priority.get(x["status"], 9), x["entity_code"]))

    status_fill = {
        "BEI_SIG_MISSING": RED, "NO_SIGNATORIES": RED,
        "SINGLE_SIG": YELLOW, "DUAL_SIGNED_DELEGATE": YELLOW, "BLANK_TEMPLATE": YELLOW,
    }

    for ri, a in enumerate(gap_rows, start=2):
        sigs_disp = " | ".join(
            f"{n} — {t}" if t else n
            for n, t in zip(a["sig_names"], a["sig_titles"])
        )
        ws.cell(row=ri, column=1, value=a["status"]).fill = status_fill.get(a["status"], YELLOW)
        ws.cell(row=ri, column=2, value=a["entity_code"])
        ws.cell(row=ri, column=3, value=a["document_type"])
        ws.cell(row=ri, column=4, value=a["ai_label"])
        ws.cell(row=ri, column=5, value=a["path"])
        ws.cell(row=ri, column=6, value=len(a["sig_names"]))
        ws.cell(row=ri, column=7, value=sigs_disp).alignment = Alignment(wrap_text=True)
        ws.cell(row=ri, column=8, value=a["note"]).alignment = Alignment(wrap_text=True)
        ws.cell(row=ri, column=9, value=a.get("source_drive_url", ""))
        for col in range(1, 10):
            ws.cell(row=ri, column=col).border = BORDER

    for col, w in zip("ABCDEFGHI", [22, 26, 22, 40, 55, 8, 60, 50, 55]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    # --- Stores sheet: Signature Status ---
    ws_stores = wb["Stores"]
    hdr_row = [ws_stores.cell(row=1, column=c).value for c in range(1, ws_stores.max_column + 1)]
    if "Signature Status" in hdr_row:
        sig_col = hdr_row.index("Signature Status") + 1
    else:
        sig_col = ws_stores.max_column + 1
        h = ws_stores.cell(row=1, column=sig_col, value="Signature Status")
        h.fill = HEADER_FILL; h.font = HEADER_FONT
        h.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        h.border = BORDER

    store_idx = hdr_row.index("Store") + 1
    own_idx = hdr_row.index("Ownership") + 1
    fa_idx = hdr_row.index("Franchise / JV Agreement") + 1

    for row in range(2, ws_stores.max_row + 1):
        store = str(ws_stores.cell(row=row, column=store_idx).value or "")
        own = str(ws_stores.cell(row=row, column=own_idx).value or "")
        fa_val = ws_stores.cell(row=row, column=fa_idx).value

        note = ""
        fill = GREEN

        # Explicit per-store notes first
        for key, (status, msg) in STORE_NOTES.items():
            if key in store.upper():
                note = msg
                fill = RED if status.startswith("MISSING") else YELLOW
                break

        if not note:
            if own == "Company Owned":
                note = "N/A (Company Owned)"
                fill = YELLOW
            elif fa_val in (None, "", "—", 0):
                note = "⚠ No contract visible under BEI corp code — verify or collect."
                fill = YELLOW
            else:
                note = "✓ Dual-signed contract on file under BEI corp folder."
                fill = GREEN

        cell = ws_stores.cell(row=row, column=sig_col, value=note)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BORDER
        cell.fill = fill

    ws_stores.column_dimensions[get_column_letter(sig_col)].width = 85

    wb.save(XLSX)
    print(f"\nSaved {XLSX}")
    print(f"  Signature_Gaps sheet: {len(gap_rows)} flagged docs")
    print(f"  Stores sheet: 'Signature Status' column (col {sig_col}) updated")


if __name__ == "__main__":
    main()
