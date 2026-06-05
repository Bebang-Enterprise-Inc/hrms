"""S258 Bridge handoff — Company Register XLSX (58 Frappe Companies with role tagging).

Merges live Frappe state with cleanroom register data. Outputs BEI-branded XLSX
with role + BIR TIN + RDO + VAT status + parent legal entity + canonical role.
"""
from __future__ import annotations
import csv
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# BEI palette
BEI_GREEN = '04400A'
BEI_GOLD = 'C8900A'
BEI_PURPLE = '801297'
BEI_GREEN_TINT = 'E6ECE7'
BEI_GOLD_LIGHT = 'F8F0D9'
BEI_PURPLE_LIGHT = 'F2E5F5'
BEI_TEXT_DARK = '1A1A1A'
BEI_TEXT_WHITE = 'FFFFFF'
BEI_BORDER = 'D4D0C8'

HEADER_FONT = Font(bold=True, color=BEI_TEXT_WHITE, size=11, name='Calibri')
HEADER_FILL = PatternFill('solid', fgColor=BEI_GREEN)
BODY_FONT = Font(color=BEI_TEXT_DARK, size=10, name='Calibri')
EVEN_FILL = PatternFill('solid', fgColor=BEI_GREEN_TINT)
ODD_FILL = PatternFill('solid', fgColor=BEI_TEXT_WHITE)
THIN = Side(style='thin', color=BEI_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def classify_role(c, name, is_group, parent):
    """Tag each Frappe Company with its canonical role."""
    if name == 'IRRESISTIBLE INFUSIONS INC.':
        return 'Holdco (top of tree)'
    if name == 'BEBANG ENTERPRISE INC.':
        return 'Head Office (BEI HQ)'
    if name == 'BEBANG KITCHEN INC.':
        return 'Commissary'
    if name == 'BEBANG FRANCHISE CORP.':
        return 'Franchisor (Fork 1)'
    if name == 'BEBANG FT INC.':
        return 'Legal Entity Parent (BFT)'
    # 4 BEI-TIN stub stores: ROBINSONS ANTIPOLO / SM MANILA / SM MEGAMALL / SM SOUTHMALL
    if name in ('ROBINSONS ANTIPOLO - BEBANG ENTERPRISE INC.',
                'SM MANILA - BEBANG ENTERPRISE INC.',
                'SM MEGAMALL - BEBANG ENTERPRISE INC.',
                'SM SOUTHMALL - BEBANG ENTERPRISE INC.'):
        return 'Store (BEI-TIN stub; files under BEI TIN)'
    # Store (has " - " in name)
    if ' - ' in name:
        return 'Store (per-store P&L)'
    # Other group entity (is_group=1) that's a legal entity parent
    if is_group:
        return 'Legal Entity Parent (group)'
    # Otherwise leaf legal entity (standalone)
    return 'Legal Entity (standalone)'


def main():
    sys.stdout.reconfigure(encoding="utf-8")

    state = json.load(open("output/s258/baseline_state.json"))
    fact_check = json.load(open("output/s258/fact_check.json"))

    # Load cleanroom register if available
    cleanroom_by_name = {}
    cr_path = Path("data/_CLEANROOM/2026-04-09_s175_coa_restructure/bei_company_register_team_completed_2026-04-10.csv")
    if cr_path.exists():
        with open(cr_path, encoding='utf-8') as f:
            for r in csv.DictReader(f):
                fn = r.get('Frappe Company Name', '').strip()
                if fn:
                    cleanroom_by_name[fn] = r

    wb = Workbook()
    wb.remove(wb.active)

    # === Sheet 1: Overview ===
    ws = wb.create_sheet("Overview")
    ws.sheet_view.showGridLines = False
    ws.merge_cells('A1:F1')
    ws['A1'] = "BEI Company Register — All 58 Frappe Companies"
    ws['A1'].font = Font(bold=True, color=BEI_TEXT_WHITE, size=16, name='Calibri')
    ws['A1'].fill = HEADER_FILL
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:F2')
    ws['A2'] = "Bridge Consulting QBO Setup Reference — Generated 2026-06-05 from live hq.bebang.ph"
    ws['A2'].font = Font(italic=True, color=BEI_TEXT_DARK, size=10, name='Calibri')
    ws['A2'].fill = PatternFill('solid', fgColor=BEI_GOLD_LIGHT)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # Role distribution
    role_counts = {}
    for r in state['rows']:
        role = classify_role(r, r['name'], r.get('is_group'), r.get('parent_company'))
        role_counts[role] = role_counts.get(role, 0) + 1

    ws['A4'] = "Role Distribution"
    ws['A4'].font = HEADER_FONT
    ws['A4'].fill = HEADER_FILL
    ws.merge_cells('A4:F4')

    ws['A5'] = "Role"
    ws['B5'] = "Count"
    ws['C5'] = "Notes"
    for col_letter in ('A', 'B', 'C'):
        c = ws[f'{col_letter}5']
        c.font = Font(bold=True, color=BEI_TEXT_DARK, size=10)
        c.fill = PatternFill('solid', fgColor=BEI_GOLD)
        c.border = BORDER
        c.alignment = Alignment(horizontal='center', vertical='center')

    role_notes = {
        'Holdco (top of tree)': 'IRRESISTIBLE INFUSIONS INC. — is_group=1, parent_company=NULL. 0 GL postings.',
        'Head Office (BEI HQ)': 'BEBANG ENTERPRISE INC. — populates JV fees + Brand Growth (4000234/35 + 4000005). 4 BEI-TIN stub stores live under BEI TIN.',
        'Commissary': 'BEBANG KITCHEN INC. — populates ONLY 4000200 BKI SALES sub-tree (deliveries + logistics) per COA-175-011.',
        'Franchisor (Fork 1)': 'BEBANG FRANCHISE CORP. — Fork 1 collection-agent model per COA-175-013; 1104200 DUE FROM BEI + 2102205 OUTPUT VAT.',
        'Legal Entity Parent (BFT)': 'BEBANG FT INC. — abbr=BFT (was BFI2). SEC TIN 663-440-106-00000 (RDO 044). Parent of AFT store.',
        'Store (BEI-TIN stub; files under BEI TIN)': 'ROA / SMM / SMMM / SMS — 4 BEI-TIN stubs. Separate Frappe Companies for management P&L; file under BEI TIN 647-243-690-00000 for BIR.',
        'Store (per-store P&L)': '49 stores total (incl. 4 BEI-TIN stubs above). Each has own Sales tree per-store P&L.',
        'Legal Entity Parent (group)': 'Multi-store legal entities like BEBANG MEGA INC., TUNGSTEN CAPITAL, TAJ FOOD CORP., etc.',
        'Legal Entity (standalone)': 'Single-store OPC or one-store legal entity.',
    }
    for i, (role, cnt) in enumerate(sorted(role_counts.items(), key=lambda x: -x[1])):
        r = 6 + i
        ws.cell(row=r, column=1, value=role)
        ws.cell(row=r, column=2, value=cnt)
        ws.cell(row=r, column=3, value=role_notes.get(role, ''))
        for col in (1, 2, 3):
            c = ws.cell(row=r, column=col)
            c.font = BODY_FONT
            c.fill = EVEN_FILL if i % 2 == 0 else ODD_FILL
            c.border = BORDER
            c.alignment = Alignment(horizontal='center' if col == 2 else 'left',
                                    vertical='top', wrap_text=True)

    # Total
    tr = 6 + len(role_counts)
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True, color=BEI_TEXT_WHITE, size=10)
    ws.cell(row=tr, column=1).fill = PatternFill('solid', fgColor=BEI_GREEN)
    ws.cell(row=tr, column=2, value=sum(role_counts.values())).font = Font(bold=True, color=BEI_TEXT_WHITE, size=10)
    ws.cell(row=tr, column=2).fill = PatternFill('solid', fgColor=BEI_GREEN)
    ws.cell(row=tr, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=tr, column=1).border = BORDER
    ws.cell(row=tr, column=2).border = BORDER

    for col, w in [('A', 42), ('B', 8), ('C', 95)]:
        ws.column_dimensions[col].width = w

    # === Sheet 2: All 58 Companies ===
    ws2 = wb.create_sheet("All 58 Companies")
    ws2.sheet_view.showGridLines = False
    headers = ["#", "Frappe Company Name", "ABBR", "Role (S258)", "Parent Company",
               "Is Group", "BIR TIN", "RDO", "VAT Status", "Total Accounts", "GL Entries", "Frappe Status (pre-S258)"]
    for j, h in enumerate(headers):
        c = ws2.cell(row=1, column=j + 1, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
    ws2.row_dimensions[1].height = 40
    ws2.freeze_panes = 'D2'

    for i, r in enumerate(state['rows'], 1):
        row = i + 1
        role = classify_role(r, r['name'], r.get('is_group'), r.get('parent_company'))
        cr = cleanroom_by_name.get(r['name'], {})
        ws2.cell(row=row, column=1, value=i)
        ws2.cell(row=row, column=2, value=r['name'])
        ws2.cell(row=row, column=3, value=r.get('abbr'))
        ws2.cell(row=row, column=4, value=role)
        ws2.cell(row=row, column=5, value=r.get('parent_company') or '')
        ws2.cell(row=row, column=6, value='Yes' if r.get('is_group') else 'No')
        ws2.cell(row=row, column=7, value=r.get('tax_id') or cr.get('TIN\n(pre-filled if known, team to verify)') or '')
        ws2.cell(row=row, column=8, value=cr.get('RDO Code') or '')
        ws2.cell(row=row, column=9, value=cr.get('VAT Status') or '')
        ws2.cell(row=row, column=10, value=r.get('total_accounts'))
        ws2.cell(row=row, column=11, value=r.get('gl_entry_count'))
        ws2.cell(row=row, column=12, value=r.get('status'))
        for j in range(1, 13):
            c = ws2.cell(row=row, column=j)
            c.font = BODY_FONT
            c.fill = EVEN_FILL if i % 2 == 0 else ODD_FILL
            c.border = BORDER
            c.alignment = Alignment(horizontal='left' if j in (2, 4, 5, 9)
                                                  else 'center', vertical='top',
                                    wrap_text=True)

    last_row = len(state['rows']) + 1

    # Conditional formatting on Role column
    for role_match, fill_color in [
        ('Head Office (BEI HQ)', BEI_PURPLE_LIGHT),
        ('Commissary', BEI_GOLD_LIGHT),
        ('Franchisor (Fork 1)', BEI_GOLD_LIGHT),
        ('Holdco (top of tree)', BEI_PURPLE_LIGHT),
    ]:
        ws2.conditional_formatting.add(
            f'D2:D{last_row}',
            CellIsRule(operator='equal', formula=[f'"{role_match}"'],
                       fill=PatternFill('solid', fgColor=fill_color))
        )

    ws2.auto_filter.ref = f'A1:L{last_row}'

    for col, w in [('A', 5), ('B', 50), ('C', 8), ('D', 28), ('E', 30),
                   ('F', 8), ('G', 18), ('H', 8), ('I', 12), ('J', 12),
                   ('K', 12), ('L', 12)]:
        ws2.column_dimensions[col].width = w

    # === Sheet 3: Per-Company P&L Population Rules ===
    ws3 = wb.create_sheet("P&L Population Rules")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells('A1:D1')
    ws3['A1'] = "Per-Role P&L Population Rules — which sub-tree gets postings"
    ws3['A1'].font = Font(bold=True, color=BEI_TEXT_WHITE, size=13)
    ws3['A1'].fill = HEADER_FILL
    ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 28

    ws3.merge_cells('A2:D2')
    ws3['A2'] = "Butch's 27-account Sales tree is structurally identical on all 58 Companies. Each Company populates the sub-tree appropriate to its role."
    ws3['A2'].font = Font(italic=True, color=BEI_TEXT_DARK, size=10)
    ws3['A2'].fill = PatternFill('solid', fgColor=BEI_GOLD_LIGHT)
    ws3['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws3.row_dimensions[2].height = 32

    rules = [
        ("Role", "Sales sub-trees POPULATED with postings", "Sales sub-trees STRUCTURAL ONLY (empty)", "Decision lock"),
        ("Head Office (BEI HQ)",
         "4000234 MARKETING FEES + 4000235 E-COMMERCE FEES (JV fees per JV Gabaldon §8.1/§9.1); 4000005 BRAND GROWTH FEE INCOME (PHP 2M one-time, JV §2)",
         "4000100 STORE SALES tree, 4000200 BKI SALES tree, 4000230 FEES tree (except 234/235)",
         "COA-175-007 (JV permanent to BEI); COA-175-016 (Brand Growth point-in-time)"),
        ("Commissary (BKI)",
         "4000200 BKI SALES sub-tree ONLY: 4000210 DELIVERIES + 4000220 LOGISTICS (4000221 + 4000222)",
         "4000100 STORE SALES, 4000120 ONLINE SALES, 4000230 FEES — all empty on BKI",
         "COA-175-011 (BKI = Commissary, no in-store/online/franchise income)"),
        ("Franchisor (BFC)",
         "4000230 FEES sub-tree: 4000231 ROYALTY + 4000232 MGMT + 4000233 FRANCHISE + 4000234 MARKETING + 4000235 E-COMMERCE",
         "4000100, 4000200 — empty on BFC (Franchisor never sells retail or runs commissary)",
         "COA-175-008 (franchise fees → BFC); COA-175-013 (Fork 1 collection-agent)"),
        ("Store (per-store P&L)",
         "4000110 IN-STORE SALES + 4000121 BEI WEBSITE + 4000122 FOOD PANDA + 4000123 GRAB (per-store retail + online aggregator)",
         "4000200 BKI SALES, 4000230 FEES — empty on stores (Commissary + Franchisor revenue, not store-level)",
         "Sam directive 2026-06-04 (per-store P&L); COA-175-003 (uniform tree, role-populated)"),
        ("Store (BEI-TIN stub)",
         "Same as Store above — 4000110, 4000121, 4000122, 4000123 — separate Frappe Company per-store P&L",
         "4000200, 4000230 empty. **Note:** BIR filing under BEI TIN 647-243-690-00000, NOT standalone.",
         "COA-175-025 (4 BEI-TIN stubs: ROA/SMM/SMMM/SMS — INTERNAL MGMT REPORTING)"),
        ("Holdco (III)",
         "(None — III is is_group=1 holdco, 0 GL postings expected)",
         "All Sales sub-trees structural only on III. III's role is hierarchical parent only.",
         "COA-175-003 (uniform tree); live audit confirms III = 0 GL"),
        ("Legal Entity Parent (multi-store)",
         "(None at parent level — postings happen at per-store child Companies)",
         "All sub-trees structural; consolidation via parent_company chain in QBO.",
         "Per-store P&L architecture (Sam directive)"),
    ]
    for i, row in enumerate(rules):
        rr = 4 + i
        for j, val in enumerate(row):
            c = ws3.cell(row=rr, column=j + 1, value=val)
            if i == 0:
                c.font = HEADER_FONT
                c.fill = HEADER_FILL
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                c.font = BODY_FONT
                c.fill = EVEN_FILL if i % 2 == 0 else ODD_FILL
                c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            c.border = BORDER
        ws3.row_dimensions[rr].height = 90

    for col, w in [('A', 24), ('B', 50), ('C', 50), ('D', 38)]:
        ws3.column_dimensions[col].width = w

    # === Sheet 4: 27-account Sales tree (Butch verbatim) ===
    ws4 = wb.create_sheet("27-Acct Sales Tree (Butch)")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells('A1:E1')
    ws4['A1'] = "Butch's 27-account Canonical Sales Tree — COA-175-001 (locked 2026-04-08)"
    ws4['A1'].font = Font(bold=True, color=BEI_TEXT_WHITE, size=13)
    ws4['A1'].fill = HEADER_FILL
    ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws4.row_dimensions[1].height = 28

    sales_tree = [
        # (number, name, is_group, parent, sub-tree, notes)
        ("Account #", "Account Name", "Group?", "Parent", "Sub-tree", "Notes / Population"),
        ("4000000", "SALES", "Group", "(Income root)", "ROOT", "Top of Sales hierarchy"),
        ("4000100", "STORE SALES", "Group", "4000000 SALES", "STORE", "Per-store retail rollup"),
        ("4000110", "IN-STORE SALES", "Leaf", "4000100", "STORE", "Walk-in POS revenue per store"),
        ("4000120", "ONLINE SALES", "Group", "4000100", "STORE", "Online aggregator rollup per store"),
        ("4000121", "BEI WEBSITE", "Leaf", "4000120", "STORE", "Orders from bebang.ph website"),
        ("4000122", "FOOD PANDA", "Leaf", "4000120", "STORE", "FoodPanda gross sales per store"),
        ("4000123", "GRAB", "Leaf", "4000120", "STORE", "GrabFood gross sales per store"),
        ("4000200", "BKI SALES", "Group", "4000000 SALES", "COMMISSARY", "BKI only — paired SI to stores"),
        ("4000210", "DELIVERIES", "Leaf", "4000200", "COMMISSARY", "BKI→store goods deliveries"),
        ("4000220", "LOGISTICS", "Group", "4000200", "COMMISSARY", "Delivery + logistics income"),
        ("4000221", "DELIVERY INCOME", "Leaf", "4000220", "COMMISSARY", "Per-bill delivery fee"),
        ("4000222", "LOGISTICS INCOME", "Leaf", "4000220", "COMMISSARY", "3PL re-bill, route fees"),
        ("4000230", "FEES", "Group", "4000000 SALES", "FRANCHISE/JV", "Franchisor + JV fees rollup"),
        ("4000231", "ROYALTY FEES", "Leaf", "4000230", "FRANCHISE", "BFC 7% royalty per signed FA §XIII.A"),
        ("4000232", "MANAGEMENT FEES", "Leaf", "4000230", "FRANCHISE", "BFC 2.5% per Mgmt §5.1"),
        ("4000233", "FRANCHISE FEES", "Leaf", "4000230", "FRANCHISE", "BFC PHP 2M one-time per FA §I.A"),
        ("4000234", "MARKETING FEES", "Leaf", "4000230", "JV+FRANCHISE", "5% — BEI (JV §8.1) + BFC (FA §XI.A)"),
        ("4000235", "E-COMMERCE FEES", "Leaf", "4000230", "JV+FRANCHISE", "5% — BEI (JV §9.1) + BFC (FA §XI.I)"),
        ("4000005", "BRAND GROWTH FEE INCOME", "Leaf", "(BEI only)", "JV", "BEI extension per COA-175-016. PHP 2M one-time JV §2."),
        ("4000900", "DISCOUNTS AND PROMO", "Group", "4000000 SALES", "CONTRA", "Contra-revenue group per COA-175-002"),
        ("4000901", "SALES DISCOUNT — FREE HALOHALO", "Leaf", "4000900", "CONTRA", "Discount type 1"),
        ("4000902", "SALES DISCOUNTS — PWDS", "Leaf", "4000900", "CONTRA", "BIR-mandated PWD discount"),
        ("4000903", "SALES DISCOUNTS — SENIOR CITIZENS", "Leaf", "4000900", "CONTRA", "BIR-mandated senior discount"),
        ("4000904", "SALES DISCOUNTS — PROMO", "Leaf", "4000900", "CONTRA", "Marketing promo discounts"),
        ("4000905", "SALES DISCOUNTS — EMPLOYEE", "Leaf", "4000900", "CONTRA", "Staff/employee discounts"),
        ("4000906", "SALES DISCOUNTS — REFUND", "Leaf", "4000900", "CONTRA", "Refund-induced contra"),
        ("4000907", "SALES DISCOUNTS — GIFT CARDS", "Leaf", "4000900", "CONTRA", "Gift card redemption contra"),
        ("4000908", "SALES DISCOUNTS — OTHERS", "Leaf", "4000900", "CONTRA", "Catch-all"),
    ]
    for i, row in enumerate(sales_tree):
        rr = 3 + i
        for j, val in enumerate(row):
            c = ws4.cell(row=rr, column=j + 1, value=val)
            if i == 0:
                c.font = HEADER_FONT
                c.fill = HEADER_FILL
                c.alignment = Alignment(horizontal='center', vertical='center')
            else:
                c.font = BODY_FONT
                c.fill = EVEN_FILL if i % 2 == 0 else ODD_FILL
                c.alignment = Alignment(horizontal='left' if j in (1, 5) else 'center',
                                        vertical='center', wrap_text=True)
            c.border = BORDER

    for col, w in [('A', 12), ('B', 35), ('C', 8), ('D', 22), ('E', 18), ('F', 50)]:
        ws4.column_dimensions[col].width = w

    out = Path("output/s258/bridge_handoff/COMPANY_REGISTER.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"[OK] Wrote {out}")
    print(f"     Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
