"""Export all BIR Tax Return Receipt Confirmations from james.tamaca@bebang.ph
(migrated from deleted accounting@ mailboxes) as .eml files + index XLSX + ZIPs.

Output: tmp/bir_confirmations/
  - eml/<year>/<date>_<msgid>.eml
  - BIR_CONFIRMATIONS_INDEX.xlsx
  - BIR_confirmations_<year>.zip
"""
from __future__ import annotations
import base64
import email
import email.policy
import re
import sys
import zipfile
from collections import Counter
from datetime import datetime
from pathlib import Path

from google.oauth2 import service_account
from googleapiclient.discovery import build

OUT = Path("tmp/bir_confirmations")
EML = OUT / "eml"

TIN_RE = re.compile(r"\b(\d{3}-\d{3}-\d{3}-?\d{0,5})\b")
FORM_RE = re.compile(r"\b(0605|1601-?E?Q|1601-?C|1602|1603|1604-?CF?|1604-?E|1606|1700|1701|1701-?Q|1702|1702-?Q|1702-?RT|2000|2550-?M|2550-?Q|2551-?Q|2552|2553)\b", re.I)
PERIOD_RE = re.compile(r"(?:period|return period|for the period)[:\s]*([0-9/\-]{6,10})", re.I)


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    creds = service_account.Credentials.from_service_account_file(
        "credentials/task-manager-service.json",
        scopes=["https://www.googleapis.com/auth/gmail.readonly"],
    ).with_subject("james.tamaca@bebang.ph")
    gmail = build("gmail", "v1", credentials=creds, cache_discovery=False)

    q = 'subject:"Tax Return Receipt Confirmation"'
    ids = []
    token = None
    while True:
        res = gmail.users().messages().list(userId="me", q=q, maxResults=500, pageToken=token).execute()
        ids.extend(m["id"] for m in res.get("messages", []))
        token = res.get("nextPageToken")
        if not token:
            break
    print(f"Found {len(ids)} confirmation messages")

    EML.mkdir(parents=True, exist_ok=True)
    rows = []
    for i, mid in enumerate(ids, 1):
        if i % 100 == 0:
            print(f"  {i}/{len(ids)}")
        raw = gmail.users().messages().get(userId="me", id=mid, format="raw").execute()
        data = base64.urlsafe_b64decode(raw["raw"] + "=" * ((4 - len(raw["raw"]) % 4) % 4))
        msg = email.message_from_bytes(data, policy=email.policy.default)
        ts = int(raw["internalDate"]) // 1000
        dt = datetime.fromtimestamp(ts)
        year = dt.strftime("%Y")
        fname = f"{dt.strftime('%Y-%m-%d_%H%M')}_{mid}.eml"
        ydir = EML / year
        ydir.mkdir(exist_ok=True)
        (ydir / fname).write_bytes(data)

        # Parse body for TIN / form / period
        body = ""
        try:
            b = msg.get_body(preferencelist=("plain", "html"))
            if b:
                body = b.get_content()
                body = re.sub(r"<[^>]+>", " ", body)  # strip html tags if any
        except Exception:
            pass
        tins = sorted(set(TIN_RE.findall(body)))[:3]
        forms = sorted(set(f.upper().replace("M", "M").replace(" ", "") for f in FORM_RE.findall(body)))[:3]
        period_m = PERIOD_RE.search(body)
        rows.append({
            "date": dt.strftime("%Y-%m-%d %H:%M"),
            "year": year,
            "to": str(msg.get("To", ""))[:120],
            "subject": str(msg.get("Subject", ""))[:120],
            "tin": "; ".join(tins),
            "form": "; ".join(forms),
            "period": period_m.group(1) if period_m else "",
            "eml_file": f"{year}/{fname}",
            "gmail_id": mid,
        })

    # Index XLSX
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    BEI_GREEN, BEI_TINT, DARK, WHITE, BRD = "04400A", "E6ECE7", "1A1A1A", "FFFFFF", "D4D0C8"
    wb = Workbook()
    ws = wb.active
    ws.title = "BIR Confirmations"
    headers = ["#", "Date", "To (original recipient)", "Subject", "TIN(s)", "Form(s)", "Period", "EML file", "Gmail ID"]
    thin = Side(style="thin", color=BRD)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
        c.fill = PatternFill("solid", fgColor=BEI_GREEN)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    rows.sort(key=lambda r: r["date"])
    for i, r in enumerate(rows, 1):
        vals = [i, r["date"], r["to"], r["subject"], r["tin"], r["form"], r["period"], r["eml_file"], r["gmail_id"]]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=i + 1, column=j, value=v)
            c.font = Font(color=DARK, size=10, name="Calibri")
            c.fill = PatternFill("solid", fgColor=BEI_TINT if i % 2 == 0 else WHITE)
            c.border = border
    widths = [6, 17, 45, 40, 20, 14, 12, 38, 20]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(rows) + 1}"
    wb.save(OUT / "BIR_CONFIRMATIONS_INDEX.xlsx")
    print(f"Index: {len(rows)} rows")

    # ZIP per year
    by_year = Counter(r["year"] for r in rows)
    for year in sorted(by_year):
        zp = OUT / f"BIR_confirmations_{year}.zip"
        with zipfile.ZipFile(zp, "w", zipfile.ZIP_DEFLATED) as zf:
            for f in sorted((EML / year).iterdir()):
                zf.write(f, f"{year}/{f.name}")
        print(f"  {zp.name}: {by_year[year]} emails, {zp.stat().st_size // 1024} KB")

    # Recipient summary for the README
    to_counts = Counter()
    for r in rows:
        to_lower = r["to"].lower()
        for addr in ("accounting@bebang.ph", "accounting.bei@bebang.ph", "accounting.hq@bebang.ph",
                     "accounting.shaw@bebang.ph", "storesaccounting@bebang.ph", "cost.accounting@bebang.ph",
                     "finance@bebang.ph"):
            if addr in to_lower:
                to_counts[addr] += 1
    print("\nRecipient totals:")
    for a, n in to_counts.most_common():
        print(f"  {n:4d}  {a}")
    print(f"\nDate range: {rows[0]['date']} .. {rows[-1]['date']}")


if __name__ == "__main__":
    main()
