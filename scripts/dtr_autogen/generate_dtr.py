"""BEI automated DTR generator (Phase 1: biometric-derivable values only).

Clones each store's existing DTR workbook, fills ONLY the "Worked (Day)" column
of employee rows (keyed by 9xxxxxx Bio ID), blanks the HR-input columns
(Tardy/UT/Absence/EL/SL/VL) to None so emptiness signals "HR to fill", and
appends one "Punch Detail (AUTO)" sheet with per-shift-day punch spans.

Fidelity: the base workbook is copied byte-for-byte first, then opened with
keep_links=False and ONLY value cells are touched. Rows/columns/styles/sheets
are never restructured (one sheet is appended).

Usage:
    python scripts/dtr_autogen/generate_dtr.py --from 2026-05-16 --to 2026-05-31
    [--punches-csv tmp/dtr_audit/punches_may16_31.csv]
    [--templates-dir tmp/dtr_audit/files/may_2026]
    [--out-dir tmp/dtr_audit/agents/dtr_autogen/output]
"""
import argparse
import csv
import glob
import os
import shutil
import sys
import warnings

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from dtr_common import (  # noqa: E402
    FILE_NUM_TO_ADMS_STORE, PH_REGULAR_HOLIDAYS_2026, ROVING_FILE_NUMS,
    coerce_bio, file_num, group_shift_days, in_window, load_punches, norm,
    parse_date,
)

warnings.simplefilter("ignore")

EMP_MASTER = "data/_FINAL/EMPLOYEE_MASTER.csv"


# ---------- template structure detection ----------

def pick_dtrdata_sheet(wb):
    """Return the sheet whose normalized name is exactly 'dtrdata'.

    Prefer the exact raw name 'DTRData' on a tie (e.g. file 48 has both
    'DTRData' and 'DTRData.'); otherwise the first normalized match.
    """
    matches = [s for s in wb.sheetnames if norm(s) == "dtrdata"]
    if not matches:
        return None
    for m in matches:
        if m == "DTRData":
            return m
    return matches[0]


def find_header_row(ws):
    """Row index whose cells include a 'Worked' header. Search rows 1..14."""
    for r in range(1, 15):
        for c in range(1, min(ws.max_column, 20) + 1):
            if "worked" in norm(ws.cell(row=r, column=c).value):
                return r
    return None


def find_columns(ws, hdr_row):
    """Locate (bio_col, worked_col) by header text on hdr_row.

    bio_col may be None when the template has no 'New Employee ID' column
    (e.g. file 48). worked_col is required.
    """
    bio_col = worked_col = None
    for c in range(1, ws.max_column + 1):
        h = norm(ws.cell(row=hdr_row, column=c).value)
        if h.startswith("newemployeeid"):
            bio_col = c
        if h.startswith("worked"):  # 'Worked (Day)' / 'Worked Days'
            worked_col = c
    # Fallback: infer bio column as the one with the most 9xxxxxx data cells.
    if bio_col is None:
        best, best_count = None, 0
        for c in range(1, ws.max_column + 1):
            cnt = sum(
                1 for r in range(hdr_row + 1, ws.max_row + 1)
                if coerce_bio(ws.cell(row=r, column=c).value) is not None
            )
            if cnt > best_count:
                best, best_count = c, cnt
        if best_count >= 2:  # need a real column of bio ids, not a stray
            bio_col = best
    return bio_col, worked_col


def collect_template_employees(ws, hdr_row, bio_col):
    """Return {bio_id: row_index} for every employee row (any section).

    Scans all rows below the header; a row is an employee row iff its bio_col
    cell coerces to a 9xxxxxx int. Totals rows (=SUM / #REF!), labels and blanks
    are skipped automatically.
    """
    emps = {}
    for r in range(hdr_row + 1, ws.max_row + 1):
        bio = coerce_bio(ws.cell(row=r, column=bio_col).value)
        if bio is not None:
            emps[bio] = r  # last wins on dup; dups are not expected
    return emps


# ---------- employee master store mapping (extras section only) ----------

def build_store_to_filenum():
    """Map a normalized ADMS store_name -> file number for reverse lookup."""
    return {norm(name): num for num, name in FILE_NUM_TO_ADMS_STORE.items()}


def load_emp_store_map(store_to_filenum):
    """Return {bio_id: file_num} from EMPLOYEE_MASTER store_location.

    Best-effort fuzzy match of the messy store_location text onto a file number.
    Used only to populate the 'PUNCHED BUT NOT IN TEMPLATE' extras section.
    """
    # alias hints for store_location variants that don't normalize cleanly
    alias = {
        "ayalamarketmarket": "03", "marketmarket": "03",
        "luckychinatown": "13", "lct": "13",
        "robinsonsantipolo": "14", "robinsonantipolo": "14",
        "megaworldpaseocenter": "20", "paseocenter": "20", "paseo": "20",
        "pitxterminal": "21", "pitx": "21",
        "ayalauptc": "27", "ayalauptowncenter": "27", "uptc": "27",
        "smsjdm": "28",
        "robinsonsgeneraltrias": "32", "robinsongeneraltrias": "32",
        "robinsongentri": "32", "robinsongentrias": "32",
        "dverdecalamba": "47",
        "smstarosa": "48", "smstarosa.": "48",
        "robinsondasmarinas": "49", "robinsondasma": "49",
        "naiaterminal3": "50", "naiat3": "50",
        "ortigasestancia": "52", "estancia": "52",
        "smmallofasia": "25", "smmoa": "25",
        "smgrandcentral": "16", "grandcentral": "16",
        "uptownbgc": "44", "mytown": None, "vistamalltaguig": None,
    }
    out = {}
    with open(EMP_MASTER, encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            bio = coerce_bio(row.get("new_attendance_device_id"))
            if bio is None:
                continue
            loc = norm(row.get("store_location"))
            if not loc:
                continue
            fn = store_to_filenum.get(loc)
            if fn is None:
                fn = alias.get(loc)
            if fn:
                out[bio] = fn
    return out


# ---------- punch detail sheet ----------

DETAIL_HEADERS = [
    "Bio ID", "Name", "Date (PHT, shift-day)", "First punch", "Last punch",
    "# punches", "Span (hr)", "Raw OT (hr, span-9 floor 0)", "Holiday tag",
    "Devices/stores punched",
]


def emp_name(ws, row, bio_col):
    """Best-effort 'Last, First' from the template row (columns right of bio)."""
    if not row:
        return ""
    parts = []
    for c in (bio_col + 1, bio_col + 2):
        v = ws.cell(row=row, column=c).value
        if isinstance(v, str) and v.strip() and v.strip() != "-----":
            parts.append(v.strip())
    return ", ".join(parts)


def append_punch_detail(wb, rows_for_detail, no_bio_note):
    ws = wb.create_sheet(title="Punch Detail (AUTO)")
    ws.append(DETAIL_HEADERS)
    if no_bio_note:
        ws.append([no_bio_note])
    for r in rows_for_detail:
        ws.append(r)


# ---------- per-file generation ----------

def generate_one(template_path, out_dir, by_pin, flat_punches, emp_store_map,
                 from_date, to_date, period_label):
    fname = os.path.basename(template_path)
    fnum = file_num(fname)
    stem = os.path.splitext(fname)[0]
    out_path = os.path.join(out_dir, f"{stem}__AUTO-DRAFT {period_label}.xlsx")

    # Fidelity: clone the file first, then edit the clone in place.
    shutil.copyfile(template_path, out_path)
    wb = openpyxl.load_workbook(out_path, keep_links=False)

    sheet = pick_dtrdata_sheet(wb)
    result = {"file": fname, "out": os.path.basename(out_path)}
    if sheet is None:
        wb.save(out_path)
        wb.close()
        result["status"] = "NO_DTRDATA_SHEET"
        return result
    ws = wb[sheet]

    hdr_row = find_header_row(ws)
    if hdr_row is None:
        wb.save(out_path)
        wb.close()
        result["status"] = "NO_HEADER_ROW"
        return result
    bio_col, worked_col = find_columns(ws, hdr_row)

    no_bio_note = None
    template_emps = {}
    if bio_col is None:
        no_bio_note = ("NO BIO ID COLUMN in this template -- cannot auto-fill "
                       "Worked. All values left for HR.")
    else:
        template_emps = collect_template_employees(ws, hdr_row, bio_col)

    # HR-input columns to blank (relative to header text).
    hr_input_cols = []
    for c in range(1, ws.max_column + 1):
        h = norm(ws.cell(row=hdr_row, column=c).value)
        if h in ("tardymin", "uthr", "absenceday", "elday", "slday", "vlday"):
            hr_input_cols.append(c)

    # Compute Worked + shift-days per template employee.
    detail_rows = []
    filled = 0
    no_punch = 0
    for bio, row in sorted(template_emps.items()):
        name = emp_name(ws, row, bio_col)
        punches = [(dt, st) for (dt, st) in by_pin.get(str(bio), [])
                   if in_window(dt, from_date, to_date)]
        # Blank HR-input cells on this row (signals "HR to fill").
        for c in hr_input_cols:
            ws.cell(row=row, column=c).value = None
        if not punches:
            ws.cell(row=row, column=worked_col).value = None  # empty, not 0
            no_punch += 1
            detail_rows.append([bio, name, "NO PUNCHES", "", "", 0, "", "", "", ""])
            continue
        days = group_shift_days(punches)
        # only count shift-days whose label date is inside the window
        days = [d for d in days if from_date <= parse_date(d["date"]) <= to_date]
        ws.cell(row=row, column=worked_col).value = len(days)
        filled += 1
        for d in days:
            span_hr = (d["last"] - d["first"]).total_seconds() / 3600.0
            raw_ot = max(0.0, span_hr - 9.0)
            stores = sorted({st for (_dt, st) in d["punches"] if st})
            tag = PH_REGULAR_HOLIDAYS_2026.get(d["date"], "")
            detail_rows.append([
                bio, name, d["date"],
                d["first"].strftime("%H:%M"), d["last"].strftime("%H:%M"),
                len(d["punches"]), round(span_hr, 2), round(raw_ot, 2),
                tag, ", ".join(stores),
            ])

    # Extras: PINs assigned to THIS store (per EMPLOYEE_MASTER) that punched in
    # the window but are NOT in the template. Skip for roving/office files.
    extras = []
    if fnum and fnum not in ROVING_FILE_NUMS and bio_col is not None:
        punched_pins = {pin for (pin, dt, _st) in flat_punches
                        if in_window(dt, from_date, to_date)}
        for bio_str in punched_pins:
            bio = coerce_bio(bio_str)
            if bio is None:
                continue
            if emp_store_map.get(bio) == fnum and bio not in template_emps:
                cnt = sum(1 for (dt, _st) in by_pin.get(bio_str, [])
                          if in_window(dt, from_date, to_date))
                extras.append((bio, cnt))
    if extras:
        detail_rows.append([])
        detail_rows.append(["PUNCHED BUT NOT IN TEMPLATE", "", "", "", "",
                            "", "", "", "", ""])
        for bio, cnt in sorted(extras):
            detail_rows.append([bio, "(per EMPLOYEE_MASTER store)", "", "", "",
                               cnt, "", "", "", "punches in window"])

    append_punch_detail(wb, detail_rows, no_bio_note)
    wb.save(out_path)
    wb.close()

    result.update({
        "status": "OK",
        "sheet": sheet,
        "hdr_row": hdr_row,
        "bio_col": bio_col,
        "worked_col": worked_col,
        "template_emps": len(template_emps),
        "worked_filled": filled,
        "no_punch": no_punch,
        "extras": len(extras),
    })
    return result


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--from", dest="from_date", required=True, help="PHT YYYY-MM-DD")
    ap.add_argument("--to", dest="to_date", required=True, help="PHT YYYY-MM-DD")
    ap.add_argument("--punches-csv", default="tmp/dtr_audit/punches_may16_31.csv")
    ap.add_argument("--templates-dir", default="tmp/dtr_audit/files/may_2026")
    ap.add_argument("--out-dir",
                    default="tmp/dtr_audit/agents/dtr_autogen/output")
    args = ap.parse_args()

    from_date = parse_date(args.from_date)
    to_date = parse_date(args.to_date)
    period_label = f"{args.from_date}_to_{args.to_date}"
    os.makedirs(args.out_dir, exist_ok=True)

    by_pin, flat = load_punches(args.punches_csv)
    emp_store_map = load_emp_store_map(build_store_to_filenum())

    templates = sorted(glob.glob(os.path.join(args.templates_dir, "*.xlsx")))
    results = []
    for t in templates:
        try:
            res = generate_one(t, args.out_dir, by_pin, flat, emp_store_map,
                               from_date, to_date, period_label)
        except Exception as e:  # keep going; record the failure
            res = {"file": os.path.basename(t), "status": "ERROR", "err": repr(e)}
        results.append(res)
        print(f"{res.get('status','?'):16} {res['file'][:42]:42} "
              f"emps={res.get('template_emps','-')} "
              f"filled={res.get('worked_filled','-')} "
              f"nopunch={res.get('no_punch','-')} "
              f"extras={res.get('extras','-')}")

    import json
    summ = os.path.join(args.out_dir, "_generation_summary.json")
    with open(summ, "w", encoding="utf-8") as fh:
        json.dump(results, fh, indent=2)
    ok = sum(1 for r in results if r.get("status") == "OK")
    print(f"\nGenerated {ok}/{len(results)} workbooks -> {args.out_dir}")
    print(f"Summary: {summ}")


if __name__ == "__main__":
    main()
