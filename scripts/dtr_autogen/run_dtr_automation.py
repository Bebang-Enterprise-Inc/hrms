"""BEI automated DTR production runner (Cloud Run / cron entrypoint).

One self-contained job: derive the cut-off from the run date -> pull punches from
Supabase -> fetch the canonical bio-keyed store templates from Drive -> generate
DTRs (calls generate_dtr.py) -> finalize (clean names + period title) -> zip ->
email to HR via Gmail (service-account domain-wide delegation).

Schedule (Cloud Scheduler, Asia/Manila):
    01:00 on the 16th -> cut-off 1..15 of the current month
    01:00 on the 1st  -> cut-off 16..EOM of the PREVIOUS month

Credentials (identical code path local vs cloud):
    - Service-account JSON: env DTR_SA_JSON, else credentials/task-manager-service.json
      (in Cloud Run, mount the Secret Manager secret at that path).
    - Supabase key: env SUPABASE_SERVICE_ROLE_KEY, else Doppler (local only).

Safety flags:
    --from / --to        explicit window (YYYY-MM-DD); overrides date derivation
    --to-email ADDR      recipient (default hr@bebang.ph)
    --send-to-self       send to sam@bebang.ph instead of --to-email (test)
    --no-send            generate + zip only, do not email
"""
import argparse
import base64
import calendar
import csv
import datetime as dt
import glob
import io
import json
import os
import re
import subprocess
import sys
import time
import urllib.parse
import urllib.request
import warnings
from email.message import EmailMessage

import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from generate_dtr import pick_dtrdata_sheet, find_header_row, find_columns  # noqa: E402

warnings.simplefilter("ignore")
PHT = dt.timezone(dt.timedelta(hours=8))
SUPABASE_URL = "https://csnniykjrychgajfrgua.supabase.co"
TIMEKEEPING_ROOT = "17tL5KPQ1DDHVOdOQbcs5kp3eMpPaGDZ2"  # "Timekeeping DTR and Payroll Files"
SA_JSON = os.environ.get("DTR_SA_JSON", "credentials/task-manager-service.json")
SENDER = "sam@bebang.ph"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
AUTO_PARENT_NAME = "Auto-Generated DTRs"  # dedicated parent so auto-drafts never mix with HR's canonical batches
CREATE_NO_WINDOW = 0x08000000 if sys.platform == "win32" else 0
MONTHS = ["", "JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE",
          "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER"]


# ---------- cut-off window ----------

def derive_window(today):
    """Return (from_date, to_date) per the 1st/16th schedule, based on PHT date."""
    if today.day == 16:
        return dt.date(today.year, today.month, 1), dt.date(today.year, today.month, 15)
    if today.day == 1:
        prev_end = today - dt.timedelta(days=1)
        last = calendar.monthrange(prev_end.year, prev_end.month)[1]
        return dt.date(prev_end.year, prev_end.month, 16), dt.date(prev_end.year, prev_end.month, last)
    raise SystemExit(f"Run date {today} is not the 1st or 16th; pass --from/--to to run manually.")


def period_label(frm, to):
    return f"{MONTHS[frm.month]} {frm.day:02d}-{to.day:02d}, {frm.year}"


# ---------- credentials ----------

def supabase_key():
    k = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if k:
        return k.strip()
    out = subprocess.run(
        ["doppler", "secrets", "get", "SUPABASE_SERVICE_ROLE_KEY", "--plain",
         "--project", "bei-erp", "--config", "dev"],
        capture_output=True, text=True, creationflags=CREATE_NO_WINDOW).stdout.strip()
    if not out:
        raise SystemExit("No SUPABASE_SERVICE_ROLE_KEY (env or Doppler).")
    return out


def sa_creds(scopes, subject=SENDER):
    return service_account.Credentials.from_service_account_file(
        SA_JSON, scopes=scopes).with_subject(subject)


# ---------- pull punches ----------

def pull_punches(frm, to, out_csv):
    """PHT [frm 00:00 .. to+1 00:00) -> UTC bounds; page Supabase REST to CSV."""
    t0 = dt.datetime(frm.year, frm.month, frm.day, tzinfo=PHT).astimezone(dt.timezone.utc)
    t1 = (dt.datetime(to.year, to.month, to.day, tzinfo=PHT) + dt.timedelta(days=1)).astimezone(dt.timezone.utc)
    key = supabase_key()
    rows, offset = [], 0
    while True:
        params = urllib.parse.urlencode({
            "select": "pin,event_time,device_sn,store_name,status_code",
            "event_time": "gte." + t0.isoformat(),
            "and": "(event_time.lt." + t1.isoformat() + ")",
            "order": "event_time.asc", "limit": "1000", "offset": str(offset)})
        req = urllib.request.Request(SUPABASE_URL + "/rest/v1/attendance_punches?" + params,
                                     headers={"apikey": key, "Authorization": "Bearer " + key})
        with urllib.request.urlopen(req) as resp:
            batch = json.loads(resp.read())
        rows.extend(batch)
        if len(batch) < 1000:
            break
        offset += 1000
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["pin", "event_time", "device_sn", "store_name", "status_code"])
        w.writeheader()
        w.writerows(rows)
    return len(rows)


# ---------- fetch canonical bio-keyed templates ----------

def _drive():
    return build("drive", "v3", credentials=sa_creds(
        ["https://www.googleapis.com/auth/drive.readonly"]), cache_discovery=False)


def _children(drive, fid, only_folders=False):
    q = f"'{fid}' in parents and trashed=false"
    if only_folders:
        q += " and mimeType='application/vnd.google-apps.folder'"
    items, tok = [], None
    while True:
        resp = drive.files().list(q=q, fields="nextPageToken, files(id,name,mimeType,modifiedTime)",
                                  supportsAllDrives=True, includeItemsFromAllDrives=True,
                                  pageSize=200, pageToken=tok).execute()
        items.extend(resp.get("files", []))
        tok = resp.get("nextPageToken")
        if not tok:
            break
    return items


def fetch_templates(out_dir):
    """Download the latest 'WITH UPDATED BIO ID' consolidated batch (bio-keyed templates).

    Picks the most recently modified Timekeeping subfolder whose name contains
    'UPDATED BIO ID', gathers xlsx from it and its subfolders, dedupes by the
    leading store-number prefix (latest modifiedTime wins). NEVER store self-submissions.
    """
    os.makedirs(out_dir, exist_ok=True)
    drive = _drive()
    subs = _children(drive, TIMEKEEPING_ROOT, only_folders=True)
    canon = [s for s in subs if "updated bio id" in s["name"].lower()]
    if not canon:
        raise SystemExit("No 'WITH UPDATED BIO ID' canonical template folder found in Timekeeping root.")
    batch = sorted(canon, key=lambda s: s["modifiedTime"])[-1]
    # gather files from the batch folder + any subfolders (B1/B2/B3 etc.)
    folders = [batch["id"]] + [c["id"] for c in _children(drive, batch["id"], only_folders=True)]
    files = []
    for fid in folders:
        files += [f for f in _children(drive, fid)
                  if f["name"].lower().endswith(".xlsx")
                  and "payroll instructions" not in f["name"].lower()]
    by_prefix = {}
    for f in files:
        m = re.match(r"^(\d+)_", f["name"])
        keyp = m.group(1) if m else f["name"]
        if keyp not in by_prefix or f["modifiedTime"] > by_prefix[keyp]["modifiedTime"]:
            by_prefix[keyp] = f
    for f in by_prefix.values():
        safe = re.sub(r"[^\w.\- ]", "_", f["name"])
        req = drive.files().get_media(fileId=f["id"], supportsAllDrives=True)
        fh = io.BytesIO()
        dl = MediaIoBaseDownload(fh, req)
        done = False
        while not done:
            _, done = dl.next_chunk()
        with open(os.path.join(out_dir, safe), "wb") as fo:
            fo.write(fh.getvalue())
    return batch["name"], len(by_prefix)


# ---------- finalize (clean names + period title) ----------

def _store_name(ws, hdr, base, num):
    v = ws.cell(row=hdr + 1, column=1).value
    if isinstance(v, str) and v.strip() and not v.strip().replace(".", "").isdigit():
        return v.strip()
    stem = base[len(num) + 1:] if base[:len(num)] == num else base
    return (re.split(r"_DTR|_ ?20\d\d|__AUTO", stem)[0]).strip(" _") or os.path.splitext(base)[0]


def finalize(src_dir, dst_dir, frm, to):
    os.makedirs(dst_dir, exist_ok=True)
    label = period_label(frm, to)
    fname_period = f"{MONTHS[frm.month].title()} {frm.day}-{to.day} {frm.year}"
    count = 0
    for p in sorted(glob.glob(os.path.join(src_dir, "*__AUTO-DRAFT*.xlsx"))):
        base = os.path.basename(p)
        num = (re.match(r"\s*(\d{2})", base) or [None, "00"])[1]
        wb = openpyxl.load_workbook(p, keep_links=False)
        sheet = pick_dtrdata_sheet(wb)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        hdr = find_header_row(ws) if sheet else None
        store = _store_name(ws, hdr, base, num) if hdr else os.path.splitext(base)[0]
        if hdr and hdr - 1 >= 1:
            pc = ws.cell(row=hdr - 1, column=1)
            if pc.value in (None, ""):
                pc.value = label
        safe_store = re.sub(r"[^\w .\-]", "_", store)
        out = os.path.join(dst_dir, f"{num}_{safe_store}_DTR_{fname_period} (AUTO).xlsx")
        wb.save(out)
        wb.close()
        count += 1
    return count


def summarize(gen_summary_json):
    s = json.load(open(gen_summary_json, encoding="utf-8"))
    ok = [r for r in s if r.get("status") == "OK"]
    te = sum(r.get("template_emps", 0) for r in ok)
    fl = sum(r.get("worked_filled", 0) for r in ok)
    return {"workbooks": len(ok), "emps": te, "filled": fl,
            "pct": (100.0 * fl / te) if te else 0.0,
            "no_punch": sum(r.get("no_punch", 0) for r in ok)}


# ---------- Drive delivery (per-cutoff folder, idempotent upload, share, link) ----------

def _drive_rw():
    return build("drive", "v3", credentials=sa_creds(
        ["https://www.googleapis.com/auth/drive"]), cache_discovery=False)


def _with_backoff(fn, tries=4):
    for i in range(tries):
        try:
            return fn()
        except HttpError as e:
            if e.resp.status not in (403, 429, 500, 503) or i == tries - 1:
                raise
            time.sleep(2 ** i)


def find_or_create_folder(drive, parent_id, name):
    q = (f"'{parent_id}' in parents and name = {json.dumps(name)} and trashed=false "
         "and mimeType='application/vnd.google-apps.folder'")
    hits = drive.files().list(q=q, fields="files(id)", supportsAllDrives=True,
                              includeItemsFromAllDrives=True).execute().get("files", [])
    if hits:
        return hits[0]["id"]
    body = {"name": name, "mimeType": "application/vnd.google-apps.folder", "parents": [parent_id]}
    return drive.files().create(body=body, fields="id", supportsAllDrives=True).execute()["id"]


def upload_or_update(drive, folder_id, path):
    """Idempotent: update the file in place if a same-named one exists, else create.

    Keeps the xlsx format (byte-fidelity); Drive renders/edits it in Google Sheets.
    """
    name = os.path.basename(path)
    q = f"'{folder_id}' in parents and name = {json.dumps(name)} and trashed=false"
    hits = drive.files().list(q=q, fields="files(id)", supportsAllDrives=True,
                              includeItemsFromAllDrives=True).execute().get("files", [])
    media = MediaFileUpload(path, mimetype=XLSX_MIME, resumable=False)
    if hits:
        _with_backoff(lambda: drive.files().update(
            fileId=hits[0]["id"], media_body=media, supportsAllDrives=True).execute())
    else:
        _with_backoff(lambda: drive.files().create(
            body={"name": name, "parents": [folder_id]}, media_body=media,
            fields="id", supportsAllDrives=True).execute())


def share_folder(drive, folder_id, email, role="writer"):
    drive.permissions().create(
        fileId=folder_id, sendNotificationEmail=False,
        body={"type": "user", "role": role, "emailAddress": email},
        supportsAllDrives=True).execute()


def send_link_email(to_addr, subject, body):
    creds = sa_creds(["https://www.googleapis.com/auth/gmail.send"], subject=SENDER)
    gmail = build("gmail", "v1", credentials=creds, cache_discovery=False)
    msg = EmailMessage()
    msg["To"], msg["From"], msg["Subject"] = to_addr, SENDER, subject
    msg.set_content(body)
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    return gmail.users().messages().send(userId="me", body={"raw": raw}).execute().get("id")


# ---------- orchestration ----------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--from", dest="frm")
    ap.add_argument("--to", dest="to")
    ap.add_argument("--to-email", default="hr@bebang.ph")
    ap.add_argument("--send-to-self", action="store_true")
    ap.add_argument("--no-send", action="store_true")
    ap.add_argument("--workdir", default="tmp/dtr_autogen_run")
    args = ap.parse_args()

    if args.frm and args.to:
        frm = dt.date.fromisoformat(args.frm)
        to = dt.date.fromisoformat(args.to)
    else:
        frm, to = derive_window(dt.datetime.now(PHT).date())
    label = period_label(frm, to)
    print(f"[1/6] Cut-off: {frm} .. {to}  ({label})")

    wd = args.workdir
    punches = os.path.join(wd, "punches.csv")
    tdir = os.path.join(wd, "templates")
    gdir = os.path.join(wd, "gen")
    fdir = os.path.join(wd, "final")
    for d in (wd, gdir):
        os.makedirs(d, exist_ok=True)

    n = pull_punches(frm, to, punches)
    print(f"[2/6] Pulled {n} punches -> {punches}")

    batch_name, n_tpl = fetch_templates(tdir)
    print(f"[3/6] Templates: {n_tpl} from '{batch_name}'")

    here = os.path.dirname(os.path.abspath(__file__))
    subprocess.run([sys.executable, os.path.join(here, "generate_dtr.py"),
                    "--from", frm.isoformat(), "--to", to.isoformat(),
                    "--punches-csv", punches, "--templates-dir", tdir, "--out-dir", gdir],
                   check=True, creationflags=CREATE_NO_WINDOW)
    print(f"[4/6] Generated -> {gdir}")

    nf = finalize(gdir, fdir, frm, to)
    stats = summarize(os.path.join(gdir, "_generation_summary.json"))
    print(f"[5/6] Finalized {nf} workbooks | {stats['filled']}/{stats['emps']} "
          f"Worked auto-filled ({stats['pct']:.1f}%)")

    # 6. deliver to a per-cut-off Drive folder (idempotent) + email only the link
    drive = _drive_rw()
    parent = os.environ.get("DTR_DRIVE_PARENT") or find_or_create_folder(
        drive, TIMEKEEPING_ROOT, AUTO_PARENT_NAME)
    folder_id = find_or_create_folder(drive, parent, f"DTR_{label} (AUTO)")
    for p in sorted(glob.glob(os.path.join(fdir, "*.xlsx"))):
        upload_or_update(drive, folder_id, p)
    link = f"https://drive.google.com/drive/folders/{folder_id}"
    print(f"[6/6] Uploaded {nf} workbooks -> {link}")

    if args.no_send:
        print("       --no-send: not emailing.")
        return
    recipient = SENDER if args.send_to_self else args.to_email
    if not args.send_to_self:
        share_folder(drive, folder_id, recipient, "writer")
    subject = f"BEI DTR (AUTO) — {label}"
    body = (
        f"The automated DTR for cut-off {label} is ready in Google Drive:\n\n"
        f"    {link}\n\n"
        f"Stores: {stats['workbooks']}  |  Employees: {stats['emps']}\n"
        f"Worked(Day) auto-filled from biometric logs: {stats['filled']} ({stats['pct']:.1f}%)\n"
        f"Left blank for HR (no punches / new hire / enrollment gap): {stats['no_punch']}\n\n"
        "Open the folder, then each store's workbook in Google Sheets. Only Worked(Day) is "
        "filled from the bio machine; Tardy/UT/Absence/leave are left blank for HR to complete "
        "in place. A 'Punch Detail (AUTO)' tab shows each person's daily in/out spans.\n\n"
        "Generated automatically by the BEI DTR system.\n")
    mid = send_link_email(recipient, subject, body)
    print(f"       Emailed folder link -> {recipient} (message id {mid})")


if __name__ == "__main__":
    main()
