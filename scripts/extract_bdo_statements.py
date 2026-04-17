"""Extract BDO bank statements from 3 tabs (BEI / BSI / BKI).

Challenge: headers repeat every ~10 rows (one per printed page).
Approach: scan row-by-row, skip metadata + header + footer rows, capture transaction rows.

Output: one CSV per account + combined CSV with source tab tagged.
"""
from pathlib import Path
import openpyxl
import pandas as pd

SRC = Path('F:/Downloads/BDO Bank Statements Head Office Accounts_ Jan 16 to Apr17 2026.xlsx')
OUT_DIR = Path('CEO/CashFlow/bdo_statements')
OUT_DIR.mkdir(parents=True, exist_ok=True)

COL_DATE = 2
COL_BRANCH = 4
COL_DESC = 7
COL_DEBIT = 8
COL_CREDIT = 9
COL_BALANCE = 10
COL_CHECK = 12

def is_date_str(v):
    if v is None:
        return False
    s = str(v).strip()
    return len(s) == 10 and s[2] == '/' and s[5] == '/'

def is_header_row(row):
    """Detect the repeated 'POSTING DATE / DEBIT / CREDIT' header rows."""
    return (
        str(row[COL_DATE-1] or '').strip() == 'POSTING DATE'
        or str(row[COL_DESC-1] or '').strip() == 'DESCRIPTION'
    )

def extract_tab(ws, account_name, account_number):
    """Extract transaction rows."""
    txns = []
    for r_idx in range(1, ws.max_row + 1):
        row = [ws.cell(r_idx, c).value for c in range(1, ws.max_column + 1)]
        # Skip header repeats and empty rows
        if is_header_row(row):
            continue
        if not row[COL_DATE-1]:
            continue
        # Must have valid date in col 2
        date_val = row[COL_DATE-1]
        if not is_date_str(date_val):
            continue
        # Parse amounts (values arrive as strings with thousands separators like "7,248.01")
        def parse_num(v):
            if v in (None, ''):
                return 0.0
            try:
                return float(str(v).replace(',', '').strip())
            except (ValueError, TypeError):
                return 0.0
        debit = parse_num(row[COL_DEBIT-1])
        credit = parse_num(row[COL_CREDIT-1])
        balance = parse_num(row[COL_BALANCE-1]) if row[COL_BALANCE-1] not in (None, '') else None
        txns.append({
            'account': account_name,
            'account_number': account_number,
            'source_row': r_idx,
            'date': str(date_val).strip(),
            'branch': str(row[COL_BRANCH-1] or '').strip(),
            'description': str(row[COL_DESC-1] or '').strip(),
            'debit': debit,
            'credit': credit,
            'balance': balance,
            'check_number': str(row[COL_CHECK-1] or '').strip(),
        })
    return txns

wb = openpyxl.load_workbook(SRC, data_only=True)
all_txns = []
for tab in wb.sheetnames:
    ws = wb[tab]
    # Read account metadata
    acct_num = ws.cell(12, 5).value
    acct_name = ws.cell(14, 5).value
    print(f'\n=== {tab} ===')
    print(f'  Account: {acct_name} ({acct_num})')
    print(f'  Period: {ws.cell(10, 5).value}')
    txns = extract_tab(ws, str(acct_name), str(acct_num))
    print(f'  Transactions extracted: {len(txns)}')
    total_debit = sum(t['debit'] for t in txns)
    total_credit = sum(t['credit'] for t in txns)
    print(f'  Total DEBIT  (out): ₱{total_debit:,.2f}')
    print(f'  Total CREDIT (in):  ₱{total_credit:,.2f}')
    all_txns.extend(txns)
    # Save per-tab CSV
    df_tab = pd.DataFrame(txns)
    safe_name = tab.replace(' ', '_').replace('/', '_')
    df_tab.to_csv(OUT_DIR / f'{safe_name}.csv', index=False)
    print(f'  Saved: {OUT_DIR / (safe_name + ".csv")}')

df = pd.DataFrame(all_txns)
df.to_csv(OUT_DIR / 'all_bdo_transactions.csv', index=False)
print()
print(f'TOTAL across 3 accounts: {len(df)} transactions')
print(f'Combined DEBIT:  ₱{df["debit"].sum():,.2f}')
print(f'Combined CREDIT: ₱{df["credit"].sum():,.2f}')

# Quick analysis
print()
print('=== DESCRIPTION types (all accounts) ===')
desc_counts = df['description'].value_counts().head(30)
for d, c in desc_counts.items():
    print(f'  {d[:40]:40s}  {c:5d} txns')

# Check numbers analysis
print()
print('=== CHECK NUMBER analysis ===')
# Real check numbers (not 000000000 or empty)
real_checks = df[
    (df['check_number'] != '000000000') &
    (df['check_number'] != '') &
    (df['check_number'].notna())
]
print(f'Real check numbers: {len(real_checks)} / {len(df)}')
if len(real_checks) > 0:
    print('Sample check-number txns:')
    for _, r in real_checks.head(15).iterrows():
        print(f'  {r["date"]}  chk {r["check_number"]:<15}  debit ₱{r["debit"]:>10,.2f}  desc: {r["description"][:40]}')
