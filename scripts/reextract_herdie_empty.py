"""
Re-extract 38 FAILED_EMPTY files from Herdie case verification.
Uses Gemini 2.5 Flash for PDFs, Gemini 2.0 Flash for images, Docling for DOCX.
"""
import os
import sys
import json
import time
import re
import traceback
from pathlib import Path

# ── Config ──────────────────────────────────────────────────────────────
GEMINI_API_KEY = "AIzaSyC9rY3tyqs7cfO-59gWWlpthDqCwefZOgo"
BASE = Path(r"F:\Dropbox\Projects\BEI-ERP\data\Audits\Herdie")
EXTRACTED_BASE = BASE / "_local_extracted"
REPORT_PATH = EXTRACTED_BASE / "_reextraction_report.json"

ANTI_HALLUCINATION_PROMPT = (
    "Extract ALL text from this document VERBATIM. Preserve exact structure, "
    "tables, numbers, dates, signatures, headers, footers. Do NOT interpret, "
    "summarize, paraphrase, or add any context. If text is illegible, write "
    "[ILLEGIBLE]. Output as clean markdown."
)

# ── Source mapping ──────────────────────────────────────────────────────
# Build the mapping of md_path -> source_path for all 38 files

def build_source_map(failed_entries):
    """Map each failed .md file to its original source file."""
    mapping = {}
    attach_dir = BASE / "01_Evidence" / "Commissary_Raw" / "attachments"

    # Known Commissary_Raw mappings (found by manual search)
    commissary_map = {
        "BEBANG_CONTRACT_AGREEMENT_signed..md": "BEBANG CONTRACT AGREEMENT_ signed..pdf",
        "dan_19bbbd9329a4b515_Commi_NTP_Signed.md": "dan_19bbbd9329a4b515_Commi NTP Signed.pdf",
        "dan_19bbbd9329a4b515_Commi_RFP_Signed.md": "dan_19bbbd9329a4b515_Commi RFP Signed.pdf",
        "dan_drive_Commi_NTP_Signed.md": "dan_drive_Commi NTP Signed.pdf",
        "dan_drive_Commi_RFP_Signed.md": "dan_drive_Commi RFP Signed.pdf",
        "dan_drive_FD_Commi_Contract.final_Signed.md": "dan_drive_FD Commi Contract.final Signed.pdf",
        "dan_drive_Kumsung_Quote_1.7M_Initial_Insulated_Panels.md": "dan_drive_Kumsung Quote 1.7M Initial _Insulated Panels_.pdf",
        "dan_drive_Kumsung_Quote_2.4M_Initial_Ref.md": "dan_drive_Kumsung Quote 2.4M Initial _Ref_.pdf",
    }

    for entry in failed_entries:
        md_path = Path(entry["file"])
        md_name = md_path.name

        if "Commissary_Raw" in str(md_path) and md_name in commissary_map:
            source = attach_dir / commissary_map[md_name]
            mapping[str(md_path)] = str(source)
        elif "attachments_raw" in str(md_path):
            # For email attachments: source is the same file without .md
            orig_name = md_name[:-3] if md_name.endswith(".md") else md_name
            # The source is in the non-extracted path
            rel = md_path.relative_to(EXTRACTED_BASE)
            # Remove .md extension
            rel_str = str(rel)
            if rel_str.endswith(".md"):
                rel_str = rel_str[:-3]
            source = BASE / rel_str
            if source.exists():
                mapping[str(md_path)] = str(source)
            else:
                # Try with .pdf extension for the franchise agreement
                for ext in [".pdf", ".docx", ".doc", ".xlsx"]:
                    try_path = str(source) + ext
                    if os.path.exists(try_path):
                        mapping[str(md_path)] = try_path
                        break
                else:
                    print(f"  WARNING: No source found for {md_name}")
                    mapping[str(md_path)] = None

    return mapping


# ── Extraction functions ────────────────────────────────────────────────

def extract_with_gemini_pdf(source_path):
    """Extract text from PDF using Gemini 2.5 Flash."""
    from google import genai

    client = genai.Client(api_key=GEMINI_API_KEY)

    file_size = os.path.getsize(source_path)
    print(f"    Uploading PDF ({file_size:,} bytes) to Gemini...")

    uploaded = client.files.upload(file=source_path)

    # Wait for processing
    while uploaded.state.name == "PROCESSING":
        time.sleep(2)
        uploaded = client.files.get(name=uploaded.name)

    if uploaded.state.name == "FAILED":
        raise RuntimeError(f"Gemini file processing failed for {source_path}")

    response = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=[uploaded, ANTI_HALLUCINATION_PROMPT],
    )

    # Clean up uploaded file
    try:
        client.files.delete(name=uploaded.name)
    except Exception:
        pass

    return response.text if response.text else ""


def extract_with_gemini_image(source_path):
    """Extract text from image using Gemini 2.0 Flash."""
    from google import genai
    from google.genai import types
    import mimetypes

    client = genai.Client(api_key=GEMINI_API_KEY)

    mime_type = mimetypes.guess_type(source_path)[0] or "image/png"

    with open(source_path, "rb") as f:
        image_data = f.read()

    file_size = len(image_data)
    print(f"    Sending image ({file_size:,} bytes) to Gemini 2.0 Flash...")

    response = client.models.generate_content(
        model="gemini-2.0-flash",
        contents=[
            types.Part.from_bytes(data=image_data, mime_type=mime_type),
            ANTI_HALLUCINATION_PROMPT,
        ],
    )

    return response.text if response.text else ""


def extract_with_docling(source_path):
    """Extract text from DOCX using Docling."""
    from docling.document_converter import DocumentConverter

    converter = DocumentConverter()
    result = converter.convert(source_path)
    return result.document.export_to_markdown()


def extract_with_pymupdf(source_path):
    """Extract text from DOC/PDF using PyMuPDF."""
    import fitz

    doc = fitz.open(source_path)
    pages = []
    for page_num in range(len(doc)):
        page = doc[page_num]
        text = page.get_text()
        if text.strip():
            pages.append(f"--- Page {page_num + 1} ---\n{text}")
    doc.close()
    return "\n\n".join(pages)


# ── Main logic ──────────────────────────────────────────────────────────

def get_file_type(source_path):
    """Determine file type from extension."""
    ext = os.path.splitext(source_path)[1].lower()
    if ext == ".pdf":
        return "pdf"
    elif ext in (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"):
        return "image"
    elif ext == ".docx":
        return "docx"
    elif ext == ".doc":
        return "doc"
    elif ext in (".xlsx", ".xlsm"):
        return "xlsx"
    return "unknown"


def extract_file(source_path, file_type):
    """Route extraction to the correct tool."""
    if file_type == "pdf":
        return extract_with_gemini_pdf(source_path)
    elif file_type == "image":
        return extract_with_gemini_image(source_path)
    elif file_type == "docx":
        return extract_with_docling(source_path)
    elif file_type == "doc":
        return extract_with_pymupdf(source_path)
    elif file_type == "xlsx":
        # openpyxl extraction
        import openpyxl
        wb = openpyxl.load_workbook(source_path, data_only=True)
        lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"## Sheet: {sheet_name}\n")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    lines.append(" | ".join(cells))
        return "\n".join(lines)
    else:
        raise ValueError(f"Unknown file type: {file_type}")


def write_md_file(md_path, content, source_path, file_type):
    """Write extracted content to .md file with YAML frontmatter."""
    frontmatter = f"""---
source_file: "{os.path.basename(source_path)}"
file_type: "{file_type}"
extraction_tool: "{'gemini-2.5-flash' if file_type == 'pdf' else 'gemini-2.0-flash' if file_type == 'image' else 'docling' if file_type == 'docx' else 'pymupdf' if file_type == 'doc' else 'openpyxl'}"
extraction_date: "{time.strftime('%Y-%m-%d %H:%M:%S')}"
chars_extracted: {len(content)}
reextracted: true
---

"""
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(frontmatter + content)


def main():
    print("=" * 70)
    print("Herdie Case Re-Extraction: 38 FAILED_EMPTY Files")
    print("=" * 70)

    # Load failed entries
    failed_path = EXTRACTED_BASE / "_failed_empty.json"
    with open(failed_path, "r", encoding="utf-8") as f:
        failed_entries = json.load(f)

    print(f"\nLoaded {len(failed_entries)} failed entries")

    # Build source mapping
    print("\nBuilding source file mapping...")
    source_map = build_source_map(failed_entries)

    found = sum(1 for v in source_map.values() if v is not None)
    missing = sum(1 for v in source_map.values() if v is None)
    print(f"  Found: {found}, Missing sources: {missing}")

    # Process each file
    report = []
    stats = {"success": 0, "ok_non_text": 0, "failed": 0, "no_source": 0}

    for i, entry in enumerate(failed_entries):
        md_path = entry["file"]
        md_name = os.path.basename(md_path)
        source_path = source_map.get(md_path)

        # Extract old chars count from failure message
        old_chars = 0
        for fail in entry.get("failures", []):
            m = re.search(r"only (\d+) chars", fail)
            if m:
                old_chars = int(m.group(1))

        print(f"\n[{i+1}/{len(failed_entries)}] {md_name}")

        if source_path is None:
            print(f"  SKIP: No source file found")
            report.append({
                "md_file": md_path,
                "source_file": None,
                "old_status": "FAIL",
                "old_chars": old_chars,
                "new_status": "no_source",
                "new_chars": 0,
                "error": "Source file not found",
            })
            stats["no_source"] += 1
            continue

        file_type = get_file_type(source_path)
        print(f"  Source: {os.path.basename(source_path)} ({file_type})")

        try:
            content = extract_file(source_path, file_type)
            new_chars = len(content.strip()) if content else 0

            if new_chars < 10:
                # Likely non-text document
                note = (
                    "## Extraction Note\n\n"
                    "This file appears to be a non-text document "
                    "(architectural drawing/photo/scan/email signature image). "
                    "Visual inspection required for content description."
                )
                write_md_file(md_path, note, source_path, file_type)
                new_status = "ok_non_text"
                stats["ok_non_text"] += 1
                print(f"  Result: non-text document ({new_chars} chars extracted)")
            else:
                write_md_file(md_path, content, source_path, file_type)
                new_status = "success"
                stats["success"] += 1
                print(f"  Result: SUCCESS ({old_chars} -> {new_chars} chars)")

            report.append({
                "md_file": md_path,
                "source_file": source_path,
                "file_type": file_type,
                "old_status": "FAIL",
                "old_chars": old_chars,
                "new_status": new_status,
                "new_chars": new_chars,
            })

        except Exception as e:
            error_msg = f"{type(e).__name__}: {str(e)}"
            print(f"  ERROR: {error_msg}")
            traceback.print_exc()
            report.append({
                "md_file": md_path,
                "source_file": source_path,
                "file_type": file_type,
                "old_status": "FAIL",
                "old_chars": old_chars,
                "new_status": "failed",
                "new_chars": 0,
                "error": error_msg,
            })
            stats["failed"] += 1

        # Rate limit for Gemini API
        if file_type in ("pdf", "image"):
            time.sleep(1)

    # Write report
    report_data = {
        "reextraction_date": time.strftime("%Y-%m-%d %H:%M:%S"),
        "total_files": len(failed_entries),
        "stats": stats,
        "results": report,
    }

    with open(REPORT_PATH, "w", encoding="utf-8") as f:
        json.dump(report_data, f, indent=2, ensure_ascii=False)

    # Summary
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"  Total files:   {len(failed_entries)}")
    print(f"  Success:       {stats['success']}")
    print(f"  Non-text (OK): {stats['ok_non_text']}")
    print(f"  Failed:        {stats['failed']}")
    print(f"  No source:     {stats['no_source']}")
    print(f"\nReport: {REPORT_PATH}")


if __name__ == "__main__":
    main()
