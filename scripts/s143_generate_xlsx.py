"""S143: Generate BEI-branded Excel reports for procurement team.

File 1: Supplier Compliance Report (excludes test suppliers)
File 2: Item Price & Data Quality Report
"""
import csv
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy
import shutil

# === BEI Brand Colors ===
BEI_GREEN = '04400A'
BEI_GOLD = 'C8900A'
BEI_PURPLE = '801297'
BEI_GREEN_TINT = 'E6ECE7'
BEI_GOLD_LIGHT = 'F8F0D9'
BEI_PURPLE_LIGHT = 'F2E5F5'
BEI_TEXT_DARK = '1A1A1A'
BEI_TEXT_WHITE = 'FFFFFF'
BEI_BORDER = 'D4D0C8'
BEI_RED = 'CC0000'

# === Styles ===
header_font = Font(bold=True, color=BEI_TEXT_WHITE, size=11, name='Calibri')
header_fill = PatternFill('solid', fgColor=BEI_GREEN)
subheader_font = Font(bold=True, color=BEI_TEXT_DARK, size=10, name='Calibri')
subheader_fill = PatternFill('solid', fgColor=BEI_GOLD)
body_font = Font(color=BEI_TEXT_DARK, size=10, name='Calibri')
even_fill = PatternFill('solid', fgColor=BEI_GREEN_TINT)
odd_fill = PatternFill('solid', fgColor=BEI_TEXT_WHITE)
kpi_font = Font(bold=True, color=BEI_PURPLE, size=14, name='Calibri')
kpi_fill = PatternFill('solid', fgColor=BEI_PURPLE_LIGHT)
kpi_label_font = Font(bold=True, color=BEI_TEXT_DARK, size=10, name='Calibri')
total_font = Font(bold=True, color=BEI_TEXT_WHITE, size=10, name='Calibri')
total_fill = PatternFill('solid', fgColor=BEI_GREEN)
callout_fill = PatternFill('solid', fgColor=BEI_GOLD_LIGHT)
red_font = Font(bold=True, color=BEI_RED, size=10, name='Calibri')
green_font = Font(bold=True, color=BEI_GREEN, size=10, name='Calibri')
thin_border = Border(
    left=Side(style='thin', color=BEI_BORDER),
    right=Side(style='thin', color=BEI_BORDER),
    top=Side(style='thin', color=BEI_BORDER),
    bottom=Side(style='thin', color=BEI_BORDER),
)
PHP_FMT = '₱#,##0.00_);[Red](₱#,##0.00);₱"-"??_)'

# Test supplier patterns to exclude
TEST_PATTERNS = [
    r'^API Test', r'^Curl Test', r'^Cutover Readiness', r'^E2E',
    r'^Dup Bank', r'^Dup Email', r'^JS Test', r'^No Terms',
    r'^QA-', r'^S062', r'^S063', r'^s063', r'^UI Created',
    r'^CYCLE-', r'^NOTERMS', r'^NOADDR',
]

def is_test_supplier(name, code):
    for pat in TEST_PATTERNS:
        if re.search(pat, name, re.IGNORECASE) or re.search(pat, code, re.IGNORECASE):
            return True
    return False

def apply_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def apply_body_row(ws, row, cols, is_even):
    fill = even_fill if is_even else odd_fill
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = body_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')

def auto_width(ws, min_width=10, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def generate_supplier_compliance():
    """File 1: Supplier Compliance Report."""
    print("=== Generating Supplier Compliance Report ===")

    with open('output/s143/suppliers_missing_compliance.csv', 'r', encoding='utf-8') as f:
        all_rows = list(csv.DictReader(f))

    # Split real vs test
    real = [r for r in all_rows if not is_test_supplier(r['supplier_name'], r['supplier_code'])]
    test = [r for r in all_rows if is_test_supplier(r['supplier_name'], r['supplier_code'])]

    print(f"  Total: {len(all_rows)} | Real: {len(real)} | Test: {len(test)}")

    wb = Workbook()

    # === Sheet 1: Summary Dashboard ===
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = BEI_GREEN

    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "BEI Supplier Compliance Report"
    title_cell.font = Font(bold=True, color=BEI_TEXT_WHITE, size=16, name='Calibri')
    title_cell.fill = header_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:F2')
    ws['A2'].value = f"Generated: {datetime.now().strftime('%B %d, %Y')} | Sprint S143 — Procurement Defect Remediation"
    ws['A2'].font = Font(color=BEI_TEXT_DARK, size=9, name='Calibri', italic=True)
    ws['A2'].fill = callout_fill
    ws['A2'].alignment = Alignment(horizontal='center')

    # KPIs
    kpis = [
        ("Total Active Suppliers", len(real)),
        ("With TIN", sum(1 for r in real if r['tin_present'] == 'Yes')),
        ("With BIR 2307", sum(1 for r in real if r['bir_2307_present'] == 'Yes')),
        ("With SEC Cert", sum(1 for r in real if r['sec_cert_present'] == 'Yes')),
        ("With Business Permit", sum(1 for r in real if r['business_permit_present'] == 'Yes')),
        ("Fully Compliant", sum(1 for r in real if r['docs_complete'] == 'Yes')),
    ]

    row = 4
    for i, (label, value) in enumerate(kpis):
        col = (i % 3) * 2 + 1
        if i == 3:
            row = 7

        ws.cell(row=row, column=col, value=value).font = kpi_font
        ws.cell(row=row, column=col).fill = kpi_fill
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=col).border = thin_border

        ws.cell(row=row + 1, column=col, value=label).font = kpi_label_font
        ws.cell(row=row + 1, column=col).alignment = Alignment(horizontal='center')

    # Compliance gap callout
    row = 10
    ws.merge_cells(f'A{row}:F{row}')
    gap_cell = ws.cell(row=row, column=1)
    gap_pct = 0 if len(real) == 0 else round((sum(1 for r in real if r['docs_complete'] == 'Yes') / len(real)) * 100)
    gap_cell.value = f"⚠ Compliance Gap: {100 - gap_pct}% of active suppliers are missing at least one compliance document"
    gap_cell.font = Font(bold=True, color=BEI_TEXT_DARK, size=11, name='Calibri')
    gap_cell.fill = callout_fill
    gap_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[row].height = 30

    # Action items
    row = 12
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value="Required Actions for Procurement Team:").font = subheader_font
    ws.cell(row=row, column=1).fill = subheader_fill

    actions = [
        "1. Upload BIR 2307 certificates for all active suppliers",
        "2. Upload SEC Registration certificates",
        "3. Upload current Business Permits (check expiry dates)",
        "4. Verify TIN numbers are correct for suppliers marked 'No'",
        f"5. {len(test)} test/E2E suppliers should be set to Inactive or deleted",
    ]
    for i, action in enumerate(actions):
        ws.merge_cells(f'A{row + 1 + i}:F{row + 1 + i}')
        ws.cell(row=row + 1 + i, column=1, value=action).font = body_font

    auto_width(ws)

    # === Sheet 2: Supplier Detail ===
    ws2 = wb.create_sheet("Suppliers — Action Required")
    ws2.sheet_properties.tabColor = BEI_GOLD

    headers = ['#', 'Supplier Name', 'Supplier Code', 'TIN', 'BIR 2307', 'SEC Cert', 'Business Permit', 'Complete']
    for col, h in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=h)
    apply_header_row(ws2, 1, len(headers))
    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    for i, r in enumerate(real):
        row = i + 2
        ws2.cell(row=row, column=1, value=i + 1)
        ws2.cell(row=row, column=2, value=r['supplier_name'])
        ws2.cell(row=row, column=3, value=r['supplier_code'])
        ws2.cell(row=row, column=4, value=r['tin_present'])
        ws2.cell(row=row, column=5, value=r['bir_2307_present'])
        ws2.cell(row=row, column=6, value=r['sec_cert_present'])
        ws2.cell(row=row, column=7, value=r['business_permit_present'])
        ws2.cell(row=row, column=8, value=r['docs_complete'])
        apply_body_row(ws2, row, len(headers), i % 2 == 0)

        # Red highlight for "No" cells
        for col in [4, 5, 6, 7]:
            cell = ws2.cell(row=row, column=col)
            if cell.value == 'No':
                cell.font = red_font
            elif cell.value == 'Yes':
                cell.font = green_font

    auto_width(ws2)

    # === Sheet 3: Test Suppliers (reference) ===
    ws3 = wb.create_sheet("Test Suppliers (Exclude)")
    ws3.sheet_properties.tabColor = BEI_BORDER

    test_headers = ['#', 'Supplier Name', 'Supplier Code', 'Recommendation']
    for col, h in enumerate(test_headers, 1):
        ws3.cell(row=1, column=col, value=h)
    apply_header_row(ws3, 1, len(test_headers))

    for i, r in enumerate(test):
        row = i + 2
        ws3.cell(row=row, column=1, value=i + 1)
        ws3.cell(row=row, column=2, value=r['supplier_name'])
        ws3.cell(row=row, column=3, value=r['supplier_code'])
        ws3.cell(row=row, column=4, value='Set to Inactive or Delete')
        apply_body_row(ws3, row, len(test_headers), i % 2 == 0)

    auto_width(ws3)

    # Save
    path1 = 'output/s143/BEI_Supplier_Compliance_Report_S143.xlsx'
    wb.save(path1)
    print(f"  Saved: {path1}")
    return path1


def generate_item_data_quality():
    """File 2: Item Price & Data Quality Report."""
    print("\n=== Generating Item Price & Data Quality Report ===")

    with open('output/s143/item_price_backfill.csv', 'r', encoding='utf-8') as f:
        items = list(csv.DictReader(f))

    with open('output/s143/sku_master_crossref.csv', 'r', encoding='utf-8') as f:
        skus = list(csv.DictReader(f))

    zero_price = [i for i in items if float(i.get('current_rate', 0) or 0) == 0]
    has_price = [i for i in items if float(i.get('current_rate', 0) or 0) > 0]

    orphans = [
        {'item_code': 'OS220', 'item_name': 'ECOTHERM 71GALLONS', 'source_po': 'PO-2026-04228'},
        {'item_code': 'OS221', 'item_name': 'ECOTHERM STANDARD FITTINGS', 'source_po': 'PO-2026-04228'},
        {'item_code': 'OS222', 'item_name': 'ECOTHERM BRACKET', 'source_po': 'PO-2026-04228'},
    ]

    print(f"  Items: {len(items)} | Zero price: {len(zero_price)} | SKUs: {len(skus)} | Orphans: {len(orphans)}")

    wb = Workbook()

    # === Sheet 1: Summary ===
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = BEI_GREEN

    ws.merge_cells('A1:F1')
    ws['A1'].value = "BEI Item Price & Data Quality Report"
    ws['A1'].font = Font(bold=True, color=BEI_TEXT_WHITE, size=16, name='Calibri')
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:F2')
    ws['A2'].value = f"Generated: {datetime.now().strftime('%B %d, %Y')} | Sprint S143 — Procurement Defect Remediation"
    ws['A2'].font = Font(color=BEI_TEXT_DARK, size=9, name='Calibri', italic=True)
    ws['A2'].fill = callout_fill
    ws['A2'].alignment = Alignment(horizontal='center')

    # KPIs
    kpis = [
        ("Total Items", len(items)),
        ("With Price", len(has_price)),
        ("Zero Price", len(zero_price)),
        ("SKU Master Items", len(skus)),
        ("SKU Matches", sum(1 for s in skus if s.get('frappe_match') != 'NO_MATCH')),
        ("Orphan Items", len(orphans)),
    ]

    row = 4
    for i, (label, value) in enumerate(kpis):
        col = (i % 3) * 2 + 1
        if i == 3:
            row = 7
        ws.cell(row=row, column=col, value=value).font = kpi_font
        ws.cell(row=row, column=col).fill = kpi_fill
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row + 1, column=col, value=label).font = kpi_label_font
        ws.cell(row=row + 1, column=col).alignment = Alignment(horizontal='center')

    # Coverage
    row = 10
    pct = round(len(has_price) / len(items) * 100) if items else 0
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value=f"Price Coverage: {pct}% ({len(has_price)}/{len(items)} items have standard_rate > 0)")
    ws.cell(row=row, column=1).font = Font(bold=True, color=BEI_TEXT_DARK, size=11, name='Calibri')
    ws.cell(row=row, column=1).fill = callout_fill
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')

    row = 12
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value="Required Actions:").font = subheader_font
    ws.cell(row=row, column=1).fill = subheader_fill
    actions = [
        f"1. Set prices for {len(zero_price)} items with zero standard_rate (see 'Items — Zero Price' tab)",
        "2. Create Item Master entries for 3 orphan items (ECOTHERM) or correct PO-2026-04228",
        f"3. All {len(skus)} SKU Master items already matched in Frappe — no action needed",
    ]
    for i, a in enumerate(actions):
        ws.merge_cells(f'A{row + 1 + i}:F{row + 1 + i}')
        ws.cell(row=row + 1 + i, column=1, value=a).font = body_font

    auto_width(ws)

    # === Sheet 2: Items with Zero Price ===
    ws2 = wb.create_sheet("Items — Zero Price")
    ws2.sheet_properties.tabColor = BEI_RED

    headers = ['#', 'Item Code', 'Item Name', 'Current Rate', 'Last PO Rate', 'UOM', 'Last PO', 'PO Date']
    for col, h in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=h)
    apply_header_row(ws2, 1, len(headers))
    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    for i, item in enumerate(zero_price):
        row = i + 2
        ws2.cell(row=row, column=1, value=i + 1)
        ws2.cell(row=row, column=2, value=item['item_code'])
        ws2.cell(row=row, column=3, value=item['item_name'])
        rate_cell = ws2.cell(row=row, column=4, value=float(item.get('current_rate', 0) or 0))
        rate_cell.number_format = PHP_FMT
        suggested = item.get('suggested_rate', '')
        sug_cell = ws2.cell(row=row, column=5, value=float(suggested) if suggested else None)
        sug_cell.number_format = PHP_FMT
        ws2.cell(row=row, column=6, value=item.get('uom', ''))
        ws2.cell(row=row, column=7, value=item.get('source_po', ''))
        ws2.cell(row=row, column=8, value=item.get('po_date', ''))
        apply_body_row(ws2, row, len(headers), i % 2 == 0)

    auto_width(ws2)

    # === Sheet 3: All Items (full list) ===
    ws3 = wb.create_sheet("All Items")
    ws3.sheet_properties.tabColor = BEI_GREEN

    all_headers = ['#', 'Item Code', 'Item Name', 'Standard Rate', 'Last PO Rate', 'UOM', 'Last PO', 'PO Date', 'Needs Update']
    for col, h in enumerate(all_headers, 1):
        ws3.cell(row=1, column=col, value=h)
    apply_header_row(ws3, 1, len(all_headers))
    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f'A1:{get_column_letter(len(all_headers))}1'

    for i, item in enumerate(items):
        row = i + 2
        ws3.cell(row=row, column=1, value=i + 1)
        ws3.cell(row=row, column=2, value=item['item_code'])
        ws3.cell(row=row, column=3, value=item['item_name'])
        rate_cell = ws3.cell(row=row, column=4, value=float(item.get('current_rate', 0) or 0))
        rate_cell.number_format = PHP_FMT
        suggested = item.get('suggested_rate', '')
        sug_cell = ws3.cell(row=row, column=5, value=float(suggested) if suggested else None)
        sug_cell.number_format = PHP_FMT
        ws3.cell(row=row, column=6, value=item.get('uom', ''))
        ws3.cell(row=row, column=7, value=item.get('source_po', ''))
        ws3.cell(row=row, column=8, value=item.get('po_date', ''))
        ws3.cell(row=row, column=9, value=item.get('needs_update', ''))
        apply_body_row(ws3, row, len(all_headers), i % 2 == 0)

        # Highlight zero price rows
        if float(item.get('current_rate', 0) or 0) == 0:
            rate_cell.font = red_font

    auto_width(ws3)

    # === Sheet 4: SKU Master Crossref ===
    ws4 = wb.create_sheet("SKU Master Crossref")
    ws4.sheet_properties.tabColor = BEI_GOLD

    sku_headers = ['#', 'CSV Item Code', 'CSV Item Name', 'Frappe Match', 'Frappe Item Code', 'Match Method']
    for col, h in enumerate(sku_headers, 1):
        ws4.cell(row=1, column=col, value=h)
    apply_header_row(ws4, 1, len(sku_headers))
    ws4.freeze_panes = 'A2'

    for i, sku in enumerate(skus):
        row = i + 2
        ws4.cell(row=row, column=1, value=i + 1)
        ws4.cell(row=row, column=2, value=sku.get('csv_item_code', ''))
        ws4.cell(row=row, column=3, value=sku.get('csv_item_name', ''))
        match_cell = ws4.cell(row=row, column=4, value=sku.get('frappe_match', ''))
        ws4.cell(row=row, column=5, value=sku.get('frappe_item_code', ''))
        ws4.cell(row=row, column=6, value=sku.get('match_method', ''))
        apply_body_row(ws4, row, len(sku_headers), i % 2 == 0)
        if sku.get('frappe_match') == 'EXACT_CODE':
            match_cell.font = green_font
        elif sku.get('frappe_match') == 'NO_MATCH':
            match_cell.font = red_font

    auto_width(ws4)

    # === Sheet 5: Orphan Items ===
    ws5 = wb.create_sheet("Orphan Items")
    ws5.sheet_properties.tabColor = BEI_RED

    orph_headers = ['#', 'Item Code', 'Item Name', 'Source PO', 'Action Required']
    for col, h in enumerate(orph_headers, 1):
        ws5.cell(row=1, column=col, value=h)
    apply_header_row(ws5, 1, len(orph_headers))

    for i, o in enumerate(orphans):
        row = i + 2
        ws5.cell(row=row, column=1, value=i + 1)
        ws5.cell(row=row, column=2, value=o['item_code'])
        ws5.cell(row=row, column=3, value=o['item_name'])
        ws5.cell(row=row, column=4, value=o['source_po'])
        ws5.cell(row=row, column=5, value='Create in Item Master or correct PO')
        apply_body_row(ws5, row, len(orph_headers), i % 2 == 0)

    auto_width(ws5)

    path2 = 'output/s143/BEI_Item_Data_Quality_Report_S143.xlsx'
    wb.save(path2)
    print(f"  Saved: {path2}")
    return path2


if __name__ == '__main__':
    import os
    os.chdir('F:/Dropbox/Projects/BEI-ERP')

    p1 = generate_supplier_compliance()
    p2 = generate_item_data_quality()

    # Copy to Downloads
    dl = 'F:/Downloads'
    os.makedirs(dl, exist_ok=True)
    shutil.copy2(p1, os.path.join(dl, os.path.basename(p1)))
    shutil.copy2(p2, os.path.join(dl, os.path.basename(p2)))
    print(f"\nCopied to {dl}")
    print("Done!")
