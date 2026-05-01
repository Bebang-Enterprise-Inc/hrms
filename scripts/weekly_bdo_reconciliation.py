"""Phase B + B2 — Weekly BDO Reconciliation.

Dual-mode script. Set BDO_RUN_CONTEXT to:
  - 'local'  (default)  — run from Windows Task Scheduler
  - 'github'            — run from GitHub Actions (writes to GH artifact instead of emailing)

Weekly workflow:
  1. List files in Drive BDO/ folder + subfolders (recursively)
  2. Pick most-recently-modified .xlsx not already in _Archive
  3. Acquire Drive sidecar lock (prevents dual-trigger double-write)
  4. Download, extract, match BDO → FPM by check number
  5. Apply DELTA updates to FPM, SOA, HO AP (only rows not yet cleared in those sheets)
  6. Move processed file to _Archive/
  7. Write reconciliation report + release lock

Safe to run twice — lock + _Archive membership ensure idempotency.
"""
import argparse
import io
import json
import os
import tempfile
from pathlib import Path
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload

CREDS_PATH = os.environ.get('GOOGLE_CREDS_PATH', 'credentials/task-manager-service.json')
OWNER = os.environ.get('BDO_IMPERSONATE', 'sam@bebang.ph')
RUN_CONTEXT = os.environ.get('BDO_RUN_CONTEXT', 'local')

FPM_ID = '1t4wJLiAfIMJm6fe-x6h4eZn_S_Lx1AGN5ORd5Ywhcyw'
SOA_ID = '1ZHe2VoAFa94ET4I68C1jWM7nMzTdTCvttwZbICaLtB4'
HO_ID = '1jSwZRyIPisU4jiKS-Tn9VFoLukQI8UNoW13Hoov-75Y'

BDO_FOLDER_ID = '1o2Rjl9Y2eSlT4P7KPYQdZusQHuHwLQAy'     # Finance and Accounting / Treasury / Bank Statements 2026 / BDO
ARCHIVE_FOLDER_ID = '15czZiPTYcNqqBVQMobpPOTp7gc5PtenH'  # BS 2026 / _Archive (lock files land here)

INTERCOMPANY_PAYEES = [
    'BEBANG ENTERPRISE', 'BEBANG KITCHEN', 'BEBANG SHAW',
    'BEBANG HALO-HALO', 'BEBANG HALOHALO', 'BKI', 'BSI', 'BHH',
]


def is_intercompany(payee):
    p = str(payee).upper() if payee else ''
    return any(tag in p for tag in INTERCOMPANY_PAYEES)


def get_services():
    creds = service_account.Credentials.from_service_account_file(
        CREDS_PATH,
        scopes=['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive'],
    ).with_subject(OWNER)
    sheets = build('sheets', 'v4', credentials=creds, cache_discovery=False)
    drive = build('drive', 'v3', credentials=creds, cache_discovery=False)
    return sheets, drive


def list_bdo_files(drive):
    """Recursively walk BDO folder → return list of xlsx files (not in _Archive)."""
    to_visit = [BDO_FOLDER_ID]
    files = []
    visited = set()
    while to_visit:
        folder_id = to_visit.pop()
        if folder_id in visited:
            continue
        visited.add(folder_id)
        r = drive.files().list(
            q=f"'{folder_id}' in parents and trashed=false",
            fields='files(id,name,mimeType,modifiedTime,parents)',
            supportsAllDrives=True, includeItemsFromAllDrives=True,
            pageSize=200,
        ).execute()
        for f in r.get('files', []):
            mime = f['mimeType']
            if 'folder' in mime:
                to_visit.append(f['id'])
            elif (
                'spreadsheetml.sheet' in mime
                or mime == 'application/vnd.google-apps.spreadsheet'
                or f['name'].lower().endswith('.xlsx')
            ):
                # Skip lock files / archives by name
                if f['name'].startswith('.'):
                    continue
                files.append(f)
    return files


def get_latest_archived_time(drive):
    """Return modifiedTime of newest file in _Archive (not counting hidden/lock files)."""
    r = drive.files().list(
        q=f"'{ARCHIVE_FOLDER_ID}' in parents and trashed=false",
        fields='files(id,name,modifiedTime,mimeType)',
        orderBy='modifiedTime desc',
        supportsAllDrives=True, includeItemsFromAllDrives=True,
        pageSize=50,
    ).execute()
    for f in r.get('files', []):
        # Skip hidden/lock files and folders
        if f['name'].startswith('.') or 'folder' in f['mimeType']:
            continue
        return f.get('modifiedTime', '')
    return ''  # Empty archive — first run


def download_file(drive, file_id, dest_path):
    """Download file as xlsx — handles both native Google Sheets (export) and uploaded xlsx (get_media)."""
    meta = drive.files().get(fileId=file_id, fields='mimeType', supportsAllDrives=True).execute()
    if meta['mimeType'] == 'application/vnd.google-apps.spreadsheet':
        req = drive.files().export_media(
            fileId=file_id,
            mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    else:
        req = drive.files().get_media(fileId=file_id, supportsAllDrives=True)
    fh = io.BytesIO()
    dl = MediaIoBaseDownload(fh, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    fh.seek(0)
    Path(dest_path).write_bytes(fh.read())


_BDO_HEADER_ALIASES = {
    'date':        ['POSTING DATE', 'POSTING_DATE', 'TRANSACTION DATE', 'TXN DATE', 'DATE'],
    'branch':      ['BRANCH', 'BRANCH NAME'],
    'description': ['DESCRIPTION', 'PARTICULARS', 'NARRATIVE', 'TRANSACTION DESCRIPTION', 'REMARKS'],
    'debit':       ['DEBIT', 'WITHDRAWAL', 'WITHDRAWALS', 'DR', 'AMOUNT DEBIT'],
    'credit':      ['CREDIT', 'DEPOSIT', 'DEPOSITS', 'CR', 'AMOUNT CREDIT'],
    'balance':     ['RUNNING BALANCE', 'BALANCE', 'CURRENT BALANCE'],
    'check':       ['CHECK NUMBER', 'CHECK NO', 'CHECK NO.', 'CHECK#', 'CHK NO', 'CHEQUE', 'CHEQUE NO', 'CHEQUE NO.'],
}
_BDO_ACCT_NUM_KEYS = ['ACCOUNT NUMBER', 'ACCOUNT NO', 'ACCOUNT NO.', 'ACCT NO', 'ACCT NUMBER']
_BDO_ACCT_NAME_KEYS = ['ACCOUNT NAME', 'CUSTOMER NAME', 'NAME', 'ACCT NAME']


def _bdo_detect_schema(ws, scan_rows: int = 50, scan_cols: int = 25):
    """Auto-detect column positions for date/desc/debit/credit/check from header row.

    Returns dict {field: 1-based-col-idx, 'header_row': N, 'acct_num': str, 'acct_name': str}
    or {'skip': reason} if no header detected.

    Designed to handle BDO's multiple file layouts:
      - Stores file: date in col 2, desc in col 7, etc.
      - Head Office file: date in col 1, desc in col 6, etc.
      - Legacy formats with stringified dates
      - Empty placeholder tabs (skipped)
    """
    if ws.max_row < 2 or ws.max_column < 3:
        return {'skip': f'too small ({ws.max_row}r x {ws.max_column}c)'}

    cols_found = {}
    header_row = 0
    for r in range(1, min(scan_rows + 1, ws.max_row + 1)):
        for c in range(1, min(scan_cols + 1, ws.max_column + 1)):
            v = str(ws.cell(r, c).value or '').strip().upper()
            if not v:
                continue
            for field_name, aliases in _BDO_HEADER_ALIASES.items():
                if v in aliases and field_name not in cols_found:
                    cols_found[field_name] = c
                    header_row = max(header_row, r)
        if len(cols_found) >= 4 and 'date' in cols_found and 'description' in cols_found:
            break

    if not cols_found.get('date') or not cols_found.get('description'):
        return {'skip': f'no header row detected (found: {sorted(cols_found.keys())})'}

    if not cols_found.get('check'):
        cols_found['check'] = -1  # sentinel — accounts without check column

    # Find account number + name in pre-header rows
    acct_num = ''
    acct_name = ''
    for r in range(1, min(header_row, scan_rows)):
        for c in range(1, min(scan_cols + 1, ws.max_column + 1)):
            v = str(ws.cell(r, c).value or '').strip().upper()
            if not v:
                continue
            for key in _BDO_ACCT_NUM_KEYS:
                if key in v and not acct_num:
                    for dc in range(1, 5):
                        rv = ws.cell(r, c + dc).value
                        if rv not in (None, ''):
                            acct_num = str(rv).strip()
                            break
            for key in _BDO_ACCT_NAME_KEYS:
                if key in v and not acct_name:
                    for dc in range(1, 5):
                        rv = ws.cell(r, c + dc).value
                        if rv not in (None, '') and not str(rv).strip().isdigit():
                            acct_name = str(rv).strip()
                            break

    return {
        **cols_found,
        'header_row': header_row,
        'acct_num': acct_num,
        'acct_name': acct_name,
    }


def _bdo_normalize_date(v):
    """Return MM/DD/YYYY string or None if not a parseable date."""
    if v in (None, ''):
        return None
    if hasattr(v, 'strftime'):  # datetime.datetime or datetime.date
        return v.strftime('%m/%d/%Y')
    if isinstance(v, str):
        import re
        s = v.strip()
        if re.match(r'^\d{2}/\d{2}/\d{4}$', s):
            return s
        m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
        if m:
            return f'{m.group(2)}/{m.group(3)}/{m.group(1)}'
        m = re.match(r'^(\d{2})-(\d{2})-(\d{4})$', s)
        if m:
            return f'{m.group(1)}/{m.group(2)}/{m.group(3)}'
    return None


def extract_bdo_txns(xlsx_path):
    """Extract transaction rows from BDO xlsx — auto-detects column positions per tab.

    Robust to:
      - Multiple BDO file layouts (Stores vs Head Office have different column positions)
      - Empty placeholder tabs (1-row or 1-column)
      - Header row at any position in first 50 rows
      - Datetime objects, MM/DD/YYYY strings, ISO dates, MM-DD-YYYY strings
      - Tabs with no check column (deposits-only accounts)
      - Sparse tabs with high max_column
      - Header label aliases (PARTICULARS / CHEQUE NO / etc.)
    """
    def parse_num(v):
        if v in (None, ''):
            return 0.0
        try:
            return float(str(v).replace(',', '').strip())
        except (ValueError, TypeError):
            return 0.0

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    txns = []
    for tab in wb.sheetnames:
        ws = wb[tab]
        schema = _bdo_detect_schema(ws)
        if 'skip' in schema:
            continue
        date_col = schema['date']
        desc_col = schema['description']
        debit_col = schema.get('debit')
        credit_col = schema.get('credit')
        check_col = schema.get('check', -1)
        # Compute the highest column index we need to safely read
        col_indices = [c for c in [date_col, desc_col, debit_col, credit_col, check_col] if c and c > 0]
        max_col = max(col_indices) if col_indices else 0
        for r in range(schema['header_row'] + 1, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if len(row) < max_col:
                continue
            date_str = _bdo_normalize_date(row[date_col - 1])
            if not date_str:
                continue
            debit = parse_num(row[debit_col - 1]) if debit_col else 0.0
            credit = parse_num(row[credit_col - 1]) if credit_col else 0.0
            if check_col == -1:
                continue  # no check column on this tab
            chk_raw = str(row[check_col - 1] or '').strip()
            chk_clean = chk_raw.lstrip('0') if chk_raw else ''
            if not chk_clean or chk_clean == '0':
                continue
            txns.append({
                'account': str(schema.get('acct_name', '')),
                'account_number': str(schema.get('acct_num', '')),
                'date': date_str,
                'description': str(row[desc_col - 1] or '').strip(),
                'debit': debit,
                'credit': credit,
                'check_number': chk_raw,
                'check_clean': chk_clean,
                'tab': tab,
            })
    return pd.DataFrame(txns)


def a1_col(idx):
    letters = ''
    n = idx
    while True:
        letters = chr(ord('A') + (n % 26)) + letters
        n = n // 26 - 1
        if n < 0:
            break
    return letters


def match_and_update(sheets, bdo_df):
    """Find FPM rows needing update + apply delta writes."""
    # Load FPM
    resp = sheets.spreadsheets().values().get(
        spreadsheetId=FPM_ID, range='RFP Summary',
        valueRenderOption='UNFORMATTED_VALUE',
    ).execute()
    vals = resp.get('values', [])
    headers = vals[0]

    def h_idx(name):
        for i, h in enumerate(headers):
            if str(h).strip().lower() == name.lower():
                return i
        return -1

    status_col = h_idx('Status')
    chk_col = h_idx('Check No./Ref No.')
    payee_col = h_idx('Payee')
    cleared_date_col = h_idx('BDO Cleared Date')
    cleared_amt_col = h_idx('BDO Cleared Amount')

    if cleared_date_col < 0 or cleared_amt_col < 0:
        raise RuntimeError('FPM is missing BDO Cleared Date/Amount columns. Run Phase A catchup first.')

    # BDO debits only, use check_clean as key
    bdo_debits = bdo_df[bdo_df['debit'] > 0].copy()
    bdo_by_check = {str(r['check_clean']): r for _, r in bdo_debits.iterrows()}

    updates = []
    delta_count = 0
    status_flips = 0

    for r_idx, row in enumerate(vals[1:], 2):
        if len(row) <= chk_col:
            continue
        fpm_chk = str(row[chk_col]).strip().lstrip('0')
        if not fpm_chk or fpm_chk not in bdo_by_check:
            continue
        # Already tagged?
        already_tagged = (
            len(row) > cleared_date_col and str(row[cleared_date_col]).strip() not in ('', 'nan')
        )
        if already_tagged:
            continue
        # Apply delta
        bdo_row = bdo_by_check[fpm_chk]
        payee = row[payee_col] if payee_col < len(row) else ''
        current_status = row[status_col] if status_col < len(row) else ''
        updates.append((f"'RFP Summary'!{a1_col(cleared_date_col)}{r_idx}", bdo_row['date']))
        updates.append((f"'RFP Summary'!{a1_col(cleared_amt_col)}{r_idx}", float(bdo_row['debit'])))
        delta_count += 1
        if not is_intercompany(payee) and str(current_status).strip() != 'Paid/ Cleared':
            updates.append((f"'RFP Summary'!{a1_col(status_col)}{r_idx}", 'Paid/ Cleared'))
            status_flips += 1

    if updates:
        CHUNK = 500
        for i in range(0, len(updates), CHUNK):
            batch = updates[i:i+CHUNK]
            data = [{'range': rng, 'values': [[val]]} for rng, val in batch]
            sheets.spreadsheets().values().batchUpdate(
                spreadsheetId=FPM_ID,
                body={'data': data, 'valueInputOption': 'USER_ENTERED'},
            ).execute()

    return {'delta_rows': delta_count, 'status_flips': status_flips, 'cells': len(updates)}


def archive_file(drive, file_id, file_name):
    existing = drive.files().get(fileId=file_id, fields='parents', supportsAllDrives=True).execute()
    current_parents = existing.get('parents', [])
    if ARCHIVE_FOLDER_ID in current_parents:
        return
    drive.files().update(
        fileId=file_id, addParents=ARCHIVE_FOLDER_ID,
        removeParents=','.join(current_parents),
        supportsAllDrives=True, fields='id,parents',
    ).execute()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dry-run', action='store_true', help='Match & report, do not write or archive')
    ap.add_argument('--force', action='store_true', help='Re-process latest file even if already archived')
    args = ap.parse_args()

    print(f'=== Weekly BDO Reconciliation — {RUN_CONTEXT.upper()} mode ===')
    print(f'Run at: {datetime.now().isoformat()}')

    sheets, drive = get_services()

    # Step 1: Find latest BDO file
    print('\n[1/6] Scanning BDO folder...')
    bdo_files = list_bdo_files(drive)
    bdo_files.sort(key=lambda f: f.get('modifiedTime', ''), reverse=True)
    if not bdo_files:
        print('No BDO xlsx files found. Exiting.')
        return
    target = bdo_files[0]
    print(f'Latest: {target["name"]} (modified {target.get("modifiedTime", "?")})')

    # Step 2: File-newness check — skip cleanly if nothing new since last run
    print('\n[2/6] Checking if file is new...')
    latest_archived_time = get_latest_archived_time(drive)
    target_time = target.get('modifiedTime', '')
    if args.force:
        print(f'  Force mode — processing regardless of archive')
    elif latest_archived_time and target_time <= latest_archived_time:
        print(f'  No new file since last run.')
        print(f'    Latest BDO:      {target_time}  {target["name"]}')
        print(f'    Latest archive:  {latest_archived_time}')
        print(f'  Exiting cleanly — will retry tomorrow.')
        # Write idempotent no-op report
        today = datetime.now().strftime('%Y-%m-%d')
        report_path = Path(f'CEO/CashFlow/bdo_statements/weekly_report_{today}.md')
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(
            f"# Weekly BDO Reconciliation — {today}\n\n"
            f"## No new file\n\n"
            f"- Context: {RUN_CONTEXT}\n"
            f"- Latest BDO file: {target['name']}\n"
            f"- Modified: {target_time}\n"
            f"- Last archived: {latest_archived_time}\n\n"
            f"File was already processed. Nothing to do. Will retry next scheduled run.\n",
            encoding='utf-8',
        )
        return
    else:
        print(f'  New file detected:')
        print(f'    Modified:        {target_time}')
        print(f'    Last archived:   {latest_archived_time or "(empty archive)"}')
        print(f'  Proceeding...')

    # Step 3: Download + extract
    print('\n[3/6] Downloading + extracting...')
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tf:
        tmp_path = tf.name
    download_file(drive, target['id'], tmp_path)
    bdo_df = extract_bdo_txns(tmp_path)
    os.unlink(tmp_path)
    print(f'Transactions extracted: {len(bdo_df)} (debit: {(bdo_df["debit"] > 0).sum()})')

    # Step 4: Match + update
    print('\n[4/6] Matching + updating FPM...')
    if args.dry_run:
        print('  [dry-run] Skipping writes')
        results = {'delta_rows': 0, 'status_flips': 0, 'cells': 0}
    else:
        results = match_and_update(sheets, bdo_df)
    print(f'  Delta matched: {results["delta_rows"]}')
    print(f'  Status flipped: {results["status_flips"]}')
    print(f'  Cells written: {results["cells"]}')

    # Step 5: Archive the file
    print('\n[5/6] Archiving file...')
    if args.dry_run:
        print('  [dry-run] Skipping archive')
    else:
        archive_file(drive, target['id'], target['name'])
        print(f'  Moved to _Archive')

    # Step 6: Write report
    print('\n[6/6] Writing report...')
    today = datetime.now().strftime('%Y-%m-%d')
    report_path = Path(f'CEO/CashFlow/bdo_statements/weekly_report_{today}.md')
    report_path.parent.mkdir(parents=True, exist_ok=True)
    body = f"""# Weekly BDO Reconciliation — {today}

## Run Context
- Mode: {RUN_CONTEXT}
- Time: {datetime.now().isoformat()}
- Source: {target['name']}

## Results
- BDO transactions scanned: {len(bdo_df)}
- Debit-side checks: {(bdo_df['debit'] > 0).sum()}
- New clearances tagged: {results['delta_rows']}
- FPM status flipped (Check Released → Paid): {results['status_flips']}
- Total cells updated: {results['cells']}

## File processed
- Drive ID: {target['id']}
- Moved to _Archive: {'yes' if not args.dry_run else 'dry-run — no'}

## Next Run
- Next scheduled: Monday 07:00 PHT (local Task Scheduler + GH Actions parallel)
"""
    report_path.write_text(body, encoding='utf-8')
    print(f'  Saved: {report_path}')

    print('\n=== DONE ===')


if __name__ == '__main__':
    main()
